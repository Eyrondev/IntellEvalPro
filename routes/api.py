"""
API routes blueprint for IntellEvalPro
Handles all API endpoints that return JSON data
"""
from flask import Blueprint, request, session, current_app
from models import Faculty, Student, Evaluation, get_db_connection
from utils.json_encoder import jsonify
from utils import login_required
from datetime import datetime

# Create blueprint
api_bp = Blueprint('api', __name__, url_prefix='/api')


@api_bp.route('/session-status')
@login_required
def session_status():
    """
    Check session status and return remaining time
    Used for auto-logout warning and session refresh
    """
    try:
        last_activity = session.get('last_activity')
        if not last_activity:
            return jsonify({
                'success': False,
                'message': 'No session activity found',
                'logged_in': False
            }), 401
        
        last_activity_time = datetime.fromisoformat(last_activity)
        current_time = datetime.now()
        time_elapsed = current_time - last_activity_time
        
        # Calculate remaining time (1 hour = 3600 seconds)
        remaining_seconds = 3600 - int(time_elapsed.total_seconds())
        
        if remaining_seconds <= 0:
            return jsonify({
                'success': False,
                'message': 'Session expired',
                'logged_in': False,
                'remaining_seconds': 0
            }), 401
        
        # Update last activity (this call itself counts as activity)
        session['last_activity'] = current_time.isoformat()
        
        return jsonify({
            'success': True,
            'logged_in': True,
            'remaining_seconds': remaining_seconds,
            'user': {
                'username': session.get('username'),
                'role': session.get('role'),
                'first_name': session.get('first_name'),
                'last_name': session.get('last_name')
            }
        })
        
    except Exception as e:
        print(f"Error checking session status: {e}")
        return jsonify({
            'success': False,
            'message': 'Error checking session',
            'error': str(e)
        }), 500


@api_bp.route('/refresh-session', methods=['POST'])
@login_required
def refresh_session():
    """
    Refresh session activity timestamp
    Called by client-side JavaScript to keep session alive during user activity
    """
    try:
        # Update last activity time
        session['last_activity'] = datetime.now().isoformat()
        
        return jsonify({
            'success': True,
            'message': 'Session refreshed',
            'timestamp': session['last_activity']
        })
        
    except Exception as e:
        print(f"Error refreshing session: {e}")
        return jsonify({
            'success': False,
            'message': 'Error refreshing session',
            'error': str(e)
        }), 500


@api_bp.route('/evaluation/start', methods=['POST'])
@login_required
def start_evaluation_timer():
    """
    Start evaluation timer for a student
    Returns session information and time limit
    """
    try:
        data = request.get_json()
        evaluation_id = data.get('evaluation_id')
        
        if not evaluation_id:
            return jsonify({
                'success': False,
                'error': 'Evaluation ID is required'
            }), 400
        
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({
                'success': False,
                'error': 'User not logged in'
            }), 401
        
        conn = get_db_connection()
        if not conn:
            return jsonify({
                'success': False,
                'error': 'Database connection failed'
            }), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Check if evaluation exists and belongs to user
        cursor.execute("""
            SELECT e.*, ep.time_limit_minutes
            FROM evaluations e
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE e.evaluation_id = %s AND e.student_id = (
                SELECT id FROM std_info WHERE user_id = %s
            )
        """, (evaluation_id, user_id))
        
        evaluation = cursor.fetchone()
        
        if not evaluation:
            cursor.close()
            conn.close()
            return jsonify({
                'success': False,
                'error': 'Evaluation not found or access denied'
            }), 404
        
        # Check if evaluation is already completed
        if evaluation['status'] == 'Completed':
            cursor.close()
            conn.close()
            return jsonify({
                'success': False,
                'error': 'This evaluation is already completed'
            }), 400
        
        # Get timer settings from timer_settings table
        cursor.execute("""
            SELECT enabled, time_limit
            FROM timer_settings
            WHERE setting_id = 1
            LIMIT 1
        """)
        
        timer_config = cursor.fetchone()
        
        # Determine time limit and timer state
        timer_enabled = timer_config and timer_config.get('enabled')
        if timer_enabled:
            # Use configured time limit from timer_settings
            time_limit = timer_config.get('time_limit', 30)
        else:
            # Timer disabled - set to None (no timer)
            time_limit = None
        
        # Check for existing timer session
        cursor.execute("""
            SELECT * FROM evaluation_timer_sessions
            WHERE evaluation_id = %s AND user_id = %s
            ORDER BY start_time DESC
            LIMIT 1
        """, (evaluation_id, user_id))
        
        existing_session = cursor.fetchone()
        
        # If timer is disabled, skip timer logic completely
        if not timer_enabled or time_limit is None:
            cursor.close()
            conn.close()
            return jsonify({
                'success': True,
                'timer_enabled': False,
                'session_id': None,
                'message': 'Timer is disabled. You have unlimited time to complete this evaluation.'
            })
        
        if existing_session:
            # Check if session has expired
            start_time = existing_session['start_time']
            elapsed = (datetime.now() - start_time).total_seconds()
            remaining = (time_limit * 60) - elapsed
            
            if remaining <= 0:
                # Session expired
                cursor.close()
                conn.close()
                return jsonify({
                    'success': False,
                    'expired': True,
                    'error': 'Evaluation time has expired'
                }), 400
            
            # Resume existing session
            cursor.close()
            conn.close()
            return jsonify({
                'success': True,
                'resumed': True,
                'session_id': existing_session['session_id'],
                'start_time': start_time.isoformat(),
                'time_limit': time_limit,
                'remaining_seconds': int(remaining)
            })
        
        # Create new timer session
        cursor.execute("""
            INSERT INTO evaluation_timer_sessions
            (evaluation_id, user_id, start_time, time_limit_minutes)
            VALUES (%s, %s, NOW(), %s)
        """, (evaluation_id, user_id, time_limit))
        
        session_id = cursor.lastrowid
        conn.commit()
        
        # Get the start time
        cursor.execute("""
            SELECT start_time FROM evaluation_timer_sessions
            WHERE session_id = %s
        """, (session_id,))
        
        start_time = cursor.fetchone()['start_time']
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'resumed': False,
            'session_id': session_id,
            'start_time': start_time.isoformat(),
            'time_limit': time_limit,
            'remaining_seconds': time_limit * 60
        })
        
    except Exception as e:
        print(f"Error starting evaluation timer: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/evaluation/check-time/<int:session_id>', methods=['GET'])
@login_required
def check_evaluation_time(session_id):
    """
    Check remaining time for an evaluation session
    Used for server-side validation
    """
    try:
        user_id = session.get('user_id')
        
        conn = get_db_connection()
        if not conn:
            return jsonify({
                'success': False,
                'error': 'Database connection failed'
            }), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Get session info
        cursor.execute("""
            SELECT * FROM evaluation_timer_sessions
            WHERE session_id = %s AND user_id = %s
        """, (session_id, user_id))
        
        timer_session = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not timer_session:
            return jsonify({
                'success': False,
                'error': 'Session not found'
            }), 404
        
        # Calculate remaining time
        start_time = timer_session['start_time']
        time_limit = timer_session['time_limit_minutes']
        
        elapsed = (datetime.now() - start_time).total_seconds()
        remaining = (time_limit * 60) - elapsed
        
        if remaining <= 0:
            return jsonify({
                'success': True,
                'status': 'expired',
                'elapsed_seconds': int(elapsed),
                'remaining_seconds': 0
            })
        
        return jsonify({
            'success': True,
            'status': 'active',
            'elapsed_seconds': int(elapsed),
            'remaining_seconds': int(remaining)
        })
        
    except Exception as e:
        print(f"Error checking evaluation time: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/evaluation/mark-expired', methods=['POST'])
@login_required
def mark_evaluation_expired():
    """
    Mark evaluation as expired (timer ran out)
    Does NOT submit responses - they are discarded
    """
    try:
        user_id = session.get('user_id')
        data = request.get_json()
        session_id = data.get('session_id')
        
        if not session_id:
            return jsonify({
                'success': False,
                'error': 'Session ID is required'
            }), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({
                'success': False,
                'error': 'Database connection failed'
            }), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Get the timer session
        cursor.execute("""
            SELECT ets.evaluation_id, e.status
            FROM evaluation_timer_sessions ets
            JOIN evaluations e ON ets.evaluation_id = e.evaluation_id
            WHERE ets.session_id = %s AND ets.user_id = %s
        """, (session_id, user_id))
        
        session_info = cursor.fetchone()
        
        if not session_info:
            cursor.close()
            conn.close()
            return jsonify({
                'success': False,
                'error': 'Session not found'
            }), 404
        
        evaluation_id = session_info['evaluation_id']
        
        # Mark timer session as completed (expired)
        cursor.execute("""
            UPDATE evaluation_timer_sessions
            SET completed_at = NOW()
            WHERE session_id = %s
        """, (session_id,))
        
        # Update evaluation status to Expired (no responses saved)
        cursor.execute("""
            UPDATE evaluations
            SET status = 'Expired'
            WHERE evaluation_id = %s
        """, (evaluation_id,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation marked as expired'
        })
        
    except Exception as e:
        print(f"Error marking evaluation as expired: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/ai-support-chat', methods=['POST'])
@login_required
def ai_support_chat():
    """
    AI Support Chatbot endpoint
    Uses Gemini AI to provide intelligent help and support
    """
    try:
        from utils.ai_support import get_ai_response, is_system_related_question
        
        data = request.get_json()
        user_message = data.get('message', '').strip()
        
        if not user_message:
            return jsonify({
                'success': False,
                'error': 'Message is required'
            }), 400
        
        # Quick filter for obviously off-topic questions
        if not is_system_related_question(user_message):
            return jsonify({
                'success': True,
                'reply': "I'm the IntellEvalPro AI Assistant and I can only help with questions about this evaluation system. Please ask me about system features, navigation, evaluations, or troubleshooting. 😊"
            })
        
        # Get AI response
        success, response = get_ai_response(user_message)
        
        if success:
            return jsonify({
                'success': True,
                'reply': response
            })
        else:
            return jsonify({
                'success': False,
                'reply': response
            })
            
    except ImportError:
        # AI support not available (Gemini API not installed)
        return jsonify({
            'success': True,
            'reply': "The AI Assistant feature is currently unavailable. Please refer to the FAQ sections above or contact your administrator for assistance."
        })
    except Exception as e:
        print(f"Error in AI support chat: {e}")
        return jsonify({
            'success': False,
            'error': 'An error occurred while processing your request'
        }), 500


@api_bp.route('/faculty/next-number')
@login_required
def get_next_faculty_number():
    """Generate the next faculty number in format NC-FAC-XX"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get the latest faculty number with NC-FAC- pattern
        cursor.execute("""
            SELECT faculty_number 
            FROM faculty 
            WHERE faculty_number LIKE 'NC-FAC-%' 
            ORDER BY faculty_number DESC 
            LIMIT 1
        """)
        result = cursor.fetchone()
        
        if result:
            # Extract the number part and increment
            last_number = result['faculty_number']
            # Extract number after NC-FAC-
            try:
                num_part = last_number.split('NC-FAC-')[1]
                next_num = int(num_part) + 1
            except (IndexError, ValueError):
                # If parsing fails, start from 1
                next_num = 1
        else:
            # No faculty with NC-FAC- pattern, start from 1
            next_num = 1
        
        # Format with leading zeros (minimum 2 digits)
        next_faculty_number = f"NC-FAC-{next_num:02d}"
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'next_number': next_faculty_number
        })
        
    except Exception as e:
        print(f"Error generating next faculty number: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/faculty')
@login_required
def get_faculty():
    """Get all faculty members with program information"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if we should include archived faculty (default: exclude)
        include_archived = request.args.get('include_archived', 'false').lower() == 'true'
        
        # Get all faculty with program info, class counts, and average ratings
        archive_filter = "" if include_archived else "WHERE f.is_archived = FALSE"
        
        query = f"""
            SELECT f.*, 
                   p.name as department_name,
                   p.program_code,
                   COUNT(DISTINCT cs.section_id) as total_classes,
                   ROUND(AVG(er.rating), 2) as avg_rating,
                   COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                   GROUP_CONCAT(DISTINCT s.subject_code ORDER BY s.subject_code SEPARATOR ',') as workload_codes
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN subjects s ON cs.subject_id = s.subject_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id AND e.status = 'Completed'
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            {archive_filter}
            GROUP BY f.faculty_id
            ORDER BY f.last_name, f.first_name
        """
        cursor.execute(query)
        faculty_list = cursor.fetchall()
        
        # Get statistics (only for non-archived faculty)
        stats_query = """
            SELECT 
                COUNT(DISTINCT f.faculty_id) as total_faculty,
                COUNT(DISTINCT cs.section_id) as total_classes,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                ROUND(AVG(er.rating), 2) as avg_rating
            FROM faculty f
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id AND e.status = 'Completed'
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            WHERE f.is_archived = FALSE
        """
        cursor.execute(stats_query)
        stats = cursor.fetchone()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': faculty_list,
            'stats': {
                'total_faculty': stats['total_faculty'] or 0,
                'total_classes': stats['total_classes'] or 0,
                'avg_rating': float(stats['avg_rating']) if stats['avg_rating'] else 0,
                'total_evaluations': stats['total_evaluations'] or 0
            }
        })
        
    except Exception as e:
        print(f"Error getting faculty: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/faculty', methods=['POST'])
@login_required
def create_faculty():
    """Create a new faculty member"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Validate required fields
        if not data.get('first_name') or not data.get('last_name') or not data.get('faculty_number'):
            return jsonify({'success': False, 'error': 'First name, last name, and faculty number are required'}), 400
        
        # Check if faculty number already exists
        cursor.execute("SELECT faculty_id FROM faculty WHERE faculty_number = %s", (data['faculty_number'],))
        if cursor.fetchone():
            return jsonify({'success': False, 'error': 'Faculty number already exists'}), 400
        
        # Get program_id from department_name if provided
        program_id = None
        if data.get('department_name'):
            cursor.execute("SELECT program_id FROM programs WHERE name = %s", (data['department_name'],))
            program = cursor.fetchone()
            if program:
                program_id = program['program_id']
        
        # Insert new faculty
        insert_query = """
            INSERT INTO faculty 
            (faculty_number, first_name, last_name, email, program_id, status, specialization)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        
        insert_values = (
            data['faculty_number'],
            data['first_name'],
            data['last_name'],
            data.get('email', ''),
            program_id,
            data.get('status', 'Active'),
            data.get('specialization', '')
        )
        
        cursor.execute(insert_query, insert_values)
        conn.commit()
        
        faculty_id = cursor.lastrowid
        
        return jsonify({
            'success': True,
            'message': 'Faculty created successfully',
            'faculty_id': faculty_id
        }), 201
        
    except Exception as e:
        conn.rollback()
        print(f"Error creating faculty: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/faculty/<int:faculty_id>', methods=['PUT'])
@login_required
def update_faculty(faculty_id):
    """Update faculty details"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Check if faculty exists
        cursor.execute("SELECT faculty_id FROM faculty WHERE faculty_id = %s", (faculty_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Faculty not found'}), 404
        
        # Get program_id from department_name if provided
        program_id = None
        if 'department_name' in data and data['department_name']:
            cursor.execute("SELECT program_id FROM programs WHERE name = %s", (data['department_name'],))
            program = cursor.fetchone()
            if program:
                program_id = program['program_id']
        
        # Build update query dynamically
        update_fields = []
        update_values = []
        
        allowed_fields = {
            'first_name': 'first_name',
            'last_name': 'last_name',
            'faculty_number': 'faculty_number',
            'email': 'email',
            'specialization': 'specialization'
        }
        
        for field_key, db_column in allowed_fields.items():
            if field_key in data and data[field_key] is not None:
                update_fields.append(f"{db_column} = %s")
                update_values.append(data[field_key])
        
        # Add program_id if we found one
        if program_id:
            update_fields.append("program_id = %s")
            update_values.append(program_id)
        
        if not update_fields:
            return jsonify({'success': False, 'error': 'No fields to update'}), 400
        
        update_values.append(faculty_id)
        update_query = f"UPDATE faculty SET {', '.join(update_fields)} WHERE faculty_id = %s"
        
        cursor.execute(update_query, update_values)
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Faculty updated successfully'
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error updating faculty: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/faculty/<int:faculty_id>/archive', methods=['POST'])
@login_required
def archive_faculty(faculty_id):
    """Archive a faculty member"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if faculty exists
        cursor.execute("SELECT faculty_id, first_name, last_name, is_archived FROM faculty WHERE faculty_id = %s", (faculty_id,))
        faculty = cursor.fetchone()
        
        if not faculty:
            return jsonify({'success': False, 'error': 'Faculty not found'}), 404
        
        if faculty['is_archived']:
            return jsonify({'success': False, 'error': 'Faculty is already archived'}), 400
        
        # Archive the faculty member
        cursor.execute("UPDATE faculty SET is_archived = TRUE WHERE faculty_id = %s", (faculty_id,))
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': f"{faculty['first_name']} {faculty['last_name']} has been archived successfully"
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error archiving faculty: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/faculty/<int:faculty_id>/unarchive', methods=['POST'])
@login_required
def unarchive_faculty(faculty_id):
    """Unarchive a faculty member"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if faculty exists
        cursor.execute("SELECT faculty_id, first_name, last_name, is_archived FROM faculty WHERE faculty_id = %s", (faculty_id,))
        faculty = cursor.fetchone()
        
        if not faculty:
            return jsonify({'success': False, 'error': 'Faculty not found'}), 404
        
        if not faculty['is_archived']:
            return jsonify({'success': False, 'error': 'Faculty is not archived'}), 400
        
        # Unarchive the faculty member
        cursor.execute("UPDATE faculty SET is_archived = FALSE WHERE faculty_id = %s", (faculty_id,))
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': f"{faculty['first_name']} {faculty['last_name']} has been restored successfully"
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error unarchiving faculty: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/students')
@login_required
def get_students():
    """Get all students with user information"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if we should include archived students (default: exclude)
        include_archived = request.args.get('include_archived', 'false').lower() == 'true'
        
        # Check if we should exclude students already assigned to sections
        exclude_assigned = request.args.get('exclude_assigned', 'false').lower() == 'true'
        
        # Build archive filter
        archive_filter = "" if include_archived else "WHERE s.is_archived = FALSE"
        
        # Build section assignment filter
        section_filter = ""
        if exclude_assigned:
            section_filter = """
                AND s.id NOT IN (
                    SELECT student_id 
                    FROM section_students 
                    WHERE status = 'Active'
                )
            """
        
        # Get all students with user information and current section
        query = f"""
            SELECT s.id,
                   s.std_Number as student_number,
                   s.std_Firstname as first_name,
                   s.std_Surname as last_name,
                   s.std_Middlename as middle_name,
                   s.std_Birthdate as birthdate,
                   s.std_Age as age,
                   s.std_Address as address,
                   s.std_Gender as gender,
                   s.std_EmailAdd as email,
                   s.std_ContactNum as contact_number,
                   s.std_Course as program,
                   p.program_id,
                   s.std_Level as year_level,
                   s.std_Status as status,
                   s.is_archived,
                   s.created_at as enrollment_date,
                   u.username,
                   u.is_active as user_active,
                   sec.section_id as current_section_id,
                   sec.section_code as current_section_code,
                   sec.section_name as current_section_name
            FROM std_info s
            LEFT JOIN users u ON s.user_id = u.user_id
            LEFT JOIN programs p ON s.std_Course COLLATE utf8mb4_unicode_ci = p.name COLLATE utf8mb4_unicode_ci
            LEFT JOIN section_students ss ON s.id = ss.student_id AND ss.status = 'Active'
            LEFT JOIN sections sec ON ss.section_id = sec.section_id
            {archive_filter}
            {section_filter}
            ORDER BY s.std_Surname, s.std_Firstname
        """
        cursor.execute(query)
        student_list = cursor.fetchall()
        
        # Get statistics (only for non-archived students)
        stats_query = """
            SELECT 
                COUNT(DISTINCT s.id) as total_students,
                COUNT(DISTINCT CASE WHEN s.std_Status = 'Enrolled' THEN s.id END) as active_students,
                COUNT(DISTINCT s.std_Course) as total_programs,
                COUNT(DISTINCT CASE WHEN s.user_id IS NOT NULL THEN s.id END) as students_with_accounts
            FROM std_info s
            WHERE s.is_archived = FALSE
        """
        cursor.execute(stats_query)
        stats = cursor.fetchone()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': student_list,
            'stats': {
                'total_students': stats['total_students'] or 0,
                'active_students': stats['active_students'] or 0,
                'total_programs': stats['total_programs'] or 0,
                'students_with_accounts': stats['students_with_accounts'] or 0
            }
        })
        
    except Exception as e:
        print(f"Error getting students: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/students', methods=['POST'])
@login_required
def create_student():
    """Create a new student"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Validate required fields
        required_fields = ['student_number', 'first_name', 'last_name', 
                         'gender', 'birthdate', 'age', 'address', 'program_id', 'year_level']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                'success': False, 
                'error': f'Missing required fields: {", ".join(missing_fields)}'
            }), 400
        
        # Check if student number already exists
        cursor.execute(
            "SELECT id FROM std_info WHERE std_Number = %s", 
            (data['student_number'],)
        )
        if cursor.fetchone():
            return jsonify({
                'success': False, 
                'error': f'Student number {data["student_number"]} already exists'
            }), 400
        
        # Get program name from program_id
        cursor.execute(
            "SELECT name FROM programs WHERE program_id = %s", 
            (data['program_id'],)
        )
        program = cursor.fetchone()
        if not program:
            return jsonify({
                'success': False, 
                'error': 'Invalid program ID'
            }), 400
        
        # Map frontend fields to database columns
        insert_query = """
            INSERT INTO std_info (
                std_Number, std_Surname, std_Firstname, std_Middlename,
                std_Birthdate, std_Age, std_Address, std_Gender,
                std_EmailAdd, std_ContactNum, std_Level, std_Course, std_Status
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        values = (
            data['student_number'],
            data['last_name'],
            data['first_name'],
            data.get('middle_name', ''),
            data['birthdate'],
            int(data['age']),
            data['address'],
            data['gender'],
            data.get('email', ''),
            data.get('contact_number', ''),
            data['year_level'],
            program['name'],
            data.get('status', 'Enrolled')
        )
        
        cursor.execute(insert_query, values)
        student_id = cursor.lastrowid
        
        # Automatically create user account for the student
        # Password format: surname + student_number (e.g., "Smith20241234")
        from utils.security import generate_password_hash
        from datetime import datetime
        
        username = data['student_number']  # Use student number as username
        password = data['last_name'] + data['student_number']  # surname + student_number
        hashed_password = generate_password_hash(password)
        email = data.get('email', f"{data['student_number']}@student.edu")
        
        # Check if user already exists
        cursor.execute("SELECT user_id FROM users WHERE username = %s", (username,))
        existing_user = cursor.fetchone()
        
        if not existing_user:
            # Create new user account
            now = datetime.now()
            user_insert_query = """
                INSERT INTO users (username, password, email, first_name, last_name, 
                                 role, is_active, is_verified, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            user_values = (
                username,
                hashed_password,
                email,
                data['first_name'],
                data['last_name'],
                'student',
                1,  # is_active
                1,  # is_verified
                now,
                now
            )
            cursor.execute(user_insert_query, user_values)
            user_id = cursor.lastrowid
            
            # Update std_info with user_id to link student record to user account
            cursor.execute("UPDATE std_info SET user_id = %s WHERE id = %s", (user_id, student_id))
            
            # Log activity
            log_activity(
                user_id=session.get('user_id'),
                user_name=session.get('username'),
                user_role=session.get('role'),
                activity_type='create',
                description=f"Created student and user account: {username}",
                ip_address=request.remote_addr,
                target_user=username
            )
        
        # Assign to section only if explicitly selected by admin
        section_id = data.get('section_id')  # Get explicit section selection from form
        
        if section_id:
            # Admin explicitly selected a section - assign student to it
            cursor.execute("""
                INSERT INTO section_students (section_id, student_id, status, assigned_date)
                VALUES (%s, %s, 'Active', NOW())
            """, (section_id, student_id))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Student created successfully with user account',
            'student_id': student_id,
            'username': username,
            'default_password': password  # Send back for admin to inform student
        })
        
    except Exception as e:
        print(f"Error creating student: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/students/<int:student_id>', methods=['PUT'])
@login_required
def update_student(student_id):
    """Update an existing student"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Check if student exists
        cursor.execute("SELECT id FROM std_info WHERE id = %s", (student_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Student not found'}), 404
        
        # Check if student number is being changed and if it already exists
        if data.get('student_number'):
            cursor.execute(
                "SELECT id FROM std_info WHERE std_Number = %s AND id != %s", 
                (data['student_number'], student_id)
            )
            if cursor.fetchone():
                return jsonify({
                    'success': False, 
                    'error': f'Student number {data["student_number"]} already exists'
                }), 400
            
            # Also check if the student number (username) is already taken in users table
            cursor.execute(
                """SELECT user_id FROM users WHERE username = %s 
                   AND user_id != (SELECT user_id FROM std_info WHERE id = %s)""", 
                (data['student_number'], student_id)
            )
            if cursor.fetchone():
                return jsonify({
                    'success': False, 
                    'error': f'Username {data["student_number"]} is already taken by another user'
                }), 400
        
        # Get program name from program_id if provided
        program_name = None
        if data.get('program_id'):
            cursor.execute(
                "SELECT name FROM programs WHERE program_id = %s", 
                (data['program_id'],)
            )
            program = cursor.fetchone()
            if not program:
                return jsonify({'success': False, 'error': 'Invalid program ID'}), 400
            program_name = program['name']
        
        # Build update query dynamically for provided fields
        update_fields = []
        values = []
        
        field_mapping = {
            'student_number': 'std_Number',
            'last_name': 'std_Surname',
            'first_name': 'std_Firstname',
            'middle_name': 'std_Middlename',
            'birthdate': 'std_Birthdate',
            'age': 'std_Age',
            'address': 'std_Address',
            'gender': 'std_Gender',
            'email': 'std_EmailAdd',
            'contact_number': 'std_ContactNum',
            'year_level': 'std_Level',
            'status': 'std_Status'
        }
        
        for frontend_field, db_field in field_mapping.items():
            if frontend_field in data:
                update_fields.append(f"{db_field} = %s")
                values.append(data[frontend_field])
        
        # Add program if provided
        if program_name:
            update_fields.append("std_Course = %s")
            values.append(program_name)
        
        if not update_fields:
            return jsonify({'success': False, 'error': 'No fields to update'}), 400
        
        # Add student_id to values for WHERE clause
        values.append(student_id)
        
        update_query = f"""
            UPDATE std_info 
            SET {', '.join(update_fields)}
            WHERE id = %s
        """
        
        cursor.execute(update_query, values)
        
        # Update corresponding user information if the student has a linked user account
        cursor.execute("SELECT user_id FROM std_info WHERE id = %s", (student_id,))
        student_record = cursor.fetchone()
        
        if student_record and student_record['user_id']:
            user_id = student_record['user_id']
            
            # Build update query for users table
            user_update_fields = []
            user_values = []
            
            # Map student fields to user fields
            user_field_mapping = {
                'student_number': 'username',  # Student number maps to username
                'first_name': 'first_name',
                'last_name': 'last_name',
                'email': 'email'
            }
            
            for student_field, user_field in user_field_mapping.items():
                if student_field in data:
                    user_update_fields.append(f"{user_field} = %s")
                    user_values.append(data[student_field])
            
            # Always update the updated_at timestamp
            user_update_fields.append("updated_at = NOW()")
            
            if len(user_values) > 0:  # Only update if there are fields to update
                user_values.append(user_id)
                
                user_update_query = f"""
                    UPDATE users 
                    SET {', '.join(user_update_fields)}
                    WHERE user_id = %s
                """
                
                cursor.execute(user_update_query, user_values)
        
        # Handle section assignment only if explicitly provided
        section_id = data.get('section_id')
        
        if section_id:
            # Admin explicitly selected a section - update section assignment
            # First deactivate all current assignments
            cursor.execute(
                "UPDATE section_students SET status = 'Inactive' WHERE student_id = %s",
                (student_id,)
            )
            
            # Check if assignment to this section already exists
            cursor.execute("""
                SELECT id FROM section_students 
                WHERE section_id = %s AND student_id = %s
            """, (section_id, student_id))
            
            existing = cursor.fetchone()
            if existing:
                # Reactivate existing assignment
                cursor.execute("""
                    UPDATE section_students 
                    SET status = 'Active', assigned_date = NOW()
                    WHERE id = %s
                """, (existing['id'],))
            else:
                # Create new assignment
                cursor.execute("""
                    INSERT INTO section_students (section_id, student_id, status, assigned_date)
                    VALUES (%s, %s, 'Active', NOW())
                """, (section_id, student_id))
        
        conn.commit()
        cursor.close()
        
        # Determine message based on whether user was also updated
        message = 'Student updated successfully'
        if student_record and student_record['user_id']:
            message = 'Student and user account updated successfully'
        
        return jsonify({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        print(f"Error updating student: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/students/<int:student_id>/archive', methods=['PUT'])
@login_required
def archive_student(student_id):
    """Archive a student"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if student exists
        cursor.execute("SELECT id FROM std_info WHERE id = %s", (student_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Student not found'}), 404
        
        # Archive the student
        cursor.execute(
            "UPDATE std_info SET is_archived = TRUE WHERE id = %s",
            (student_id,)
        )
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Student archived successfully'
        })
        
    except Exception as e:
        print(f"Error archiving student: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/students/<int:student_id>/unarchive', methods=['PUT'])
@login_required
def unarchive_student(student_id):
    """Unarchive a student"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if student exists
        cursor.execute("SELECT id FROM std_info WHERE id = %s", (student_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Student not found'}), 404
        
        # Unarchive the student
        cursor.execute(
            "UPDATE std_info SET is_archived = FALSE WHERE id = %s",
            (student_id,)
        )
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Student restored successfully'
        })
        
    except Exception as e:
        print(f"Error unarchiving student: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/students/<int:student_id>/evaluations')
@login_required
def get_student_evaluations(student_id):
    """Get evaluation history for a specific student"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all evaluations for this student with related information
        query = """
            SELECT e.evaluation_id,
                   e.period_id,
                   e.section_id,
                   e.status,
                   e.completion_time,
                   e.start_time,
                   ep.title as period_title,
                   ep.start_date as period_start,
                   ep.end_date as period_end,
                   cs.section_name,
                   s.title as subject_name,
                   s.subject_code,
                   CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                   f.faculty_number
            FROM evaluations e
            INNER JOIN evaluation_periods ep ON e.period_id = ep.period_id
            INNER JOIN class_sections cs ON e.section_id = cs.section_id
            INNER JOIN subjects s ON cs.subject_id = s.subject_id
            INNER JOIN faculty f ON cs.faculty_id = f.faculty_id
            WHERE e.student_id = %s
            ORDER BY e.completion_time DESC, e.created_at DESC
        """
        cursor.execute(query, (student_id,))
        evaluations = cursor.fetchall()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'evaluations': evaluations
        })
        
    except Exception as e:
        print(f"Error getting student evaluations: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/evaluations/<int:evaluation_id>/reset', methods=['POST'])
@login_required
def reset_evaluation(evaluation_id):
    """Reset an evaluation to allow student to retake it"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # First, check if evaluation exists
        cursor.execute("SELECT * FROM evaluations WHERE evaluation_id = %s", (evaluation_id,))
        evaluation = cursor.fetchone()
        
        if not evaluation:
            return jsonify({'success': False, 'error': 'Evaluation not found'}), 404
        
        # Delete all evaluation responses for this evaluation
        cursor.execute("DELETE FROM evaluation_responses WHERE evaluation_id = %s", (evaluation_id,))
        
        # Reset the evaluation status to Pending
        update_query = """
            UPDATE evaluations 
            SET status = 'Pending',
                start_time = NULL,
                completion_time = NULL,
                updated_at = NOW()
            WHERE evaluation_id = %s
        """
        cursor.execute(update_query, (evaluation_id,))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation has been reset successfully'
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error resetting evaluation: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/departments')
def get_departments():
    """Get all programs (alias for backward compatibility - departments are now programs) - public access for signup"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all programs with subject counts
        query = """
            SELECT p.program_id,
                   p.program_id as department_id,
                   p.program_code,
                   p.name,
                   p.description,
                   COUNT(DISTINCT s.subject_id) as total_subjects,
                   COUNT(DISTINCT f.faculty_id) as total_faculty
            FROM programs p
            LEFT JOIN subjects s ON p.program_id = s.program_id
            LEFT JOIN faculty f ON p.program_id = f.program_id
            GROUP BY p.program_id
            ORDER BY p.program_code
        """
        cursor.execute(query)
        programs_list = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'data': programs_list
        })
        
    except Exception as e:
        print(f"Error fetching programs: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/programs')
def get_programs():
    """Alias for /api/departments - public access for signup"""
    return get_departments()


@api_bp.route('/subjects')
@login_required
def get_subjects():
    """Get all subjects with program information"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all subjects with program info and student count
        query = """
            SELECT s.subject_id,
                   s.subject_code,
                   s.title as subject_name,
                   s.description,
                   s.units,
                   p.name as department_name,
                   p.program_id as department_id,
                   p.program_code,
                   COUNT(DISTINCT cs.section_ref_id) as section_count,
                   COUNT(DISTINCT cs.faculty_id) as total_faculty,
                   COUNT(DISTINCT e.student_id) as student_count
            FROM subjects s
            LEFT JOIN programs p ON s.program_id = p.program_id
            LEFT JOIN class_sections cs ON s.subject_id = cs.subject_id
            LEFT JOIN enrollments e ON cs.section_id = e.section_id AND e.status = 'Enrolled'
            GROUP BY s.subject_id
            ORDER BY s.subject_code
        """
        cursor.execute(query)
        subjects_list = cursor.fetchall()
        
        # Get comprehensive statistics including total units
        stats_query = """
            SELECT 
                COUNT(DISTINCT s.subject_id) as total_subjects,
                COUNT(DISTINCT s.program_id) as total_departments,
                SUM(s.units) as total_units,
                COUNT(DISTINCT cs.section_id) as total_sections,
                COUNT(DISTINCT cs.faculty_id) as faculty_assigned,
                COUNT(DISTINCT e.enrollment_id) as total_enrollments
            FROM subjects s
            LEFT JOIN class_sections cs ON s.subject_id = cs.subject_id
            LEFT JOIN enrollments e ON cs.section_id = e.section_id
        """
        cursor.execute(stats_query)
        stats = cursor.fetchone()
        
        return jsonify({
            'success': True,
            'data': subjects_list,
            'stats': stats
        })
        
    except Exception as e:
        print(f"Error fetching subjects: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/subjects/<int:subject_id>/sections')
@login_required
def get_subject_sections(subject_id):
    """Get all sections for a specific subject with faculty and enrollment information"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get subject information first
        subject_query = """
            SELECT s.subject_id, s.subject_code, s.title as subject_name, s.units,
                   d.name as department_name, c.name as course_name
            FROM subjects s
            LEFT JOIN departments d ON s.department_id = d.department_id
            LEFT JOIN courses c ON d.course_id = c.course_id
            WHERE s.subject_id = %s
        """
        cursor.execute(subject_query, (subject_id,))
        subject = cursor.fetchone()
        
        if not subject:
            return jsonify({'success': False, 'error': 'Subject not found'}), 404
        
        # Get all sections for this subject with faculty and enrollment info
        sections_query = """
            SELECT cs.section_id,
                   cs.section_name,
                   cs.schedule,
                   cs.room,
                   f.faculty_id,
                   CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                   f.email as faculty_email,
                   CONCAT(ay.year_code, ' - ', at.term_name) as term_name,
                   at.acad_term_id,
                   COUNT(DISTINCT e.enrollment_id) as enrolled_students,
                   COUNT(DISTINCT ev.evaluation_id) as total_evaluations,
                   COUNT(DISTINCT CASE WHEN ev.status = 'Completed' THEN ev.evaluation_id END) as completed_evaluations
            FROM class_sections cs
            LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
            LEFT JOIN academic_terms at ON cs.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            LEFT JOIN enrollments e ON cs.section_id = e.section_id
            LEFT JOIN evaluations ev ON cs.section_id = ev.section_id
            WHERE cs.subject_id = %s
            GROUP BY cs.section_id
            ORDER BY cs.section_name
        """
        cursor.execute(sections_query, (subject_id,))
        sections = cursor.fetchall()
        
        # Calculate statistics
        total_sections = len(sections)
        total_enrolled = sum(section['enrolled_students'] for section in sections)
        
        return jsonify({
            'success': True,
            'subject': subject,
            'sections': sections,
            'stats': {
                'total_sections': total_sections,
                'total_enrolled': total_enrolled
            }
        })
        
    except Exception as e:
        print(f"Error fetching subject sections: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/subjects', methods=['POST'])
@login_required
def create_subject():
    """Create a new subject"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        
        # Validate required fields - accept both program_id and department_id for compatibility
        required_fields = ['subject_code', 'title', 'units']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        # Get program_id (accept either program_id or department_id)
        program_id = data.get('program_id') or data.get('department_id')
        if not program_id:
            return jsonify({'success': False, 'error': 'Missing required field: program_id'}), 400
        
        cursor = conn.cursor(dictionary=True)
        
        # Check if subject code already exists
        cursor.execute("SELECT subject_id FROM subjects WHERE subject_code = %s", (data['subject_code'],))
        if cursor.fetchone():
            return jsonify({'success': False, 'error': 'Subject code already exists'}), 400
        
        # Insert new subject
        insert_query = """
            INSERT INTO subjects (subject_code, title, description, units, program_id)
            VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(insert_query, (
            data['subject_code'],
            data['title'],
            data.get('description', ''),
            data['units'],
            program_id
        ))
        
        conn.commit()
        subject_id = cursor.lastrowid
        
        return jsonify({
            'success': True,
            'message': 'Subject created successfully',
            'subject_id': subject_id
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error creating subject: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/subjects/<int:subject_id>', methods=['PUT'])
@login_required
def update_subject(subject_id):
    """Update an existing subject"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Check if subject exists
        cursor.execute("SELECT subject_id FROM subjects WHERE subject_id = %s", (subject_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Subject not found'}), 404
        
        # Check if subject code is being changed and if it conflicts
        if 'subject_code' in data:
            cursor.execute(
                "SELECT subject_id FROM subjects WHERE subject_code = %s AND subject_id != %s",
                (data['subject_code'], subject_id)
            )
            if cursor.fetchone():
                return jsonify({'success': False, 'error': 'Subject code already exists'}), 400
        
        # Build update query dynamically
        update_fields = []
        update_values = []
        
        # Map department_id to program_id for backward compatibility
        allowed_fields = ['subject_code', 'title', 'description', 'units']
        for field in allowed_fields:
            if field in data:
                update_fields.append(f"{field} = %s")
                update_values.append(data[field])
        
        # Handle program_id (accept both program_id and department_id)
        program_id = data.get('program_id') or data.get('department_id')
        if program_id:
            update_fields.append("program_id = %s")
            update_values.append(program_id)
        
        if not update_fields:
            return jsonify({'success': False, 'error': 'No fields to update'}), 400
        
        update_values.append(subject_id)
        update_query = f"UPDATE subjects SET {', '.join(update_fields)} WHERE subject_id = %s"
        
        cursor.execute(update_query, update_values)
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Subject updated successfully'
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error updating subject: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/subjects/<int:subject_id>', methods=['DELETE'])
@login_required
def delete_subject(subject_id):
    """Delete a subject"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if subject exists
        cursor.execute("SELECT subject_id, subject_code, title FROM subjects WHERE subject_id = %s", (subject_id,))
        subject = cursor.fetchone()
        if not subject:
            return jsonify({'success': False, 'error': 'Subject not found'}), 404
        
        # Check if subject has any sections (optional - uncomment to prevent deletion)
        # cursor.execute("SELECT COUNT(*) as count FROM class_sections WHERE subject_id = %s", (subject_id,))
        # if cursor.fetchone()['count'] > 0:
        #     return jsonify({'success': False, 'error': 'Cannot delete subject with existing sections'}), 400
        
        # Delete the subject
        cursor.execute("DELETE FROM subjects WHERE subject_id = %s", (subject_id,))
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': f'Subject {subject["subject_code"]} deleted successfully'
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error deleting subject: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/subjects/<int:subject_id>/classes')
@login_required
def get_subject_classes(subject_id):
    """Get all class sections for a specific subject"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all classes for this subject
        cursor.execute("""
            SELECT 
                cs.section_id,
                cs.section_ref_id,
                cs.section_name,
                cs.schedule,
                cs.room,
                CONCAT(COALESCE(f.first_name, ''), ' ', COALESCE(f.last_name, '')) as faculty_name,
                CASE 
                    WHEN ay.year_code IS NOT NULL AND at.term_name IS NOT NULL 
                    THEN CONCAT(ay.year_code, ' - ', at.term_name)
                    ELSE NULL
                END as term_name
            FROM class_sections cs
            LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
            LEFT JOIN academic_terms at ON cs.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE cs.subject_id = %s
            ORDER BY cs.section_name
        """, (subject_id,))
        
        classes = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'classes': classes
        })
        
    except Exception as e:
        print(f"Error getting subject classes: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/subjects/<int:subject_id>/students')
@login_required
def get_subject_students(subject_id):
    """Get all students enrolled in any class of a specific subject"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all students enrolled in classes for this subject
        cursor.execute("""
            SELECT 
                e.enrollment_id,
                s.id as student_id,
                s.std_Number,
                CONCAT(s.std_Firstname, ' ', s.std_Surname) as student_name,
                cs.section_name,
                e.status,
                e.enrollment_date
            FROM enrollments e
            INNER JOIN std_info s ON e.student_id = s.id
            INNER JOIN class_sections cs ON e.section_id = cs.section_id
            WHERE cs.subject_id = %s
            ORDER BY s.std_Surname, s.std_Firstname
        """, (subject_id,))
        
        students = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'students': students
        })
        
    except Exception as e:
        print(f"Error getting subject students: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/classes/<int:class_id>/available-students')
@login_required
def get_available_students(class_id):
    """Get students not yet enrolled in a specific class"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get students not enrolled in this class
        cursor.execute("""
            SELECT 
                s.id,
                s.std_Number,
                s.std_Firstname,
                s.std_Surname,
                s.std_Level,
                s.std_Course
            FROM std_info s
            WHERE s.std_Status = 'Enrolled'
            AND s.is_archived = 0
            AND s.id NOT IN (
                SELECT student_id 
                FROM enrollments 
                WHERE section_id = %s 
                AND status = 'Enrolled'
            )
            ORDER BY s.std_Surname, s.std_Firstname
        """, (class_id,))
        
        students = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'students': students
        })
        
    except Exception as e:
        print(f"Error getting available students: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/enrollments', methods=['POST'])
@login_required
def create_enrollment():
    """Enroll a student in a class"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    student_id = data.get('student_id')
    section_id = data.get('section_id')
    status = data.get('status', 'Enrolled')
    
    if not student_id or not section_id:
        return jsonify({'success': False, 'error': 'Missing required fields'}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if student is already enrolled in this class
        cursor.execute("""
            SELECT enrollment_id 
            FROM enrollments 
            WHERE student_id = %s AND section_id = %s
        """, (student_id, section_id))
        
        if cursor.fetchone():
            return jsonify({'success': False, 'error': 'Student is already enrolled in this class'}), 400
        
        # Insert enrollment
        cursor.execute("""
            INSERT INTO enrollments (student_id, section_id, status, enrollment_date)
            VALUES (%s, %s, %s, NOW())
        """, (student_id, section_id, status))
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Student enrolled successfully',
            'enrollment_id': cursor.lastrowid
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error creating enrollment: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/enrollments/<int:enrollment_id>', methods=['DELETE'])
@login_required
def delete_enrollment(enrollment_id):
    """Remove a student from a class"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if enrollment exists
        cursor.execute("SELECT enrollment_id FROM enrollments WHERE enrollment_id = %s", (enrollment_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Enrollment not found'}), 404
        
        # Delete enrollment
        cursor.execute("DELETE FROM enrollments WHERE enrollment_id = %s", (enrollment_id,))
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Student removed from class successfully'
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error deleting enrollment: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections')
@login_required
def get_sections():
    """Get all class sections with subject, faculty, and term information"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all sections with related information
        query = """
            SELECT 
                cs.section_id,
                cs.section_name,
                cs.schedule,
                cs.room,
                cs.subject_id,
                cs.faculty_id,
                cs.acad_term_id,
                s.subject_code,
                s.title as subject_title,
                CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                CONCAT(ay.year_code, ' - ', at.term_name) as term_name,
                at.is_current as is_current_term,
                COUNT(DISTINCT e.student_id) as enrolled_students
            FROM class_sections cs
            LEFT JOIN subjects s ON cs.subject_id = s.subject_id
            LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
            LEFT JOIN academic_terms at ON cs.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            LEFT JOIN enrollments e ON cs.section_id = e.section_id
            GROUP BY cs.section_id
            ORDER BY at.is_current DESC, s.subject_code, cs.section_name
        """
        cursor.execute(query)
        sections = cursor.fetchall()
        
        # Get statistics
        stats_query = """
            SELECT 
                COUNT(DISTINCT cs.section_id) as total_sections,
                COUNT(DISTINCT cs.subject_id) as unique_subjects,
                COUNT(DISTINCT cs.faculty_id) as active_faculty,
                COUNT(DISTINCT CASE WHEN at.is_current = TRUE THEN cs.section_id END) as current_term_sections
            FROM class_sections cs
            LEFT JOIN academic_terms at ON cs.acad_term_id = at.acad_term_id
        """
        cursor.execute(stats_query)
        stats = cursor.fetchone()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': sections,
            'stats': {
                'total_sections': stats['total_sections'] or 0,
                'unique_subjects': stats['unique_subjects'] or 0,
                'active_faculty': stats['active_faculty'] or 0,
                'current_term_sections': stats['current_term_sections'] or 0
            }
        })
        
    except Exception as e:
        print(f"Error getting sections: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections', methods=['POST'])
@login_required
def create_section():
    """Create a new class section"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['subject_id', 'faculty_id', 'acad_term_id', 'section_name']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        cursor = conn.cursor(dictionary=True)
        
        # Insert new section
        insert_query = """
            INSERT INTO class_sections 
            (subject_id, faculty_id, acad_term_id, section_name, schedule, room)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(insert_query, (
            data['subject_id'],
            data['faculty_id'],
            data['acad_term_id'],
            data['section_name'],
            data.get('schedule', ''),
            data.get('room', '')
        ))
        
        conn.commit()
        section_id = cursor.lastrowid
        
        return jsonify({
            'success': True,
            'message': 'Section created successfully',
            'section_id': section_id
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error creating section: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections/<int:section_id>', methods=['PUT'])
@login_required
def update_section(section_id):
    """Update an existing section"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Check if section exists
        cursor.execute("SELECT * FROM class_sections WHERE section_id = %s", (section_id,))
        section = cursor.fetchone()
        
        if not section:
            cursor.close()
            return jsonify({'success': False, 'error': 'Section not found'}), 404
        
        # Build update query dynamically based on provided fields
        update_fields = []
        update_values = []
        
        if 'subject_id' in data:
            update_fields.append('subject_id = %s')
            update_values.append(data['subject_id'])
        if 'faculty_id' in data:
            update_fields.append('faculty_id = %s')
            update_values.append(data['faculty_id'])
        if 'acad_term_id' in data:
            update_fields.append('acad_term_id = %s')
            update_values.append(data['acad_term_id'])
        if 'section_name' in data:
            update_fields.append('section_name = %s')
            update_values.append(data['section_name'])
        if 'schedule' in data:
            update_fields.append('schedule = %s')
            update_values.append(data['schedule'])
        if 'room' in data:
            update_fields.append('room = %s')
            update_values.append(data['room'])
        
        if not update_fields:
            cursor.close()
            return jsonify({'success': False, 'error': 'No fields to update'}), 400
        
        # Add section_id to values for WHERE clause
        update_values.append(section_id)
        
        # Execute update
        update_query = f"""
            UPDATE class_sections 
            SET {', '.join(update_fields)}
            WHERE section_id = %s
        """
        cursor.execute(update_query, update_values)
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Section updated successfully'
        })
        
    except Exception as e:
        print(f"Error updating section: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections/<int:section_id>', methods=['DELETE'])
@login_required
def delete_section(section_id):
    """Delete a section"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if section exists
        cursor.execute("SELECT * FROM class_sections WHERE section_id = %s", (section_id,))
        section = cursor.fetchone()
        
        if not section:
            cursor.close()
            return jsonify({'success': False, 'error': 'Section not found'}), 404
        
        # Check if section has enrollments
        cursor.execute(
            "SELECT COUNT(*) as count FROM enrollments WHERE section_id = %s", 
            (section_id,)
        )
        enrollment_count = cursor.fetchone()['count']
        
        if enrollment_count > 0:
            cursor.close()
            return jsonify({
                'success': False, 
                'error': f'Cannot delete section with {enrollment_count} enrolled student(s). Please remove enrollments first.'
            }), 400
        
        # Check if section has evaluations
        cursor.execute(
            "SELECT COUNT(*) as count FROM evaluations WHERE section_id = %s", 
            (section_id,)
        )
        evaluation_count = cursor.fetchone()['count']
        
        if evaluation_count > 0:
            cursor.close()
            return jsonify({
                'success': False, 
                'error': f'Cannot delete section with {evaluation_count} evaluation(s). Please remove evaluations first.'
            }), 400
        
        # Delete the section
        cursor.execute("DELETE FROM class_sections WHERE section_id = %s", (section_id,))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Section deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting section: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/terms')
@login_required
def get_terms():
    """Get all academic terms with their academic years"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                at.acad_term_id as term_id,
                at.term_name,
                at.start_date,
                at.end_date,
                at.is_current,
                ay.acad_year_id,
                ay.year_code,
                ay.year_name
            FROM academic_terms at
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            ORDER BY at.start_date DESC
        """)
        terms = cursor.fetchall()
        
        return jsonify({'success': True, 'data': terms})
        
    except Exception as e:
        print(f"Error fetching terms: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/student/dashboard-data')
@login_required
def student_dashboard_data():
    """Get student dashboard data including pending evaluations"""
    if session.get('role') != 'student':
        return jsonify({'error': 'Unauthorized'}), 403
    
    user_id = session.get('user_id')
    student = Student.get_by_user_id(user_id)
    
    if not student:
        return jsonify({'error': 'Student not found'}), 404
    
    # Get pending evaluation count
    pending_count = Student.get_pending_evaluation_count(user_id)
    
    # Get evaluation history
    evaluations = Evaluation.get_student_evaluations(student['id'])
    
    return jsonify({
        'student': student,
        'pending_count': pending_count,
        'evaluations': evaluations
    })


@api_bp.route('/student/evaluation-counts')
@login_required
def student_evaluation_counts():
    """Get evaluation counts for student"""
    if session.get('role') != 'student':
        return jsonify({'error': 'Unauthorized'}), 403
    
    user_id = session.get('user_id')
    pending_count = Student.get_pending_evaluation_count(user_id)
    
    return jsonify({
        'pending': pending_count,
        'completed': 0  # TODO: Implement completed count
    })


@api_bp.route('/evaluation-periods', methods=['GET'])
@login_required
def get_evaluation_periods():
    """Get all evaluation periods"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM evaluation_periods 
            ORDER BY start_date DESC
        """)
        periods = cursor.fetchall()
        cursor.close()
        return jsonify(periods)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/academic-year', methods=['GET'])
def get_academic_year():
    """Get current academic year information"""
    conn = get_db_connection()
    if not conn:
        return jsonify({
            'success': True,
            'year': '2025-2026',
            'semester': '1st Semester',
            'status': 'Active'
        })
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM evaluation_periods 
            WHERE status = 'Active' 
            LIMIT 1
        """)
        period = cursor.fetchone()
        cursor.close()
        
        if period:
            title = period.get('title', '')
            year = '2025-2026'
            semester = '1st Semester'
            
            # Parse title for year and semester
            if '2026' in title or '2025' in title:
                year = '2025-2026'
            if 'first' in title.lower() or '1st' in title.lower():
                semester = '1st Semester'
            elif 'second' in title.lower() or '2nd' in title.lower():
                semester = '2nd Semester'
            
            return jsonify({
                'success': True,
                'year': year,
                'semester': semester,
                'full_title': title,
                'status': 'Active'
            })
    except Exception as e:
        print(f"Error getting academic year: {e}")
    finally:
        conn.close()
    
    # Return fallback data
    return jsonify({
        'success': True,
        'year': '2025-2026',
        'semester': '1st Semester',
        'status': 'Active'
    })


@api_bp.route('/student/profile', methods=['GET', 'PUT'])
@login_required
def student_profile():
    """Get or update student profile information"""
    if session.get('role') != 'student':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    user_id = session.get('user_id')
    
    if request.method == 'GET':
        try:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            
            # Get student info with user data
            cursor.execute("""
                SELECT 
                    u.user_id, u.username, u.first_name, u.last_name, u.role,
                    s.id, s.std_Number as student_number, s.std_Middlename as middle_name, 
                    s.std_Suffix as suffix, s.std_Gender as gender, s.std_Age as age, 
                    s.std_ContactNum as contact_number, s.std_Address as address, 
                    s.std_EmailAdd as email, s.std_Course as program, 
                    s.std_Level as year_level, s.std_Status as status
                FROM users u
                LEFT JOIN std_info s ON u.user_id = s.user_id
                WHERE u.user_id = %s
            """, (user_id,))
            
            profile = cursor.fetchone()
            cursor.close()
            conn.close()
            
            if not profile:
                return jsonify({'success': False, 'message': 'Profile not found'}), 404
            
            return jsonify({
                'success': True,
                'profile': {
                    'first_name': profile['first_name'],
                    'last_name': profile['last_name'],
                    'middle_name': profile.get('middle_name', ''),
                    'suffix': profile.get('suffix', ''),
                    'student_number': profile.get('student_number', 'N/A'),
                    'email': profile.get('email', 'N/A'),
                    'gender': profile.get('gender', ''),
                    'age': profile.get('age', ''),
                    'contact_number': profile.get('contact_number', ''),
                    'address': profile.get('address', ''),
                    'program': profile.get('program', 'N/A'),
                    'year_level': profile.get('year_level', 'N/A'),
                    'status': profile.get('status', 'Active')
                }
            })
            
        except Exception as e:
            print(f"Error getting student profile: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Update user table
            cursor.execute("""
                UPDATE users 
                SET first_name = %s, last_name = %s
                WHERE user_id = %s
            """, (data.get('first_name'), data.get('last_name'), user_id))
            
            # Update std_info table
            cursor.execute("""
                UPDATE std_info 
                SET middle_name = %s, suffix = %s, gender = %s, age = %s,
                    contact_number = %s, address = %s, email = %s
                WHERE user_id = %s
            """, (
                data.get('middle_name'), data.get('suffix'), data.get('gender'),
                data.get('age'), data.get('contact_number'), data.get('address'),
                data.get('email'), user_id
            ))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            # Update session data
            session['first_name'] = data.get('first_name')
            session['last_name'] = data.get('last_name')
            
            return jsonify({'success': True, 'message': 'Profile updated successfully'})
            
        except Exception as e:
            print(f"Error updating student profile: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/student/enrollments')
@login_required
def student_enrollments():
    """Get student's current course enrollments"""
    if session.get('role') != 'student':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        user_id = session.get('user_id')
        
        # Get student_id from std_info
        cursor.execute("SELECT id FROM std_info WHERE user_id = %s", (user_id,))
        student_record = cursor.fetchone()
        
        if not student_record:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'enrollments': []})
        
        student_id = student_record['id']
        
        # Get enrollments with course and faculty info
        cursor.execute("""
            SELECT 
                c.course_code,
                c.title as course_title,
                c.units,
                CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                cs.schedule,
                cs.room
            FROM enrollments e
            JOIN class_sections cs ON e.section_id = cs.section_id
            JOIN courses c ON cs.course_id = c.course_id
            JOIN faculty f ON cs.faculty_id = f.faculty_id
            WHERE e.student_id = %s
            ORDER BY c.course_code
        """, (student_id,))
        
        enrollments = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'enrollments': enrollments
        })
        
    except Exception as e:
        print(f"Error getting student enrollments: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'enrollments': [], 'message': str(e)})


@api_bp.route('/student/activity-stats')
@login_required
def student_activity_stats():
    """Get student's evaluation activity statistics"""
    if session.get('role') != 'student':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        user_id = session.get('user_id')
        
        # Get student_id from std_info
        cursor.execute("SELECT id FROM std_info WHERE user_id = %s", (user_id,))
        student_record = cursor.fetchone()
        
        if not student_record:
            cursor.close()
            conn.close()
            return jsonify({
                'success': True,
                'stats': {
                    'total_evaluations': 0,
                    'completed_evaluations': 0,
                    'pending_evaluations': 0
                }
            })
        
        student_id = student_record['id']
        
        # Get evaluation statistics
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN status = 'Completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN status IN ('Pending', 'In Progress') THEN 1 ELSE 0 END) as pending
            FROM evaluations
            WHERE student_id = %s
        """, (student_id,))
        
        stats = cursor.fetchone()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'stats': {
                'total_evaluations': stats['total'] or 0,
                'completed_evaluations': stats['completed'] or 0,
                'pending_evaluations': stats['pending'] or 0
            }
        })
        
    except Exception as e:
        print(f"Error getting activity stats: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': True,
            'stats': {
                'total_evaluations': 0,
                'completed_evaluations': 0,
                'pending_evaluations': 0
            }
        })


# ===========================
# Guidance API Endpoints
# ===========================

@api_bp.route('/guidance/students')
@login_required
def guidance_students():
    """Get all students with evaluation status for guidance counselor"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get all students with their evaluation status
        cursor.execute("""
            SELECT 
                s.id,
                s.std_Number as student_number,
                s.std_Firstname as first_name,
                s.std_Surname as last_name,
                s.std_EmailAdd as email,
                s.std_Course as department,
                s.std_Status as enrollment_status,
                s.created_at,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                SUM(CASE WHEN e.status = 'Completed' THEN 1 ELSE 0 END) as completed_evaluations,
                CASE 
                    WHEN COUNT(DISTINCT e.evaluation_id) = 0 THEN 'No Evaluations'
                    WHEN SUM(CASE WHEN e.status = 'Completed' THEN 1 ELSE 0 END) = COUNT(DISTINCT e.evaluation_id) THEN 'Completed'
                    WHEN SUM(CASE WHEN e.status = 'In Progress' THEN 1 ELSE 0 END) > 0 THEN 'In Progress'
                    ELSE 'Pending'
                END as status
            FROM std_info s
            LEFT JOIN evaluations e ON s.id = e.student_id
            GROUP BY s.id, s.std_Number, s.std_Firstname, s.std_Surname, 
                     s.std_EmailAdd, s.std_Course, s.std_Status, s.created_at
            ORDER BY s.std_Surname, s.std_Firstname
        """)
        
        students = cursor.fetchall()
        cursor.close()
        conn.close()
        
        # Convert datetime objects to strings
        for student in students:
            if student.get('created_at'):
                student['created_at'] = student['created_at'].isoformat() if hasattr(student['created_at'], 'isoformat') else str(student['created_at'])
        
        return jsonify({
            'success': True,
            'students': students
        })
        
    except Exception as e:
        print(f"Error getting students for guidance: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/faculty')
@login_required
def guidance_faculty():
    """Get all faculty members for guidance counselor"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                f.faculty_id,
                f.first_name,
                f.last_name,
                f.faculty_number,
                f.status,
                f.rank,
                p.name as department_name
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            WHERE f.is_archived = 0
            ORDER BY f.last_name, f.first_name
        """)
        
        faculty = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'data': faculty
        })
        
    except Exception as e:
        print(f"Error getting faculty for guidance: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/faculty-list')
@login_required
def guidance_faculty_list():
    """Get detailed faculty list with subjects and statistics for faculty management"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get all faculty with their department and subject count
        cursor.execute("""
            SELECT 
                f.faculty_id,
                f.first_name,
                f.last_name,
                f.faculty_number,
                f.email,
                f.status,
                f.rank,
                p.name as department_name,
                COUNT(DISTINCT cs.section_id) as subject_count
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            WHERE f.is_archived = 0
            GROUP BY f.faculty_id, f.first_name, f.last_name, f.faculty_number, 
                     f.email, f.status, f.rank, p.name
            ORDER BY f.last_name, f.first_name
        """)
        
        faculty_list = cursor.fetchall()
        
        # Get subjects for each faculty member
        for faculty in faculty_list:
            cursor.execute("""
                SELECT DISTINCT
                    s.subject_code,
                    s.title as subject_name
                FROM class_sections cs
                INNER JOIN subjects s ON cs.subject_id = s.subject_id
                WHERE cs.faculty_id = %s
                ORDER BY s.subject_code
            """, (faculty['faculty_id'],))
            
            faculty['subjects'] = cursor.fetchall()
        
        # Get statistics
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT f.faculty_id) as total_faculty,
                COUNT(DISTINCT CASE WHEN f.status = 'active' THEN f.faculty_id END) as active_faculty,
                COUNT(DISTINCT f.program_id) as total_departments,
                COUNT(DISTINCT cs.subject_id) as total_subjects
            FROM faculty f
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            WHERE f.is_archived = 0
        """)
        
        stats = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'faculty': faculty_list,
            'stats': {
                'total': stats['total_faculty'] or 0,
                'active': stats['active_faculty'] or 0,
                'departments': stats['total_departments'] or 0,
                'subjects': stats['total_subjects'] or 0
            }
        })
        
    except Exception as e:
        print(f"Error getting faculty list for management: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/student-evaluations/<student_number>')
@login_required
def guidance_student_evaluations(student_number):
    """Get evaluation history for a specific student"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get student evaluations with faculty and subject details
        cursor.execute("""
            SELECT 
                e.evaluation_id,
                e.status,
                e.completion_time as submitted_at,
                e.created_at,
                CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                s.subject_code as course_code,
                s.title as course_title,
                CONCAT(ay.year_code, ' - ', at.term_name) as period_name,
                YEAR(ay.start_date) as academic_year,
                at.term_name as semester
            FROM evaluations e
            INNER JOIN std_info si ON e.student_id = si.id
            INNER JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
            LEFT JOIN subjects s ON cs.subject_id = s.subject_id
            LEFT JOIN academic_terms at ON cs.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE si.std_Number = %s
            ORDER BY e.completion_time DESC, e.created_at DESC
        """, (student_number,))
        
        evaluations = cursor.fetchall()
        cursor.close()
        conn.close()
        
        # Convert datetime objects to strings
        for evaluation in evaluations:
            if evaluation.get('completion_time'):
                evaluation['completion_time'] = evaluation['completion_time'].isoformat() if hasattr(evaluation['completion_time'], 'isoformat') else str(evaluation['completion_time'])
            if evaluation.get('created_at'):
                evaluation['created_at'] = evaluation['created_at'].isoformat() if hasattr(evaluation['created_at'], 'isoformat') else str(evaluation['created_at'])
        
        return jsonify({
            'success': True,
            'evaluations': evaluations
        })
        
    except Exception as e:
        print(f"Error getting student evaluations: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/allow-retake', methods=['POST'])
@login_required
def guidance_allow_retake():
    """Allow a student to retake evaluation(s)"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.json
        student_number = data.get('student_number')
        evaluation_ids = data.get('evaluation_ids', [])  # Array of evaluation IDs
        reason = data.get('reason')
        
        # Support legacy single faculty_id parameter
        faculty_id = data.get('faculty_id')
        
        if not student_number or not reason:
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        if not evaluation_ids and not faculty_id:
            return jsonify({'success': False, 'message': 'Either evaluation_ids or faculty_id is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get student ID and name
        cursor.execute("SELECT id, std_Firstname, std_Surname FROM std_info WHERE std_Number = %s", (student_number,))
        student = cursor.fetchone()
        
        if not student:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        
        student_id = student['id']
        student_name = f"{student['std_Firstname']} {student['std_Surname']}"
        
        retake_count = 0
        subject_names = []
        
        # Handle multiple evaluation IDs (new approach)
        if evaluation_ids:
            for eval_id in evaluation_ids:
                # Get evaluation details for logging
                cursor.execute("""
                    SELECT e.evaluation_id, sub.title, sub.subject_code,
                           CONCAT(f.first_name, ' ', f.last_name) as faculty_name
                    FROM evaluations e
                    INNER JOIN class_sections cs ON e.section_id = cs.section_id
                    INNER JOIN subjects sub ON cs.subject_id = sub.subject_id
                    INNER JOIN faculty f ON cs.faculty_id = f.faculty_id
                    WHERE e.evaluation_id = %s AND e.student_id = %s
                """, (eval_id, student_id))
                
                evaluation = cursor.fetchone()
                
                if evaluation:
                    # Delete existing evaluation responses (ratings) to clear previous answers
                    cursor.execute("""
                        DELETE FROM evaluation_responses 
                        WHERE evaluation_id = %s
                    """, (eval_id,))
                    
                    # Delete existing comments to avoid duplication on retake
                    cursor.execute("""
                        DELETE FROM comments 
                        WHERE evaluation_id = %s
                    """, (eval_id,))
                    
                    # Delete timer session to allow fresh start (prevents re-expiring)
                    cursor.execute("""
                        DELETE FROM evaluation_timer_sessions 
                        WHERE evaluation_id = %s
                    """, (eval_id,))
                    
                    # Update evaluation to allow retake
                    cursor.execute("""
                        UPDATE evaluations 
                        SET status = 'Pending', 
                            completion_time = NULL,
                            start_time = NULL,
                            updated_at = NOW()
                        WHERE evaluation_id = %s
                    """, (eval_id,))
                    
                    retake_count += 1
                    subject_names.append(f"{evaluation['subject_code']} - {evaluation['title']}")
                    
                    # Log this retake activity
                    log_activity(
                        user_id=session.get('user_id'),
                        user_name=f"{session.get('first_name')} {session.get('last_name')}",
                        user_role='guidance',
                        activity_type='retake',
                        description=f"Allowed retake for {student_name} - {evaluation['subject_code']}: {evaluation['title']} (Faculty: {evaluation['faculty_name']}). Previous responses and comments deleted.",
                        reason=reason,
                        target_user=student_name,
                        ip_address=request.remote_addr,
                        additional_data={'evaluation_id': eval_id, 'student_number': student_number}
                    )
        
        # Handle single faculty_id (legacy support)
        elif faculty_id:
            cursor.execute("""
                SELECT e.evaluation_id, sub.title, sub.subject_code,
                       CONCAT(f.first_name, ' ', f.last_name) as faculty_name
                FROM evaluations e
                INNER JOIN class_sections cs ON e.section_id = cs.section_id
                INNER JOIN subjects sub ON cs.subject_id = sub.subject_id
                INNER JOIN faculty f ON cs.faculty_id = f.faculty_id
                WHERE e.student_id = %s AND cs.faculty_id = %s
                ORDER BY e.created_at DESC
                LIMIT 1
            """, (student_id, faculty_id))
            
            evaluation = cursor.fetchone()
            
            if evaluation:
                # Delete existing evaluation responses (ratings) to clear previous answers
                cursor.execute("""
                    DELETE FROM evaluation_responses 
                    WHERE evaluation_id = %s
                """, (evaluation['evaluation_id'],))
                
                # Delete existing comments to avoid duplication on retake
                cursor.execute("""
                    DELETE FROM comments 
                    WHERE evaluation_id = %s
                """, (evaluation['evaluation_id'],))
                
                # Delete timer session to allow fresh start (prevents re-expiring)
                cursor.execute("""
                    DELETE FROM evaluation_timer_sessions 
                    WHERE evaluation_id = %s
                """, (evaluation['evaluation_id'],))
                
                # Update existing evaluation to allow retake
                cursor.execute("""
                    UPDATE evaluations 
                    SET status = 'Pending', 
                        completion_time = NULL,
                        start_time = NULL,
                        updated_at = NOW()
                    WHERE evaluation_id = %s
                """, (evaluation['evaluation_id'],))
                
                retake_count = 1
                subject_names.append(f"{evaluation['subject_code']} - {evaluation['title']}")
                
                # Log this retake activity
                log_activity(
                    user_id=session.get('user_id'),
                    user_name=f"{session.get('first_name')} {session.get('last_name')}",
                    user_role='guidance',
                    activity_type='retake',
                    description=f"Allowed retake for {student_name} - {evaluation['subject_code']}: {evaluation['title']} (Faculty: {evaluation['faculty_name']}). Previous responses and comments deleted.",
                    reason=reason,
                    target_user=student_name,
                    ip_address=request.remote_addr,
                    additional_data={'faculty_id': faculty_id, 'student_number': student_number}
                )
            else:
                cursor.close()
                conn.close()
                return jsonify({'success': False, 'message': 'No evaluation found for this student and faculty'}), 404
        
        if retake_count == 0:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'No evaluations were updated'}), 400
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Retake permission granted for {retake_count} evaluation(s)',
            'retake_count': retake_count,
            'subjects': subject_names
        })
        
    except Exception as e:
        print(f"Error allowing retake: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/admin/allow-retake', methods=['POST'])
@login_required
def admin_allow_retake():
    """Allow a student to retake evaluation(s) - Admin version"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.json
        student_number = data.get('student_number')
        evaluation_ids = data.get('evaluation_ids', [])
        reason = data.get('reason')
        
        if not student_number or not reason:
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        if not evaluation_ids:
            return jsonify({'success': False, 'message': 'No evaluation IDs provided'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get student ID and name
        cursor.execute("SELECT id, std_Firstname, std_Surname FROM std_info WHERE std_Number = %s", (student_number,))
        student = cursor.fetchone()
        
        if not student:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        
        student_id = student['id']
        student_name = f"{student['std_Firstname']} {student['std_Surname']}"
        
        retake_count = 0
        subject_names = []
        
        # Handle multiple evaluation IDs
        for eval_id in evaluation_ids:
            # Get evaluation details for logging
            cursor.execute("""
                SELECT e.evaluation_id, sub.title, sub.subject_code,
                       CONCAT(f.first_name, ' ', f.last_name) as faculty_name
                FROM evaluations e
                INNER JOIN class_sections cs ON e.section_id = cs.section_id
                INNER JOIN subjects sub ON cs.subject_id = sub.subject_id
                INNER JOIN faculty f ON cs.faculty_id = f.faculty_id
                WHERE e.evaluation_id = %s AND e.student_id = %s
            """, (eval_id, student_id))
            
            evaluation = cursor.fetchone()
            
            if evaluation:
                # Delete existing evaluation responses (ratings)
                cursor.execute("""
                    DELETE FROM evaluation_responses 
                    WHERE evaluation_id = %s
                """, (eval_id,))
                
                # Delete existing comments
                cursor.execute("""
                    DELETE FROM comments 
                    WHERE evaluation_id = %s
                """, (eval_id,))
                
                # Delete timer session to allow fresh start
                cursor.execute("""
                    DELETE FROM evaluation_timer_sessions 
                    WHERE evaluation_id = %s
                """, (eval_id,))
                
                # Update evaluation to allow retake
                cursor.execute("""
                    UPDATE evaluations 
                    SET status = 'Pending', 
                        completion_time = NULL,
                        start_time = NULL,
                        updated_at = NOW()
                    WHERE evaluation_id = %s
                """, (eval_id,))
                
                retake_count += 1
                subject_names.append(f"{evaluation['subject_code']} - {evaluation['title']}")
                
                # Log this retake activity
                log_activity(
                    user_id=session.get('user_id'),
                    user_name=f"{session.get('first_name')} {session.get('last_name')}",
                    user_role='admin',
                    activity_type='retake',
                    description=f"Allowed retake for {student_name} - {evaluation['subject_code']}: {evaluation['title']} (Faculty: {evaluation['faculty_name']}). Previous responses and comments deleted.",
                    reason=reason,
                    target_user=student_name,
                    ip_address=request.remote_addr,
                    additional_data={'evaluation_id': eval_id, 'student_number': student_number}
                )
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Retake permission granted for {retake_count} evaluation(s)',
            'retake_count': retake_count,
            'subjects': subject_names
        })
        
    except Exception as e:
        print(f"Error allowing retake: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/faculty-evaluations')
@login_required
def guidance_faculty_evaluations():
    """Get faculty evaluation statistics and status for guidance counselor"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get faculty with their evaluation statistics
        cursor.execute("""
            SELECT 
                f.faculty_id,
                f.first_name,
                f.last_name,
                f.faculty_number,
                p.name as department_name,
                '' as college_name,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                SUM(CASE WHEN e.status = 'Completed' THEN 1 ELSE 0 END) as completed_evaluations,
                SUM(CASE WHEN e.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress_evaluations,
                SUM(CASE WHEN e.status = 'Pending' THEN 1 ELSE 0 END) as pending_evaluations,
                AVG(CASE WHEN er.overall_rating IS NOT NULL THEN er.overall_rating ELSE NULL END) as average_rating
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
            LEFT JOIN (
                SELECT evaluation_id, AVG(rating) as overall_rating
                FROM evaluation_responses
                GROUP BY evaluation_id
            ) er ON e.evaluation_id = er.evaluation_id
            WHERE f.is_archived = 0
            GROUP BY f.faculty_id, f.first_name, f.last_name, f.faculty_number, p.name
            ORDER BY f.last_name, f.first_name
        """)
        
        faculty_list = cursor.fetchall()
        
        # Get overall statistics
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT f.faculty_id) as total_faculty,
                COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as total_completed,
                COUNT(DISTINCT CASE WHEN e.status = 'Pending' THEN e.evaluation_id END) as total_pending,
                AVG(CASE WHEN er.overall_rating IS NOT NULL THEN er.overall_rating ELSE NULL END) as overall_average_rating
            FROM faculty f
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
            LEFT JOIN (
                SELECT evaluation_id, AVG(rating) as overall_rating
                FROM evaluation_responses
                GROUP BY evaluation_id
            ) er ON e.evaluation_id = er.evaluation_id
        """)
        
        statistics = cursor.fetchone()
        
        # Get current evaluation period with academic term info
        cursor.execute("""
            SELECT 
                ep.title as period_name, 
                CONCAT(ay.year_code, ' - ', at.term_name) as academic_year,
                ep.start_date, 
                ep.end_date
            FROM evaluation_periods ep
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE ep.status = 'Active'
            ORDER BY ep.start_date DESC
            LIMIT 1
        """)
        
        current_period = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        # Process the data
        for faculty in faculty_list:
            # Calculate response rate
            if faculty['total_evaluations'] and faculty['total_evaluations'] > 0:
                faculty['response_rate'] = round((faculty['completed_evaluations'] / faculty['total_evaluations']) * 100, 1)
            else:
                faculty['response_rate'] = 0
            
            # Round average rating
            if faculty['average_rating']:
                faculty['average_rating'] = round(faculty['average_rating'], 1)
            else:
                faculty['average_rating'] = 0
            
            # Determine status
            if faculty['total_evaluations'] == 0:
                faculty['status'] = 'No Evaluations'
            elif faculty['completed_evaluations'] == faculty['total_evaluations']:
                faculty['status'] = 'Completed'
            elif faculty['in_progress_evaluations'] > 0:
                faculty['status'] = 'In Progress'
            else:
                faculty['status'] = 'Pending'
        
        # Process statistics
        if statistics:
            if statistics['overall_average_rating']:
                statistics['overall_average_rating'] = round(statistics['overall_average_rating'], 1)
            else:
                statistics['overall_average_rating'] = 0
        
        # Process current period dates
        if current_period:
            if current_period.get('start_date'):
                current_period['start_date'] = current_period['start_date'].isoformat() if hasattr(current_period['start_date'], 'isoformat') else str(current_period['start_date'])
            if current_period.get('end_date'):
                current_period['end_date'] = current_period['end_date'].isoformat() if hasattr(current_period['end_date'], 'isoformat') else str(current_period['end_date'])
        
        return jsonify({
            'success': True,
            'faculty': faculty_list,
            'statistics': statistics,
            'current_period': current_period
        })
        
    except Exception as e:
        print(f"Error getting faculty evaluations: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/evaluation-periods-legacy')
@login_required
def guidance_evaluation_periods_legacy():
    """[DEPRECATED] Legacy endpoint - Use /api/guidance/evaluation-periods instead"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get all evaluation periods with academic term info
        cursor.execute("""
            SELECT 
                ep.period_id,
                ep.title as period_name,
                ep.start_date,
                ep.end_date,
                ep.created_at,
                CONCAT(ay.year_code, ' - ', at.term_name) as academic_year,
                at.acad_term_id
            FROM evaluation_periods ep
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            ORDER BY ep.start_date DESC
        """)
        
        periods = cursor.fetchall()
        
        # Get statistics for each period
        for period in periods:
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT evaluation_id) as total_evaluations,
                    SUM(CASE WHEN status = 'Completed' THEN 1 ELSE 0 END) as completed_evaluations,
                    SUM(CASE WHEN status = 'In Progress' THEN 1 ELSE 0 END) as in_progress_evaluations,
                    SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending_evaluations
                FROM evaluations
                WHERE period_id = %s
            """, (period['period_id'],))
            
            stats = cursor.fetchone()
            if stats:
                period['total_evaluations'] = stats['total_evaluations'] or 0
                period['completed_evaluations'] = stats['completed_evaluations'] or 0
                period['in_progress_evaluations'] = stats['in_progress_evaluations'] or 0
                period['pending_evaluations'] = stats['pending_evaluations'] or 0
            else:
                period['total_evaluations'] = 0
                period['completed_evaluations'] = 0
                period['in_progress_evaluations'] = 0
                period['pending_evaluations'] = 0
        
        cursor.close()
        conn.close()
        
        # Process dates and calculate additional info
        from datetime import datetime
        now = datetime.now()
        current_period = None
        
        for period in periods:
            # Convert datetime objects to strings
            if period.get('start_date'):
                period['start_date'] = period['start_date'].isoformat() if hasattr(period['start_date'], 'isoformat') else str(period['start_date'])
            if period.get('end_date'):
                period['end_date'] = period['end_date'].isoformat() if hasattr(period['end_date'], 'isoformat') else str(period['end_date'])
            if period.get('created_at'):
                period['created_at'] = period['created_at'].isoformat() if hasattr(period['created_at'], 'isoformat') else str(period['created_at'])
            
            # Calculate duration and status
            try:
                start = datetime.fromisoformat(period['start_date'].replace('Z', '+00:00'))
                end = datetime.fromisoformat(period['end_date'].replace('Z', '+00:00'))
                period['duration_days'] = (end - start).days
                
                # Determine status based on dates
                if now < start:
                    period['status'] = 'Pending'
                    period['days_remaining'] = (start - now).days
                    period['progress_percent'] = 0
                elif now > end:
                    period['status'] = 'Closed'
                    period['days_remaining'] = 0
                    period['progress_percent'] = 100
                else:
                    period['status'] = 'Active'
                    period['days_remaining'] = (end - now).days
                    total_duration = (end - start).days
                    elapsed = (now - start).days
                    period['progress_percent'] = min(100, max(0, round((elapsed / total_duration) * 100)))
                    
                    # Set as current period if active
                    if current_period is None:
                        current_period = period
            except:
                period['duration_days'] = 0
                period['days_remaining'] = 0
                period['progress_percent'] = 0
                period['status'] = 'Unknown'
            
            # Calculate response rate
            if period['total_evaluations'] and period['total_evaluations'] > 0:
                period['response_rate'] = round((period['completed_evaluations'] / period['total_evaluations']) * 100, 1)
            else:
                period['response_rate'] = 0
        
        return jsonify({
            'success': True,
            'periods': periods,
            'current_period': current_period
        })
        
    except Exception as e:
        print(f"Error getting evaluation periods: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/performance-trends')
@login_required
def guidance_performance_trends():
    """Get performance trends data for departments or faculty across evaluation periods"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        trend_type = request.args.get('trend_type', 'department')  # 'department' or 'faculty'
        period_id = request.args.get('period_id', 'all')
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get evaluation periods for x-axis labels
        if period_id == 'all':
            cursor.execute("""
                SELECT 
                    ep.period_id,
                    ep.title as period_name,
                    CONCAT(ay.year_code, ' - ', at.term_name) as full_name
                FROM evaluation_periods ep
                LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
                LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
                ORDER BY ep.start_date ASC
                LIMIT 10
            """)
        else:
            cursor.execute("""
                SELECT 
                    ep.period_id,
                    ep.title as period_name,
                    CONCAT(ay.year_code, ' - ', at.term_name) as full_name
                FROM evaluation_periods ep
                LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
                LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
                WHERE ep.period_id = %s
            """, (period_id,))
        
        periods = cursor.fetchall()
        
        if not periods:
            return jsonify({
                'success': True,
                'labels': [],
                'datasets': [],
                'message': 'No evaluation periods found'
            })
        
        # Create labels from period names
        labels = [p['period_name'] or p['full_name'] for p in periods]
        period_ids = [p['period_id'] for p in periods]
        
        datasets = []
        
        if trend_type == 'department':
            # Get average ratings by department for each period
            cursor.execute("""
                SELECT DISTINCT p.program_id, p.name as program_name
                FROM programs p
                INNER JOIN faculty f ON p.program_id = f.program_id
                WHERE f.status = 'Active'
                ORDER BY p.name
                LIMIT 10
            """)
            departments = cursor.fetchall()
            
            # Color palette for departments
            colors = [
                'rgb(59, 130, 246)',   # blue
                'rgb(16, 185, 129)',   # green
                'rgb(249, 115, 22)',   # orange
                'rgb(168, 85, 247)',   # purple
                'rgb(236, 72, 153)',   # pink
                'rgb(234, 179, 8)',    # yellow
                'rgb(14, 165, 233)',   # sky
                'rgb(139, 92, 246)',   # violet
                'rgb(34, 197, 94)',    # emerald
                'rgb(239, 68, 68)'     # red
            ]
            
            for idx, dept in enumerate(departments):
                data = []
                for period_id in period_ids:
                    # Get average rating for this department in this period
                    cursor.execute("""
                        SELECT AVG(er.rating) as avg_rating
                        FROM evaluation_responses er
                        INNER JOIN evaluations e ON er.evaluation_id = e.evaluation_id
                        INNER JOIN class_sections cs ON e.section_id = cs.section_id
                        INNER JOIN faculty f ON cs.faculty_id = f.faculty_id
                        WHERE f.program_id = %s AND e.period_id = %s AND e.status = 'Completed'
                    """, (dept['program_id'], period_id))
                    result = cursor.fetchone()
                    avg_rating = float(result['avg_rating']) if result and result['avg_rating'] else None
                    data.append(round(avg_rating, 2) if avg_rating else None)
                
                color = colors[idx % len(colors)]
                datasets.append({
                    'label': dept['program_name'],
                    'data': data,
                    'borderColor': color,
                    'backgroundColor': color.replace('rgb', 'rgba').replace(')', ', 0.1)'),
                    'tension': 0.4,
                    'fill': True,
                    'borderWidth': 2
                })
        
        elif trend_type == 'faculty':
            # Get top 10 faculty by average rating across all periods
            cursor.execute("""
                SELECT 
                    f.faculty_id,
                    CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                    AVG(er.rating) as overall_avg
                FROM faculty f
                INNER JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                INNER JOIN evaluations e ON cs.section_id = e.section_id
                INNER JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                WHERE f.status = 'Active' AND e.status = 'Completed'
                GROUP BY f.faculty_id, f.first_name, f.last_name
                ORDER BY overall_avg DESC
                LIMIT 10
            """)
            faculty_list = cursor.fetchall()
            
            # Color palette for faculty
            colors = [
                'rgb(59, 130, 246)',   # blue
                'rgb(16, 185, 129)',   # green
                'rgb(249, 115, 22)',   # orange
                'rgb(168, 85, 247)',   # purple
                'rgb(236, 72, 153)',   # pink
                'rgb(234, 179, 8)',    # yellow
                'rgb(14, 165, 233)',   # sky
                'rgb(139, 92, 246)',   # violet
                'rgb(34, 197, 94)',    # emerald
                'rgb(239, 68, 68)'     # red
            ]
            
            for idx, faculty in enumerate(faculty_list):
                data = []
                for period_id in period_ids:
                    # Get average rating for this faculty in this period
                    cursor.execute("""
                        SELECT AVG(er.rating) as avg_rating
                        FROM evaluation_responses er
                        INNER JOIN evaluations e ON er.evaluation_id = e.evaluation_id
                        INNER JOIN class_sections cs ON e.section_id = cs.section_id
                        WHERE cs.faculty_id = %s AND e.period_id = %s AND e.status = 'Completed'
                    """, (faculty['faculty_id'], period_id))
                    result = cursor.fetchone()
                    avg_rating = float(result['avg_rating']) if result and result['avg_rating'] else None
                    data.append(round(avg_rating, 2) if avg_rating else None)
                
                color = colors[idx % len(colors)]
                datasets.append({
                    'label': faculty['faculty_name'],
                    'data': data,
                    'borderColor': color,
                    'backgroundColor': color.replace('rgb', 'rgba').replace(')', ', 0.1)'),
                    'tension': 0.4,
                    'fill': True,
                    'borderWidth': 2
                })
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'labels': labels,
            'datasets': datasets
        })
        
    except Exception as e:
        print(f"Error getting performance trends: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# ACADEMIC TERMS (ACADEMIC YEARS) ENDPOINTS
# ============================================================

@api_bp.route('/current-term')
@login_required
def get_current_term():
    """Get current academic term information"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get current academic year
        cursor.execute("""
            SELECT acad_year_id, year_code, year_name, start_date, end_date, is_current
            FROM academic_years
            WHERE is_current = 1
            LIMIT 1
        """)
        
        current_year = cursor.fetchone()
        
        if current_year:
            # Get current term/semester for this year
            cursor.execute("""
                SELECT acad_term_id, term_name, term_code, start_date, end_date, is_current
                FROM academic_terms
                WHERE acad_year_id = %s AND is_current = 1
                LIMIT 1
            """, (current_year['acad_year_id'],))
            current_term = cursor.fetchone()
            
            cursor.close()
            
            return jsonify({
                'success': True,
                'data': {
                    'acad_year_id': current_year['acad_year_id'],
                    'year_code': current_year['year_code'],
                    'year_name': current_year['year_name'],
                    'acad_term_id': current_term['acad_term_id'] if current_term else None,
                    'term_name': current_term['term_name'] if current_term else None,
                    'start_date': str(current_year['start_date']),
                    'end_date': str(current_year['end_date'])
                }
            })
        else:
            cursor.close()
            # No current year set, return default
            return jsonify({
                'success': True,
                'data': {
                    'acad_year_id': None,
                    'year_code': 'No Active Year',
                    'year_name': 'No Active Academic Year',
                    'acad_term_id': None,
                    'term_name': None,
                    'start_date': None,
                    'end_date': None
                }
            })
        
    except Exception as e:
        print(f"Error getting current term: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/academic-years')
@login_required
def get_academic_years():
    """Get all academic years with statistics"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all academic years with statistics
        query = """
            SELECT 
                ay.*,
                COUNT(DISTINCT at.acad_term_id) as total_terms,
                COUNT(DISTINCT cs.section_id) as total_sections,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as completed_evaluations,
                COUNT(DISTINCT ep.period_id) as evaluation_periods,
                CASE
                    WHEN ay.is_current = 1 THEN 'Active'
                    WHEN CURDATE() < ay.start_date THEN 'Upcoming'
                    WHEN CURDATE() > ay.end_date THEN 'Completed'
                    ELSE 'Active'
                END as status,
                CASE 
                    WHEN COUNT(DISTINCT e.evaluation_id) > 0 
                    THEN ROUND((COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) * 100.0 / COUNT(DISTINCT e.evaluation_id)), 1)
                    ELSE 0 
                END as completion_rate
            FROM academic_years ay
            LEFT JOIN academic_terms at ON ay.acad_year_id = at.acad_year_id
            LEFT JOIN class_sections cs ON at.acad_term_id = cs.acad_term_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
            LEFT JOIN evaluation_periods ep ON at.acad_term_id = ep.acad_term_id
            GROUP BY ay.acad_year_id
            ORDER BY ay.start_date DESC
        """
        cursor.execute(query)
        academic_years = cursor.fetchall()
        
        # For each academic year, get its terms
        for year in academic_years:
            cursor.execute("""
                SELECT 
                    acad_term_id,
                    acad_year_id,
                    term_name,
                    term_code,
                    start_date,
                    end_date,
                    is_current
                FROM academic_terms
                WHERE acad_year_id = %s
                ORDER BY term_code
            """, (year['acad_year_id'],))
            year['terms'] = cursor.fetchall()
            
            # Convert dates to strings for JSON serialization
            for term in year['terms']:
                if term.get('start_date'):
                    term['start_date'] = str(term['start_date'])
                if term.get('end_date'):
                    term['end_date'] = str(term['end_date'])
        
        # Get overall statistics
        stats_query = """
            SELECT 
                COUNT(DISTINCT ay.acad_year_id) as total_years,
                COUNT(DISTINCT CASE WHEN at.is_current = 1 THEN at.acad_term_id END) as active_terms,
                COUNT(DISTINCT ep.period_id) as total_periods,
                (SELECT year_code FROM academic_years WHERE is_current = 1 LIMIT 1) as current_year
            FROM academic_years ay
            LEFT JOIN academic_terms at ON ay.acad_year_id = at.acad_year_id
            LEFT JOIN evaluation_periods ep ON at.acad_term_id = ep.acad_term_id
        """
        cursor.execute(stats_query)
        stats = cursor.fetchone()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': academic_years,
            'stats': stats
        })
        
    except Exception as e:
        print(f"Error getting academic years: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/academic-years', methods=['POST'])
@login_required
def create_academic_year():
    """Create a new academic term"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Validate required fields
        if not data.get('year_code') or not data.get('start_date') or not data.get('end_date'):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        # If setting as active, deactivate all other years and terms first
        if data.get('is_current'):
            cursor.execute("UPDATE academic_years SET is_current = 0")
            cursor.execute("UPDATE academic_terms SET is_current = 0")
        
        # Insert new academic year
        insert_query = """
            INSERT INTO academic_years (year_code, year_name, start_date, end_date, is_current)
            VALUES (%s, %s, %s, %s, %s)
        """
        year_name = f"Academic Year {data['year_code']}"
        cursor.execute(insert_query, (
            data['year_code'],
            year_name,
            data['start_date'],
            data['end_date'],
            1 if data.get('is_current') else 0
        ))
        
        acad_year_id = cursor.lastrowid
        
        # Create default semesters (1st and 2nd Semester)
        from datetime import datetime, timedelta
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d')
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d')
        mid_date = start_date + timedelta(days=150)  # ~5 months
        
        # 1st Semester
        cursor.execute("""
            INSERT INTO academic_terms 
            (acad_year_id, term_name, term_code, start_date, end_date, is_current)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (acad_year_id, '1st Semester', '1ST_SEM', data['start_date'], 
              mid_date.strftime('%Y-%m-%d'), 1 if data.get('is_current') else 0))
        
        # 2nd Semester
        cursor.execute("""
            INSERT INTO academic_terms 
            (acad_year_id, term_name, term_code, start_date, end_date, is_current)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (acad_year_id, '2nd Semester', '2ND_SEM', 
              (mid_date + timedelta(days=1)).strftime('%Y-%m-%d'), 
              data['end_date'], 0))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Academic year and semesters created successfully',
            'acad_year_id': acad_year_id
        })
        
    except Exception as e:
        print(f"Error creating academic year: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/academic-years/<int:acad_year_id>', methods=['PUT'])
@login_required
def update_academic_year(acad_year_id):
    """Update an existing academic year"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Check if year exists
        cursor.execute("SELECT acad_year_id FROM academic_years WHERE acad_year_id = %s", (acad_year_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Academic year not found'}), 404
        
        # If setting as current, deactivate all other years and terms first
        if data.get('is_current') == 1:
            cursor.execute("UPDATE academic_years SET is_current = 0")
            cursor.execute("UPDATE academic_terms SET is_current = 0")
            # Activate the first semester of this year
            cursor.execute("""
                UPDATE academic_terms SET is_current = 1 
                WHERE acad_year_id = %s AND term_code = '1ST_SEM'
            """, (acad_year_id,))
        
        # Build update query dynamically
        update_fields = []
        values = []
        
        if 'year_code' in data:
            update_fields.append("year_code = %s")
            values.append(data['year_code'])
            update_fields.append("year_name = %s")
            values.append(f"Academic Year {data['year_code']}")
        if 'start_date' in data:
            update_fields.append("start_date = %s")
            values.append(data['start_date'])
        if 'end_date' in data:
            update_fields.append("end_date = %s")
            values.append(data['end_date'])
        if 'is_current' in data:
            update_fields.append("is_current = %s")
            values.append(data['is_current'])
        
        if not update_fields:
            return jsonify({'success': False, 'error': 'No fields to update'}), 400
        
        values.append(acad_year_id)
        update_query = f"""
            UPDATE academic_years 
            SET {', '.join(update_fields)}
            WHERE acad_year_id = %s
        """
        
        cursor.execute(update_query, values)
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Academic year updated successfully'
        })
        
    except Exception as e:
        print(f"Error updating academic year: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/academic-years/<int:acad_year_id>', methods=['DELETE'])
@login_required
def delete_academic_year(acad_year_id):
    """Delete an academic year (cascade deletes terms and periods)"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if year exists
        cursor.execute("SELECT acad_year_id, is_current FROM academic_years WHERE acad_year_id = %s", (acad_year_id,))
        year = cursor.fetchone()
        if not year:
            return jsonify({'success': False, 'error': 'Academic year not found'}), 404
        
        # Don't allow deleting current year
        if year['is_current']:
            return jsonify({'success': False, 'error': 'Cannot delete the current academic year'}), 400
        
        # Check for associated sections
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM class_sections cs
            INNER JOIN academic_terms at ON cs.acad_term_id = at.acad_term_id
            WHERE at.acad_year_id = %s
        """, (acad_year_id,))
        if cursor.fetchone()['count'] > 0:
            return jsonify({'success': False, 'error': 'Cannot delete academic year with existing class sections'}), 400
        
        # Foreign keys will cascade delete academic_terms and their evaluation_periods
        cursor.execute("DELETE FROM academic_years WHERE acad_year_id = %s", (acad_year_id,))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Academic year deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting academic year: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/academic-years/current-overview')
@login_required  
def get_current_year_overview():
    """Get detailed overview of the current academic year"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get current academic year
        cursor.execute("""
            SELECT * FROM academic_years 
            WHERE is_current = 1 
            LIMIT 1
        """)
        current_year = cursor.fetchone()
        
        if not current_year:
            return jsonify({'success': False, 'error': 'No current academic year set'}), 404
        
        # Get evaluation periods for all terms in current year
        cursor.execute("""
            SELECT 
                ep.*,
                at.term_name,
                at.term_code,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as completed_evaluations
            FROM evaluation_periods ep
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            LEFT JOIN evaluations e ON ep.period_id = e.period_id
            WHERE at.acad_year_id = %s
            GROUP BY ep.period_id
            ORDER BY at.term_code, ep.start_date
        """, (current_year['acad_year_id'],))
        evaluation_periods = cursor.fetchall()
        
        # Get statistics for current year
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT s.id) as total_students,
                COUNT(DISTINCT f.faculty_id) as active_faculty,
                COUNT(DISTINCT sub.subject_id) as total_courses,
                CASE 
                    WHEN COUNT(DISTINCT e.evaluation_id) > 0 
                    THEN ROUND((COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) * 100.0 / COUNT(DISTINCT e.evaluation_id)), 1)
                    ELSE 0 
                END as evaluation_rate
            FROM academic_years ay
            LEFT JOIN academic_terms at ON ay.acad_year_id = at.acad_year_id
            LEFT JOIN class_sections cs ON at.acad_term_id = cs.acad_term_id
            LEFT JOIN std_info s ON s.is_archived = 0
            LEFT JOIN faculty f ON f.is_archived = 0
            LEFT JOIN subjects sub ON cs.subject_id = sub.subject_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
            WHERE ay.acad_year_id = %s
        """, (current_year['acad_year_id'],))
        statistics = cursor.fetchone()
        
        # Add display status and color for each evaluation period
        from datetime import datetime
        now = datetime.now()
        for period in evaluation_periods:
            start_date = period['start_date']
            end_date = period['end_date']
            
            # Convert to datetime if they're date objects
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
            if isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S')
            
            if period['status'] == 'Active':
                period['display_status'] = 'Active'
                period['status_color'] = 'green'
            elif period['status'] == 'Closed':
                period['display_status'] = 'Closed'
                period['status_color'] = 'blue'
            elif now < start_date:
                period['display_status'] = 'Upcoming'
                period['status_color'] = 'gray'
            elif now > end_date:
                period['display_status'] = 'Completed'
                period['status_color'] = 'blue'
            else:
                period['display_status'] = 'Pending'
                period['status_color'] = 'yellow'
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'current_year': current_year,
            'evaluation_periods': evaluation_periods,
            'statistics': statistics
        })
        
    except Exception as e:
        print(f"Error getting current year overview: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ============================================================
# ACADEMIC TERMS (SEMESTERS) MANAGEMENT ENDPOINTS
# ============================================================

@api_bp.route('/academic-terms/<int:acad_year_id>')
@login_required
def get_academic_terms(acad_year_id):
    """Get all academic terms/semesters for a specific academic year"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all terms for this academic year
        cursor.execute("""
            SELECT 
                at.acad_term_id,
                at.acad_year_id,
                at.term_name,
                at.term_code,
                at.start_date,
                at.end_date,
                at.is_current,
                ay.year_code,
                COUNT(DISTINCT ep.period_id) as evaluation_periods_count
            FROM academic_terms at
            INNER JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            LEFT JOIN evaluation_periods ep ON at.acad_term_id = ep.acad_term_id
            WHERE at.acad_year_id = %s
            GROUP BY at.acad_term_id
            ORDER BY at.start_date ASC
        """, (acad_year_id,))
        
        terms = cursor.fetchall()
        cursor.close()
        
        # Convert dates to strings
        for term in terms:
            if term.get('start_date'):
                term['start_date'] = str(term['start_date'])
            if term.get('end_date'):
                term['end_date'] = str(term['end_date'])
        
        return jsonify({
            'success': True,
            'data': terms
        })
        
    except Exception as e:
        print(f"Error getting academic terms: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/academic-terms', methods=['POST'])
@login_required
def create_academic_term():
    """Create a new academic term/semester"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['acad_year_id', 'term_name', 'term_code', 'start_date', 'end_date']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        cursor = conn.cursor(dictionary=True)
        
        # Check if term already exists for this year
        cursor.execute("""
            SELECT acad_term_id FROM academic_terms
            WHERE acad_year_id = %s AND term_code = %s
        """, (data['acad_year_id'], data['term_code']))
        
        existing_term = cursor.fetchone()
        if existing_term:
            cursor.close()
            return jsonify({'success': False, 'error': 'A term with this code already exists for this academic year'}), 400
        
        # Insert new term
        cursor.execute("""
            INSERT INTO academic_terms 
            (acad_year_id, term_name, term_code, start_date, end_date, is_current)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            data['acad_year_id'],
            data['term_name'],
            data['term_code'],
            data['start_date'],
            data['end_date'],
            data.get('is_current', 0)
        ))
        
        acad_term_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Academic term created successfully',
            'acad_term_id': acad_term_id
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error creating academic term: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/academic-terms/<int:acad_term_id>', methods=['PUT'])
@login_required
def update_academic_term(acad_term_id):
    """Update an academic term/semester"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Build update query dynamically
        update_fields = []
        update_values = []
        
        allowed_fields = ['term_name', 'term_code', 'start_date', 'end_date', 'is_current']
        for field in allowed_fields:
            if field in data:
                update_fields.append(f"{field} = %s")
                update_values.append(data[field])
        
        if not update_fields:
            cursor.close()
            return jsonify({'success': False, 'error': 'No fields to update'}), 400
        
        update_values.append(acad_term_id)
        
        cursor.execute(f"""
            UPDATE academic_terms 
            SET {', '.join(update_fields)}
            WHERE acad_term_id = %s
        """, update_values)
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Academic term updated successfully'
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error updating academic term: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/academic-terms/<int:acad_term_id>', methods=['DELETE'])
@login_required
def delete_academic_term(acad_term_id):
    """Delete an academic term/semester"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if term has evaluation periods
        cursor.execute("""
            SELECT COUNT(*) as count FROM evaluation_periods
            WHERE acad_term_id = %s
        """, (acad_term_id,))
        
        result = cursor.fetchone()
        if result['count'] > 0:
            cursor.close()
            return jsonify({'success': False, 'error': 'Cannot delete term with existing evaluation periods'}), 400
        
        # Delete the term
        cursor.execute("DELETE FROM academic_terms WHERE acad_term_id = %s", (acad_term_id,))
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Academic term deleted successfully'
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error deleting academic term: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ============================================================
# EVALUATION PERIODS MANAGEMENT ENDPOINTS
# ============================================================

def update_period_status_auto():
    """Automatically update evaluation period status based on dates"""
    conn = get_db_connection()
    if not conn:
        return
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get periods that will transition to 'Active' (for email notifications)
        cursor.execute("""
            SELECT period_id, title, start_date, end_date
            FROM evaluation_periods 
            WHERE CURDATE() >= start_date 
            AND CURDATE() <= end_date 
            AND status != 'Active'
        """)
        periods_to_activate = cursor.fetchall()
        
        # Update status to 'Active' for periods that have started but not ended
        cursor.execute("""
            UPDATE evaluation_periods 
            SET status = 'Active'
            WHERE CURDATE() >= start_date 
            AND CURDATE() <= end_date 
            AND status != 'Active'
        """)
        
        # Update status to 'Closed' for periods that have ended
        # Database schema uses: 'Pending', 'Active', 'Closed', 'Canceled'
        cursor.execute("""
            UPDATE evaluation_periods 
            SET status = 'Closed'
            WHERE CURDATE() > end_date 
            AND status != 'Closed'
            AND status != 'Canceled'
        """)
        
        # Update status to 'Pending' for periods that haven't started
        cursor.execute("""
            UPDATE evaluation_periods 
            SET status = 'Pending'
            WHERE CURDATE() < start_date 
            AND status NOT IN ('Pending', 'Canceled')
        """)
        
        conn.commit()
        
        # Automatic email notifications removed - use manual "Send Email Notifications" button
        if periods_to_activate:
            print(f"✅ {len(periods_to_activate)} period(s) automatically activated.")
            print(f"📧 Use the 'Send Email Notifications' button to notify users.")
            
            # Still sync evaluations for newly activated periods
            for period in periods_to_activate:
                try:
                    cursor.execute("""
                        INSERT INTO evaluations (period_id, section_id, student_id, status, created_at)
                        SELECT DISTINCT
                            %s as period_id,
                            cs.section_id,
                            ss.student_id,
                            'Pending' as status,
                            NOW() as created_at
                        FROM section_students ss
                        INNER JOIN class_sections cs ON (
                            ss.section_id = cs.section_id 
                            OR ss.section_id = cs.section_ref_id
                        )
                        WHERE ss.status = 'Active'
                        AND cs.faculty_id IS NOT NULL
                        AND NOT EXISTS (
                            SELECT 1 FROM evaluations ev 
                            WHERE ev.period_id = %s 
                            AND ev.section_id = cs.section_id 
                            AND ev.student_id = ss.student_id
                        )
                    """, (period['period_id'], period['period_id']))
                    
                    evaluations_created = cursor.rowcount
                    conn.commit()
                    print(f"Synced evaluations for period {period['period_id']}: {evaluations_created} evaluation(s) created")
                except Exception as sync_error:
                    print(f"Error syncing evaluations for period {period['period_id']}: {sync_error}")
        
        cursor.close()
    except Exception as e:
        print(f"Error updating period status: {e}")
        conn.rollback()
    finally:
        conn.close()


@api_bp.route('/evaluation-periods-admin')
@login_required
def get_evaluation_periods_admin():
    """Get all evaluation periods with statistics for admin management"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Auto-update statuses based on dates
    update_period_status_auto()
    
    # Get filter parameter
    academic_year_id = request.args.get('academic_year_id', type=int)
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Build WHERE clause based on filters
        where_conditions = ["COALESCE(ep.is_archived, 0) = 0"]
        query_params = []
        
        if academic_year_id:
            where_conditions.append("ay.acad_year_id = %s")
            query_params.append(academic_year_id)
        
        where_clause = " AND ".join(where_conditions)
        
        # Get all non-archived evaluation periods with term info and statistics
        query = f"""
            SELECT 
                ep.*,
                at.term_name as term_title,
                at.term_code,
                ay.year_code,
                ay.acad_year_id,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE 
                    WHEN e.status = 'Completed' OR e.completion_time IS NOT NULL 
                    THEN e.evaluation_id 
                END) as completed_evaluations,
                COUNT(DISTINCT CASE WHEN e.status = 'Pending' THEN e.evaluation_id END) as pending_evaluations,
                COUNT(DISTINCT e.student_id) as total_students,
                DATEDIFF(ep.end_date, CURDATE()) as days_remaining,
                CASE 
                    WHEN CURDATE() < ep.start_date THEN 'Pending'
                    WHEN CURDATE() > ep.end_date THEN 'Closed'
                    WHEN CURDATE() BETWEEN ep.start_date AND ep.end_date THEN 'Active'
                    ELSE ep.status
                END as computed_status
            FROM evaluation_periods ep
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            LEFT JOIN evaluations e ON ep.period_id = e.period_id
            WHERE {where_clause}
            GROUP BY ep.period_id
            ORDER BY ep.start_date DESC
        """
        cursor.execute(query, query_params)
        periods = cursor.fetchall()
        
        # Get overall statistics (excluding archived)
        stats_query = """
            SELECT 
                COUNT(DISTINCT ep.period_id) as total_periods,
                COUNT(DISTINCT CASE WHEN ep.status = 'Active' THEN ep.period_id END) as active_periods,
                COUNT(DISTINCT CASE WHEN ep.status = 'Closed' THEN ep.period_id END) as completed_periods,
                COUNT(DISTINCT CASE WHEN ep.status = 'Pending' THEN ep.period_id END) as upcoming_periods,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE 
                    WHEN e.status = 'Completed' OR e.completion_time IS NOT NULL 
                    THEN e.evaluation_id 
                END) as completed_evaluations
            FROM evaluation_periods ep
            LEFT JOIN evaluations e ON ep.period_id = e.period_id
            WHERE COALESCE(ep.is_archived, 0) = 0
        """
        cursor.execute(stats_query)
        stats = cursor.fetchone()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': periods,
            'stats': {
                'total_periods': stats['total_periods'] or 0,
                'active_periods': stats['active_periods'] or 0,
                'completed_periods': stats['completed_periods'] or 0,
                'upcoming_periods': stats['upcoming_periods'] or 0,
                'total_evaluations': stats['total_evaluations'] or 0,
                'completed_evaluations': stats['completed_evaluations'] or 0
            }
        })
        
    except Exception as e:
        print(f"Error getting evaluation periods: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/evaluation-periods-admin', methods=['POST'])
@login_required
def create_evaluation_period_admin():
    """Create a new evaluation period"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Validate required fields
        required_fields = ['acad_term_id', 'title', 'start_date', 'end_date']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        # Validate dates
        from datetime import datetime
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d')
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d')
        
        if end_date <= start_date:
            return jsonify({'success': False, 'error': 'End date must be after start date'}), 400
        
        # Determine initial status based on dates
        today = datetime.now().date()
        if start_date.date() > today:
            status = 'Upcoming'
        elif start_date.date() <= today <= end_date.date():
            status = 'Active'
        else:
            status = 'Completed'
        
        # Check for overlapping periods
        cursor.execute("""
            SELECT period_id, title 
            FROM evaluation_periods 
            WHERE acad_term_id = %s 
            AND (
                (start_date <= %s AND end_date >= %s) OR
                (start_date <= %s AND end_date >= %s) OR
                (start_date >= %s AND end_date <= %s)
            )
        """, (data['acad_term_id'], data['start_date'], data['start_date'], 
              data['end_date'], data['end_date'], data['start_date'], data['end_date']))
        
        overlapping = cursor.fetchone()
        if overlapping:
            return jsonify({
                'success': False, 
                'error': f'Period overlaps with existing period: {overlapping["title"]}'
            }), 400
        
        # Insert new evaluation period
        insert_query = """
            INSERT INTO evaluation_periods 
            (acad_term_id, title, start_date, end_date, status, created_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
        """
        cursor.execute(insert_query, (
            data['acad_term_id'],
            data['title'],
            data['start_date'],
            data['end_date'],
            status
        ))
        
        period_id = cursor.lastrowid
        
        # Auto-sync evaluations for all assigned students
        # This handles both direct section_id matches and section_ref_id relationships
        try:
            cursor.execute("""
                INSERT INTO evaluations (period_id, section_id, student_id, status, created_at)
                SELECT DISTINCT
                    %s as period_id,
                    cs.section_id,
                    ss.student_id,
                    'Pending' as status,
                    NOW() as created_at
                FROM section_students ss
                INNER JOIN class_sections cs ON (
                    ss.section_id = cs.section_id 
                    OR ss.section_id = cs.section_ref_id
                )
                WHERE ss.status = 'Active'
                AND cs.faculty_id IS NOT NULL
                AND NOT EXISTS (
                    SELECT 1 FROM evaluations ev 
                    WHERE ev.period_id = %s 
                    AND ev.section_id = cs.section_id 
                    AND ev.student_id = ss.student_id
                )
            """, (period_id, period_id))
            
            evaluations_created = cursor.rowcount
            conn.commit()
        except Exception as sync_error:
            print(f"Error syncing evaluations: {sync_error}")
            evaluations_created = 0
        
        # Email notifications removed - use manual "Send Email Notifications" button instead
        print(f"✅ Evaluation period created with status: {status}")
        print(f"📧 Automatic email notifications disabled. Use 'Send Email Notifications' button.")
        
        cursor.close()
        
        response_data = {
            'success': True,
            'message': 'Evaluation period created successfully',
            'period_id': period_id,
            'status': status,
            'evaluations_created': evaluations_created
        }
        
        # Note: Email notifications are sent in background thread
        # No job tracking in simplified version
        
        return jsonify(response_data)
        
    except ValueError as e:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        print(f"Error creating evaluation period: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            conn.close()


@api_bp.route('/evaluation-periods-admin/<int:period_id>', methods=['PUT'])
@login_required
def update_evaluation_period_admin(period_id):
    """Update an existing evaluation period"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Check if period exists
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            return jsonify({'success': False, 'error': 'Evaluation period not found'}), 404
        
        # Validate dates if provided
        if data.get('start_date') and data.get('end_date'):
            from datetime import datetime
            start_date = datetime.strptime(data['start_date'], '%Y-%m-%d')
            end_date = datetime.strptime(data['end_date'], '%Y-%m-%d')
            
            if end_date <= start_date:
                return jsonify({'success': False, 'error': 'End date must be after start date'}), 400
        
        # Build update query dynamically
        update_fields = []
        update_values = []
        
        allowed_fields = ['acad_term_id', 'title', 'start_date', 'end_date', 'status']
        for field in allowed_fields:
            if field in data:
                update_fields.append(f"{field} = %s")
                update_values.append(data[field])
        
        if not update_fields:
            return jsonify({'success': False, 'error': 'No fields to update'}), 400
        
        update_values.append(period_id)
        update_query = f"""
            UPDATE evaluation_periods 
            SET {', '.join(update_fields)}, updated_at = NOW()
            WHERE period_id = %s
        """
        
        cursor.execute(update_query, update_values)
        conn.commit()
        
        # Auto-update status if dates changed
        update_period_status_auto()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation period updated successfully'
        })
        
    except ValueError as e:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        print(f"Error updating evaluation period: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/evaluation-periods-admin/<int:period_id>/archive', methods=['PUT'])
@login_required
def archive_evaluation_period_admin(period_id):
    """Archive an evaluation period (mark as archived)"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if period exists
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            return jsonify({'success': False, 'error': 'Evaluation period not found'}), 404
        
        # Don't allow archiving active periods
        if period['status'] == 'Active':
            return jsonify({'success': False, 'error': 'Cannot archive an active evaluation period'}), 400
        
        # Add is_archived column if it doesn't exist
        try:
            cursor.execute("ALTER TABLE evaluation_periods ADD COLUMN is_archived TINYINT(1) NOT NULL DEFAULT 0")
            conn.commit()
        except Exception:
            # Column might already exist, ignore error
            pass
        
        # Archive the period (mark as archived)
        cursor.execute("UPDATE evaluation_periods SET is_archived = 1 WHERE period_id = %s", (period_id,))
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation period archived successfully'
        })
        
    except Exception as e:
        print(f"Error archiving evaluation period: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/evaluation-periods-archived')
@login_required
def get_archived_evaluation_periods():
    """Get all archived evaluation periods"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get archived evaluation periods with term info and statistics
        query = """
            SELECT 
                ep.*,
                CONCAT(ay.year_code, ' - ', at.term_name) as term_title,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE 
                    WHEN e.status = 'Completed' OR e.completion_time IS NOT NULL 
                    THEN e.evaluation_id 
                END) as completed_evaluations,
                COUNT(DISTINCT CASE WHEN e.status = 'Pending' THEN e.evaluation_id END) as pending_evaluations,
                COUNT(DISTINCT e.student_id) as total_students
            FROM evaluation_periods ep
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            LEFT JOIN evaluations e ON ep.period_id = e.period_id
            WHERE ep.is_archived = 1
            GROUP BY ep.period_id
            ORDER BY ep.updated_at DESC
        """
        cursor.execute(query)
        periods = cursor.fetchall()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': periods
        })
        
    except Exception as e:
        print(f"Error getting archived evaluation periods: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/evaluation-periods-admin/<int:period_id>/unarchive', methods=['PUT'])
@login_required
def unarchive_evaluation_period_admin(period_id):
    """Unarchive (restore) an evaluation period"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if period exists and is archived
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s AND is_archived = 1", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            return jsonify({'success': False, 'error': 'Archived evaluation period not found'}), 404
        
        # Unarchive the period
        cursor.execute("UPDATE evaluation_periods SET is_archived = 0 WHERE period_id = %s", (period_id,))
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation period restored successfully'
        })
        
    except Exception as e:
        print(f"Error unarchiving evaluation period: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/evaluation-periods-admin/<int:period_id>', methods=['DELETE'])
@login_required
def delete_evaluation_period_admin(period_id):
    """Delete an evaluation period (only if no evaluations exist)"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if period exists
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            return jsonify({'success': False, 'error': 'Evaluation period not found'}), 404
        
        # Check if period has evaluations
        cursor.execute(
            "SELECT COUNT(*) as count FROM evaluations WHERE period_id = %s", 
            (period_id,)
        )
        eval_count = cursor.fetchone()['count']
        
        if eval_count > 0:
            return jsonify({
                'success': False, 
                'error': f'Cannot delete period with {eval_count} existing evaluation(s)'
            }), 400
        
        # Delete the period
        cursor.execute("DELETE FROM evaluation_periods WHERE period_id = %s", (period_id,))
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation period deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting evaluation period: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sync-evaluations/<int:period_id>', methods=['POST'])
@login_required
def sync_evaluations_for_period(period_id):
    """Manually sync/create evaluations for all enrolled students in a specific period"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if period exists
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            return jsonify({'success': False, 'error': 'Evaluation period not found'}), 404
        
        # Create evaluations for all assigned students who don't have one yet
        cursor.execute("""
            INSERT INTO evaluations (period_id, section_id, student_id, status, created_at)
            SELECT DISTINCT
                %s as period_id,
                cs.section_id,
                ss.student_id,
                'Pending' as status,
                NOW() as created_at
            FROM section_students ss
            INNER JOIN class_sections cs ON (
                ss.section_id = cs.section_id 
                OR ss.section_id = cs.section_ref_id
            )
            WHERE ss.status = 'Active'
            AND cs.faculty_id IS NOT NULL
            AND NOT EXISTS (
                SELECT 1 FROM evaluations ev 
                WHERE ev.period_id = %s 
                AND ev.section_id = cs.section_id 
                AND ev.student_id = ss.student_id
            )
        """, (period_id, period_id))
        
        evaluations_created = cursor.rowcount
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': f'Successfully synced evaluations. {evaluations_created} new evaluation(s) created.',
            'evaluations_created': evaluations_created
        })
        
    except Exception as e:
        print(f"Error syncing evaluations: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sync-all-evaluations', methods=['POST'])
@login_required
def sync_all_evaluations():
    """Sync evaluations for all active evaluation periods"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all active or upcoming periods
        cursor.execute("""
            SELECT period_id FROM evaluation_periods 
            WHERE status IN ('Active', 'Upcoming')
        """)
        periods = cursor.fetchall()
        
        total_created = 0
        
        for period in periods:
            period_id = period['period_id']
            
            # Create evaluations for all assigned students who don't have one yet
            cursor.execute("""
                INSERT INTO evaluations (period_id, section_id, student_id, status, created_at)
                SELECT DISTINCT
                    %s as period_id,
                    cs.section_id,
                    ss.student_id,
                    'Pending' as status,
                    NOW() as created_at
                FROM section_students ss
                INNER JOIN class_sections cs ON (
                    ss.section_id = cs.section_id 
                    OR ss.section_id = cs.section_ref_id
                )
                WHERE ss.status = 'Active'
                AND cs.faculty_id IS NOT NULL
                AND NOT EXISTS (
                    SELECT 1 FROM evaluations ev 
                    WHERE ev.period_id = %s 
                    AND ev.section_id = cs.section_id 
                    AND ev.student_id = ss.student_id
                )
            """, (period_id, period_id))
            
            total_created += cursor.rowcount
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': f'Successfully synced all evaluations. {total_created} new evaluation(s) created across {len(periods)} period(s).',
            'evaluations_created': total_created,
            'periods_synced': len(periods)
        })
        
    except Exception as e:
        print(f"Error syncing all evaluations: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/test-sync/<int:period_id>', methods=['GET'])
@login_required
def test_sync_evaluations(period_id):
    """Test what evaluations would be created without actually creating them"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get period info
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            return jsonify({'success': False, 'error': 'Evaluation period not found'}), 404
        
        # Test query - shows what WOULD be created
        cursor.execute("""
            SELECT DISTINCT
                %s as period_id,
                cs.section_id,
                ss.student_id,
                'Pending' as status,
                CONCAT(si.std_Firstname, ' ', si.std_Surname) as student_name,
                cs.section_name,
                CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                s.subject_code,
                s.title as subject_title,
                ss.section_id as student_section_id,
                cs.section_ref_id
            FROM section_students ss
            INNER JOIN std_info si ON ss.student_id = si.id
            INNER JOIN class_sections cs ON (
                ss.section_id = cs.section_id 
                OR ss.section_id = cs.section_ref_id
            )
            INNER JOIN subjects s ON cs.subject_id = s.subject_id
            INNER JOIN faculty f ON cs.faculty_id = f.faculty_id
            WHERE ss.status = 'Active'
            AND cs.faculty_id IS NOT NULL
            AND NOT EXISTS (
                SELECT 1 FROM evaluations ev 
                WHERE ev.period_id = %s 
                AND ev.section_id = cs.section_id 
                AND ev.student_id = ss.student_id
            )
            ORDER BY student_name, subject_code
        """, (period_id, period_id))
        
        potential_evaluations = cursor.fetchall()
        cursor.close()
        
        return jsonify({
            'success': True,
            'period': {
                'id': period['period_id'],
                'title': period['title'],
                'status': period['status']
            },
            'potential_evaluations': potential_evaluations,
            'count': len(potential_evaluations),
            'message': f'Would create {len(potential_evaluations)} evaluation(s) if synced'
        })
        
    except Exception as e:
        print(f"Error testing sync: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ============================================================================
# GUIDANCE EVALUATION PERIOD MANAGEMENT (GUIDANCE-SPECIFIC ENDPOINTS)
# ============================================================================

@api_bp.route('/guidance/evaluation-periods')
@login_required
def get_evaluation_periods_guidance():
    """Get all evaluation periods with statistics for guidance management"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Auto-update statuses based on dates
    try:
        update_period_status_auto()
    except Exception as update_error:
        pass  # Non-fatal error, continue
    
    # Get filter parameter
    academic_year_id = request.args.get('academic_year_id', type=int)
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Build WHERE clause based on filters
        where_conditions = ["COALESCE(ep.is_archived, 0) = 0"]
        query_params = []
        
        if academic_year_id:
            where_conditions.append("ay.acad_year_id = %s")
            query_params.append(academic_year_id)
        
        where_clause = " AND ".join(where_conditions)
        
        # Get all non-archived evaluation periods with term info and statistics
        query = f"""
            SELECT 
                ep.*,
                at.term_name as term_title,
                at.term_code,
                ay.year_code,
                ay.acad_year_id,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE 
                    WHEN e.status = 'Completed' OR e.completion_time IS NOT NULL 
                    THEN e.evaluation_id 
                END) as completed_evaluations,
                COUNT(DISTINCT CASE WHEN e.status = 'Pending' THEN e.evaluation_id END) as pending_evaluations,
                COUNT(DISTINCT e.student_id) as total_students,
                DATEDIFF(ep.end_date, CURDATE()) as days_remaining,
                CASE 
                    WHEN CURDATE() < ep.start_date THEN 'Pending'
                    WHEN CURDATE() > ep.end_date THEN 'Closed'
                    WHEN CURDATE() BETWEEN ep.start_date AND ep.end_date THEN 'Active'
                    ELSE ep.status
                END as computed_status
            FROM evaluation_periods ep
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            LEFT JOIN evaluations e ON ep.period_id = e.period_id
            WHERE {where_clause}
            GROUP BY ep.period_id
            ORDER BY ep.start_date DESC
        """
        cursor.execute(query, query_params)
        
        periods = cursor.fetchall()
        
        # Get overall statistics (excluding archived)
        stats_query = """
            SELECT 
                COUNT(DISTINCT ep.period_id) as total_periods,
                COUNT(DISTINCT CASE WHEN ep.status = 'Active' THEN ep.period_id END) as active_periods,
                COUNT(DISTINCT CASE WHEN ep.status = 'Closed' THEN ep.period_id END) as completed_periods,
                COUNT(DISTINCT CASE WHEN ep.status = 'Pending' THEN ep.period_id END) as upcoming_periods,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE 
                    WHEN e.status = 'Completed' OR e.completion_time IS NOT NULL 
                    THEN e.evaluation_id 
                END) as completed_evaluations
            FROM evaluation_periods ep
            LEFT JOIN evaluations e ON ep.period_id = e.period_id
            WHERE COALESCE(ep.is_archived, 0) = 0
        """
        cursor.execute(stats_query)
        
        stats = cursor.fetchone()
        
        if stats is None:
            stats = {
                'total_periods': 0,
                'active_periods': 0,
                'completed_periods': 0,
                'upcoming_periods': 0,
                'total_evaluations': 0,
                'completed_evaluations': 0
            }
        
        cursor.close()
        
        stats_dict = {
            'total_periods': stats.get('total_periods', 0) or 0,
            'active_periods': stats.get('active_periods', 0) or 0,
            'completed_periods': stats.get('completed_periods', 0) or 0,
            'upcoming_periods': stats.get('upcoming_periods', 0) or 0,
            'total_evaluations': stats.get('total_evaluations', 0) or 0,
            'completed_evaluations': stats.get('completed_evaluations', 0) or 0
        }
        
        return jsonify({
            'success': True,
            'data': periods,
            'stats': stats_dict
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/guidance/evaluation-periods', methods=['POST'])
@login_required
def create_evaluation_period_guidance():
    """Create a new evaluation period (Guidance version)"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Validate required fields
        required_fields = ['acad_term_id', 'title', 'start_date', 'end_date']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        # Validate dates
        from datetime import datetime
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d')
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d')
        
        if end_date <= start_date:
            return jsonify({'success': False, 'error': 'End date must be after start date'}), 400
        
        # Determine initial status based on dates
        today = datetime.now().date()
        if start_date.date() > today:
            status = 'Upcoming'
        elif start_date.date() <= today <= end_date.date():
            status = 'Active'
        else:
            status = 'Completed'
        
        # Check for overlapping periods
        cursor.execute("""
            SELECT period_id, title 
            FROM evaluation_periods 
            WHERE acad_term_id = %s 
            AND (
                (start_date <= %s AND end_date >= %s) OR
                (start_date <= %s AND end_date >= %s) OR
                (start_date >= %s AND end_date <= %s)
            )
        """, (data['acad_term_id'], data['start_date'], data['start_date'], 
              data['end_date'], data['end_date'], data['start_date'], data['end_date']))
        
        overlapping = cursor.fetchone()
        if overlapping:
            return jsonify({
                'success': False, 
                'error': f'Period overlaps with existing period: {overlapping["title"]}'
            }), 400
        
        # Insert new evaluation period
        insert_query = """
            INSERT INTO evaluation_periods 
            (acad_term_id, title, start_date, end_date, status, created_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
        """
        cursor.execute(insert_query, (
            data['acad_term_id'],
            data['title'],
            data['start_date'],
            data['end_date'],
            status
        ))
        
        period_id = cursor.lastrowid
        
        # Auto-sync evaluations for all assigned students
        try:
            cursor.execute("""
                INSERT INTO evaluations (period_id, section_id, student_id, status, created_at)
                SELECT DISTINCT
                    %s as period_id,
                    cs.section_id,
                    ss.student_id,
                    'Pending' as status,
                    NOW() as created_at
                FROM section_students ss
                INNER JOIN class_sections cs ON (
                    ss.section_id = cs.section_id 
                    OR ss.section_id = cs.section_ref_id
                )
                WHERE ss.status = 'Active'
                AND cs.faculty_id IS NOT NULL
                AND NOT EXISTS (
                    SELECT 1 FROM evaluations ev 
                    WHERE ev.period_id = %s 
                    AND ev.section_id = cs.section_id 
                    AND ev.student_id = ss.student_id
                )
            """, (period_id, period_id))
            
            evaluations_created = cursor.rowcount
            conn.commit()
        except Exception as sync_error:
            print(f"Error syncing evaluations: {sync_error}")
            evaluations_created = 0
        
        # Email notifications removed - use manual "Send Email Notifications" button instead
        print(f"✅ Evaluation period created with status: {status}")
        print(f"📧 Automatic email notifications disabled. Use 'Send Email Notifications' button.")
        
        cursor.close()
        
        response_data = {
            'success': True,
            'message': 'Evaluation period created successfully',
            'period_id': period_id,
            'status': status,
            'evaluations_created': evaluations_created
        }
        
        return jsonify(response_data)
        
    except ValueError as e:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        print(f"Error creating evaluation period: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            conn.close()


@api_bp.route('/email-job-status/<job_id>', methods=['GET'])
@login_required
def get_email_job_status(job_id):
    """Get the status of an email sending job"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    if not hasattr(current_app, 'email_jobs') or job_id not in current_app.email_jobs:
        return jsonify({'success': False, 'error': 'Job not found'}), 404
    
    job_info = current_app.email_jobs[job_id]
    return jsonify({
        'success': True,
        'job_id': job_id,
        'status': job_info['status'],
        'total': job_info['total'],
        'sent': job_info['sent'],
        'failed': job_info['failed'],
        'progress_percentage': round((job_info['sent'] + job_info['failed']) / job_info['total'] * 100) if job_info['total'] > 0 else 0
    })


@api_bp.route('/send-evaluation-notifications/<int:period_id>', methods=['POST'])
@login_required
def send_evaluation_notifications(period_id):
    """Manually send email notifications for an evaluation period"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get the evaluation period details
        cursor.execute("""
            SELECT ep.*, ay.year_code, at.term_name
            FROM evaluation_periods ep
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE ep.period_id = %s
        """, (period_id,))
        
        period = cursor.fetchone()
        
        if not period:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Evaluation period not found'}), 404
        
        # Get all users with valid email addresses
        cursor.execute("""
            SELECT user_id, email, first_name, last_name, role
            FROM users
            WHERE email IS NOT NULL 
            AND email != ''
            AND email LIKE '%@%'
            AND role IN ('student', 'faculty', 'guidance', 'admin')
            AND is_active = 1
        """)
        
        all_users = cursor.fetchall()
        cursor.close()
        
        if not all_users:
            conn.close()
            return jsonify({
                'success': False,
                'error': 'No users with valid email addresses found'
            }), 400
        
        # Import email utility
        from utils.email_utils import send_evaluation_start_notification
        from threading import Thread
        import uuid
        import time
        
        # Generate unique job ID for tracking
        email_job_id = str(uuid.uuid4())
        
        # Initialize email jobs tracking if not exists
        if not hasattr(current_app, 'email_jobs'):
            current_app.email_jobs = {}
        
        current_app.email_jobs[email_job_id] = {
            'status': 'processing',
            'total': len(all_users),
            'sent': 0,
            'failed': 0,
            'started_at': time.time()
        }
        
        # Get the current app context for background thread
        app = current_app._get_current_object()
        
        # Function to send emails in background with app context
        def send_notification_emails():
            import concurrent.futures
            from threading import Lock
            
            stats_lock = Lock()
            
            def send_to_user(user):
                # Each thread needs its own app context
                with app.app_context():
                    try:
                        user_name = f"{user['first_name']} {user['last_name']}"
                        result = send_evaluation_start_notification(
                            user['email'],
                            user_name,
                            period['title'],
                            period['start_date'].strftime('%Y-%m-%d'),
                            period['end_date'].strftime('%Y-%m-%d')
                        )
                        with stats_lock:
                            if result:
                                app.email_jobs[email_job_id]['sent'] += 1
                                print(f"   ✅ Email sent: {user['email']}")
                            else:
                                app.email_jobs[email_job_id]['failed'] += 1
                                print(f"   ❌ Email failed: {user['email']}")
                        return result
                    except Exception as email_error:
                        with stats_lock:
                            app.email_jobs[email_job_id]['failed'] += 1
                        print(f"   ❌ Exception: {user['email']}: {str(email_error)}")
                        return False
            
            # Use ThreadPoolExecutor for parallel email sending (5 concurrent)
            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                list(executor.map(send_to_user, all_users))
            
            # Mark job as complete
            with app.app_context():
                app.email_jobs[email_job_id]['status'] = 'completed'
                app.email_jobs[email_job_id]['completed_at'] = time.time()
                print(f"\n📧 Email Summary: {app.email_jobs[email_job_id]['sent']} sent, {app.email_jobs[email_job_id]['failed']} failed")
        
        # Start email sending in background thread
        thread = Thread(target=send_notification_emails)
        thread.daemon = True
        thread.start()
        
        print(f"🚀 Started email notification job for period '{period['title']}' (Job ID: {email_job_id})")
        print(f"📊 Sending to {len(all_users)} users")
        
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Email notifications are being sent to {len(all_users)} users',
            'job_id': email_job_id,
            'total_recipients': len(all_users)
        })
        
    except Exception as e:
        print(f"Error sending evaluation notifications: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/guidance/evaluation-periods/<int:period_id>', methods=['PUT'])
@login_required
def update_evaluation_period_guidance(period_id):
    """Update an existing evaluation period (Guidance version)"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Check if period exists
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            return jsonify({'success': False, 'error': 'Evaluation period not found'}), 404
        
        # Validate dates if provided
        if data.get('start_date') and data.get('end_date'):
            from datetime import datetime
            start_date = datetime.strptime(data['start_date'], '%Y-%m-%d')
            end_date = datetime.strptime(data['end_date'], '%Y-%m-%d')
            
            if end_date <= start_date:
                return jsonify({'success': False, 'error': 'End date must be after start date'}), 400
        
        # Build update query dynamically
        update_fields = []
        update_values = []
        
        allowed_fields = ['acad_term_id', 'title', 'start_date', 'end_date', 'status']
        for field in allowed_fields:
            if field in data:
                update_fields.append(f"{field} = %s")
                update_values.append(data[field])
        
        if not update_fields:
            return jsonify({'success': False, 'error': 'No fields to update'}), 400
        
        update_values.append(period_id)
        update_query = f"""
            UPDATE evaluation_periods 
            SET {', '.join(update_fields)}, updated_at = NOW()
            WHERE period_id = %s
        """
        
        cursor.execute(update_query, update_values)
        conn.commit()
        
        # Auto-update status if dates changed
        update_period_status_auto()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation period updated successfully'
        })
        
    except ValueError as e:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        print(f"Error updating evaluation period: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/guidance/evaluation-periods/<int:period_id>/archive', methods=['PUT'])
@login_required
def archive_evaluation_period_guidance(period_id):
    """Archive an evaluation period (Guidance version)"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if period exists
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            return jsonify({'success': False, 'error': 'Evaluation period not found'}), 404
        
        # Don't allow archiving active periods
        if period['status'] == 'Active':
            return jsonify({'success': False, 'error': 'Cannot archive an active evaluation period'}), 400
        
        # Add is_archived column if it doesn't exist
        try:
            cursor.execute("ALTER TABLE evaluation_periods ADD COLUMN is_archived TINYINT(1) NOT NULL DEFAULT 0")
            conn.commit()
        except Exception:
            # Column might already exist, ignore error
            pass
        
        # Archive the period (mark as archived)
        cursor.execute("UPDATE evaluation_periods SET is_archived = 1 WHERE period_id = %s", (period_id,))
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation period archived successfully'
        })
        
    except Exception as e:
        print(f"Error archiving evaluation period: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/guidance/evaluation-periods-archived')
@login_required
def get_archived_evaluation_periods_guidance():
    """Get all archived evaluation periods (Guidance version)"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get archived evaluation periods with term info and statistics
        query = """
            SELECT 
                ep.*,
                CONCAT(ay.year_code, ' - ', at.term_name) as term_title,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE 
                    WHEN e.status = 'Completed' OR e.completion_time IS NOT NULL 
                    THEN e.evaluation_id 
                END) as completed_evaluations,
                COUNT(DISTINCT CASE WHEN e.status = 'Pending' THEN e.evaluation_id END) as pending_evaluations,
                COUNT(DISTINCT e.student_id) as total_students
            FROM evaluation_periods ep
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            LEFT JOIN evaluations e ON ep.period_id = e.period_id
            WHERE ep.is_archived = 1
            GROUP BY ep.period_id
            ORDER BY ep.updated_at DESC
        """
        cursor.execute(query)
        periods = cursor.fetchall()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': periods
        })
        
    except Exception as e:
        print(f"Error getting archived evaluation periods: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/guidance/evaluation-periods/<int:period_id>/unarchive', methods=['PUT'])
@login_required
def unarchive_evaluation_period_guidance(period_id):
    """Unarchive (restore) an evaluation period (Guidance version)"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if period exists and is archived
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s AND is_archived = 1", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            return jsonify({'success': False, 'error': 'Archived evaluation period not found'}), 404
        
        # Unarchive the period
        cursor.execute("UPDATE evaluation_periods SET is_archived = 0 WHERE period_id = %s", (period_id,))
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation period restored successfully'
        })
        
    except Exception as e:
        print(f"Error unarchiving evaluation period: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/guidance/evaluation-periods/<int:period_id>', methods=['DELETE'])
@login_required
def delete_evaluation_period_guidance(period_id):
    """Delete an evaluation period (Guidance version - only if no evaluations exist)"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if period exists
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            return jsonify({'success': False, 'error': 'Evaluation period not found'}), 404
        
        # Check if period has evaluations
        cursor.execute(
            "SELECT COUNT(*) as count FROM evaluations WHERE period_id = %s", 
            (period_id,)
        )
        eval_count = cursor.fetchone()['count']
        
        if eval_count > 0:
            return jsonify({
                'success': False, 
                'error': f'Cannot delete period with {eval_count} existing evaluation(s)'
            }), 400
        
        # Delete the period
        cursor.execute("DELETE FROM evaluation_periods WHERE period_id = %s", (period_id,))
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation period deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting evaluation period: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/guidance/sync-all-evaluations', methods=['POST'])
@login_required
def sync_all_evaluations_guidance():
    """Sync evaluations for all active evaluation periods (Guidance version)"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all active or upcoming periods
        cursor.execute("""
            SELECT period_id FROM evaluation_periods 
            WHERE status IN ('Active', 'Upcoming')
        """)
        periods = cursor.fetchall()
        
        total_created = 0
        
        for period in periods:
            period_id = period['period_id']
            
            # Create evaluations for all assigned students who don't have one yet
            cursor.execute("""
                INSERT INTO evaluations (period_id, section_id, student_id, status, created_at)
                SELECT DISTINCT
                    %s as period_id,
                    cs.section_id,
                    ss.student_id,
                    'Pending' as status,
                    NOW() as created_at
                FROM section_students ss
                INNER JOIN class_sections cs ON (
                    ss.section_id = cs.section_id 
                    OR ss.section_id = cs.section_ref_id
                )
                WHERE ss.status = 'Active'
                AND cs.faculty_id IS NOT NULL
                AND NOT EXISTS (
                    SELECT 1 FROM evaluations ev 
                    WHERE ev.period_id = %s 
                    AND ev.section_id = cs.section_id 
                    AND ev.student_id = ss.student_id
                )
            """, (period_id, period_id))
            
            total_created += cursor.rowcount
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': f'Successfully synced all evaluations. {total_created} new evaluation(s) created across {len(periods)} period(s).',
            'evaluations_created': total_created,
            'periods_synced': len(periods)
        })
        
    except Exception as e:
        print(f"Error syncing all evaluations: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ============================================================================
# SECTIONS MASTER TABLE MANAGEMENT
# ============================================================================

@api_bp.route('/sections-master')
def get_sections_master():
    """Get all active (non-disabled) sections from the sections master table - public access for signup"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all active (non-disabled) sections with program information and student count
        query = """
            SELECT 
                s.section_id,
                s.section_code,
                s.section_name,
                s.program_id,
                s.year_level,
                s.is_disable,
                s.created_at,
                s.updated_at,
                p.name as program_name,
                p.program_code,
                COUNT(DISTINCT cs.section_id) as class_count,
                COUNT(DISTINCT ss.student_id) as total_students
            FROM sections s
            LEFT JOIN programs p ON s.program_id = p.program_id
            LEFT JOIN class_sections cs ON s.section_id = cs.section_ref_id
            LEFT JOIN section_students ss ON s.section_id = ss.section_id AND ss.status = 'Active'
            WHERE s.is_disable = 0
            GROUP BY s.section_id
            ORDER BY s.program_id, s.year_level, s.section_code
        """
        cursor.execute(query)
        sections = cursor.fetchall()
        
        # Get statistics (only active sections)
        stats_query = """
            SELECT 
                COUNT(DISTINCT s.section_id) as total_sections,
                COUNT(DISTINCT s.program_id) as total_programs,
                COUNT(DISTINCT cs.section_id) as active_classes
            FROM sections s
            LEFT JOIN class_sections cs ON s.section_id = cs.section_ref_id
            WHERE s.is_disable = 0
        """
        cursor.execute(stats_query)
        stats = cursor.fetchone()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': sections,
            'stats': {
                'total_sections': stats['total_sections'] or 0,
                'total_programs': stats['total_programs'] or 0,
                'active_classes': stats['active_classes'] or 0
            }
        })
        
    except Exception as e:
        print(f"Error getting sections: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections-master', methods=['POST'])
@login_required
def create_section_master():
    """Create a new section in the master table"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['section_code', 'section_name', 'program_id', 'year_level']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        # Validate year_level
        year_level = int(data['year_level'])
        if year_level < 1 or year_level > 4:
            return jsonify({'success': False, 'error': 'Year level must be between 1 and 4'}), 400
        
        cursor = conn.cursor(dictionary=True)
        
        # Check if section_code already exists
        cursor.execute("SELECT section_id FROM sections WHERE section_code = %s", (data['section_code'],))
        if cursor.fetchone():
            cursor.close()
            return jsonify({'success': False, 'error': 'Section code already exists'}), 400
        
        # Insert new section
        insert_query = """
            INSERT INTO sections 
            (section_code, section_name, program_id, year_level)
            VALUES (%s, %s, %s, %s)
        """
        cursor.execute(insert_query, (
            data['section_code'],
            data['section_name'],
            data['program_id'],
            year_level
        ))
        
        conn.commit()
        section_id = cursor.lastrowid
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Section created successfully',
            'section_id': section_id
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error creating section: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections-master/<int:section_id>', methods=['PUT'])
@login_required
def update_section_master(section_id):
    """Update an existing section in the master table with class and student sync"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Check if section exists
        cursor.execute("SELECT * FROM sections WHERE section_id = %s", (section_id,))
        section = cursor.fetchone()
        
        if not section:
            cursor.close()
            return jsonify({'success': False, 'error': 'Section not found'}), 404
        
        # Build update query dynamically based on provided fields
        update_fields = []
        update_values = []
        
        if 'section_code' in data:
            # Check if new section_code already exists (excluding current section)
            cursor.execute(
                "SELECT section_id FROM sections WHERE section_code = %s AND section_id != %s", 
                (data['section_code'], section_id)
            )
            if cursor.fetchone():
                cursor.close()
                return jsonify({'success': False, 'error': 'Section code already exists'}), 400
            update_fields.append('section_code = %s')
            update_values.append(data['section_code'])
        
        if 'section_name' in data:
            update_fields.append('section_name = %s')
            update_values.append(data['section_name'])
        
        if 'program_id' in data:
            update_fields.append('program_id = %s')
            update_values.append(data['program_id'])
        
        if 'year_level' in data:
            year_level = int(data['year_level'])
            if year_level < 1 or year_level > 4:
                cursor.close()
                return jsonify({'success': False, 'error': 'Year level must be between 1 and 4'}), 400
            update_fields.append('year_level = %s')
            update_values.append(year_level)
        
        if update_fields:
            # Add section_id to values for WHERE clause
            update_values.append(section_id)
            
            # Execute update
            update_query = f"""
                UPDATE sections 
                SET {', '.join(update_fields)}
                WHERE section_id = %s
            """
            cursor.execute(update_query, update_values)
        
        # Handle class assignments and student sync
        students_synced = 0
        if 'class_ids' in data and isinstance(data['class_ids'], list):
            class_ids = data['class_ids']
            
            # Get current class assignments (section_id in class_sections is the class ID)
            cursor.execute("""
                SELECT section_id FROM class_sections WHERE section_ref_id = %s
            """, (section_id,))
            current_classes = [row['section_id'] for row in cursor.fetchall()]
            
            # Determine classes to add and remove
            classes_to_add = [cid for cid in class_ids if cid not in current_classes]
            classes_to_remove = [cid for cid in current_classes if cid not in class_ids]
            
            # Remove old class assignments (set section_ref_id to NULL)
            if classes_to_remove:
                placeholders = ','.join(['%s'] * len(classes_to_remove))
                cursor.execute(f"""
                    UPDATE class_sections 
                    SET section_ref_id = NULL
                    WHERE section_id IN ({placeholders})
                """, classes_to_remove)
            
            # Add new class assignments (update section_ref_id)
            if classes_to_add:
                for class_id in classes_to_add:
                    cursor.execute("""
                        UPDATE class_sections 
                        SET section_ref_id = %s 
                        WHERE section_id = %s
                    """, (section_id, class_id))
            
            # Count students from all assigned classes
            if class_ids:
                # Get all unique students from the assigned classes
                # enrollments.section_id refers to class_sections.section_id (the class)
                placeholders = ','.join(['%s'] * len(class_ids))
                cursor.execute(f"""
                    SELECT DISTINCT e.student_id
                    FROM enrollments e
                    INNER JOIN std_info s ON e.student_id = s.id
                    WHERE e.section_id IN ({placeholders})
                    AND e.status = 'Enrolled'
                """, class_ids)
                
                students = cursor.fetchall()
                students_synced = len(students)
                
                # Students are already linked via enrollments table
                # No need to update std_info as it doesn't have section_id column
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Section updated successfully',
            'students_synced': students_synced
        })
        
    except Exception as e:
        print(f"Error updating section: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections-master/<int:section_id>', methods=['DELETE'])
@login_required
def delete_section_master(section_id):
    """Disable a section (soft disable) from the master table"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if section exists
        cursor.execute("SELECT * FROM sections WHERE section_id = %s", (section_id,))
        section = cursor.fetchone()
        
        if not section:
            cursor.close()
            return jsonify({'success': False, 'error': 'Section not found'}), 404
        
        # Disable the section (soft disable) instead of hard delete
        cursor.execute(
            "UPDATE sections SET is_disable = 1, updated_at = NOW() WHERE section_id = %s", 
            (section_id,)
        )
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Section disabled successfully'
        })
        
    except Exception as e:
        print(f"Error disabling section: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections-master/disabled', methods=['GET'])
@login_required
def get_disabled_sections():
    """Get all disabled sections"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all disabled sections
        query = """
            SELECT 
                s.section_id,
                s.section_code,
                s.section_name,
                s.program_id,
                s.year_level,
                s.is_disable,
                s.created_at,
                s.updated_at,
                p.name as program_name,
                p.program_code
            FROM sections s
            LEFT JOIN programs p ON s.program_id = p.program_id
            WHERE s.is_disable = 1
            ORDER BY s.updated_at DESC
        """
        cursor.execute(query)
        sections = cursor.fetchall()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': sections
        })
        
    except Exception as e:
        print(f"Error getting disabled sections: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections-master/<int:section_id>/enable', methods=['PUT'])
@login_required
def enable_section_master(section_id):
    """Enable (restore) a disabled section"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if section exists and is disabled
        cursor.execute("SELECT * FROM sections WHERE section_id = %s", (section_id,))
        section = cursor.fetchone()
        
        if not section:
            cursor.close()
            return jsonify({'success': False, 'error': 'Section not found'}), 404
        
        if not section.get('is_disable'):
            cursor.close()
            return jsonify({'success': False, 'error': 'Section is not disabled'}), 400
        
        # Enable the section
        cursor.execute(
            "UPDATE sections SET is_disable = 0, updated_at = NOW() WHERE section_id = %s", 
            (section_id,)
        )
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Section enabled successfully'
        })
        
    except Exception as e:
        print(f"Error enabling section: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()



@api_bp.route('/sections-master/<int:section_id>/students', methods=['GET'])
@login_required
def get_section_students(section_id):
    """Get all students directly assigned to a section"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if section exists
        cursor.execute("SELECT section_id FROM sections WHERE section_id = %s", (section_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Section not found'}), 404
        
        # Get all students directly assigned to this section
        cursor.execute("""
            SELECT 
                s.id,
                s.std_Number,
                s.std_Firstname,
                s.std_Surname,
                s.std_Level,
                s.std_Course,
                s.std_Status,
                ss.assigned_date
            FROM std_info s
            INNER JOIN section_students ss ON s.id = ss.student_id
            WHERE ss.section_id = %s
            AND ss.status = 'Active'
            AND s.std_Status = 'Enrolled'
            AND s.is_archived = 0
            ORDER BY s.std_Surname, s.std_Firstname
        """, (section_id,))
        
        students = cursor.fetchall()
        cursor.close()
        
        return jsonify({
            'success': True,
            'students': students
        })
        
    except Exception as e:
        print(f"Error getting section students: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections-master/<int:section_id>/students', methods=['POST'])
@login_required
def add_student_to_section(section_id):
    """Add a student directly to a section (simple assignment without classes)"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        
        if not student_id:
            return jsonify({'success': False, 'error': 'Student ID is required'}), 400
        
        cursor = conn.cursor(dictionary=True)
        
        # Check if section exists
        cursor.execute("SELECT section_id FROM sections WHERE section_id = %s", (section_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Section not found'}), 404
        
        # Check if student exists
        cursor.execute("SELECT id FROM std_info WHERE id = %s AND std_Status = 'Enrolled'", (student_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Student not found or not enrolled'}), 404
        
        # Check if student is already assigned to this section
        cursor.execute("""
            SELECT id FROM section_students 
            WHERE section_id = %s AND student_id = %s
        """, (section_id, student_id))
        
        if cursor.fetchone():
            cursor.close()
            return jsonify({
                'success': False, 
                'error': 'Student is already assigned to this section'
            }), 400
        
        # Add student to section
        cursor.execute("""
            INSERT INTO section_students (section_id, student_id, status, assigned_date)
            VALUES (%s, %s, 'Active', NOW())
        """, (section_id, student_id))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Student added to section successfully'
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error adding student to section: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections-master/<int:section_id>/students/<int:student_id>', methods=['DELETE'])
@login_required
def remove_student_from_section(section_id, student_id):
    """Remove a student directly from a section"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if section exists
        cursor.execute("SELECT section_id FROM sections WHERE section_id = %s", (section_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Section not found'}), 404
        
        # Delete student from section
        cursor.execute("""
            DELETE FROM section_students
            WHERE section_id = %s AND student_id = %s
        """, (section_id, student_id))
        
        removed_count = cursor.rowcount
        
        if removed_count == 0:
            cursor.close()
            return jsonify({
                'success': False,
                'error': 'Student not found in this section'
            }), 404
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Student removed from section successfully'
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error removing student from section: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/sections/<int:section_id>/classes', methods=['GET'])
@login_required
def get_section_classes(section_id):
    """Get available and assigned classes for a section"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if section exists
        cursor.execute("SELECT * FROM sections WHERE section_id = %s", (section_id,))
        section = cursor.fetchone()
        
        if not section:
            cursor.close()
            return jsonify({'success': False, 'error': 'Section not found'}), 404
        
        # Get assigned classes for this section
        # Note: class_sections.section_id is the class ID, section_ref_id references sections table
        cursor.execute("""
            SELECT 
                cs.section_id as class_id,
                cs.subject_id,
                s.subject_code as course_code,
                s.title as course_title,
                CONCAT(COALESCE(f.first_name, ''), ' ', COALESCE(f.last_name, '')) as faculty_name,
                cs.schedule,
                cs.room,
                (SELECT COUNT(*) FROM enrollments WHERE section_id = cs.section_id) as student_count
            FROM class_sections cs
            INNER JOIN subjects s ON cs.subject_id = s.subject_id
            LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
            WHERE cs.section_ref_id = %s
            ORDER BY s.subject_code
        """, (section_id,))
        assigned_classes = cursor.fetchall()
        
        # Get all available classes (not yet assigned to this section)
        cursor.execute("""
            SELECT 
                cs.section_id as class_id,
                cs.subject_id,
                s.subject_code as course_code,
                s.title as course_title,
                CONCAT(COALESCE(f.first_name, ''), ' ', COALESCE(f.last_name, '')) as faculty_name,
                cs.schedule,
                cs.room,
                (SELECT COUNT(*) FROM enrollments WHERE section_id = cs.section_id) as student_count
            FROM class_sections cs
            INNER JOIN subjects s ON cs.subject_id = s.subject_id
            LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
            WHERE (cs.section_ref_id IS NULL OR cs.section_ref_id != %s)
            ORDER BY s.subject_code
        """, (section_id,))
        available_classes = cursor.fetchall()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'assigned_classes': assigned_classes,
            'available_classes': available_classes
        })
        
    except Exception as e:
        print(f"Error getting section classes: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ============================================================================
# CLASSES MANAGEMENT (CLASS_SECTIONS with SECTION_REF_ID)
# ============================================================================

@api_bp.route('/classes')
@login_required
def get_classes():
    """Get all class sections with subject, section, faculty, and term information"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all classes with related information (using section_ref_id)
        query = """
            SELECT 
                cs.section_id,
                cs.section_ref_id,
                cs.section_name,
                cs.schedule,
                cs.room,
                cs.subject_id,
                cs.faculty_id,
                cs.acad_term_id,
                s.subject_code,
                s.title as subject_title,
                sec.section_code,
                sec.section_name as section_full_name,
                sec.year_level,
                CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                CONCAT(ay.year_code, ' - ', at.term_name) as term_name,
                at.is_current as is_current_term,
                COUNT(DISTINCT e.student_id) as enrolled_students
            FROM class_sections cs
            LEFT JOIN subjects s ON cs.subject_id = s.subject_id
            LEFT JOIN sections sec ON cs.section_ref_id = sec.section_id
            LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
            LEFT JOIN academic_terms at ON cs.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            LEFT JOIN enrollments e ON cs.section_id = e.section_id
            GROUP BY cs.section_id
            ORDER BY at.is_current DESC, sec.section_code, s.subject_code
        """
        cursor.execute(query)
        classes = cursor.fetchall()
        
        # Get statistics
        stats_query = """
            SELECT 
                COUNT(DISTINCT cs.section_id) as total_classes,
                COUNT(DISTINCT cs.subject_id) as unique_subjects,
                COUNT(DISTINCT cs.faculty_id) as active_faculty,
                COUNT(DISTINCT CASE WHEN at.is_current = TRUE THEN cs.section_id END) as current_term_classes
            FROM class_sections cs
            LEFT JOIN academic_terms at ON cs.acad_term_id = at.acad_term_id
        """
        cursor.execute(stats_query)
        stats = cursor.fetchone()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': classes,
            'stats': {
                'total_classes': stats['total_classes'] or 0,
                'unique_subjects': stats['unique_subjects'] or 0,
                'active_faculty': stats['active_faculty'] or 0,
                'current_term_classes': stats['current_term_classes'] or 0
            }
        })
        
    except Exception as e:
        print(f"Error getting classes: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/classes', methods=['POST'])
@login_required
def create_class():
    """Create a new class section"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['subject_id', 'section_ref_id', 'faculty_id', 'acad_term_id']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        cursor = conn.cursor(dictionary=True)
        
        # Get section_name from sections table for backward compatibility
        cursor.execute("SELECT section_code FROM sections WHERE section_id = %s", (data['section_ref_id'],))
        section = cursor.fetchone()
        if not section:
            cursor.close()
            return jsonify({'success': False, 'error': 'Invalid section selected'}), 400
        
        # Insert new class
        insert_query = """
            INSERT INTO class_sections 
            (subject_id, faculty_id, acad_term_id, section_ref_id, section_name, schedule, room)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(insert_query, (
            data['subject_id'],
            data['faculty_id'],
            data['acad_term_id'],
            data['section_ref_id'],
            section['section_code'],  # Store section_code as section_name for backward compatibility
            data.get('schedule', ''),
            data.get('room', '')
        ))
        
        conn.commit()
        class_id = cursor.lastrowid
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Class created successfully',
            'class_id': class_id
        })
        
    except Exception as e:
        conn.rollback()
        print(f"Error creating class: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/classes/<int:class_id>', methods=['PUT'])
@login_required
def update_class(class_id):
    """Update an existing class section"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        data = request.get_json()
        cursor = conn.cursor(dictionary=True)
        
        # Check if class exists
        cursor.execute("SELECT * FROM class_sections WHERE section_id = %s", (class_id,))
        class_item = cursor.fetchone()
        
        if not class_item:
            cursor.close()
            return jsonify({'success': False, 'error': 'Class not found'}), 404
        
        # Build update query dynamically based on provided fields
        update_fields = []
        update_values = []
        
        if 'subject_id' in data:
            update_fields.append('subject_id = %s')
            update_values.append(data['subject_id'])
        if 'faculty_id' in data:
            update_fields.append('faculty_id = %s')
            update_values.append(data['faculty_id'])
        if 'acad_term_id' in data:
            update_fields.append('acad_term_id = %s')
            update_values.append(data['acad_term_id'])
        if 'section_ref_id' in data:
            # Get section_code for backward compatibility
            cursor.execute("SELECT section_code FROM sections WHERE section_id = %s", (data['section_ref_id'],))
            section = cursor.fetchone()
            if section:
                update_fields.append('section_ref_id = %s')
                update_values.append(data['section_ref_id'])
                update_fields.append('section_name = %s')
                update_values.append(section['section_code'])
        if 'schedule' in data:
            update_fields.append('schedule = %s')
            update_values.append(data['schedule'])
        if 'room' in data:
            update_fields.append('room = %s')
            update_values.append(data['room'])
        
        if not update_fields:
            cursor.close()
            return jsonify({'success': False, 'error': 'No fields to update'}), 400
        
        # Add class_id to values for WHERE clause
        update_values.append(class_id)
        
        # Execute update
        update_query = f"""
            UPDATE class_sections 
            SET {', '.join(update_fields)}
            WHERE section_id = %s
        """
        cursor.execute(update_query, update_values)
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Class updated successfully'
        })
        
    except Exception as e:
        print(f"Error updating class: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/classes/<int:class_id>', methods=['DELETE'])
@login_required
def delete_class(class_id):
    """Delete a class section"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if class exists
        cursor.execute("SELECT * FROM class_sections WHERE section_id = %s", (class_id,))
        class_item = cursor.fetchone()
        
        if not class_item:
            cursor.close()
            return jsonify({'success': False, 'error': 'Class not found'}), 404
        
        # Check if class has enrollments
        cursor.execute(
            "SELECT COUNT(*) as count FROM enrollments WHERE section_id = %s", 
            (class_id,)
        )
        enrollment_count = cursor.fetchone()['count']
        
        if enrollment_count > 0:
            cursor.close()
            return jsonify({
                'success': False, 
                'error': f'Cannot delete class with {enrollment_count} enrolled student(s). Please remove enrollments first.'
            }), 400
        
        # Check if class has evaluations
        cursor.execute(
            "SELECT COUNT(*) as count FROM evaluations WHERE section_id = %s", 
            (class_id,)
        )
        evaluation_count = cursor.fetchone()['count']
        
        if evaluation_count > 0:
            cursor.close()
            return jsonify({
                'success': False, 
                'error': f'Cannot delete class with {evaluation_count} evaluation(s). Please remove evaluations first.'
            }), 400
        
        # Delete the class
        cursor.execute("DELETE FROM class_sections WHERE section_id = %s", (class_id,))
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Class deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting class: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/dashboard/stats')
@login_required
def get_dashboard_stats():
    """Get comprehensive dashboard statistics"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get total users count
        cursor.execute("""
            SELECT COUNT(*) as total_users
            FROM users
            WHERE is_active = TRUE
        """)
        total_users = cursor.fetchone()['total_users'] or 0
        
        # Get faculty count
        cursor.execute("""
            SELECT COUNT(*) as total_faculty
            FROM faculty
            WHERE is_archived = FALSE
        """)
        total_faculty = cursor.fetchone()['total_faculty'] or 0
        
        # Get enrolled students count
        cursor.execute("""
            SELECT COUNT(DISTINCT s.id) as enrolled_students
            FROM std_info s
            WHERE s.std_Status = 'Enrolled' AND s.is_archived = FALSE
        """)
        enrolled_students = cursor.fetchone()['enrolled_students'] or 0
        
        # Get completed evaluations count and completion rate
        cursor.execute("""
            SELECT 
                COUNT(CASE WHEN status = 'Completed' THEN 1 END) as completed,
                COUNT(*) as total,
                ROUND((COUNT(CASE WHEN status = 'Completed' THEN 1 END) / COUNT(*)) * 100, 1) as completion_rate
            FROM evaluations
        """)
        eval_stats = cursor.fetchone()
        completed_evaluations = eval_stats['completed'] or 0
        completion_rate = float(eval_stats['completion_rate'] or 0)
        
        # Get evaluation progress by week (last 6 weeks)
        cursor.execute("""
            SELECT 
                WEEK(completion_time) as week_num,
                DATE_FORMAT(completion_time, '%b %d') as week_label,
                COUNT(*) as count
            FROM evaluations
            WHERE status = 'Completed' 
                AND completion_time >= DATE_SUB(NOW(), INTERVAL 6 WEEK)
            GROUP BY WEEK(completion_time), DATE_FORMAT(completion_time, '%b %d')
            ORDER BY completion_time
            LIMIT 6
        """)
        weekly_progress = cursor.fetchall()
        
        # Get cumulative counts for progress chart
        cumulative_data = []
        cumulative_count = 0
        for item in weekly_progress:
            cumulative_count += item['count']
            cumulative_data.append({
                'week_label': item['week_label'],
                'count': cumulative_count
            })
        
        # Get faculty rating distribution
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN avg_rating >= 4.5 THEN '5 Stars'
                    WHEN avg_rating >= 3.5 THEN '4 Stars'
                    WHEN avg_rating >= 2.5 THEN '3 Stars'
                    WHEN avg_rating >= 1.5 THEN '2 Stars'
                    ELSE '1 Star'
                END as rating_category,
                COUNT(*) as faculty_count
            FROM (
                SELECT 
                    f.faculty_id,
                    ROUND(AVG(er.rating), 2) as avg_rating
                FROM faculty f
                LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                LEFT JOIN evaluations e ON cs.section_id = e.section_id AND e.status = 'Completed'
                LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                WHERE f.is_archived = FALSE
                GROUP BY f.faculty_id
                HAVING avg_rating IS NOT NULL
            ) as faculty_ratings
            GROUP BY rating_category
            ORDER BY FIELD(rating_category, '5 Stars', '4 Stars', '3 Stars', '2 Stars', '1 Star')
        """)
        rating_distribution = cursor.fetchall()
        
        # Get department performance
        cursor.execute("""
            SELECT 
                p.name as department,
                p.program_code,
                COUNT(DISTINCT f.faculty_id) as faculty_count,
                -- Count total student evaluations that are completed
                COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as completed_evals,
                -- Count all student evaluations (completed + pending)
                COUNT(DISTINCT e.evaluation_id) as total_evals,
                -- Average rating from completed evaluations only
                ROUND(AVG(CASE WHEN e.status = 'Completed' THEN er.rating END), 2) as avg_rating,
                -- Calculate completion percentage: completed student evaluations / total student evaluations
                CASE 
                    WHEN COUNT(DISTINCT e.evaluation_id) > 0 THEN
                        ROUND((COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) / 
                               COUNT(DISTINCT e.evaluation_id)) * 100, 0)
                    ELSE 0
                END as completion_percentage,
                CASE 
                    WHEN COUNT(DISTINCT f.faculty_id) = 0 THEN 'No Faculty'
                    WHEN COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) = 0 THEN 'Not Started'
                    WHEN ROUND((COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) / 
                               NULLIF(COUNT(DISTINCT e.evaluation_id), 0)) * 100, 0) >= 80 THEN 'Active'
                    WHEN ROUND((COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) / 
                               NULLIF(COUNT(DISTINCT e.evaluation_id), 0)) * 100, 0) >= 50 THEN 'Active'
                    ELSE 'Attention Needed'
                END as status
            FROM programs p
            LEFT JOIN faculty f ON p.program_id = f.program_id AND f.is_archived = FALSE
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            GROUP BY p.program_id, p.name, p.program_code
            HAVING faculty_count > 0
            ORDER BY completion_percentage DESC, avg_rating DESC
            LIMIT 10
        """)
        department_performance = cursor.fetchall()
        
        # Add pending_evals to the results
        for dept in department_performance:
            dept['pending_evals'] = dept['total_evals'] - dept['completed_evals']
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': {
                'overview': {
                    'total_users': total_users,
                    'total_faculty': total_faculty,
                    'enrolled_students': enrolled_students,
                    'completed_evaluations': completed_evaluations,
                    'completion_rate': completion_rate
                },
                'evaluation_progress': cumulative_data,
                'rating_distribution': rating_distribution,
                'department_performance': department_performance
            }
        })
        
    except Exception as e:
        print(f"Error getting dashboard stats: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/dashboard/departments', methods=['GET'])
@login_required
def get_dashboard_departments():
    """Get list of all departments/programs for dashboard filtering"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                program_id,
                name as department_name,
                program_code
            FROM programs
            WHERE program_id IN (
                SELECT DISTINCT program_id 
                FROM faculty 
                WHERE is_archived = FALSE AND program_id IS NOT NULL
            )
            ORDER BY name
        """)
        departments = cursor.fetchall()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': departments
        })
        
    except Exception as e:
        print(f"Error getting departments: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/dashboard/rating-distribution', methods=['GET'])
@login_required
def get_dashboard_rating_distribution():
    """Get faculty rating distribution, optionally filtered by department"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get department filter from query params
        program_id = request.args.get('program_id', None)
        
        # Build query based on filter
        if program_id and program_id != 'all':
            query = """
                SELECT 
                    CASE 
                        WHEN avg_rating >= 4.5 THEN '5 Stars'
                        WHEN avg_rating >= 3.5 THEN '4 Stars'
                        WHEN avg_rating >= 2.5 THEN '3 Stars'
                        WHEN avg_rating >= 1.5 THEN '2 Stars'
                        ELSE '1 Star'
                    END as rating_category,
                    COUNT(*) as faculty_count
                FROM (
                    SELECT 
                        f.faculty_id,
                        ROUND(AVG(er.rating), 2) as avg_rating
                    FROM faculty f
                    LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                    LEFT JOIN evaluations e ON cs.section_id = e.section_id AND e.status = 'Completed'
                    LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                    WHERE f.is_archived = FALSE AND f.program_id = %s
                    GROUP BY f.faculty_id
                    HAVING avg_rating IS NOT NULL
                ) as faculty_ratings
                GROUP BY rating_category
                ORDER BY FIELD(rating_category, '5 Stars', '4 Stars', '3 Stars', '2 Stars', '1 Star')
            """
            cursor.execute(query, (program_id,))
        else:
            query = """
                SELECT 
                    CASE 
                        WHEN avg_rating >= 4.5 THEN '5 Stars'
                        WHEN avg_rating >= 3.5 THEN '4 Stars'
                        WHEN avg_rating >= 2.5 THEN '3 Stars'
                        WHEN avg_rating >= 1.5 THEN '2 Stars'
                        ELSE '1 Star'
                    END as rating_category,
                    COUNT(*) as faculty_count
                FROM (
                    SELECT 
                        f.faculty_id,
                        ROUND(AVG(er.rating), 2) as avg_rating
                    FROM faculty f
                    LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                    LEFT JOIN evaluations e ON cs.section_id = e.section_id AND e.status = 'Completed'
                    LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                    WHERE f.is_archived = FALSE
                    GROUP BY f.faculty_id
                    HAVING avg_rating IS NOT NULL
                ) as faculty_ratings
                GROUP BY rating_category
                ORDER BY FIELD(rating_category, '5 Stars', '4 Stars', '3 Stars', '2 Stars', '1 Star')
            """
            cursor.execute(query)
        
        rating_distribution = cursor.fetchall()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': rating_distribution
        })
        
    except Exception as e:
        print(f"Error getting rating distribution: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/guidance/academic-years')
@login_required
def guidance_academic_years():
    """Get all academic years for filtering"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                acad_year_id as academic_year_id,
                year_code,
                year_name,
                start_date,
                end_date,
                is_current
            FROM academic_years
            ORDER BY start_date DESC
        """)
        
        academic_years = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'academic_years': academic_years
        })
        
    except Exception as e:
        print(f"Error getting academic years: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/periods-by-year/<int:year_id>')
@login_required
def guidance_periods_by_year(year_id):
    """Get evaluation periods for a specific academic year"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                ep.period_id,
                ep.title as period_name,
                ep.start_date,
                ep.end_date,
                ep.status,
                CASE 
                    WHEN CURDATE() BETWEEN ep.start_date AND ep.end_date THEN 1
                    ELSE 0
                END as is_active
            FROM evaluation_periods ep
            INNER JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE at.acad_year_id = %s AND COALESCE(ep.is_archived, 0) = 0
            ORDER BY ep.start_date DESC
        """, (year_id,))
        
        periods = cursor.fetchall()
        
        # Format dates
        for period in periods:
            if period['start_date']:
                period['start_date'] = period['start_date'].strftime('%Y-%m-%d')
            if period['end_date']:
                period['end_date'] = period['end_date'].strftime('%Y-%m-%d')
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'periods': periods
        })
        
    except Exception as e:
        print(f"Error getting periods: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/evaluation-results')
@login_required
def guidance_evaluation_results():
    """Get comprehensive evaluation results with faculty ratings and statistics"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get filter parameters
        academic_year_id = request.args.get('academic_year_id', type=int)
        period_id = request.args.get('period_id', type=int)
        
        # Validate parameters
        if not academic_year_id or not period_id:
            return jsonify({
                'success': False, 
                'message': 'Academic year and period are required'
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verify academic year and period exist
        cursor.execute("""
            SELECT ay.year_name, ep.title as period_name
            FROM academic_years ay
            JOIN academic_terms at ON ay.acad_year_id = at.acad_year_id
            JOIN evaluation_periods ep ON at.acad_term_id = ep.acad_term_id
            WHERE ay.acad_year_id = %s AND ep.period_id = %s
        """, (academic_year_id, period_id))
        
        period_info = cursor.fetchone()
        if not period_info:
            return jsonify({
                'success': False,
                'message': 'Invalid academic year or period'
            }), 404
        
        # Get faculty evaluation results with average ratings for the selected period
        cursor.execute("""
            SELECT 
                f.faculty_id,
                f.first_name,
                f.last_name,
                f.faculty_number,
                p.name as department_name,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as completed_evaluations,
                AVG(er.rating) as overall_rating,
                COUNT(DISTINCT er.response_id) as total_responses
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            WHERE f.is_archived = 0
                AND e.period_id = %s
            GROUP BY f.faculty_id, f.first_name, f.last_name, f.faculty_number, p.name
            HAVING total_evaluations > 0
            ORDER BY overall_rating DESC
        """, (period_id,))
        
        faculty_results = cursor.fetchall()
        
        # Calculate overall statistics
        total_faculty = len(faculty_results)
        overall_avg_rating = 0
        total_completed = 0
        total_evaluations = 0
        
        for faculty in faculty_results:
            if faculty['overall_rating']:
                overall_avg_rating += faculty['overall_rating']
            total_completed += faculty['completed_evaluations'] or 0
            total_evaluations += faculty['total_evaluations'] or 0
        
        if total_faculty > 0:
            overall_avg_rating = round(overall_avg_rating / total_faculty, 2)
        
        response_rate = round((total_completed / total_evaluations * 100), 1) if total_evaluations > 0 else 0
        
        # Find top performer and needs attention
        top_performer = None
        needs_attention_count = 0
        top_performers = []
        needs_attention = []
        
        for faculty in faculty_results:
            rating = faculty['overall_rating'] or 0
            faculty['overall_rating'] = round(rating, 1) if rating else 0
            
            # Response rate calculation
            if faculty['total_evaluations'] and faculty['total_evaluations'] > 0:
                faculty['response_rate'] = round((faculty['completed_evaluations'] / faculty['total_evaluations']) * 100, 1)
            else:
                faculty['response_rate'] = 0
            
            # Categorize faculty - adjusted thresholds for 1-5 rating scale
            if rating >= 4.0:  # Changed from 4.5 to 4.0
                top_performers.append({
                    'name': f"{faculty['first_name']} {faculty['last_name']}",
                    'rating': round(rating, 1)
                })
            elif rating < 3.0 and rating > 0:  # Changed from 3.5 to 3.0
                needs_attention.append({
                    'name': f"{faculty['first_name']} {faculty['last_name']}",
                    'rating': round(rating, 1)
                })
                needs_attention_count += 1
        
        # Sort and limit top performers and needs attention
        top_performers = sorted(top_performers, key=lambda x: x['rating'], reverse=True)[:3]
        needs_attention = sorted(needs_attention, key=lambda x: x['rating'])[:3]
        
        top_performer_name = top_performers[0]['name'] if top_performers else 'N/A'
        
        # Get rating distribution for selected period
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN AVG(er.rating) >= 4.5 THEN '5 Stars'
                    WHEN AVG(er.rating) >= 3.5 THEN '4 Stars'
                    WHEN AVG(er.rating) >= 2.5 THEN '3 Stars'
                    WHEN AVG(er.rating) >= 1.5 THEN '2 Stars'
                    ELSE '1 Star'
                END as rating_category,
                COUNT(DISTINCT f.faculty_id) as count
            FROM faculty f
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            WHERE f.is_archived = 0 
                AND e.status = 'Completed'
                AND e.period_id = %s
            GROUP BY f.faculty_id
            HAVING AVG(er.rating) IS NOT NULL
        """, (period_id,))
        
        rating_dist_raw = cursor.fetchall()
        rating_distribution = {'1 Star': 0, '2 Stars': 0, '3 Stars': 0, '4 Stars': 0, '5 Stars': 0}
        
        for item in rating_dist_raw:
            rating_distribution[item['rating_category']] = item['count']
        
        # Get department performance (average by department) for selected period
        cursor.execute("""
            SELECT 
                p.name as department_name,
                AVG(er.rating) as avg_rating,
                COUNT(DISTINCT f.faculty_id) as faculty_count
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            WHERE f.is_archived = 0 
                AND e.status = 'Completed'
                AND e.period_id = %s
            GROUP BY p.name
            HAVING AVG(er.rating) IS NOT NULL
            ORDER BY avg_rating DESC
        """, (period_id,))
        
        department_performance = cursor.fetchall()
        
        for dept in department_performance:
            dept['avg_rating'] = round(dept['avg_rating'], 2) if dept['avg_rating'] else 0
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'statistics': {
                'overall_rating': overall_avg_rating,
                'response_rate': response_rate,
                'top_performer': top_performer_name,
                'needs_attention_count': needs_attention_count
            },
            'faculty_results': faculty_results,
            'top_performers': top_performers,
            'needs_attention': needs_attention,
            'rating_distribution': rating_distribution,
            'department_performance': department_performance,
            'period_info': {
                'year_name': period_info['year_name'],
                'period_name': period_info['period_name']
            }
        })
        
    except Exception as e:
        print(f"Error getting evaluation results: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/faculty-evaluation-details/<int:faculty_id>')
@login_required
def faculty_evaluation_details(faculty_id):
    """Get detailed faculty evaluation results including anonymous student comments"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get period filter from query parameters
        academic_year_id = request.args.get('academic_year_id', type=int)
        period_id = request.args.get('period_id', type=int)
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Build WHERE clause for period filtering
        period_filter = ""
        period_params = [faculty_id]
        
        if period_id:
            period_filter = "AND e.period_id = %s"
            period_params.append(period_id)
        
        # Get faculty basic information
        query = f"""
            SELECT 
                f.faculty_id,
                f.first_name,
                f.last_name,
                f.faculty_number,
                p.name as department_name,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as completed_evaluations
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
            WHERE f.faculty_id = %s {period_filter}
            GROUP BY f.faculty_id, f.first_name, f.last_name, f.faculty_number, p.name
        """
        
        cursor.execute(query, tuple(period_params))
        
        faculty = cursor.fetchone()
        
        if not faculty:
            return jsonify({'success': False, 'message': 'Faculty not found'}), 404
        
        # Calculate overall rating with period filter
        rating_params = [faculty_id]
        if period_id:
            rating_params.append(period_id)
            
        rating_query = f"""
            SELECT AVG(er.rating) as overall_rating
            FROM evaluation_responses er
            JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            WHERE cs.faculty_id = %s AND e.status = 'Completed' {period_filter}
        """
        
        cursor.execute(rating_query, tuple(rating_params))
        
        rating_result = cursor.fetchone()
        faculty['overall_rating'] = round(rating_result['overall_rating'], 1) if rating_result['overall_rating'] else 0
        
        # Calculate response rate
        if faculty['total_evaluations'] > 0:
            faculty['response_rate'] = round((faculty['completed_evaluations'] / faculty['total_evaluations']) * 100, 1)
        else:
            faculty['response_rate'] = 0
        
        # Get anonymous student comments and ratings with period filter
        # Comments are stored in a separate table linked by evaluation_id
        comments_params = [faculty_id]
        if period_id:
            comments_params.append(period_id)
            
        comments_query = f"""
            SELECT 
                AVG(er.rating) as rating,
                c.comment_text as comment,
                c.sentiment,
                e.completion_time as date
            FROM comments c
            JOIN evaluations e ON c.evaluation_id = e.evaluation_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            WHERE cs.faculty_id = %s 
            AND e.status = 'Completed'
            {period_filter}
            AND c.comment_text IS NOT NULL
            AND c.comment_text != ''
            GROUP BY c.comment_id, c.comment_text, c.sentiment, e.completion_time
            ORDER BY e.completion_time DESC
        """
        
        cursor.execute(comments_query, tuple(comments_params))
        
        comments = cursor.fetchall()
        
        # Get rating breakdown by category with period filter
        breakdown_params = [faculty_id]
        if period_id:
            breakdown_params.append(period_id)
            
        breakdown_query = f"""
            SELECT 
                ec.name as category,
                AVG(er.rating) as average,
                (AVG(er.rating) / 5.0 * 100) as percentage
            FROM evaluation_responses er
            JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            JOIN evaluation_criteria ecr ON er.criteria_id = ecr.criteria_id
            JOIN evaluation_categories ec ON ecr.category_id = ec.category_id
            WHERE cs.faculty_id = %s 
            AND e.status = 'Completed'
            {period_filter}
            GROUP BY ec.category_id, ec.name
            ORDER BY average DESC
        """
        
        cursor.execute(breakdown_query, tuple(breakdown_params))
        
        rating_breakdown_list = cursor.fetchall()
        
        # Convert rating breakdown to dictionary
        rating_breakdown = {}
        for item in rating_breakdown_list:
            rating_breakdown[item['category']] = {
                'average': round(item['average'], 1) if item['average'] else 0,
                'percentage': round(item['percentage'], 1) if item['percentage'] else 0
            }
        
        # Get evaluation period info - use selected period if provided
        if period_id:
            cursor.execute("""
                SELECT ep.title
                FROM evaluation_periods ep
                WHERE ep.period_id = %s
                LIMIT 1
            """, (period_id,))
        else:
            cursor.execute("""
                SELECT ep.title
                FROM evaluation_periods ep
                WHERE ep.status = 'Active'
                LIMIT 1
            """)
        
        period = cursor.fetchone()
        evaluation_period = period['title'] if period else 'Current Evaluation Period'
        
        # Get existing recommendation for this faculty (most recent)
        # Note: faculty_recommendations table doesn't have period_id column
        cursor.execute("""
            SELECT 
                r.recommendation_text,
                r.created_at,
                r.updated_at,
                CONCAT(u.first_name, ' ', u.last_name) as counselor_name
            FROM faculty_recommendations r
            LEFT JOIN users u ON r.counselor_id = u.user_id
            WHERE r.faculty_id = %s
            ORDER BY r.updated_at DESC
            LIMIT 1
        """, (faculty_id,))
        
        recommendation = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'faculty': faculty,
            'comments': comments,
            'rating_breakdown': rating_breakdown,
            'evaluation_period': evaluation_period,
            'recommendation': recommendation
        })
        
    except Exception as e:
        print(f"Error getting faculty evaluation details: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/faculty-subjects/<int:faculty_id>')
@login_required
def guidance_get_faculty_subjects(faculty_id):
    """Get subjects taught by a specific faculty member"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Get distinct subjects taught by the faculty
        cursor.execute("""
            SELECT DISTINCT s.subject_id, s.subject_code, s.title as subject_name
            FROM subjects s
            JOIN class_sections cs ON s.subject_id = cs.subject_id
            WHERE cs.faculty_id = %s
            ORDER BY s.subject_code
        """, (faculty_id,))
        
        subjects = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'subjects': subjects if subjects else []
        })
        
    except Exception as e:
        print(f"Error getting faculty subjects: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/faculty-sections')
@login_required
def guidance_get_faculty_sections():
    """Get sections for a specific faculty member and subject"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        faculty_id = request.args.get('faculty_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        
        if not faculty_id or not subject_id:
            return jsonify({'success': False, 'message': 'Faculty ID and Subject ID are required'}), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Get sections taught by the faculty for the specific subject
        cursor.execute("""
            SELECT cs.section_id, sec.section_name, sec.section_code
            FROM class_sections cs
            JOIN sections sec ON cs.section_ref_id = sec.section_id
            WHERE cs.faculty_id = %s AND cs.subject_id = %s AND sec.is_disable = 0
            ORDER BY sec.section_name
        """, (faculty_id, subject_id))
        
        sections = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'sections': sections if sections else []
        })
        
    except Exception as e:
        print(f"Error getting faculty sections: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/academic-years')
@login_required
def guidance_get_academic_years():
    """Get all academic years"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT acad_year_id as academic_year_id, year_name, year_code, start_date, end_date, is_current
            FROM academic_years
            ORDER BY year_code DESC
        """)
        
        academic_years = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'academic_years': academic_years if academic_years else []
        })
        
    except Exception as e:
        print(f"Error getting academic years: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/evaluation-criteria-results/<int:faculty_id>')
@login_required
def evaluation_criteria_results(faculty_id):
    """Get detailed evaluation results with vote distribution per criterion"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        period_id = request.args.get('period_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        section_id = request.args.get('section_id', type=int)  # Optional section filter
        academic_year_id = request.args.get('academic_year_id', type=int)
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Build filters
        filters = []
        params = [faculty_id]
        
        if period_id:
            filters.append("e.period_id = %s")
            params.append(period_id)
        
        if subject_id:
            filters.append("cs.subject_id = %s")
            params.append(subject_id)
        
        if section_id:
            filters.append("cs.section_id = %s")
            params.append(section_id)
        
        if academic_year_id:
            filters.append("at.acad_year_id = %s")
            params.append(academic_year_id)
        
        period_filter = ""
        if filters:
            period_filter = "AND " + " AND ".join(filters)
        
        period_params = params
        
        # Get faculty info
        cursor.execute("""
            SELECT f.faculty_id, f.first_name, f.last_name, f.faculty_number,
                   p.name as department_name
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            WHERE f.faculty_id = %s
        """, (faculty_id,))
        
        faculty = cursor.fetchone()
        if not faculty:
            return jsonify({'success': False, 'message': 'Faculty not found'}), 404
        
        # Get period status if period_id is provided
        period_status = 'Active'  # Default status
        if period_id:
            cursor.execute("""
                SELECT status FROM evaluation_periods WHERE period_id = %s
            """, (period_id,))
            period_result = cursor.fetchone()
            if period_result:
                period_status = period_result['status']
        
        # Get overall rating
        cursor.execute(f"""
            SELECT AVG(er.rating) as overall_rating
            FROM evaluation_responses er
            JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE cs.faculty_id = %s AND e.status = 'Completed' {period_filter}
        """, tuple(period_params))
        
        rating_result = cursor.fetchone()
        faculty['overall_rating'] = round(rating_result['overall_rating'], 2) if rating_result['overall_rating'] else 0
        
        # Get vote distribution per criterion grouped by category
        cursor.execute(f"""
            SELECT 
                ec.category_id,
                ec.name as category_name,
                ecr.criteria_id,
                ecr.description as criteria_description,
                ecr.`order`,
                COUNT(er.response_id) as total_responses,
                SUM(CASE WHEN er.rating = 5 THEN 1 ELSE 0 END) as votes_5,
                SUM(CASE WHEN er.rating = 4 THEN 1 ELSE 0 END) as votes_4,
                SUM(CASE WHEN er.rating = 3 THEN 1 ELSE 0 END) as votes_3,
                SUM(CASE WHEN er.rating = 2 THEN 1 ELSE 0 END) as votes_2,
                SUM(CASE WHEN er.rating = 1 THEN 1 ELSE 0 END) as votes_1,
                AVG(er.rating) as mean_rating
            FROM evaluation_criteria ecr
            JOIN evaluation_categories ec ON ecr.category_id = ec.category_id
            LEFT JOIN evaluation_responses er ON ecr.criteria_id = er.criteria_id
            LEFT JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            LEFT JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE (cs.faculty_id = %s OR cs.faculty_id IS NULL) 
              AND e.status = 'Completed'
              {period_filter}
            GROUP BY ec.category_id, ec.name, ecr.criteria_id, ecr.description, ecr.`order`
            ORDER BY ec.category_id, ecr.`order`, ecr.criteria_id
        """, tuple(period_params))
        
        criteria_results = cursor.fetchall()
        
        # Helper function to determine remarks from mean
        def get_remarks(mean):
            if mean >= 4.5:
                return "Outstanding"
            elif mean >= 3.5:
                return "Very Satisfactory"
            elif mean >= 2.5:
                return "Satisfactory"
            elif mean >= 1.5:
                return "Fair"
            else:
                return "Poor"
        
        # Group by category
        categories = {}
        for row in criteria_results:
            category_id = row['category_id']
            
            if category_id not in categories:
                categories[category_id] = {
                    'category_id': category_id,
                    'category_name': row['category_name'],
                    'criteria': []
                }
            
            # Calculate mean and remarks
            mean = round(row['mean_rating'], 2) if row['mean_rating'] else 0
            
            criterion_data = {
                'criteria_id': row['criteria_id'],
                'description': row['criteria_description'],
                'votes': {
                    '5': row['votes_5'] or 0,
                    '4': row['votes_4'] or 0,
                    '3': row['votes_3'] or 0,
                    '2': row['votes_2'] or 0,
                    '1': row['votes_1'] or 0
                },
                'total_responses': row['total_responses'] or 0,
                'mean': mean,
                'remarks': get_remarks(mean) if mean > 0 else 'No Data'
            }
            
            categories[category_id]['criteria'].append(criterion_data)
        
        # Convert to list
        categories_list = list(categories.values())
        
        # Get unique comments from comments table (excluding empty/null comments)
        conn_comments = get_db_connection()
        cursor_comments = conn_comments.cursor(dictionary=True)
        
        cursor_comments.execute(f"""
            SELECT DISTINCT c.comment_text
            FROM comments c
            JOIN evaluations e ON c.evaluation_id = e.evaluation_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE cs.faculty_id = %s 
              AND e.status = 'Completed'
              AND c.comment_text IS NOT NULL 
              AND TRIM(c.comment_text) != ''
              {period_filter}
            ORDER BY c.comment_text
        """, tuple(period_params))
        
        comments_results = cursor_comments.fetchall()
        unique_comments = [row['comment_text'] for row in comments_results if row['comment_text']]
        
        cursor_comments.close()
        conn_comments.close()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'faculty': faculty,
            'categories': categories_list,
            'comments': unique_comments,
            'period_status': period_status
        })
        
    except Exception as e:
        print(f"Error getting evaluation criteria results: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/export-evaluation-pdf/<int:faculty_id>')
@login_required
def export_evaluation_pdf(faculty_id):
    """Export evaluation results to PDF"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from flask import send_file
        import os
        from datetime import datetime
        
        period_id = request.args.get('period_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        section_id = request.args.get('section_id', type=int)
        academic_year_id = request.args.get('academic_year_id', type=int)
        
        # Get signature position parameters (as percentages for X and Y)
        sig_faculty_y = request.args.get('sig_faculty_y', type=float, default=8.33)  # Y: ~50px / 600px * 100
        sig_faculty_x = request.args.get('sig_faculty_x', type=float, default=10.0)  # X: left side default
        
        sig_dean_y = request.args.get('sig_dean_y', type=float, default=33.33)  # Y: ~200px / 600px * 100
        sig_dean_x = request.args.get('sig_dean_x', type=float, default=10.0)  # X: left side default
        
        sig_president_y = request.args.get('sig_president_y', type=float, default=58.33)  # Y: ~350px / 600px * 100
        sig_president_x = request.args.get('sig_president_x', type=float, default=10.0)  # X: left side default
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get faculty info
        cursor.execute("""
            SELECT f.faculty_id, f.first_name, f.last_name, f.faculty_number
            FROM faculty f
            WHERE f.faculty_id = %s
        """, (faculty_id,))
        faculty = cursor.fetchone()
        
        if not faculty:
            return jsonify({'success': False, 'message': 'Faculty not found'}), 404
        
        # Get subject info
        cursor.execute("SELECT subject_code, title FROM subjects WHERE subject_id = %s", (subject_id,))
        subject = cursor.fetchone()
        
        # Get section info if section_id is provided
        section = None
        if section_id:
            cursor.execute("SELECT section_name FROM class_sections WHERE section_id = %s", (section_id,))
            section = cursor.fetchone()
        
        # Get period and year info
        cursor.execute("""
            SELECT ep.title as period_name, ay.year_name
            FROM evaluation_periods ep
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE ep.period_id = %s
        """, (period_id,))
        period_info = cursor.fetchone()
        
        # Build filters
        filters = []
        params = [faculty_id]
        
        if period_id:
            filters.append("e.period_id = %s")
            params.append(period_id)
        if subject_id:
            filters.append("cs.subject_id = %s")
            params.append(subject_id)
        if section_id:
            filters.append("cs.section_id = %s")
            params.append(section_id)
        if academic_year_id:
            filters.append("at.acad_year_id = %s")
            params.append(academic_year_id)
        
        period_filter = ""
        if filters:
            period_filter = "AND " + " AND ".join(filters)
        
        # Get evaluation data (same query as the main results)
        cursor.execute(f"""
            SELECT 
                c.category_id,
                c.name as category_name,
                c.display_order as category_order,
                cr.criteria_id as criterion_id,
                cr.description,
                cr.order as criterion_order,
                er.rating,
                e.evaluation_id
            FROM evaluation_categories c
            LEFT JOIN evaluation_criteria cr ON c.category_id = cr.category_id
            LEFT JOIN evaluation_responses er ON cr.criteria_id = er.criteria_id
            LEFT JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            LEFT JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE cs.faculty_id = %s 
              AND e.status = 'Completed'
              {period_filter}
            ORDER BY c.display_order, cr.order
        """, tuple(params))
        
        results = cursor.fetchall()
        
        # Process data into categories
        categories = {}
        for row in results:
            category_id = row['category_id']
            if category_id not in categories:
                categories[category_id] = {
                    'category_id': category_id,
                    'category_name': row['category_name'],
                    'category_order': row['category_order'],
                    'criteria': []
                }
            
            if row['criterion_id']:
                criterion_id = row['criterion_id']
                existing_criterion = next((c for c in categories[category_id]['criteria'] if c['criterion_id'] == criterion_id), None)
                
                if not existing_criterion:
                    criterion_data = {
                        'criterion_id': criterion_id,
                        'description': row['description'],
                        'criterion_order': row['criterion_order'],
                        'votes': {'5': 0, '4': 0, '3': 0, '2': 0, '1': 0},
                        'total_responses': 0,
                        'mean': 0,
                        'remarks': 'No Data'
                    }
                    categories[category_id]['criteria'].append(criterion_data)
                    existing_criterion = criterion_data
                
                if row['rating']:
                    rating_str = str(int(row['rating']))
                    existing_criterion['votes'][rating_str] = existing_criterion['votes'].get(rating_str, 0) + 1
                    existing_criterion['total_responses'] += 1
        
        # Calculate means
        for category in categories.values():
            for criterion in category['criteria']:
                if criterion['total_responses'] > 0:
                    total_score = sum(int(rating) * count for rating, count in criterion['votes'].items())
                    criterion['mean'] = total_score / criterion['total_responses']
                    
                    if criterion['mean'] >= 4.50:
                        criterion['remarks'] = 'OUTSTANDING'
                    elif criterion['mean'] >= 3.50:
                        criterion['remarks'] = 'HIGHLY SATISFACTORY'
                    elif criterion['mean'] >= 2.50:
                        criterion['remarks'] = 'SATISFACTORY'
                    elif criterion['mean'] >= 1.50:
                        criterion['remarks'] = 'NEEDS IMPROVEMENT'
                    else:
                        criterion['remarks'] = 'POOR'
        
        categories_list = list(categories.values())
        
        # Calculate overall rating
        total_mean = 0
        total_criteria = 0
        for category in categories_list:
            for criterion in category['criteria']:
                if criterion['total_responses'] > 0:
                    total_mean += criterion['mean']
                    total_criteria += 1
        
        overall_mean = total_mean / total_criteria if total_criteria > 0 else 0
        
        if overall_mean >= 4.50:
            overall_remarks = 'OUTSTANDING'
        elif overall_mean >= 3.50:
            overall_remarks = 'HIGHLY SATISFACTORY'
        elif overall_mean >= 2.50:
            overall_remarks = 'SATISFACTORY'
        elif overall_mean >= 1.50:
            overall_remarks = 'NEEDS IMPROVEMENT'
        else:
            overall_remarks = 'POOR'
        
        # Get comments
        cursor.execute(f"""
            SELECT DISTINCT c.comment_text
            FROM comments c
            JOIN evaluations e ON c.evaluation_id = e.evaluation_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE cs.faculty_id = %s 
              AND e.status = 'Completed'
              AND c.comment_text IS NOT NULL 
              AND TRIM(c.comment_text) != ''
              {period_filter}
            ORDER BY c.comment_text
        """, tuple(params))
        
        comments_results = cursor.fetchall()
        unique_comments = [row['comment_text'] for row in comments_results if row['comment_text']]
        
        cursor.close()
        conn.close()
        
        # Generate PDF
        filename = f"Evaluation_Report_{faculty['last_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join('static', 'reports', filename)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Add logo and header
        logo_path = os.path.join('static', 'images', 'nclogo.png')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=0.8*inch, height=0.8*inch)
            elements.append(logo)
        
        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=12, textColor=colors.black, spaceAfter=6, alignment=TA_CENTER, fontName='Times-Bold')
        elements.append(Paragraph('NORZAGARAY COLLEGE', title_style))
        elements.append(Paragraph('Municipal Compound, Norzagaray, Bulacan', ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER)))
        elements.append(Paragraph('GUIDANCE AND COUNSELING CENTER', ParagraphStyle('Subtitle2', parent=styles['Normal'], fontSize=10, fontName='Times-Bold', alignment=TA_CENTER, spaceAfter=6)))
        elements.append(Paragraph('FACULTY TEACHING PERFORMANCE EVALUATION SUMMARY REPORT', ParagraphStyle('Subtitle3', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, spaceAfter=3)))
        
        if period_info:
            # Remove any status text (Active), (Closed), etc. from period name using regex
            import re
            period_name = re.sub(r'\s*\([^)]*\)\s*$', '', period_info['period_name']).strip()
            elements.append(Paragraph(f"{period_name}, A.Y. {period_info['year_name']}", ParagraphStyle('Period', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, spaceAfter=6)))
        
        # Add section info to print header if section is selected
        if section:
            elements.append(Paragraph(f"Section: {section['section_name']}", ParagraphStyle('Section', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, spaceAfter=12)))
        else:
            elements.append(Spacer(1, 0.1*inch))
        
        elements.append(Spacer(1, 0.2*inch))
        
        # Faculty info
        elements.append(Paragraph(f"<b>Faculty Name:</b> {faculty['first_name']} {faculty['last_name']}", styles['Normal']))
        if subject:
            elements.append(Paragraph(f"<b>Subject:</b> {subject['subject_code']} - {subject['title']}", styles['Normal']))
        if section:
            elements.append(Paragraph(f"<b>Section:</b> {section['section_name']}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Build results tables
        for category in categories_list:
            # Category header
            elements.append(Paragraph(f"<b>{category['category_name']}</b>", ParagraphStyle('CategoryHeader', parent=styles['Heading2'], fontSize=11, fontName='Times-Bold', spaceAfter=6)))
            
            # Criteria table
            table_data = [['No.', 'Performance Indicators', '5', '4', '3', '2', '1', 'Mean', 'Remarks']]
            
            # Create paragraph styles for table cells
            cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=8, leading=10)
            remarks_style = ParagraphStyle('RemarksStyle', parent=styles['Normal'], fontSize=7, leading=9, alignment=TA_CENTER)
            
            for idx, criterion in enumerate(category['criteria'], 1):
                # Wrap remarks text in Paragraph for text wrapping
                remarks_text = criterion['remarks'] if criterion['total_responses'] > 0 else 'NO DATA'
                
                table_data.append([
                    str(idx),
                    Paragraph(criterion['description'], cell_style),
                    str(criterion['votes']['5']),
                    str(criterion['votes']['4']),
                    str(criterion['votes']['3']),
                    str(criterion['votes']['2']),
                    str(criterion['votes']['1']),
                    f"{criterion['mean']:.2f}" if criterion['total_responses'] > 0 else 'N/A',
                    Paragraph(remarks_text, remarks_style)
                ])
            
            # Adjusted column widths to give more space to remarks while keeping everything on one page
            table = Table(table_data, colWidths=[0.3*inch, 2.8*inch, 0.35*inch, 0.35*inch, 0.35*inch, 0.35*inch, 0.35*inch, 0.5*inch, 1.0*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 0.15*inch))
        
        # Comments section
        if unique_comments:
            elements.append(Paragraph('<b>COMMENTS:</b>', ParagraphStyle('CommentsHeader', parent=styles['Heading3'], fontSize=10, fontName='Times-Bold', spaceAfter=6)))
            for idx, comment in enumerate(unique_comments, 1):
                elements.append(Paragraph(f"{idx}. {comment}", ParagraphStyle('Comment', parent=styles['Normal'], fontSize=9, leftIndent=20, spaceAfter=4)))
            elements.append(Spacer(1, 0.2*inch))
        
        # Rating scale
        elements.append(Paragraph('<b>RATING SCALE:</b>', ParagraphStyle('RatingHeader', parent=styles['Heading3'], fontSize=10, fontName='Times-Bold', spaceAfter=6)))
        rating_data = [
            ['Rating', 'Equivalent', 'Total:', f"{total_mean:.2f}"],
            ['4.50 - 5.00', 'OUTSTANDING', 'Rating:', f"{overall_mean:.2f}"],
            ['3.50 - 4.49', 'HIGHLY SATISFACTORY', 'Remarks:', overall_remarks],
            ['2.50 - 3.49', 'SATISFACTORY', '', ''],
            ['1.50 - 2.49', 'NEEDS IMPROVEMENT', '', ''],
            ['1.00 - 1.49', 'POOR', '', '']
        ]
        
        rating_table = Table(rating_data, colWidths=[1.5*inch, 2*inch, 1.2*inch, 1.5*inch])
        rating_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('SPAN', (2, 3), (3, 5)),
        ]))
        
        elements.append(rating_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Signatures with custom positioning
        # Convert percentage to actual spacing (max available space ~4 inches after rating table)
        max_signature_space = 4.0  # inches available for signature positioning
        
        # Calculate vertical spacing based on Y percentages
        faculty_spacing = (sig_faculty_y / 100.0) * max_signature_space
        dean_spacing = ((sig_dean_y - sig_faculty_y) / 100.0) * max_signature_space
        president_spacing = ((sig_president_y - sig_dean_y) / 100.0) * max_signature_space
        
        # Determine horizontal alignment based on X percentage
        # 0-33% = LEFT, 33-66% = CENTER, 66-100% = RIGHT
        def get_alignment(x_percent):
            print(f"X Percent: {x_percent}")
            if x_percent < 33:
                print("  -> LEFT alignment")
                return TA_LEFT
            elif x_percent < 66:
                print("  -> CENTER alignment")
                return TA_CENTER
            else:
                print("  -> RIGHT alignment")
                return 2  # TA_RIGHT value
        
        print(f"\n=== Signature Positioning Debug ===")
        print(f"Faculty X: {sig_faculty_x}%, Y: {sig_faculty_y}%")
        print(f"Dean X: {sig_dean_x}%, Y: {sig_dean_y}%")
        print(f"President X: {sig_president_x}%, Y: {sig_president_y}%")
        
        faculty_align = get_alignment(sig_faculty_x)
        dean_align = get_alignment(sig_dean_x)
        president_align = get_alignment(sig_president_x)
        
        print(f"Final alignments - Faculty: {faculty_align}, Dean: {dean_align}, President: {president_align}")
        print(f"=== End Debug ===\n")
        
        # Faculty Signature
        elements.append(Spacer(1, faculty_spacing * inch))
        elements.append(Paragraph('_' * 50, ParagraphStyle('SigLine1', parent=styles['Normal'], alignment=faculty_align)))
        elements.append(Paragraph('<b>SIGNATURE OF FACULTY</b>', ParagraphStyle('Sig', parent=styles['Normal'], fontSize=9, spaceAfter=12, alignment=faculty_align)))
        
        # Dean Signature
        elements.append(Spacer(1, dean_spacing * inch))
        elements.append(Paragraph('_' * 50, ParagraphStyle('SigLine2', parent=styles['Normal'], alignment=dean_align)))
        elements.append(Paragraph('<b>SIGNATURE OF COLLEGE DEAN</b>', ParagraphStyle('Sig2', parent=styles['Normal'], fontSize=9, spaceAfter=12, alignment=dean_align)))
        
        # President Signature
        elements.append(Spacer(1, president_spacing * inch))
        elements.append(Paragraph('<b>NOTED BY:</b>', ParagraphStyle('Noted', parent=styles['Normal'], fontSize=9, fontName='Times-Bold', spaceAfter=6, alignment=president_align)))
        elements.append(Paragraph('<b>MA. LIBERTY DG. PASCUAL, Ph.D.</b>', ParagraphStyle('Name', parent=styles['Normal'], fontSize=10, fontName='Times-Bold', alignment=president_align)))
        elements.append(Paragraph('College President', ParagraphStyle('Title', parent=styles['Normal'], alignment=president_align)))
        
        # Build PDF
        doc.build(elements)
        
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/pdf')
        
    except Exception as e:
        print(f"Error exporting to PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/export-evaluation-excel/<int:faculty_id>')
@login_required
def export_evaluation_excel(faculty_id):
    """Export evaluation results to Excel"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from flask import send_file
        import os
        from datetime import datetime
        
        period_id = request.args.get('period_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        section_id = request.args.get('section_id', type=int)
        academic_year_id = request.args.get('academic_year_id', type=int)
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get faculty info
        cursor.execute("""
            SELECT f.faculty_id, f.first_name, f.last_name, f.faculty_number
            FROM faculty f
            WHERE f.faculty_id = %s
        """, (faculty_id,))
        faculty = cursor.fetchone()
        
        if not faculty:
            return jsonify({'success': False, 'message': 'Faculty not found'}), 404
        
        # Get subject info
        cursor.execute("SELECT subject_code, title FROM subjects WHERE subject_id = %s", (subject_id,))
        subject = cursor.fetchone()
        
        # Get section info if section_id is provided
        section = None
        if section_id:
            cursor.execute("SELECT section_name FROM class_sections WHERE section_id = %s", (section_id,))
            section = cursor.fetchone()
        
        # Get period and year info
        cursor.execute("""
            SELECT ep.title as period_name, ay.year_name
            FROM evaluation_periods ep
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE ep.period_id = %s
        """, (period_id,))
        period_info = cursor.fetchone()
        
        # Build filters (same as PDF)
        filters = []
        params = [faculty_id]
        
        if period_id:
            filters.append("e.period_id = %s")
            params.append(period_id)
        if subject_id:
            filters.append("cs.subject_id = %s")
            params.append(subject_id)
        if section_id:
            filters.append("cs.section_id = %s")
            params.append(section_id)
        if academic_year_id:
            filters.append("at.acad_year_id = %s")
            params.append(academic_year_id)
        
        period_filter = ""
        if filters:
            period_filter = "AND " + " AND ".join(filters)
        
        # Get evaluation data
        cursor.execute(f"""
            SELECT 
                c.category_id,
                c.name as category_name,
                c.display_order as category_order,
                cr.criteria_id as criterion_id,
                cr.description,
                cr.order as criterion_order,
                er.rating,
                e.evaluation_id
            FROM evaluation_categories c
            LEFT JOIN evaluation_criteria cr ON c.category_id = cr.category_id
            LEFT JOIN evaluation_responses er ON cr.criteria_id = er.criteria_id
            LEFT JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            LEFT JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE cs.faculty_id = %s 
              AND e.status = 'Completed'
              {period_filter}
            ORDER BY c.display_order, cr.order
        """, tuple(params))
        
        results = cursor.fetchall()
        
        # Process data
        categories = {}
        for row in results:
            category_id = row['category_id']
            if category_id not in categories:
                categories[category_id] = {
                    'category_id': category_id,
                    'category_name': row['category_name'],
                    'category_order': row['category_order'],
                    'criteria': []
                }
            
            if row['criterion_id']:
                criterion_id = row['criterion_id']
                existing_criterion = next((c for c in categories[category_id]['criteria'] if c['criterion_id'] == criterion_id), None)
                
                if not existing_criterion:
                    criterion_data = {
                        'criterion_id': criterion_id,
                        'description': row['description'],
                        'criterion_order': row['criterion_order'],
                        'votes': {'5': 0, '4': 0, '3': 0, '2': 0, '1': 0},
                        'total_responses': 0,
                        'mean': 0,
                        'remarks': 'No Data'
                    }
                    categories[category_id]['criteria'].append(criterion_data)
                    existing_criterion = criterion_data
                
                if row['rating']:
                    rating_str = str(int(row['rating']))
                    existing_criterion['votes'][rating_str] = existing_criterion['votes'].get(rating_str, 0) + 1
                    existing_criterion['total_responses'] += 1
        
        # Calculate means
        total_mean = 0
        total_criteria = 0
        
        for category in categories.values():
            for criterion in category['criteria']:
                if criterion['total_responses'] > 0:
                    total_score = sum(int(rating) * count for rating, count in criterion['votes'].items())
                    criterion['mean'] = total_score / criterion['total_responses']
                    total_mean += criterion['mean']
                    total_criteria += 1
                    
                    if criterion['mean'] >= 4.50:
                        criterion['remarks'] = 'OUTSTANDING'
                    elif criterion['mean'] >= 3.50:
                        criterion['remarks'] = 'HIGHLY SATISFACTORY'
                    elif criterion['mean'] >= 2.50:
                        criterion['remarks'] = 'SATISFACTORY'
                    elif criterion['mean'] >= 1.50:
                        criterion['remarks'] = 'NEEDS IMPROVEMENT'
                    else:
                        criterion['remarks'] = 'POOR'
        
        categories_list = list(categories.values())
        
        overall_mean = total_mean / total_criteria if total_criteria > 0 else 0
        
        if overall_mean >= 4.50:
            overall_remarks = 'OUTSTANDING'
        elif overall_mean >= 3.50:
            overall_remarks = 'HIGHLY SATISFACTORY'
        elif overall_mean >= 2.50:
            overall_remarks = 'SATISFACTORY'
        elif overall_mean >= 1.50:
            overall_remarks = 'NEEDS IMPROVEMENT'
        else:
            overall_remarks = 'POOR'
        
        # Get comments
        cursor.execute(f"""
            SELECT DISTINCT c.comment_text
            FROM comments c
            JOIN evaluations e ON c.evaluation_id = e.evaluation_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE cs.faculty_id = %s 
              AND e.status = 'Completed'
              AND c.comment_text IS NOT NULL 
              AND TRIM(c.comment_text) != ''
              {period_filter}
            ORDER BY c.comment_text
        """, tuple(params))
        
        comments_results = cursor.fetchall()
        unique_comments = [row['comment_text'] for row in comments_results if row['comment_text']]
        
        cursor.close()
        conn.close()
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evaluation Report"
        
        # Styles
        header_font = Font(name='Times New Roman', size=14, bold=True)
        subheader_font = Font(name='Times New Roman', size=11, bold=True)
        normal_font = Font(name='Times New Roman', size=10)
        bold_font = Font(name='Times New Roman', size=10, bold=True)
        
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        gray_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add logo at the top
        row = 1
        try:
            from openpyxl.drawing.image import Image as XLImage
            logo_path = os.path.join('static', 'images', 'nclogo.png')
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                # Make logo bigger (100x100 pixels)
                img.width = 100
                img.height = 100
                # Position the logo in column C-D area (left side, as shown in the format)
                ws.add_image(img, f'C{row}')
                # Set row heights to accommodate larger logo - reduced spacing to lower the logo
                ws.row_dimensions[row].height = 75
                ws.row_dimensions[row + 1].height = 3
                ws.row_dimensions[row + 2].height = 3
                # Add less space for the image to position it lower
                row += 3
        except Exception as e:
            print(f"Could not add logo: {e}")
            pass
        
        # Header section
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'NORZAGARAY COLLEGE'
        cell.font = Font(name='Times New Roman', size=12, bold=True)
        cell.alignment = center_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'Municipal Compound, Norzagaray, Bulacan'
        cell.font = normal_font
        cell.alignment = center_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'GUIDANCE AND COUNSELING CENTER'
        cell.font = Font(name='Times New Roman', size=10, bold=True)
        cell.alignment = center_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'FACULTY TEACHING PERFORMANCE EVALUATION SUMMARY REPORT'
        cell.font = Font(name='Times New Roman', size=9)
        cell.alignment = center_alignment
        row += 1
        
        if period_info:
            # Remove any status text (Active), (Closed), etc. from period name using regex
            import re
            period_name = re.sub(r'\s*\([^)]*\)\s*$', '', period_info['period_name']).strip()
            ws.merge_cells(f'A{row}:I{row}')
            cell = ws[f'A{row}']
            cell.value = f"{period_name}, A.Y. {period_info['year_name']}"
            cell.font = Font(name='Times New Roman', size=9)
            cell.alignment = center_alignment
            row += 1
        
        # Add section info to print header if section is selected
        if section:
            ws.merge_cells(f'A{row}:I{row}')
            cell = ws[f'A{row}']
            cell.value = f"Section: {section['section_name']}"
            cell.font = Font(name='Times New Roman', size=9)
            cell.alignment = center_alignment
            row += 1
        
        row += 1  # Empty row
        
        # Faculty info
        ws[f'A{row}'] = f"Faculty Name: {faculty['first_name']} {faculty['last_name']}"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        if subject:
            ws[f'A{row}'] = f"Subject: {subject['subject_code']} - {subject['title']}"
            ws[f'A{row}'].font = bold_font
            row += 1
        
        if section:
            ws[f'A{row}'] = f"Section: {section['section_name']}"
            ws[f'A{row}'].font = bold_font
            row += 1
        
        row += 1  # Empty row
        
        # Categories and criteria
        for category in categories_list:
            # Category header
            ws.merge_cells(f'A{row}:I{row}')
            cell = ws[f'A{row}']
            cell.value = category['category_name']
            cell.font = subheader_font
            cell.fill = gray_fill
            cell.alignment = left_alignment
            cell.border = thin_border
            row += 1
            
            # Table headers
            headers = ['No.', 'Performance Indicators', '5', '4', '3', '2', '1', 'Mean', 'Remarks']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = bold_font
                cell.fill = gray_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            row += 1
            
            # Criteria data
            for idx, criterion in enumerate(category['criteria'], 1):
                ws.cell(row=row, column=1).value = idx
                ws.cell(row=row, column=2).value = criterion['description']
                ws.cell(row=row, column=3).value = criterion['votes']['5']
                ws.cell(row=row, column=4).value = criterion['votes']['4']
                ws.cell(row=row, column=5).value = criterion['votes']['3']
                ws.cell(row=row, column=6).value = criterion['votes']['2']
                ws.cell(row=row, column=7).value = criterion['votes']['1']
                ws.cell(row=row, column=8).value = f"{criterion['mean']:.2f}" if criterion['total_responses'] > 0 else 'N/A'
                ws.cell(row=row, column=9).value = criterion['remarks']
                
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.border = thin_border
                    if col == 2:
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = center_alignment
                
                row += 1
            
            row += 1  # Empty row between categories
        
        # Comments
        if unique_comments:
            ws.merge_cells(f'A{row}:I{row}')
            cell = ws[f'A{row}']
            cell.value = 'COMMENTS:'
            cell.font = subheader_font
            cell.alignment = left_alignment
            row += 1
            
            for idx, comment in enumerate(unique_comments, 1):
                ws.merge_cells(f'A{row}:I{row}')
                cell = ws[f'A{row}']
                cell.value = f"{idx}. {comment}"
                cell.font = normal_font
                cell.alignment = left_alignment
                row += 1
            
            row += 1
        
        # Rating scale
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'RATING SCALE:'
        cell.font = subheader_font
        cell.alignment = left_alignment
        row += 1
        
        # Rating table
        rating_headers = [['Rating', 'Equivalent', 'Total:', f"{total_mean:.2f}"]]
        rating_data = [
            ['4.50 - 5.00', 'OUTSTANDING', 'Rating:', f"{overall_mean:.2f}"],
            ['3.50 - 4.49', 'HIGHLY SATISFACTORY', 'Remarks:', overall_remarks],
            ['2.50 - 3.49', 'SATISFACTORY', '', ''],
            ['1.50 - 2.49', 'NEEDS IMPROVEMENT', '', ''],
            ['1.00 - 1.49', 'POOR', '', '']
        ]
        
        for data_row in rating_headers + rating_data:
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = bold_font if row == (row - len(rating_data) - len(rating_headers) + 1) else normal_font
                cell.fill = gray_fill if row == (row - len(rating_data) - len(rating_headers) + 1) else PatternFill()
                cell.alignment = center_alignment
                cell.border = thin_border
            row += 1
        
        row += 2  # Empty rows before signatures
        
        # Signature section
        # Signature of Faculty
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = '________________________________'
        cell.alignment = left_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'SIGNATURE OF FACULTY'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 2
        
        # Signature of College Dean
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = '________________________________'
        cell.alignment = left_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'SIGNATURE OF COLLEGE DEAN'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 2
        
        # Noted by section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'NOTED BY:'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 2
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'MA. LIBERTY DG. PASCUAL, Ph.D.'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'College President'
        cell.font = normal_font
        cell.alignment = left_alignment
        row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 20
        
        # Configure page setup for printing
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 0  # Fit all rows on pages
        ws.page_setup.fitToWidth = 1   # Fit to one page wide
        
        # Set print margins (in inches)
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        
        # Set print area (from A1 to last used cell)
        ws.print_area = f'A1:I{row}'
        
        # Center on page when printing
        ws.page_setup.horizontalCentered = True
        
        # Print gridlines for better readability
        ws.print_options.gridLines = False
        ws.print_options.headings = False
        
        # Scale to fit (80% scale for better readability)
        ws.page_setup.scale = 85
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        
        # Print title rows (repeat header on each page if document spans multiple pages)
        # This will repeat the first 6 rows (logo and headers) on each printed page
        ws.print_title_rows = '1:6'
        
        # Set print quality
        ws.page_setup.printQuality = 600
        
        # Save file
        filename = f"Evaluation_Report_{faculty['last_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('static', 'reports', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        print(f"Error exporting to Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/department-analysis/<int:department_id>')
@login_required
def department_analysis(department_id):
    """Get department-level evaluation analysis with ISO 25010 criteria breakdown"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        academic_year_id = request.args.get('academic_year_id', type=int)
        period_id = request.args.get('period_id', type=int)
        
        if not academic_year_id or not period_id:
            return jsonify({'success': False, 'message': 'Academic year and period are required'}), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        try:
            cursor = conn.cursor(dictionary=True)
            
            # Get department info
            cursor.execute("""
                SELECT program_id, name, program_code
                FROM programs
                WHERE program_id = %s
            """, (department_id,))
            department = cursor.fetchone()
            
            if not department:
                return jsonify({'success': False, 'message': 'Department not found'}), 404
            
            # Get period status
            period_status = 'Active'  # Default status
            cursor.execute("""
                SELECT status FROM evaluation_periods WHERE period_id = %s
            """, (period_id,))
            period_result = cursor.fetchone()
            if period_result:
                period_status = period_result['status']
            
            # Get vote distribution per criterion for ALL faculty in this department
            cursor.execute("""
                SELECT 
                    ec.category_id,
                    ec.name as category_name,
                    ecr.criteria_id,
                    ecr.description as criteria_description,
                    ecr.`order`,
                    COUNT(er.response_id) as total_responses,
                    SUM(CASE WHEN er.rating = 5 THEN 1 ELSE 0 END) as votes_5,
                    SUM(CASE WHEN er.rating = 4 THEN 1 ELSE 0 END) as votes_4,
                    SUM(CASE WHEN er.rating = 3 THEN 1 ELSE 0 END) as votes_3,
                    SUM(CASE WHEN er.rating = 2 THEN 1 ELSE 0 END) as votes_2,
                    SUM(CASE WHEN er.rating = 1 THEN 1 ELSE 0 END) as votes_1,
                    AVG(er.rating) as mean_rating
                FROM evaluation_criteria ecr
                JOIN evaluation_categories ec ON ecr.category_id = ec.category_id
                LEFT JOIN evaluation_responses er ON ecr.criteria_id = er.criteria_id
                LEFT JOIN evaluations e ON er.evaluation_id = e.evaluation_id
                LEFT JOIN class_sections cs ON e.section_id = cs.section_id
                LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
                LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
                LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
                WHERE f.program_id = %s
                  AND at.acad_year_id = %s
                  AND e.period_id = %s
                  AND e.status = 'Completed'
                GROUP BY ec.category_id, ec.name, ecr.criteria_id, ecr.description, ecr.`order`
                ORDER BY ec.category_id, ecr.`order`, ecr.criteria_id
            """, (department_id, academic_year_id, period_id))
            
            criteria_results = cursor.fetchall()
            
            # Helper function to determine remarks from mean
            def get_remarks(mean):
                if mean >= 4.5:
                    return "Outstanding"
                elif mean >= 3.5:
                    return "Very Satisfactory"
                elif mean >= 2.5:
                    return "Satisfactory"
                elif mean >= 1.5:
                    return "Fair"
                else:
                    return "Poor"
            
            # Group by category
            categories = {}
            total_mean = 0
            total_criteria = 0
            
            for row in criteria_results:
                category_id = row['category_id']
                
                if category_id not in categories:
                    categories[category_id] = {
                        'category_id': category_id,
                        'category_name': row['category_name'],
                        'criteria': []
                    }
                
                # Calculate mean and remarks
                mean = round(row['mean_rating'], 2) if row['mean_rating'] else 0
                
                criterion_data = {
                    'criteria_id': row['criteria_id'],
                    'description': row['criteria_description'],
                    'votes': {
                        '5': row['votes_5'] or 0,
                        '4': row['votes_4'] or 0,
                        '3': row['votes_3'] or 0,
                        '2': row['votes_2'] or 0,
                        '1': row['votes_1'] or 0
                    },
                    'total_responses': row['total_responses'] or 0,
                    'mean': mean,
                    'remarks': get_remarks(mean) if mean > 0 else 'No Data'
                }
                
                categories[category_id]['criteria'].append(criterion_data)
                
                if mean > 0:
                    total_mean += mean
                    total_criteria += 1
            
            # Convert to list
            categories_list = list(categories.values())
            
            # Calculate overall statistics
            overall_mean = round(total_mean / total_criteria, 2) if total_criteria > 0 else 0
            
            # Get total evaluations count
            cursor.execute("""
                SELECT COUNT(DISTINCT e.evaluation_id) as total_evaluations
                FROM evaluations e
                JOIN class_sections cs ON e.section_id = cs.section_id
                JOIN faculty f ON cs.faculty_id = f.faculty_id
                LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
                LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
                WHERE f.program_id = %s
                  AND at.acad_year_id = %s
                  AND e.period_id = %s
                  AND e.status = 'Completed'
            """, (department_id, academic_year_id, period_id))
            
            eval_count = cursor.fetchone()
            total_evaluations = eval_count['total_evaluations'] if eval_count else 0
            
            cursor.close()
            conn.close()
            
            return jsonify({
                'success': True,
                'department': department,
                'categories': categories_list,
                'overall_mean': overall_mean,
                'total_responses': total_evaluations,
                'comments': [],  # Can be extended later if needed
                'period_status': period_status
            })
            
        finally:
            if conn.is_connected():
                conn.close()
                
    except Exception as e:
        print(f"Error in department analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/export-department-pdf/<int:department_id>')
@login_required
def export_department_pdf(department_id):
    """Export department analysis to PDF with ISO 25010 format"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from flask import send_file
        import os
        from datetime import datetime
        
        academic_year_id = request.args.get('academic_year_id', type=int)
        period_id = request.args.get('period_id', type=int)
        
        # Get signature position parameters
        sig_head_y = request.args.get('sig_head_y', type=float, default=8.33)
        sig_head_x = request.args.get('sig_head_x', type=float, default=10.0)
        sig_dean_y = request.args.get('sig_dean_y', type=float, default=33.33)
        sig_dean_x = request.args.get('sig_dean_x', type=float, default=10.0)
        sig_president_y = request.args.get('sig_president_y', type=float, default=58.33)
        sig_president_x = request.args.get('sig_president_x', type=float, default=10.0)
        
        if not academic_year_id or not period_id:
            return jsonify({'success': False, 'message': 'Academic year and period are required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get department info
        cursor.execute("""
            SELECT program_id, name, program_code
            FROM programs
            WHERE program_id = %s
        """, (department_id,))
        department = cursor.fetchone()
        
        if not department:
            return jsonify({'success': False, 'message': 'Department not found'}), 404
        
        # Get period and year info
        cursor.execute("""
            SELECT ep.title as period_name, ay.year_name
            FROM evaluation_periods ep
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE ep.period_id = %s AND at.acad_year_id = %s
        """, (period_id, academic_year_id))
        period_info = cursor.fetchone()
        
        # Get evaluation criteria with aggregated votes
        cursor.execute("""
            SELECT 
                ec.category_id,
                ec.name as category_name,
                ec.display_order as category_order,
                ecr.criteria_id,
                ecr.description as criteria_description,
                ecr.`order` as criterion_order,
                COUNT(er.response_id) as total_responses,
                SUM(CASE WHEN er.rating = 5 THEN 1 ELSE 0 END) as votes_5,
                SUM(CASE WHEN er.rating = 4 THEN 1 ELSE 0 END) as votes_4,
                SUM(CASE WHEN er.rating = 3 THEN 1 ELSE 0 END) as votes_3,
                SUM(CASE WHEN er.rating = 2 THEN 1 ELSE 0 END) as votes_2,
                SUM(CASE WHEN er.rating = 1 THEN 1 ELSE 0 END) as votes_1,
                AVG(er.rating) as mean_rating
            FROM evaluation_criteria ecr
            JOIN evaluation_categories ec ON ecr.category_id = ec.category_id
            LEFT JOIN evaluation_responses er ON ecr.criteria_id = er.criteria_id
            LEFT JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            LEFT JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE f.program_id = %s
              AND at.acad_year_id = %s
              AND e.period_id = %s
              AND e.status = 'Completed'
            GROUP BY ec.category_id, ec.name, ec.display_order, ecr.criteria_id, ecr.description, ecr.`order`
            ORDER BY ec.display_order, ecr.`order`
        """, (department_id, academic_year_id, period_id))
        
        criteria_results = cursor.fetchall()
        
        # Helper function for remarks
        def get_remarks(mean):
            if mean >= 4.5:
                return "OUTSTANDING"
            elif mean >= 3.5:
                return "HIGHLY SATISFACTORY"
            elif mean >= 2.5:
                return "SATISFACTORY"
            elif mean >= 1.5:
                return "NEEDS IMPROVEMENT"
            else:
                return "POOR"
        
        # Group by category
        categories = {}
        total_mean = 0
        total_criteria = 0
        
        for row in criteria_results:
            category_id = row['category_id']
            
            if category_id not in categories:
                categories[category_id] = {
                    'category_id': category_id,
                    'category_name': row['category_name'],
                    'category_order': row['category_order'],
                    'criteria': []
                }
            
            mean = round(row['mean_rating'], 2) if row['mean_rating'] else 0
            
            criterion_data = {
                'criteria_id': row['criteria_id'],
                'description': row['criteria_description'],
                'votes': {
                    '5': row['votes_5'] or 0,
                    '4': row['votes_4'] or 0,
                    '3': row['votes_3'] or 0,
                    '2': row['votes_2'] or 0,
                    '1': row['votes_1'] or 0
                },
                'total_responses': row['total_responses'] or 0,
                'mean': mean,
                'remarks': get_remarks(mean) if mean > 0 else 'No Data'
            }
            
            categories[category_id]['criteria'].append(criterion_data)
            
            if mean > 0:
                total_mean += mean
                total_criteria += 1
        
        categories_list = sorted(categories.values(), key=lambda x: x['category_order'])
        overall_mean = round(total_mean / total_criteria, 2) if total_criteria > 0 else 0
        overall_remarks = get_remarks(overall_mean)
        
        cursor.close()
        conn.close()
        
        # Create PDF
        filename = f"Department_Evaluation_{department['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join('static', 'reports', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=letter,
                              rightMargin=0.5*inch, leftMargin=0.5*inch,
                              topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=12,
            textColor=colors.black,
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Times-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.black,
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Times-Roman'
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            spaceAfter=4,
            fontName='Times-Roman'
        )
        
        category_style = ParagraphStyle(
            'CategoryHeader',
            parent=styles['Heading2'],
            fontSize=11,
            textColor=colors.black,
            spaceAfter=0,
            spaceBefore=10,
            fontName='Times-Bold',
            alignment=TA_LEFT
        )
        
        # Add logo if exists
        logo_path = os.path.join('static', 'images', 'nclogo.png')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=1.0*inch, height=1.0*inch)
            story.append(logo)
            story.append(Spacer(1, 0.1*inch))
        
        # Header
        story.append(Paragraph("NORZAGARAY COLLEGE", title_style))
        story.append(Paragraph("Municipal Compound, Norzagaray, Bulacan", subtitle_style))
        story.append(Paragraph("GUIDANCE AND COUNSELING CENTER", title_style))
        story.append(Paragraph("DEPARTMENT EVALUATION SUMMARY REPORT", subtitle_style))
        story.append(Paragraph(f"{period_info['period_name']}, A.Y. {period_info['year_name']}", subtitle_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Department info
        story.append(Paragraph(f"<b>Department:</b> {department['name']}", header_style))
        story.append(Paragraph(f"<b>Academic Year:</b> {period_info['year_name']}", header_style))
        story.append(Paragraph(f"<b>Period:</b> {period_info['period_name']}", header_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Categories and criteria tables
        for category in categories_list:
            # Category header
            story.append(Paragraph(category['category_name'].upper(), category_style))
            
            # Build table data
            table_data = [
                ['#', 'Criteria', '5', '4', '3', '2', '1', 'Mean', 'Remarks']
            ]
            
            for idx, criterion in enumerate(category['criteria'], 1):
                if criterion['total_responses'] > 0:
                    row = [
                        str(idx),
                        Paragraph(criterion['description'], ParagraphStyle('CriteriaText', fontSize=7, fontName='Times-Roman')),
                        str(criterion['votes']['5']),
                        str(criterion['votes']['4']),
                        str(criterion['votes']['3']),
                        str(criterion['votes']['2']),
                        str(criterion['votes']['1']),
                        f"{criterion['mean']:.2f}",
                        Paragraph(criterion['remarks'], ParagraphStyle('RemarksText', fontSize=7, fontName='Times-Bold', alignment=TA_CENTER))
                    ]
                    table_data.append(row)
            
            # Create table
            col_widths = [0.3*inch, 3.2*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.4*inch, 0.5*inch, 1.3*inch]
            table = Table(table_data, colWidths=col_widths)
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 0.15*inch))
        
        # Rating scale table
        story.append(Spacer(1, 0.2*inch))
        rating_data = [
            ['RATING', 'EQUIVALENT', 'TOTAL:', f"{total_mean:.2f}"],
            ['4.50 - 5.00', 'OUTSTANDING', 'RATING:', f"{overall_mean:.2f}"],
            ['3.50 - 4.49', 'HIGHLY SATISFACTORY', 'REMARKS:', overall_remarks],
            ['2.50 - 3.49', 'SATISFACTORY', '', ''],
            ['1.50 - 2.49', 'NEEDS IMPROVEMENT', '', ''],
            ['1.00 - 1.49', 'POOR', '', '']
        ]
        
        rating_table = Table(rating_data, colWidths=[1.5*inch, 2.0*inch, 1.2*inch, 1.8*inch])
        rating_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.white),
            ('BACKGROUND', (2, 0), (3, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(rating_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Signatures with custom positioning
        # Create signature table data with proper spacing based on positions
        page_width, page_height = letter
        usable_width = page_width - 1*inch  # Account for margins
        usable_height = 3.5 * inch  # Space available for signatures
        
        # Convert percentages to actual positions
        head_y_pos = (sig_head_y / 100) * usable_height
        dean_y_pos = (sig_dean_y / 100) * usable_height
        president_y_pos = (sig_president_y / 100) * usable_height
        
        # Sort signatures by Y position to maintain proper order
        signatures = [
            (head_y_pos, 'DEPARTMENT HEAD', sig_head_x, False),
            (dean_y_pos, 'COLLEGE DEAN', sig_dean_x, False),
            (president_y_pos, 'COLLEGE PRESIDENT', sig_president_x, True)  # True = show "NOTED BY:"
        ]
        signatures.sort(key=lambda x: x[0])
        
        # Add spacers and signatures in order
        sig_style = ParagraphStyle('Signature', fontSize=9, fontName='Times-Roman', alignment=TA_CENTER)
        sig_bold_style = ParagraphStyle('SignatureBold', fontSize=9, fontName='Times-Bold', alignment=TA_CENTER)
        
        current_pos = 0
        for y_pos, title, x_percent, show_noted_by in signatures:
            # Add spacer to reach this position
            space_needed = y_pos - current_pos
            if space_needed > 0:
                story.append(Spacer(1, space_needed))
            
            # Create a signature block
            if show_noted_by and title == 'COLLEGE PRESIDENT':
                # Special format for president with "NOTED BY:" (no signature line)
                sig_data = [
                    ['', Paragraph('<b>NOTED BY:</b>', sig_bold_style), ''],
                    ['', Paragraph('<b>MA. LIBERTY DG. PASCUAL, Ph.D.</b>', sig_bold_style), ''],
                    ['', Paragraph('College President', sig_style), '']
                ]
            else:
                sig_data = [
                    ['', '_' * 40, ''],
                    ['', Paragraph(f'<b>{title}</b>', sig_bold_style), '']
                ]
            
            # Calculate left padding based on x_percent
            left_padding = (x_percent / 100) * usable_width
            sig_table = Table(sig_data, colWidths=[left_padding, 2.5*inch, usable_width - left_padding - 2.5*inch])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('FONTNAME', (1, 0), (1, 0), 'Times-Roman'),
                ('FONTNAME', (1, 1), (1, 1), 'Times-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            
            story.append(sig_table)
            current_pos = y_pos + 0.5*inch  # Account for signature height
        
        # Build PDF
        doc.build(story)
        
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/pdf')
        
    except Exception as e:
        print(f"Error exporting department PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/export-department-excel/<int:department_id>')
@login_required
def export_department_excel(department_id):
    """Export department analysis to Excel with ISO 25010 format"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from flask import send_file
        import os
        from datetime import datetime
        
        academic_year_id = request.args.get('academic_year_id', type=int)
        period_id = request.args.get('period_id', type=int)
        
        if not academic_year_id or not period_id:
            return jsonify({'success': False, 'message': 'Academic year and period are required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get department info
        cursor.execute("""
            SELECT program_id, name, program_code
            FROM programs
            WHERE program_id = %s
        """, (department_id,))
        department = cursor.fetchone()
        
        if not department:
            return jsonify({'success': False, 'message': 'Department not found'}), 404
        
        # Get period and year info
        cursor.execute("""
            SELECT ep.title as period_name, ay.year_name
            FROM evaluation_periods ep
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE ep.period_id = %s AND at.acad_year_id = %s
        """, (period_id, academic_year_id))
        period_info = cursor.fetchone()
        
        # Get evaluation criteria with aggregated votes
        cursor.execute("""
            SELECT 
                ec.category_id,
                ec.name as category_name,
                ec.display_order as category_order,
                ecr.criteria_id,
                ecr.description as criteria_description,
                ecr.`order` as criterion_order,
                COUNT(er.response_id) as total_responses,
                SUM(CASE WHEN er.rating = 5 THEN 1 ELSE 0 END) as votes_5,
                SUM(CASE WHEN er.rating = 4 THEN 1 ELSE 0 END) as votes_4,
                SUM(CASE WHEN er.rating = 3 THEN 1 ELSE 0 END) as votes_3,
                SUM(CASE WHEN er.rating = 2 THEN 1 ELSE 0 END) as votes_2,
                SUM(CASE WHEN er.rating = 1 THEN 1 ELSE 0 END) as votes_1,
                AVG(er.rating) as mean_rating
            FROM evaluation_criteria ecr
            JOIN evaluation_categories ec ON ecr.category_id = ec.category_id
            LEFT JOIN evaluation_responses er ON ecr.criteria_id = er.criteria_id
            LEFT JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            LEFT JOIN class_sections cs ON e.section_id = cs.section_id
            LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE f.program_id = %s
              AND at.acad_year_id = %s
              AND e.period_id = %s
              AND e.status = 'Completed'
            GROUP BY ec.category_id, ec.name, ec.display_order, ecr.criteria_id, ecr.description, ecr.`order`
            ORDER BY ec.display_order, ecr.`order`
        """, (department_id, academic_year_id, period_id))
        
        criteria_results = cursor.fetchall()
        
        # Helper function for remarks
        def get_remarks(mean):
            if mean >= 4.5:
                return "OUTSTANDING"
            elif mean >= 3.5:
                return "HIGHLY SATISFACTORY"
            elif mean >= 2.5:
                return "SATISFACTORY"
            elif mean >= 1.5:
                return "NEEDS IMPROVEMENT"
            else:
                return "POOR"
        
        # Group by category
        categories = {}
        total_mean = 0
        total_criteria = 0
        
        for row in criteria_results:
            category_id = row['category_id']
            
            if category_id not in categories:
                categories[category_id] = {
                    'category_id': category_id,
                    'category_name': row['category_name'],
                    'category_order': row['category_order'],
                    'criteria': []
                }
            
            mean = round(row['mean_rating'], 2) if row['mean_rating'] else 0
            
            criterion_data = {
                'criteria_id': row['criteria_id'],
                'description': row['criteria_description'],
                'votes': {
                    '5': row['votes_5'] or 0,
                    '4': row['votes_4'] or 0,
                    '3': row['votes_3'] or 0,
                    '2': row['votes_2'] or 0,
                    '1': row['votes_1'] or 0
                },
                'total_responses': row['total_responses'] or 0,
                'mean': mean,
                'remarks': get_remarks(mean) if mean > 0 else 'No Data'
            }
            
            categories[category_id]['criteria'].append(criterion_data)
            
            if mean > 0:
                total_mean += mean
                total_criteria += 1
        
        categories_list = sorted(categories.values(), key=lambda x: x['category_order'])
        overall_mean = round(total_mean / total_criteria, 2) if total_criteria > 0 else 0
        overall_remarks = get_remarks(overall_mean)
        
        cursor.close()
        conn.close()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Department Evaluation"
        
        # Define styles
        header_font = Font(name='Times New Roman', size=14, bold=True)
        subheader_font = Font(name='Times New Roman', size=11, bold=True)
        normal_font = Font(name='Times New Roman', size=10)
        bold_font = Font(name='Times New Roman', size=10, bold=True)
        
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        gray_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add logo at the top
        row = 1
        try:
            from openpyxl.drawing.image import Image as XLImage
            logo_path = os.path.join('static', 'images', 'nclogo.png')
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                # Make logo bigger (100x100 pixels)
                img.width = 100
                img.height = 100
                # Position the logo in column C-D area
                ws.add_image(img, f'C{row}')
                # Set row heights to accommodate larger logo
                ws.row_dimensions[row].height = 75
                ws.row_dimensions[row + 1].height = 3
                ws.row_dimensions[row + 2].height = 3
                row += 3
        except Exception as e:
            print(f"Could not add logo: {e}")
            pass
        
        # Header section
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = "NORZAGARAY COLLEGE"
        cell.font = Font(name='Times New Roman', size=12, bold=True)
        cell.alignment = center_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = "Municipal Compound, Norzagaray, Bulacan"
        cell.font = normal_font
        cell.alignment = center_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = "GUIDANCE AND COUNSELING CENTER"
        cell.font = Font(name='Times New Roman', size=10, bold=True)
        cell.alignment = center_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = "DEPARTMENT EVALUATION SUMMARY REPORT"
        cell.font = Font(name='Times New Roman', size=9)
        cell.alignment = center_alignment
        row += 1
        
        # Remove any status text from period name
        import re
        period_name = re.sub(r'\s*\([^)]*\)\s*$', '', period_info['period_name']).strip()
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = f"{period_name}, A.Y. {period_info['year_name']}"
        cell.font = Font(name='Times New Roman', size=9)
        cell.alignment = center_alignment
        row += 1
        
        row += 1  # Empty row
        
        # Department info
        ws[f'A{row}'] = f"Department: {department['name']}"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        ws[f'A{row}'] = f"Academic Year: {period_info['year_name']}"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        ws[f'A{row}'] = f"Period: {period_name}"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        row += 1  # Empty row
        
        # Categories and criteria
        for category in categories_list:
            # Category header
            ws.merge_cells(f'A{row}:I{row}')
            cell = ws[f'A{row}']
            cell.value = category['category_name'].upper()
            cell.font = subheader_font
            cell.fill = gray_fill
            cell.alignment = left_alignment
            cell.border = thin_border
            row += 1
            
            # Table headers
            headers = ['No.', 'Performance Indicators', '5', '4', '3', '2', '1', 'Mean', 'Remarks']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = bold_font
                cell.fill = gray_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            row += 1
            
            # Criteria data
            for idx, criterion in enumerate(category['criteria'], 1):
                if criterion['total_responses'] > 0:
                    ws.cell(row=row, column=1).value = idx
                    ws.cell(row=row, column=2).value = criterion['description']
                    ws.cell(row=row, column=3).value = criterion['votes']['5']
                    ws.cell(row=row, column=4).value = criterion['votes']['4']
                    ws.cell(row=row, column=5).value = criterion['votes']['3']
                    ws.cell(row=row, column=6).value = criterion['votes']['2']
                    ws.cell(row=row, column=7).value = criterion['votes']['1']
                    ws.cell(row=row, column=8).value = f"{criterion['mean']:.2f}"
                    ws.cell(row=row, column=9).value = criterion['remarks']
                    
                    for col in range(1, 10):
                        cell = ws.cell(row=row, column=col)
                        cell.font = normal_font
                        cell.border = thin_border
                        if col == 2:
                            cell.alignment = left_alignment
                        else:
                            cell.alignment = center_alignment
                    
                    row += 1
            
            row += 1  # Empty row between categories
        
        # Rating scale
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'RATING SCALE:'
        cell.font = subheader_font
        cell.alignment = left_alignment
        row += 1
        
        # Rating table
        rating_headers = [['Rating', 'Equivalent', 'Total:', f"{total_mean:.2f}"]]
        rating_data = [
            ['4.50 - 5.00', 'OUTSTANDING', 'Rating:', f"{overall_mean:.2f}"],
            ['3.50 - 4.49', 'HIGHLY SATISFACTORY', 'Remarks:', overall_remarks],
            ['2.50 - 3.49', 'SATISFACTORY', '', ''],
            ['1.50 - 2.49', 'NEEDS IMPROVEMENT', '', ''],
            ['1.00 - 1.49', 'POOR', '', '']
        ]
        
        for data_row in rating_headers + rating_data:
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = bold_font if row == (row - len(rating_data) - len(rating_headers) + 1) else normal_font
                cell.fill = gray_fill if row == (row - len(rating_data) - len(rating_headers) + 1) else PatternFill()
                cell.alignment = center_alignment
                cell.border = thin_border
            row += 1
        
        row += 2  # Empty rows before signatures
        
        # Signature section
        # Signature of Faculty
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = '________________________________'
        cell.alignment = left_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'SIGNATURE OF FACULTY'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 2
        
        # Signature of College Dean
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = '________________________________'
        cell.alignment = left_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'SIGNATURE OF COLLEGE DEAN'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 2
        
        # Noted by section
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'NOTED BY:'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 2
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'MA. LIBERTY DG. PASCUAL, Ph.D.'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'College President'
        cell.font = normal_font
        cell.alignment = left_alignment
        row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 20
        
        # Save file
        filename = f"Department_Evaluation_{department['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('static', 'reports', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename, 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        print(f"Error exporting department Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/faculty-rankings')
@login_required
def faculty_rankings():
    """Get faculty rankings based on evaluation results for a specific period and academic year"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get query parameters
        academic_year_id = request.args.get('academic_year_id', type=int)
        period_id = request.args.get('period_id', type=int)
        department_id = request.args.get('department_id', type=int)  # Optional
        
        if not academic_year_id or not period_id:
            return jsonify({
                'success': False,
                'message': 'Academic year and period are required'
            }), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        try:
            cursor = conn.cursor(dictionary=True)
            
            # Build the query with optional department filter
            query = """
                SELECT 
                    f.faculty_id,
                    f.first_name,
                    f.last_name,
                    f.program_id,
                    p.name AS department_name,
                    COUNT(DISTINCT e.evaluation_id) AS total_evaluations,
                    -- Overall average rating
                    ROUND(AVG(overall_avg.avg_rating), 2) AS average_rating,
                    -- Learning Delivery (Category ID 2) 
                    ROUND(AVG(cat2.avg_rating), 2) AS learning_delivery,
                    -- Assessment of Student Learning (Category ID 4)
                    ROUND(AVG(cat4.avg_rating), 2) AS assessment_learning,
                    -- Student-Teacher Engagement (Category ID 5)
                    ROUND(AVG(cat5.avg_rating), 2) AS student_engagement
                FROM faculty f
                LEFT JOIN programs p ON f.program_id = p.program_id
                LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                LEFT JOIN evaluations e ON cs.section_id = e.section_id
                    AND e.period_id = %s
                    AND e.status = 'completed'
                LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
                LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
                -- Overall average subquery
                LEFT JOIN (
                    SELECT 
                        e.evaluation_id,
                        AVG(er.rating) as avg_rating
                    FROM evaluations e
                    JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                    WHERE e.period_id = %s AND e.status = 'completed'
                    GROUP BY e.evaluation_id
                ) overall_avg ON e.evaluation_id = overall_avg.evaluation_id
                -- Learning Delivery (Category 2) average
                LEFT JOIN (
                    SELECT 
                        e.evaluation_id,
                        AVG(er.rating) as avg_rating
                    FROM evaluations e
                    JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                    JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
                    WHERE e.period_id = %s AND e.status = 'completed' AND ec.category_id = 2
                    GROUP BY e.evaluation_id
                ) cat2 ON e.evaluation_id = cat2.evaluation_id
                -- Assessment of Student Learning (Category 4) average  
                LEFT JOIN (
                    SELECT 
                        e.evaluation_id,
                        AVG(er.rating) as avg_rating
                    FROM evaluations e
                    JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                    JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
                    WHERE e.period_id = %s AND e.status = 'completed' AND ec.category_id = 4
                    GROUP BY e.evaluation_id
                ) cat4 ON e.evaluation_id = cat4.evaluation_id
                -- Student-Teacher Engagement (Category 5) average
                LEFT JOIN (
                    SELECT 
                        e.evaluation_id,
                        AVG(er.rating) as avg_rating
                    FROM evaluations e
                    JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                    JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
                    WHERE e.period_id = %s AND e.status = 'completed' AND ec.category_id = 5
                    GROUP BY e.evaluation_id
                ) cat5 ON e.evaluation_id = cat5.evaluation_id
                WHERE f.is_archived = 0
                    AND at.acad_year_id = %s
            """
            
            params = [period_id, period_id, period_id, period_id, period_id, academic_year_id]
            
            # Add department filter if provided
            if department_id:
                query += " AND f.program_id = %s"
                params.append(department_id)
            
            query += """
                GROUP BY f.faculty_id, f.first_name, f.last_name, f.program_id, p.name
                HAVING COUNT(DISTINCT e.evaluation_id) > 0
                ORDER BY average_rating DESC, total_evaluations DESC
            """
            
            cursor.execute(query, tuple(params))
            rankings = cursor.fetchall()
            
            # Get period status
            period_status = 'Active'  # Default status
            cursor.execute("""
                SELECT status FROM evaluation_periods WHERE period_id = %s
            """, (period_id,))
            period_result = cursor.fetchone()
            if period_result:
                period_status = period_result['status']
            
            # Format the results
            result_data = []
            for faculty in rankings:
                result_data.append({
                    'faculty_id': faculty['faculty_id'],
                    'first_name': faculty['first_name'],
                    'last_name': faculty['last_name'],
                    'program_id': faculty['program_id'],
                    'department_name': faculty['department_name'],
                    'total_evaluations': faculty['total_evaluations'],
                    'average_rating': float(faculty['average_rating']) if faculty['average_rating'] else 0.0,
                    'learning_delivery': float(faculty['learning_delivery']) if faculty['learning_delivery'] else 0.0,
                    'assessment_learning': float(faculty['assessment_learning']) if faculty['assessment_learning'] else 0.0,
                    'student_engagement': float(faculty['student_engagement']) if faculty['student_engagement'] else 0.0
                })
            
            return jsonify({
                'success': True,
                'data': {
                    'rankings': result_data,
                    'total_faculty': len(result_data),
                    'period_status': period_status
                }
            })
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        print(f"Error fetching faculty rankings: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/save-faculty-recommendation', methods=['POST'])
@login_required
def save_faculty_recommendation():
    """Save or update guidance counselor recommendation for a faculty member"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        faculty_id = data.get('faculty_id')
        recommendation_text = data.get('recommendation', '').strip()
        
        if not faculty_id:
            return jsonify({'success': False, 'message': 'Faculty ID is required'}), 400
        
        if not recommendation_text:
            return jsonify({'success': False, 'message': 'Recommendation text is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        counselor_id = session.get('user_id')
        counselor_name = f"{session.get('first_name', '')} {session.get('last_name', '')}".strip()
        
        # Check if recommendation already exists
        cursor.execute("""
            SELECT recommendation_id 
            FROM faculty_recommendations 
            WHERE faculty_id = %s
        """, (faculty_id,))
        
        existing = cursor.fetchone()
        
        if existing:
            # Update existing recommendation
            cursor.execute("""
                UPDATE faculty_recommendations
                SET recommendation_text = %s,
                    counselor_id = %s,
                    updated_at = NOW()
                WHERE faculty_id = %s
            """, (recommendation_text, counselor_id, faculty_id))
        else:
            # Insert new recommendation
            cursor.execute("""
                INSERT INTO faculty_recommendations 
                (faculty_id, counselor_id, recommendation_text, created_at, updated_at)
                VALUES (%s, %s, %s, NOW(), NOW())
            """, (faculty_id, counselor_id, recommendation_text))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Recommendation saved successfully',
            'counselor_name': counselor_name
        })
        
    except Exception as e:
        print(f"Error saving faculty recommendation: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/generate-ai-recommendation', methods=['POST'])
@login_required
def generate_ai_recommendation():
    """Generate AI-powered recommendation using Gemini based on evaluation results"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        import google.generativeai as genai
        import os
        
        data = request.get_json()
        faculty_id = data.get('faculty_id')
        faculty_name = data.get('faculty_name', 'Unknown Faculty')
        overall_rating = data.get('overall_rating', 'N/A')
        total_evaluations = data.get('total_evaluations', 'N/A')
        response_rate = data.get('response_rate', 'N/A')
        comments = data.get('comments', '')
        
        if not faculty_id:
            return jsonify({'success': False, 'message': 'Faculty ID is required'}), 400
        
        # Configure Gemini API
        # Note: Make sure to set GEMINI_API_KEY in environment variables or config
        api_key = os.environ.get('GEMINI_API_KEY') or 'YOUR_GEMINI_API_KEY_HERE'
        
        if api_key == 'YOUR_GEMINI_API_KEY_HERE':
            return jsonify({
                'success': False, 
                'message': 'Gemini API key not configured. Please contact administrator.'
            }), 500
        
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        # Create comprehensive prompt for AI
        prompt = f"""
As a professional educational consultant, write a detailed and professional recommendation for a faculty member based on their evaluation results.

Faculty Information:
- Name: {faculty_name}
- Overall Rating: {overall_rating} out of 5.0
- Total Evaluations: {total_evaluations}
- Response Rate: {response_rate}

Student Feedback:
{comments if comments else 'No written comments provided by students.'}

Please provide a comprehensive recommendation that includes:
1. Overall performance assessment
2. Key strengths identified from the ratings and feedback
3. Areas for improvement (if rating is below 4.5)
4. Specific actionable recommendations
5. Professional development suggestions (if applicable)
6. Recognition of achievements (if rating is above 4.0)

Write in a professional, constructive, and supportive tone. The recommendation should be 200-300 words and suitable for official faculty evaluation records.

Format the recommendation as a cohesive paragraph without bullet points or numbered lists.
"""
        
        # Generate AI response
        response = model.generate_content(prompt)
        
        if not response or not response.text:
            raise Exception('AI generated empty response')
        
        ai_recommendation = response.text.strip()
        
        return jsonify({
            'success': True,
            'recommendation': ai_recommendation,
            'message': 'AI recommendation generated successfully'
        })
        
    except ImportError:
        return jsonify({
            'success': False,
            'message': 'Gemini AI library not installed. Please install google-generativeai package.'
        }), 500
    except Exception as e:
        print(f"Error generating AI recommendation: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False, 
            'message': f'Failed to generate AI recommendation: {str(e)}'
        }), 500


@api_bp.route('/guidance/generate-strengths-weaknesses', methods=['POST'])
@login_required
def generate_strengths_weaknesses():
    """Generate AI-powered strengths and weaknesses analysis using Gemini"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        import google.generativeai as genai
        import os
        import json
        
        data = request.get_json()
        faculty_id = data.get('faculty_id')
        faculty_name = data.get('faculty_name', 'Unknown Faculty')
        overall_rating = data.get('overall_rating', 'N/A')
        total_evaluations = data.get('total_evaluations', 'N/A')
        response_rate = data.get('response_rate', 'N/A')
        comments = data.get('comments', [])
        rating_categories = data.get('rating_categories', {})
        
        if not faculty_id:
            return jsonify({'success': False, 'message': 'Faculty ID is required'}), 400
        
        # Configure Gemini API
        api_key = os.environ.get('GEMINI_API_KEY') or 'YOUR_GEMINI_API_KEY_HERE'
        
        if api_key == 'YOUR_GEMINI_API_KEY_HERE':
            return jsonify({
                'success': False, 
                'message': 'Gemini API key not configured. Please contact administrator.'
            }), 500
        
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        # Prepare comments summary
        comments_text = ""
        if comments:
            positive_comments = [c['comment'] for c in comments if c.get('rating', 0) >= 4]
            neutral_comments = [c['comment'] for c in comments if c.get('rating', 0) == 3]
            negative_comments = [c['comment'] for c in comments if c.get('rating', 0) <= 2]
            
            if positive_comments:
                comments_text += f"\n\nPositive Comments ({len(positive_comments)}):\n" + "\n".join(f"- {c}" for c in positive_comments[:10])
            if neutral_comments:
                comments_text += f"\n\nNeutral Comments ({len(neutral_comments)}):\n" + "\n".join(f"- {c}" for c in neutral_comments[:5])
            if negative_comments:
                comments_text += f"\n\nCritical Comments ({len(negative_comments)}):\n" + "\n".join(f"- {c}" for c in negative_comments[:10])
        
        # Prepare rating categories summary
        categories_text = ""
        if rating_categories:
            categories_text = "\n\nRating Categories:\n" + "\n".join([f"- {cat}: {rating}" for cat, rating in rating_categories.items()])
        
        # Create comprehensive prompt for AI
        prompt = f"""
As an educational assessment expert, analyze the following faculty evaluation data and provide a comprehensive strengths and weaknesses analysis.

Faculty Information:
- Name: {faculty_name}
- Overall Rating: {overall_rating} out of 5.0
- Total Evaluations: {total_evaluations}
- Response Rate: {response_rate}
{categories_text}
{comments_text}

Based on this data, provide a detailed analysis in the following JSON format:
{{
    "strengths": [
        "First key strength with specific evidence",
        "Second key strength with specific evidence",
        "Third key strength with specific evidence"
    ],
    "weaknesses": [
        "First area for improvement with specific details",
        "Second area for improvement with specific details",
        "Third area for improvement with specific details"
    ],
    "insights": "A comprehensive paragraph (100-150 words) that synthesizes the overall performance, highlights patterns in student feedback, and provides context for the strengths and weaknesses identified. Include actionable insights for professional development."
}}

Guidelines:
1. Identify 3-5 specific strengths based on high ratings and positive feedback
2. Identify 2-4 areas for improvement, even if the overall rating is high (focus on growth opportunities)
3. Be specific and evidence-based - reference actual ratings or feedback patterns
4. Use professional, constructive language
5. Ensure insights paragraph is cohesive and actionable
6. If rating is below 3.0, be more detailed about improvement areas
7. If rating is above 4.5, focus on sustaining excellence and minor refinements

Return ONLY the JSON object, no additional text or formatting.
"""
        
        # Generate AI response
        response = model.generate_content(prompt)
        
        if not response or not response.text:
            raise Exception('AI generated empty response')
        
        # Parse the JSON response
        ai_text = response.text.strip()
        
        # Remove markdown code blocks if present
        if ai_text.startswith('```json'):
            ai_text = ai_text.replace('```json', '').replace('```', '').strip()
        elif ai_text.startswith('```'):
            ai_text = ai_text.replace('```', '').strip()
        
        # Parse JSON
        analysis = json.loads(ai_text)
        
        # Validate structure
        if not isinstance(analysis.get('strengths'), list):
            raise Exception('Invalid analysis format: strengths must be a list')
        if not isinstance(analysis.get('weaknesses'), list):
            raise Exception('Invalid analysis format: weaknesses must be a list')
        if not isinstance(analysis.get('insights'), str):
            raise Exception('Invalid analysis format: insights must be a string')
        
        return jsonify({
            'success': True,
            'analysis': analysis,
            'message': 'AI analysis generated successfully'
        })
        
    except ImportError:
        return jsonify({
            'success': False,
            'message': 'Gemini AI library not installed. Please install google-generativeai package.'
        }), 500
    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {str(e)}")
        print(f"AI Response: {response.text if 'response' in locals() else 'No response'}")
        return jsonify({
            'success': False, 
            'message': 'Failed to parse AI response. Please try again.'
        }), 500
    except Exception as e:
        print(f"Error generating AI analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False, 
            'message': f'Failed to generate AI analysis: {str(e)}'
        }), 500


@api_bp.route('/guidance/generate-comments-summary', methods=['POST'])
@login_required
def generate_comments_summary():
    """Generate AI-powered summary of all student comments"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        import google.generativeai as genai
        import os
        import json
        
        data = request.get_json()
        faculty_id = data.get('faculty_id')
        faculty_name = data.get('faculty_name', 'Unknown Faculty')
        overall_rating = data.get('overall_rating', 'N/A')
        total_comments = data.get('total_comments', 0)
        positive_count = data.get('positive_count', 0)
        neutral_count = data.get('neutral_count', 0)
        negative_count = data.get('negative_count', 0)
        comments = data.get('comments', [])
        
        if not faculty_id:
            return jsonify({'success': False, 'message': 'Faculty ID is required'}), 400
        
        if not comments or len(comments) == 0:
            return jsonify({'success': False, 'message': 'No comments available to summarize'}), 400
        
        # Configure Gemini API
        api_key = os.environ.get('GEMINI_API_KEY') or 'YOUR_GEMINI_API_KEY_HERE'
        
        if api_key == 'YOUR_GEMINI_API_KEY_HERE':
            return jsonify({
                'success': False, 
                'message': 'Gemini API key not configured. Please contact administrator.'
            }), 500
        
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        # Organize comments by rating
        positive_comments = [c['comment'] for c in comments if c.get('rating', 0) >= 4 and c.get('comment', '').strip()]
        neutral_comments = [c['comment'] for c in comments if c.get('rating', 0) == 3 and c.get('comment', '').strip()]
        negative_comments = [c['comment'] for c in comments if c.get('rating', 0) <= 2 and c.get('comment', '').strip()]
        
        # Prepare comments text
        comments_text = f"""
Total Comments: {total_comments}
- Positive (4-5 stars): {positive_count}
- Neutral (3 stars): {neutral_count}
- Critical (1-2 stars): {negative_count}

POSITIVE COMMENTS:
{chr(10).join(f'{i+1}. "{comment}"' for i, comment in enumerate(positive_comments[:20]))}

{f'''NEUTRAL COMMENTS:
{chr(10).join(f'{i+1}. "{comment}"' for i, comment in enumerate(neutral_comments[:10]))}
''' if neutral_comments else ''}

{f'''CRITICAL COMMENTS:
{chr(10).join(f'{i+1}. "{comment}"' for i, comment in enumerate(negative_comments[:20]))}
''' if negative_comments else ''}
"""
        
        # Create comprehensive prompt for AI
        prompt = f"""
As an educational assessment analyst, provide a comprehensive summary of the following student feedback for {faculty_name}.

Overall Rating: {overall_rating} out of 5.0

{comments_text}

Analyze these comments and provide a structured summary in the following JSON format:
{{
    "overall_sentiment": "2-3 sentences describing the overall tone and sentiment of student feedback. Be balanced and objective.",
    "key_themes": [
        "Theme 1: Brief description of a recurring topic",
        "Theme 2: Brief description of another recurring topic",
        "Theme 3: Brief description of another recurring topic"
    ],
    "common_praise": [
        "Specific aspect students appreciate (with context)",
        "Another aspect students appreciate (with context)",
        "Third aspect if applicable"
    ],
    "areas_for_improvement": [
        "Specific area mentioned by students (with context)",
        "Another area mentioned (with context)",
        "Third area if applicable"
    ],
    "actionable_insights": "2-3 sentences providing guidance counselors with actionable insights based on the feedback patterns. Focus on what this means for faculty development and student support."
}}

Guidelines:
1. Identify 3-5 key themes that emerge across all comments
2. Extract specific praise that appears frequently
3. Identify constructive criticism and areas for improvement
4. Be objective and balanced - acknowledge both strengths and weaknesses
5. Use specific examples from comments when possible
6. Provide actionable insights that counselors can use
7. Keep language professional and constructive
8. If comments are overwhelmingly positive or negative, acknowledge that pattern
9. Focus on patterns, not isolated comments

Return ONLY the JSON object, no additional text or formatting.
"""
        
        # Generate AI response
        response = model.generate_content(prompt)
        
        if not response or not response.text:
            raise Exception('AI generated empty response')
        
        # Parse the JSON response
        ai_text = response.text.strip()
        
        # Remove markdown code blocks if present
        if ai_text.startswith('```json'):
            ai_text = ai_text.replace('```json', '').replace('```', '').strip()
        elif ai_text.startswith('```'):
            ai_text = ai_text.replace('```', '').strip()
        
        # Parse JSON
        summary = json.loads(ai_text)
        
        # Validate structure
        required_keys = ['overall_sentiment', 'key_themes', 'common_praise', 'areas_for_improvement', 'actionable_insights']
        for key in required_keys:
            if key not in summary:
                raise Exception(f'Invalid summary format: missing {key}')
        
        return jsonify({
            'success': True,
            'summary': summary,
            'message': 'Comments summary generated successfully'
        })
        
    except ImportError:
        return jsonify({
            'success': False,
            'message': 'Gemini AI library not installed. Please install google-generativeai package.'
        }), 500
    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {str(e)}")
        print(f"AI Response: {response.text if 'response' in locals() else 'No response'}")
        return jsonify({
            'success': False, 
            'message': 'Failed to parse AI response. Please try again.'
        }), 500
    except Exception as e:
        print(f"Error generating comments summary: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False, 
            'message': f'Failed to generate comments summary: {str(e)}'
        }), 500


@api_bp.route('/guidance/filter-offensive-comments', methods=['POST'])
@login_required
def filter_offensive_comments():
    """
    Use Gemini AI to detect offensive language in student comments (any language)
    Returns list of comments with offensive flag
    """
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        import google.generativeai as genai
        import os
        
        data = request.get_json()
        comments = data.get('comments', [])
        
        if not comments or len(comments) == 0:
            return jsonify({'success': True, 'results': []})
        
        # Configure Gemini API
        api_key = os.environ.get('GEMINI_API_KEY') or 'YOUR_GEMINI_API_KEY_HERE'
        
        if api_key == 'YOUR_GEMINI_API_KEY_HERE':
            return jsonify({
                'success': False, 
                'message': 'Gemini API key not configured. Please contact administrator.'
            }), 500
        
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        # Prepare comments for batch analysis
        comments_text = "\n\n".join([
            f"Comment ID {comment['id']}: {comment['text']}"
            for comment in comments
        ])
        
        # Create AI prompt for offensive language detection
        prompt = f"""
You are a content moderation AI assistant. Analyze the following student feedback comments and identify which ones contain offensive, inappropriate, or harmful language in ANY LANGUAGE.

Consider the following as offensive:
- Profanity, vulgar language, or explicit content
- Hate speech, discriminatory remarks, or slurs
- Personal attacks, threats, or harassment
- Bullying, mockery, or humiliation
- Sexual content or inappropriate references
- Extremely disrespectful or degrading language

Do NOT flag comments that are:
- Constructive criticism (even if negative)
- Honest feedback about teaching quality
- Expressions of frustration without offensive language
- Simple complaints or suggestions for improvement

Comments to analyze:
{comments_text}

Return ONLY a JSON array with this exact format (no other text):
[
  {{"id": 0, "is_offensive": true}},
  {{"id": 1, "is_offensive": false}},
  ...
]

Include all comment IDs from the input. Return valid JSON only.
"""
        
        # Generate AI response
        response = model.generate_content(prompt)
        
        if not response or not response.text:
            raise Exception('AI generated empty response')
        
        ai_response = response.text.strip()
        
        # Extract JSON from response (remove markdown code blocks if present)
        if '```json' in ai_response:
            ai_response = ai_response.split('```json')[1].split('```')[0].strip()
        elif '```' in ai_response:
            ai_response = ai_response.split('```')[1].split('```')[0].strip()
        
        # Parse JSON response
        import json
        results = json.loads(ai_response)
        
        # Validate results
        if not isinstance(results, list):
            raise Exception('Invalid AI response format')
        
        return jsonify({
            'success': True,
            'results': results,
            'total_analyzed': len(comments),
            'offensive_count': sum(1 for r in results if r.get('is_offensive', False))
        })
        
    except ImportError:
        return jsonify({
            'success': False,
            'message': 'Gemini AI library not installed. Please install google-generativeai package.'
        }), 500
    except json.JSONDecodeError as e:
        print(f"JSON decode error: {str(e)}")
        print(f"AI Response: {ai_response}")
        return jsonify({
            'success': False,
            'message': 'Failed to parse AI response. Please try again.'
        }), 500
    except Exception as e:
        print(f"Error filtering offensive comments: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False, 
            'message': f'Failed to filter comments: {str(e)}'
        }), 500


@api_bp.route('/analytics/faculty-performance-trend')
@login_required
def faculty_performance_trend():
    """
    Get faculty performance trend across multiple evaluation periods
    Shows how a specific faculty member's performance changes over time
    """
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    faculty_id = request.args.get('faculty_id')
    
    if not faculty_id:
        return jsonify({'success': False, 'message': 'Faculty ID is required'}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get faculty details
        cursor.execute("""
            SELECT 
                faculty_id,
                CONCAT(first_name, ' ', last_name) as faculty_name,
                email,
                faculty_number
            FROM faculty
            WHERE faculty_id = %s AND is_archived = 0
        """, (faculty_id,))
        
        faculty = cursor.fetchone()
        
        if not faculty:
            return jsonify({'success': False, 'message': 'Faculty not found'}), 404
        
        # Get performance data across all evaluation periods
        # Fixed JOIN order to ensure faculty_id filter works correctly
        cursor.execute("""
            SELECT 
                ep.period_id,
                ep.title as period_title,
                ep.start_date,
                ep.end_date,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as completed_evaluations,
                AVG(CASE WHEN e.status = 'Completed' THEN er.rating END) as average_rating,
                COUNT(DISTINCT CASE WHEN c.comment_text IS NOT NULL AND c.comment_text != '' THEN c.comment_id END) as comment_count,
                COUNT(DISTINCT CASE WHEN e.status = 'Completed' AND er.rating >= 4 THEN er.response_id END) as positive_count,
                COUNT(DISTINCT CASE WHEN e.status = 'Completed' AND er.rating <= 2 THEN er.response_id END) as negative_count
            FROM class_sections cs
            INNER JOIN evaluations e ON cs.section_id = e.section_id
            INNER JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            LEFT JOIN comments c ON e.evaluation_id = c.evaluation_id
            WHERE cs.faculty_id = %s
                AND ep.is_archived = 0
            GROUP BY ep.period_id, ep.title, ep.start_date, ep.end_date
            HAVING completed_evaluations > 0
            ORDER BY ep.start_date ASC
        """, (faculty_id,))
        
        trend_data = cursor.fetchall()
        
        print(f"Faculty Performance Trend - Faculty ID: {faculty_id}, Records found: {len(trend_data)}")
        
        # Calculate overall score for each period (average rating * response rate weight)
        for item in trend_data:
            if item['total_evaluations'] > 0:
                response_rate = (item['completed_evaluations'] / item['total_evaluations']) * 100
                item['response_rate'] = round(response_rate, 1)
                
                # Calculate overall score (70% rating + 30% response rate normalized)
                # Convert Decimal to float to avoid type errors
                rating_score = float(item['average_rating'] or 0)
                response_score = (response_rate / 100) * 5  # Normalize to 5-point scale
                item['overall_score'] = round((rating_score * 0.7) + (response_score * 0.3), 2)
            else:
                item['response_rate'] = 0
                item['overall_score'] = 0
            
            # Format dates
            if item['start_date']:
                item['start_date'] = item['start_date'].strftime('%Y-%m-%d')
            if item['end_date']:
                item['end_date'] = item['end_date'].strftime('%Y-%m-%d')
        
        print(f"Returning {len(trend_data)} periods of data for {faculty['faculty_name']}")
        
        return jsonify({
            'success': True,
            'faculty_name': faculty['faculty_name'],
            'faculty_email': faculty['email'],
            'faculty_number': faculty['faculty_number'],
            'data': trend_data,
            'total_periods': len(trend_data)
        })
        
    except Exception as e:
        print(f"Error fetching faculty performance trend: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': 'Error fetching performance trend data',
            'error': str(e)
        }), 500
    finally:
        cursor.close()
        conn.close()


@api_bp.route('/analytics/faculty-comparison', methods=['POST'])
@login_required
def faculty_comparison():
    """
    Compare selected faculty performance between current and previous evaluation periods
    Includes AI-generated summary of the comparison
    """
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        period_id = data.get('period_id')
        faculty_ids = data.get('faculty_ids', [])
        
        if not period_id:
            return jsonify({'success': False, 'message': 'Period ID is required'}), 400
        
        if not faculty_ids or len(faculty_ids) == 0:
            return jsonify({'success': False, 'message': 'At least one faculty member must be selected'}), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Get current period details
        cursor.execute("""
            SELECT ep.period_id, ep.title, ep.start_date, ep.end_date,
                   at.acad_term_id, at.term_name,
                   ay.year_code
            FROM evaluation_periods ep
            INNER JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            INNER JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE ep.period_id = %s
        """, (period_id,))
        current_period = cursor.fetchone()
        
        if not current_period:
            return jsonify({'success': False, 'message': 'Current period not found'}), 404
        
        # Get previous period (same term type but earlier)
        cursor.execute("""
            SELECT ep.period_id, ep.title, ep.start_date, ep.end_date
            FROM evaluation_periods ep
            INNER JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            WHERE ep.start_date < %s
            ORDER BY ep.start_date DESC
            LIMIT 1
        """, (current_period['start_date'],))
        previous_period = cursor.fetchone()
        
        if not previous_period:
            return jsonify({
                'success': False,
                'message': 'No previous evaluation period found for comparison'
            }), 404
        
        # Build comparison data for each faculty
        comparisons = []
        
        for faculty_id in faculty_ids:
            # Get current period performance
            cursor.execute("""
                SELECT 
                    f.faculty_id,
                    CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                    GROUP_CONCAT(DISTINCT s.subject_code ORDER BY s.subject_code SEPARATOR ', ') as subjects,
                    COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                    COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as completed_evaluations,
                    AVG(CASE WHEN e.status = 'Completed' THEN er.rating END) as average_rating
                FROM faculty f
                LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                LEFT JOIN evaluations e ON cs.section_id = e.section_id AND e.period_id = %s
                LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                LEFT JOIN subjects s ON cs.subject_id = s.subject_id
                WHERE f.faculty_id = %s AND f.is_archived = 0
                GROUP BY f.faculty_id, f.first_name, f.last_name
            """, (period_id, faculty_id))
            current_data = cursor.fetchone()
            
            if not current_data:
                continue
            
            # Get previous period performance
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                    COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as completed_evaluations,
                    AVG(CASE WHEN e.status = 'Completed' THEN er.rating END) as average_rating
                FROM faculty f
                LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                LEFT JOIN evaluations e ON cs.section_id = e.section_id AND e.period_id = %s
                LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                WHERE f.faculty_id = %s
                GROUP BY f.faculty_id
            """, (previous_period['period_id'], faculty_id))
            previous_data = cursor.fetchone()
            
            # Calculate scores
            current_rating = float(current_data['average_rating'] or 0)
            current_response_rate = 0
            if current_data['total_evaluations'] > 0:
                current_response_rate = (current_data['completed_evaluations'] / current_data['total_evaluations']) * 100
            current_response_score = (current_response_rate / 100) * 5
            current_score = (current_rating * 0.7) + (current_response_score * 0.3)
            
            previous_rating = 0
            previous_score = 0
            if previous_data and previous_data['average_rating']:
                previous_rating = float(previous_data['average_rating'])
                previous_response_rate = 0
                if previous_data['total_evaluations'] > 0:
                    previous_response_rate = (previous_data['completed_evaluations'] / previous_data['total_evaluations']) * 100
                previous_response_score = (previous_response_rate / 100) * 5
                previous_score = (previous_rating * 0.7) + (previous_response_score * 0.3)
            
            score_change = current_score - previous_score
            
            comparisons.append({
                'faculty_id': current_data['faculty_id'],
                'faculty_name': current_data['faculty_name'],
                'subject_name': current_data['subjects'] or 'N/A',
                'current_score': round(current_score, 2),
                'previous_score': round(previous_score, 2),
                'score_change': round(score_change, 2),
                'current_rating': round(current_rating, 2),
                'previous_rating': round(previous_rating, 2),
                'current_responses': current_data['completed_evaluations'],
                'previous_responses': previous_data['completed_evaluations'] if previous_data else 0
            })
        
        # Generate AI summary
        ai_summary = generate_comparison_summary(comparisons, current_period['title'], previous_period['title'])
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'data': comparisons,
            'ai_summary': ai_summary,
            'current_period': current_period['title'],
            'previous_period': previous_period['title'],
            'total_compared': len(comparisons)
        })
        
    except Exception as e:
        print(f"Error generating faculty comparison: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': 'Error generating comparison',
            'error': str(e)
        }), 500


def generate_comparison_summary(comparisons, current_period, previous_period):
    """
    Generate AI summary of faculty performance comparison
    """
    try:
        from utils.ai_support import initialize_gemini
        
        # Prepare comparison data for AI
        improved = [c for c in comparisons if c['score_change'] > 0.2]
        declined = [c for c in comparisons if c['score_change'] < -0.2]
        stable = [c for c in comparisons if -0.2 <= c['score_change'] <= 0.2]
        
        avg_change = sum(c['score_change'] for c in comparisons) / len(comparisons) if comparisons else 0
        
        prompt = f"""
Analyze the following faculty performance comparison data between "{previous_period}" and "{current_period}":

**Overall Statistics:**
- Total Faculty Compared: {len(comparisons)}
- Average Score Change: {avg_change:.2f}
- Improved: {len(improved)} faculty members
- Declined: {len(declined)} faculty members
- Stable: {len(stable)} faculty members

**Top Improvements:**
{chr(10).join([f"- {c['faculty_name']}: {c['previous_score']:.2f} → {c['current_score']:.2f} (+{c['score_change']:.2f})" for c in sorted(improved, key=lambda x: x['score_change'], reverse=True)[:3]])}

**Notable Declines:**
{chr(10).join([f"- {c['faculty_name']}: {c['previous_score']:.2f} → {c['current_score']:.2f} ({c['score_change']:.2f})" for c in sorted(declined, key=lambda x: x['score_change'])[:3]])}

Provide a comprehensive but concise summary (150-200 words) that:
1. Highlights overall trends and patterns
2. Identifies standout performers (both positive and negative)
3. Suggests potential factors for changes
4. Provides actionable recommendations for guidance counselors

Use **bold** for emphasis on key points. Be professional and constructive.
"""
        
        model = initialize_gemini()
        response = model.generate_content(prompt)
        
        if response and response.text:
            return response.text.strip()
        else:
            return "AI summary generation is temporarily unavailable. The comparison data shows performance changes across the selected faculty members."
            
    except Exception as e:
        print(f"Error generating AI summary: {str(e)}")
        return f"Comparison Summary: {len(comparisons)} faculty members were analyzed. {len(improved)} showed improvement, {len(declined)} showed decline, and {len(stable)} remained stable."


@api_bp.route('/guidance/export-faculty-report', methods=['POST'])
@login_required
def export_faculty_report():
    """Export comprehensive faculty evaluation report as PDF or Excel"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from io import BytesIO
        from datetime import datetime
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        data = request.get_json()
        faculty_id = data.get('faculty_id')
        export_format = data.get('format', 'pdf')
        include_comments = data.get('include_comments', True)
        include_ratings = data.get('include_ratings', True)
        
        if not faculty_id:
            return jsonify({'success': False, 'message': 'Faculty ID is required'}), 400
        
        # Get faculty data
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        try:
            cursor = conn.cursor(dictionary=True)
            
            # Get faculty details
            cursor.execute("""
                SELECT 
                    f.faculty_id,
                    f.faculty_number,
                    f.first_name,
                    f.last_name,
                    f.email,
                    p.name as department_name,
                    COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                    COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as completed_evaluations
                FROM faculty f
                LEFT JOIN programs p ON f.program_id = p.program_id
                LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                LEFT JOIN evaluations e ON cs.section_id = e.section_id
                WHERE f.faculty_id = %s
                GROUP BY f.faculty_id, f.first_name, f.last_name, f.email, f.faculty_number, p.name
            """, (faculty_id,))
            
            faculty = cursor.fetchone()
            
            if not faculty:
                return jsonify({'success': False, 'message': 'Faculty not found'}), 404
            
            # Get overall rating
            cursor.execute("""
                SELECT AVG(er.rating) as overall_rating
                FROM evaluation_responses er
                JOIN evaluations e ON er.evaluation_id = e.evaluation_id
                JOIN class_sections cs ON e.section_id = cs.section_id
                WHERE cs.faculty_id = %s AND e.status = 'Completed'
            """, (faculty_id,))
            
            rating_result = cursor.fetchone()
            faculty['overall_rating'] = rating_result['overall_rating'] if rating_result and rating_result['overall_rating'] else 0
            
            if not faculty:
                return jsonify({'success': False, 'message': 'Faculty not found'}), 404
            
            # Get comments if requested
            comments = []
            if include_comments:
                cursor.execute("""
                    SELECT 
                        c.comment_text,
                        c.sentiment,
                        c.created_at,
                        AVG(er.rating) as rating
                    FROM comments c
                    JOIN evaluations e ON c.evaluation_id = e.evaluation_id
                    JOIN class_sections cs ON e.section_id = cs.section_id
                    LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                    WHERE cs.faculty_id = %s AND c.comment_text IS NOT NULL AND c.comment_text != ''
                    GROUP BY c.comment_id, c.comment_text, c.sentiment, c.created_at
                    ORDER BY c.created_at DESC
                """, (faculty_id,))
                comments = cursor.fetchall()
            
            # Get rating breakdown if requested
            rating_breakdown = {}
            if include_ratings:
                cursor.execute("""
                    SELECT 
                        ec.name as category_name,
                        AVG(er.rating) as avg_rating,
                        COUNT(er.response_id) as count
                    FROM evaluation_responses er
                    JOIN evaluations e ON er.evaluation_id = e.evaluation_id
                    JOIN class_sections cs ON e.section_id = cs.section_id
                    JOIN evaluation_criteria ecr ON er.criteria_id = ecr.criteria_id
                    JOIN evaluation_categories ec ON ecr.category_id = ec.category_id
                    WHERE cs.faculty_id = %s AND e.status = 'Completed'
                    GROUP BY ec.category_id, ec.name
                    ORDER BY ec.name
                """, (faculty_id,))
                ratings = cursor.fetchall()
                for rating in ratings:
                    rating_breakdown[rating['category_name']] = {
                        'average': float(rating['avg_rating']) if rating['avg_rating'] else 0,
                        'count': rating['count']
                    }
            
            # Get recommendation
            cursor.execute("""
                SELECT 
                    fr.recommendation_text,
                    fr.created_at,
                    CONCAT(u.first_name, ' ', u.last_name) as counselor_name
                FROM faculty_recommendations fr
                LEFT JOIN users u ON fr.counselor_id = u.user_id
                WHERE fr.faculty_id = %s
                ORDER BY fr.updated_at DESC
                LIMIT 1
            """, (faculty_id,))
            recommendation = cursor.fetchone()
            
            # Calculate response rate
            response_rate = 0
            if faculty['total_evaluations'] > 0:
                response_rate = (faculty['completed_evaluations'] / faculty['total_evaluations']) * 100
            
            # Generate report based on format
            if export_format == 'pdf':
                return generate_pdf_report(faculty, comments, rating_breakdown, recommendation, response_rate)
            else:
                return generate_excel_report(faculty, comments, rating_breakdown, recommendation, response_rate)
                
        finally:
            cursor.close()
            conn.close()
            
    except ImportError as e:
        return jsonify({
            'success': False,
            'message': f'Required library not installed: {str(e)}'
        }), 500
    except Exception as e:
        print(f"Error exporting faculty report: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Failed to export report: {str(e)}'
        }), 500


def generate_pdf_report(faculty, comments, rating_breakdown, recommendation, response_rate):
    """Generate PDF report"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from io import BytesIO
    from datetime import datetime
    from flask import send_file
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0059cc'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#0059cc'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    subheading_style = ParagraphStyle(
        'CustomSubHeading',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#004799'),
        spaceAfter=6,
        fontName='Helvetica-Bold'
    )
    
    # Title
    elements.append(Paragraph("Faculty Evaluation Report", title_style))
    elements.append(Paragraph("IntellEvalPro - Comprehensive Performance Analysis", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Faculty Information
    elements.append(Paragraph("Faculty Information", heading_style))
    
    faculty_info_data = [
        ['Faculty Name:', f"{faculty['first_name']} {faculty['last_name']}"],
        ['Faculty Number:', faculty.get('faculty_number', 'N/A')],
        ['Department:', faculty.get('department_name', 'N/A')],
        ['Email:', faculty.get('email', 'N/A')],
        ['Report Generated:', datetime.now().strftime('%B %d, %Y at %I:%M %p')]
    ]
    
    faculty_table = Table(faculty_info_data, colWidths=[2*inch, 4.5*inch])
    faculty_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E6F0FF')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0059cc')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(faculty_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Performance Summary
    elements.append(Paragraph("Performance Summary", heading_style))
    
    overall_rating = float(faculty['overall_rating']) if faculty['overall_rating'] else 0
    rating_color = colors.green if overall_rating >= 4.0 else colors.red if overall_rating < 3.0 else colors.orange
    
    performance_data = [
        ['Overall Rating', 'Total Evaluations', 'Response Rate', 'Performance Level'],
        [
            f"{overall_rating:.2f} / 5.0",
            f"{faculty['completed_evaluations']} / {faculty['total_evaluations']}",
            f"{response_rate:.1f}%",
            'Excellent' if overall_rating >= 4.5 else 'Very Good' if overall_rating >= 4.0 else 'Good' if overall_rating >= 3.0 else 'Needs Improvement'
        ]
    ]
    
    performance_table = Table(performance_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 2*inch])
    performance_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0059cc')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (0, -1), rating_color),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(performance_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Rating Breakdown
    if rating_breakdown:
        elements.append(Paragraph("Rating Breakdown by Category", heading_style))
        
        rating_data = [['Category', 'Average Rating', 'Evaluations']]
        for category, data in rating_breakdown.items():
            rating_data.append([
                category,
                f"{data['average']:.2f} / 5.0",
                str(data['count'])
            ])
        
        rating_table = Table(rating_data, colWidths=[3*inch, 2*inch, 1.5*inch])
        rating_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0059cc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(rating_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Recommendation
    if recommendation and recommendation['recommendation_text']:
        elements.append(Paragraph("Guidance Counselor Recommendation", heading_style))
        
        rec_style = ParagraphStyle(
            'Recommendation',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            alignment=TA_JUSTIFY,
            leftIndent=20,
            rightIndent=20,
            spaceAfter=10
        )
        
        elements.append(Paragraph(recommendation['recommendation_text'], rec_style))
        
        rec_info = f"<i>— {recommendation.get('counselor_name', 'Guidance Counselor')}, {datetime.fromisoformat(str(recommendation['created_at'])).strftime('%B %d, %Y')}</i>"
        elements.append(Paragraph(rec_info, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
    
    # Student Comments
    if comments:
        elements.append(PageBreak())
        elements.append(Paragraph(f"Student Feedback ({len(comments)} Comments)", heading_style))
        
        comment_style = ParagraphStyle(
            'Comment',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            leftIndent=15,
            spaceAfter=8
        )
        
        for idx, comment in enumerate(comments, 1):
            comment_rating = float(comment['rating']) if comment['rating'] else 0
            rating_stars = '★' * int(round(comment_rating)) + '☆' * (5 - int(round(comment_rating)))
            
            elements.append(Paragraph(f"<b>Comment #{idx}</b> - {rating_stars} ({comment_rating:.1f}/5.0)", subheading_style))
            elements.append(Paragraph(f'"{comment["comment_text"]}"', comment_style))
            elements.append(Spacer(1, 0.15*inch))
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("This report is confidential and intended for official use only.", footer_style))
    elements.append(Paragraph("Generated by IntellEvalPro Faculty Evaluation System", footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"Faculty_Evaluation_{faculty['last_name']}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


def generate_excel_report(faculty, comments, rating_breakdown, recommendation, response_rate):
    """Generate Excel report"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime
    from flask import send_file
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    summary_sheet = wb.create_sheet("Summary")
    ratings_sheet = wb.create_sheet("Rating Breakdown")
    comments_sheet = wb.create_sheet("Student Comments")
    
    # Define styles
    header_fill = PatternFill(start_color="0059CC", end_color="0059CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=16, color="0059CC")
    label_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    label_font = Font(bold=True, color="0059CC")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary Sheet
    summary_sheet['A1'] = "Faculty Evaluation Report"
    summary_sheet['A1'].font = title_font
    summary_sheet.merge_cells('A1:D1')
    
    summary_sheet['A2'] = "IntellEvalPro - Comprehensive Performance Analysis"
    summary_sheet.merge_cells('A2:D2')
    
    # Faculty Information
    row = 4
    summary_sheet[f'A{row}'] = "Faculty Information"
    summary_sheet[f'A{row}'].font = Font(bold=True, size=14, color="0059CC")
    row += 1
    
    faculty_info = [
        ("Faculty Name:", f"{faculty['first_name']} {faculty['last_name']}"),
        ("Faculty Number:", faculty.get('faculty_number', 'N/A')),
        ("Department:", faculty.get('department_name', 'N/A')),
        ("Email:", faculty.get('email', 'N/A')),
        ("Report Generated:", datetime.now().strftime('%B %d, %Y at %I:%M %p'))
    ]
    
    for label, value in faculty_info:
        summary_sheet[f'A{row}'] = label
        summary_sheet[f'A{row}'].fill = label_fill
        summary_sheet[f'A{row}'].font = label_font
        summary_sheet[f'A{row}'].border = border
        summary_sheet[f'B{row}'] = value
        summary_sheet[f'B{row}'].border = border
        row += 1
    
    # Performance Metrics
    row += 2
    summary_sheet[f'A{row}'] = "Performance Metrics"
    summary_sheet[f'A{row}'].font = Font(bold=True, size=14, color="0059CC")
    row += 1
    
    overall_rating = float(faculty['overall_rating']) if faculty['overall_rating'] else 0
    performance_level = 'Excellent' if overall_rating >= 4.5 else 'Very Good' if overall_rating >= 4.0 else 'Good' if overall_rating >= 3.0 else 'Needs Improvement'
    
    metrics_headers = ['Metric', 'Value']
    for col, header in enumerate(metrics_headers, 1):
        cell = summary_sheet.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    metrics = [
        ("Overall Rating", f"{overall_rating:.2f} / 5.0"),
        ("Total Evaluations", f"{faculty['completed_evaluations']} / {faculty['total_evaluations']}"),
        ("Response Rate", f"{response_rate:.1f}%"),
        ("Performance Level", performance_level)
    ]
    
    for metric, value in metrics:
        summary_sheet[f'A{row}'] = metric
        summary_sheet[f'A{row}'].border = border
        summary_sheet[f'A{row}'].font = label_font
        summary_sheet[f'B{row}'] = value
        summary_sheet[f'B{row}'].border = border
        summary_sheet[f'B{row}'].alignment = Alignment(horizontal='center')
        row += 1
    
    # Recommendation
    if recommendation and recommendation['recommendation_text']:
        row += 2
        summary_sheet[f'A{row}'] = "Guidance Counselor Recommendation"
        summary_sheet[f'A{row}'].font = Font(bold=True, size=14, color="0059CC")
        row += 1
        
        summary_sheet[f'A{row}'] = recommendation['recommendation_text']
        summary_sheet.merge_cells(f'A{row}:D{row}')
        summary_sheet[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        summary_sheet.row_dimensions[row].height = 100
        row += 1
        
        rec_info = f"— {recommendation.get('counselor_name', 'Guidance Counselor')}, {datetime.fromisoformat(str(recommendation['created_at'])).strftime('%B %d, %Y')}"
        summary_sheet[f'A{row}'] = rec_info
        summary_sheet[f'A{row}'].font = Font(italic=True)
    
    # Adjust column widths
    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 40
    
    # Rating Breakdown Sheet
    if rating_breakdown:
        ratings_sheet['A1'] = "Rating Breakdown by Category"
        ratings_sheet['A1'].font = title_font
        ratings_sheet.merge_cells('A1:C1')
        
        row = 3
        headers = ['Category', 'Average Rating', 'Number of Evaluations']
        for col, header in enumerate(headers, 1):
            cell = ratings_sheet.cell(row, col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        for category, data in rating_breakdown.items():
            ratings_sheet[f'A{row}'] = category
            ratings_sheet[f'A{row}'].border = border
            ratings_sheet[f'B{row}'] = f"{data['average']:.2f} / 5.0"
            ratings_sheet[f'B{row}'].border = border
            ratings_sheet[f'B{row}'].alignment = Alignment(horizontal='center')
            ratings_sheet[f'C{row}'] = data['count']
            ratings_sheet[f'C{row}'].border = border
            ratings_sheet[f'C{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        ratings_sheet.column_dimensions['A'].width = 30
        ratings_sheet.column_dimensions['B'].width = 20
        ratings_sheet.column_dimensions['C'].width = 25
    
    # Comments Sheet
    if comments:
        comments_sheet['A1'] = f"Student Feedback ({len(comments)} Comments)"
        comments_sheet['A1'].font = title_font
        comments_sheet.merge_cells('A1:D1')
        
        row = 3
        headers = ['#', 'Rating', 'Date', 'Comment']
        for col, header in enumerate(headers, 1):
            cell = comments_sheet.cell(row, col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        for idx, comment in enumerate(comments, 1):
            comments_sheet[f'A{row}'] = idx
            comments_sheet[f'A{row}'].border = border
            comments_sheet[f'A{row}'].alignment = Alignment(horizontal='center')
            
            rating = float(comment['rating']) if comment['rating'] else 0
            comments_sheet[f'B{row}'] = f"{rating:.1f} / 5.0"
            comments_sheet[f'B{row}'].border = border
            comments_sheet[f'B{row}'].alignment = Alignment(horizontal='center')
            
            date_str = datetime.fromisoformat(str(comment['created_at'])).strftime('%Y-%m-%d') if comment.get('created_at') else 'N/A'
            comments_sheet[f'C{row}'] = date_str
            comments_sheet[f'C{row}'].border = border
            comments_sheet[f'C{row}'].alignment = Alignment(horizontal='center')
            
            comments_sheet[f'D{row}'] = comment['comment_text']
            comments_sheet[f'D{row}'].border = border
            comments_sheet[f'D{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            comments_sheet.row_dimensions[row].height = 50
            
            row += 1
        
        comments_sheet.column_dimensions['A'].width = 8
        comments_sheet.column_dimensions['B'].width = 15
        comments_sheet.column_dimensions['C'].width = 15
        comments_sheet.column_dimensions['D'].width = 60
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Faculty_Evaluation_{faculty['last_name']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@api_bp.route('/guidance/export-all-evaluation-results', methods=['POST'])
@login_required
def export_all_evaluation_results():
    """Export all evaluation results for selected period as PDF or Excel"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO
        from datetime import datetime
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from flask import send_file
        
        data = request.get_json()
        academic_year_id = data.get('academic_year_id')
        period_id = data.get('period_id')
        export_format = data.get('format', 'pdf')
        
        if not academic_year_id or not period_id:
            return jsonify({'success': False, 'message': 'Academic year and period are required'}), 400
        
        # Get evaluation results data
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        try:
            cursor = conn.cursor(dictionary=True)
            
            # Get academic year and period info
            cursor.execute("""
                SELECT year_code, year_name FROM academic_years WHERE acad_year_id = %s
            """, (academic_year_id,))
            year_info = cursor.fetchone()
            
            cursor.execute("""
                SELECT title, start_date, end_date FROM evaluation_periods WHERE period_id = %s
            """, (period_id,))
            period_info = cursor.fetchone()
            
            if not year_info or not period_info:
                return jsonify({'success': False, 'message': 'Invalid academic year or period'}), 404
            
            # Get faculty evaluation results
            cursor.execute("""
                SELECT 
                    f.faculty_id,
                    f.faculty_number,
                    f.first_name,
                    f.last_name,
                    p.name as department_name,
                    COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                    COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.evaluation_id END) as completed_evaluations,
                    ROUND(AVG(er.rating), 2) as overall_rating
                FROM faculty f
                LEFT JOIN programs p ON f.program_id = p.program_id
                LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                LEFT JOIN evaluations e ON cs.section_id = e.section_id AND e.period_id = %s
                LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                WHERE f.is_archived = 0
                GROUP BY f.faculty_id, f.faculty_number, f.first_name, f.last_name, p.name
                HAVING total_evaluations > 0
                ORDER BY p.name, f.last_name, f.first_name
            """, (period_id,))
            
            results = cursor.fetchall()
            
            if not results:
                return jsonify({'success': False, 'message': 'No evaluation results found'}), 404
            
            # Calculate response rates
            for result in results:
                if result['total_evaluations'] > 0:
                    result['response_rate'] = round((result['completed_evaluations'] / result['total_evaluations']) * 100, 1)
                else:
                    result['response_rate'] = 0
            
            # Generate report based on format
            if export_format == 'pdf':
                return generate_all_results_pdf(results, year_info, period_info)
            else:
                return generate_all_results_excel(results, year_info, period_info)
                
        finally:
            cursor.close()
            conn.close()
            
    except ImportError as e:
        return jsonify({
            'success': False,
            'message': f'Required library not installed: {str(e)}'
        }), 500
    except Exception as e:
        print(f"Error exporting evaluation results: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Failed to export evaluation results: {str(e)}'
        }), 500


def generate_all_results_pdf(results, year_info, period_info):
    """Generate PDF report for all evaluation results"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from io import BytesIO
    from datetime import datetime
    from flask import send_file
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_text = f"Faculty Evaluation Results Report"
    title = Paragraph(title_text, styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Period info
    period_text = f"<b>Academic Year:</b> {year_info['year_code']}<br/>"
    period_text += f"<b>Evaluation Period:</b> {period_info['title']}<br/>"
    period_text += f"<b>Generated:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    period_para = Paragraph(period_text, styles['Normal'])
    elements.append(period_para)
    elements.append(Spacer(1, 0.3*inch))
    
    # Calculate summary statistics
    total_faculty = len(results)
    avg_rating = sum(r['overall_rating'] or 0 for r in results) / total_faculty if total_faculty > 0 else 0
    avg_response_rate = sum(r['response_rate'] for r in results) / total_faculty if total_faculty > 0 else 0
    
    summary_text = f"<b>Summary:</b> {total_faculty} Faculty Members | "
    summary_text += f"Average Rating: {avg_rating:.2f}/5.0 | "
    summary_text += f"Average Response Rate: {avg_response_rate:.1f}%"
    summary_para = Paragraph(summary_text, styles['Normal'])
    elements.append(summary_para)
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data
    table_data = [['Faculty ID', 'Name', 'Department', 'Rating', 'Evaluations', 'Response Rate']]
    
    for result in results:
        faculty_name = f"{result['first_name']} {result['last_name']}"
        rating_text = f"{result['overall_rating']:.2f}" if result['overall_rating'] else "N/A"
        evaluations_text = f"{result['completed_evaluations']}/{result['total_evaluations']}"
        response_rate_text = f"{result['response_rate']:.1f}%"
        
        table_data.append([
            result['faculty_number'] or 'N/A',
            faculty_name,
            result['department_name'] or 'N/A',
            rating_text,
            evaluations_text,
            response_rate_text
        ])
    
    # Create table
    table = Table(table_data, colWidths=[1*inch, 2*inch, 2*inch, 0.8*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0059cc')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"Evaluation_Results_{year_info['year_code'].replace('-', '_')}_{period_info['title'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


def generate_all_results_excel(results, year_info, period_info):
    """Generate Excel report for all evaluation results"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from datetime import datetime
    from flask import send_file
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"
    
    # Define styles
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='0059cc', end_color='0059cc', fill_type='solid')
    header_fill = PatternFill(start_color='4a90e2', end_color='4a90e2', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Faculty Evaluation Results Report"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 25
    
    # Period info
    ws['A2'] = f"Academic Year: {year_info['year_code']}"
    ws['A2'].font = Font(name='Calibri', size=10, bold=True)
    ws['A3'] = f"Evaluation Period: {period_info['title']}"
    ws['A3'].font = Font(name='Calibri', size=10, bold=True)
    ws['A4'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A4'].font = Font(name='Calibri', size=10)
    
    # Summary statistics
    total_faculty = len(results)
    avg_rating = sum(r['overall_rating'] or 0 for r in results) / total_faculty if total_faculty > 0 else 0
    avg_response_rate = sum(r['response_rate'] for r in results) / total_faculty if total_faculty > 0 else 0
    
    ws['A6'] = "Summary Statistics"
    ws['A6'].font = Font(name='Calibri', size=12, bold=True)
    ws['A7'] = "Total Faculty Members:"
    ws['B7'] = total_faculty
    ws['A8'] = "Average Rating:"
    ws['B8'] = f"{avg_rating:.2f} / 5.0"
    ws['A9'] = "Average Response Rate:"
    ws['B9'] = f"{avg_response_rate:.1f}%"
    
    # Table headers
    row = 11
    headers = ['Faculty ID', 'Name', 'Department', 'Overall Rating', 'Evaluations Completed', 'Response Rate']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    row += 1
    for result in results:
        ws.cell(row, 1).value = result['faculty_number'] or 'N/A'
        ws.cell(row, 2).value = f"{result['first_name']} {result['last_name']}"
        ws.cell(row, 3).value = result['department_name'] or 'N/A'
        ws.cell(row, 4).value = f"{result['overall_rating']:.2f}" if result['overall_rating'] else "N/A"
        ws.cell(row, 5).value = f"{result['completed_evaluations']}/{result['total_evaluations']}"
        ws.cell(row, 6).value = f"{result['response_rate']:.1f}%"
        
        # Apply styling
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Evaluation_Results_{year_info['year_code'].replace('-', '_')}_{period_info['title'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@api_bp.route('/guidance/questionnaires')
@login_required
def guidance_questionnaires():
    """Get all questionnaires with categories and criteria counts"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get all questionnaires (using evaluation_categories as questionnaire proxy)
        # Since the database uses categories directly, we'll treat each category as a questionnaire
        cursor.execute("""
            SELECT 
                c.category_id as questionnaire_id,
                c.name,
                c.description,
                'Active' as status,
                c.created_at,
                COUNT(DISTINCT cr.criteria_id) as criteria_count,
                1 as category_count
            FROM evaluation_categories c
            LEFT JOIN evaluation_criteria cr ON c.category_id = cr.category_id
            GROUP BY c.category_id, c.name, c.description, c.created_at
            ORDER BY c.created_at DESC
        """)
        
        questionnaires = cursor.fetchall()
        
        # Get total statistics
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT category_id) as total_categories,
                COUNT(DISTINCT criteria_id) as total_criteria
            FROM evaluation_criteria
        """)
        
        stats_raw = cursor.fetchone()
        
        statistics = {
            'total_questionnaires': len(questionnaires),
            'active_questionnaires': len([q for q in questionnaires if q['status'] == 'Active']),
            'total_categories': stats_raw['total_categories'] if stats_raw else 0,
            'total_criteria': stats_raw['total_criteria'] if stats_raw else 0
        }
        
        # Get all criteria for reference
        cursor.execute("""
            SELECT 
                criteria_id,
                category_id,
                description,
                `order`,
                created_at
            FROM evaluation_criteria
            ORDER BY `order`
        """)
        
        criteria = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Format dates
        for q in questionnaires:
            if q.get('created_at'):
                q['created_at'] = q['created_at'].isoformat() if hasattr(q['created_at'], 'isoformat') else str(q['created_at'])
        
        for c in criteria:
            if c.get('created_at'):
                c['created_at'] = c['created_at'].isoformat() if hasattr(c['created_at'], 'isoformat') else str(c['created_at'])
        
        return jsonify({
            'success': True,
            'questionnaires': questionnaires,
            'criteria': criteria,
            'statistics': statistics
        })
        
    except Exception as e:
        print(f"Error getting questionnaires: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/categories')
@login_required
def guidance_categories():
    """Get all evaluation categories with criteria counts"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Check if we should include archived categories
        include_archived = request.args.get('include_archived', 'false').lower() == 'true'
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Build query with optional archive filter
        archive_filter = "" if include_archived else "WHERE c.is_archived = FALSE"
        
        # Get all categories with criteria counts
        query = f"""
            SELECT 
                c.category_id,
                c.name,
                c.description,
                c.is_archived,
                c.display_order,
                c.created_at,
                COUNT(DISTINCT cr.criteria_id) as criteria_count
            FROM evaluation_categories c
            LEFT JOIN evaluation_criteria cr ON c.category_id = cr.category_id
            {archive_filter}
            GROUP BY c.category_id, c.name, c.description, c.is_archived, c.display_order, c.created_at
            ORDER BY c.display_order ASC, c.name
        """
        
        cursor.execute(query)
        categories = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Format dates
        for cat in categories:
            if cat.get('created_at'):
                cat['created_at'] = cat['created_at'].isoformat() if hasattr(cat['created_at'], 'isoformat') else str(cat['created_at'])
        
        return jsonify({
            'success': True,
            'categories': categories
        })
        
    except Exception as e:
        print(f"Error getting categories: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/categories', methods=['POST'])
@login_required
def create_category():
    """Create a new evaluation category"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        
        # Validate required fields
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        weight = data.get('weight', 1.0)
        
        if not name:
            return jsonify({'success': False, 'message': 'Category name is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if category with same name already exists
        cursor.execute("SELECT category_id FROM evaluation_categories WHERE name = %s", (name,))
        existing = cursor.fetchone()
        
        if existing:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'A category with this name already exists'}), 400
        
        # Insert new category
        cursor.execute("""
            INSERT INTO evaluation_categories (name, description, weight)
            VALUES (%s, %s, %s)
        """, (name, description, weight))
        
        category_id = cursor.lastrowid
        conn.commit()
        
        # Get the newly created category
        cursor.execute("""
            SELECT 
                category_id,
                name,
                description,
                weight,
                created_at
            FROM evaluation_categories
            WHERE category_id = %s
        """, (category_id,))
        
        new_category = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        # Format date
        if new_category and new_category.get('created_at'):
            new_category['created_at'] = new_category['created_at'].isoformat() if hasattr(new_category['created_at'], 'isoformat') else str(new_category['created_at'])
        
        return jsonify({
            'success': True,
            'message': 'Category created successfully',
            'category': new_category
        }), 201
        
    except Exception as e:
        print(f"Error creating category: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/criteria', methods=['GET'])
@login_required
def get_all_criteria():
    """Get all evaluation criteria or criteria for specific category"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    category_id = request.args.get('category_id', type=int)
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        if category_id:
            # Get criteria for specific category
            cursor.execute("""
                SELECT 
                    cr.*,
                    c.name as category_name
                FROM evaluation_criteria cr
                JOIN evaluation_categories c ON cr.category_id = c.category_id
                WHERE cr.category_id = %s
                ORDER BY cr.`order`, cr.criteria_id
            """, (category_id,))
        else:
            # Get all criteria
            cursor.execute("""
                SELECT 
                    cr.*,
                    c.name as category_name
                FROM evaluation_criteria cr
                JOIN evaluation_categories c ON cr.category_id = c.category_id
                ORDER BY c.name, cr.`order`, cr.criteria_id
            """)
        
        criteria = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Format dates
        for crit in criteria:
            if crit.get('created_at'):
                crit['created_at'] = crit['created_at'].isoformat() if hasattr(crit['created_at'], 'isoformat') else str(crit['created_at'])
            if crit.get('updated_at'):
                crit['updated_at'] = crit['updated_at'].isoformat() if hasattr(crit['updated_at'], 'isoformat') else str(crit['updated_at'])
        
        return jsonify({
            'success': True,
            'criteria': criteria
        })
        
    except Exception as e:
        print(f"Error getting criteria: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/criteria', methods=['POST'])
@login_required
def create_criterion():
    """Create a new evaluation criterion/question"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        
        category_id = data.get('category_id')
        description = data.get('description', '').strip()
        order = data.get('order', 0)
        
        # Validation
        if not category_id:
            return jsonify({'success': False, 'message': 'Category ID is required'}), 400
        
        if not description:
            return jsonify({'success': False, 'message': 'Question description is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if category exists
        cursor.execute("SELECT category_id FROM evaluation_categories WHERE category_id = %s", (category_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Category not found'}), 404
        
        # If order not specified, get next available order number
        if order == 0:
            cursor.execute("""
                SELECT COALESCE(MAX(`order`), 0) + 1 as next_order 
                FROM evaluation_criteria 
                WHERE category_id = %s
            """, (category_id,))
            result = cursor.fetchone()
            order = result['next_order'] if result else 1
        
        # Insert new criterion
        cursor.execute("""
            INSERT INTO evaluation_criteria (category_id, description, `order`)
            VALUES (%s, %s, %s)
        """, (category_id, description, order))
        
        criteria_id = cursor.lastrowid
        conn.commit()
        
        # Get the created criterion
        cursor.execute("""
            SELECT cr.*, c.name as category_name
            FROM evaluation_criteria cr
            JOIN evaluation_categories c ON cr.category_id = c.category_id
            WHERE cr.criteria_id = %s
        """, (criteria_id,))
        
        criterion = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        # Format dates
        if criterion.get('created_at'):
            criterion['created_at'] = criterion['created_at'].isoformat() if hasattr(criterion['created_at'], 'isoformat') else str(criterion['created_at'])
        if criterion.get('updated_at'):
            criterion['updated_at'] = criterion['updated_at'].isoformat() if hasattr(criterion['updated_at'], 'isoformat') else str(criterion['updated_at'])
        
        return jsonify({
            'success': True,
            'message': 'Question created successfully',
            'criterion': criterion
        })
        
    except Exception as e:
        print(f"Error creating criterion: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/criteria/<int:criteria_id>', methods=['PUT'])
@login_required
def update_criterion(criteria_id):
    """Update an existing evaluation criterion/question"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        
        description = data.get('description', '').strip()
        order = data.get('order')
        
        # Validation
        if not description:
            return jsonify({'success': False, 'message': 'Question description is required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if criterion exists
        cursor.execute("SELECT criteria_id FROM evaluation_criteria WHERE criteria_id = %s", (criteria_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Question not found'}), 404
        
        # Update criterion
        if order is not None:
            cursor.execute("""
                UPDATE evaluation_criteria 
                SET description = %s, `order` = %s
                WHERE criteria_id = %s
            """, (description, order, criteria_id))
        else:
            cursor.execute("""
                UPDATE evaluation_criteria 
                SET description = %s
                WHERE criteria_id = %s
            """, (description, criteria_id))
        
        conn.commit()
        
        # Get the updated criterion
        cursor.execute("""
            SELECT cr.*, c.name as category_name
            FROM evaluation_criteria cr
            JOIN evaluation_categories c ON cr.category_id = c.category_id
            WHERE cr.criteria_id = %s
        """, (criteria_id,))
        
        criterion = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        # Format dates
        if criterion.get('created_at'):
            criterion['created_at'] = criterion['created_at'].isoformat() if hasattr(criterion['created_at'], 'isoformat') else str(criterion['created_at'])
        if criterion.get('updated_at'):
            criterion['updated_at'] = criterion['updated_at'].isoformat() if hasattr(criterion['updated_at'], 'isoformat') else str(criterion['updated_at'])
        
        return jsonify({
            'success': True,
            'message': 'Question updated successfully',
            'criterion': criterion
        })
        
    except Exception as e:
        print(f"Error updating criterion: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/criteria/<int:criteria_id>', methods=['DELETE'])
@login_required
def delete_criterion(criteria_id):
    """Delete an evaluation criterion/question"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if criterion exists
        cursor.execute("SELECT criteria_id, category_id FROM evaluation_criteria WHERE criteria_id = %s", (criteria_id,))
        criterion = cursor.fetchone()
        
        if not criterion:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Question not found'}), 404
        
        # Check if criterion has responses
        cursor.execute("""
            SELECT COUNT(*) as response_count 
            FROM evaluation_responses 
            WHERE criteria_id = %s
        """, (criteria_id,))
        
        result = cursor.fetchone()
        has_responses = result['response_count'] > 0 if result else False
        
        if has_responses:
            cursor.close()
            conn.close()
            return jsonify({
                'success': False, 
                'message': 'Cannot delete question with existing student responses. Archive it instead.'
            }), 400
        
        # Delete the criterion
        cursor.execute("DELETE FROM evaluation_criteria WHERE criteria_id = %s", (criteria_id,))
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Question deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting criterion: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/criteria/reorder', methods=['PUT'])
@login_required
def reorder_criteria():
    """Reorder evaluation criteria/questions"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        criteria_order = data.get('criteria_order', [])  # Array of {criteria_id, order}
        
        if not criteria_order:
            return jsonify({'success': False, 'message': 'No criteria order provided'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Update order for each criterion
        for item in criteria_order:
            criteria_id = item.get('criteria_id')
            order = item.get('order')
            
            if criteria_id and order is not None:
                cursor.execute("""
                    UPDATE evaluation_criteria 
                    SET `order` = %s 
                    WHERE criteria_id = %s
                """, (order, criteria_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Questions reordered successfully'
        })
        
    except Exception as e:
        print(f"Error reordering criteria: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/evaluation/questions', methods=['GET'])
@login_required
def get_evaluation_questions():
    """Get all evaluation questions grouped by category for student evaluation form"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get all non-archived categories with their criteria ordered
        cursor.execute("""
            SELECT 
                c.category_id,
                c.name as category_name,
                c.description as category_description,
                c.weight,
                cr.criteria_id,
                cr.description as criteria_description,
                cr.`order`
            FROM evaluation_categories c
            LEFT JOIN evaluation_criteria cr ON c.category_id = cr.category_id
            WHERE c.is_archived = FALSE
            ORDER BY c.category_id, cr.`order`, cr.criteria_id
        """)
        
        rows = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Group criteria by category
        categories = {}
        for row in rows:
            category_id = row['category_id']
            
            if category_id not in categories:
                categories[category_id] = {
                    'category_id': category_id,
                    'category_name': row['category_name'],
                    'category_description': row['category_description'],
                    'weight': float(row['weight']) if row['weight'] else 1.0,
                    'criteria': []
                }
            
            # Add criteria if exists (skip if no criteria for category)
            if row['criteria_id']:
                categories[category_id]['criteria'].append({
                    'criteria_id': row['criteria_id'],
                    'description': row['criteria_description'],
                    'order': row['order']
                })
        
        # Convert to list
        categories_list = list(categories.values())
        
        return jsonify({
            'success': True,
            'categories': categories_list,
            'total_questions': sum(len(cat['criteria']) for cat in categories_list)
        })
        
    except Exception as e:
        print(f"Error getting evaluation questions: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/categories/<int:category_id>', methods=['PUT'])
@login_required
def update_category(category_id):
    """Update a category's information"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        
        # Validate at least one field to update
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        weight = data.get('weight')
        
        if not name and not description and weight is None:
            return jsonify({'success': False, 'message': 'No fields to update'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if category exists
        cursor.execute("SELECT * FROM evaluation_categories WHERE category_id = %s", (category_id,))
        category = cursor.fetchone()
        
        if not category:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Category not found'}), 404
        
        # Check if new name conflicts with existing category
        if name and name != category['name']:
            cursor.execute("SELECT category_id FROM evaluation_categories WHERE name = %s AND category_id != %s", (name, category_id))
            existing = cursor.fetchone()
            
            if existing:
                cursor.close()
                conn.close()
                return jsonify({'success': False, 'message': 'A category with this name already exists'}), 400
        
        # Build update query dynamically
        update_fields = []
        update_values = []
        
        if name:
            update_fields.append("name = %s")
            update_values.append(name)
        
        if description:
            update_fields.append("description = %s")
            update_values.append(description)
        
        if weight is not None:
            update_fields.append("weight = %s")
            update_values.append(weight)
        
        if update_fields:
            update_values.append(category_id)
            query = f"UPDATE evaluation_categories SET {', '.join(update_fields)} WHERE category_id = %s"
            cursor.execute(query, tuple(update_values))
            conn.commit()
        
        # Get updated category
        cursor.execute("""
            SELECT 
                category_id,
                name,
                description,
                weight,
                created_at,
                updated_at
            FROM evaluation_categories
            WHERE category_id = %s
        """, (category_id,))
        
        updated_category = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        # Format dates
        if updated_category:
            if updated_category.get('created_at'):
                updated_category['created_at'] = updated_category['created_at'].isoformat() if hasattr(updated_category['created_at'], 'isoformat') else str(updated_category['created_at'])
            if updated_category.get('updated_at'):
                updated_category['updated_at'] = updated_category['updated_at'].isoformat() if hasattr(updated_category['updated_at'], 'isoformat') else str(updated_category['updated_at'])
        
        return jsonify({
            'success': True,
            'message': 'Category updated successfully',
            'category': updated_category
        })
        
    except Exception as e:
        print(f"Error updating category: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/categories/<int:category_id>', methods=['DELETE'])
@login_required
def delete_category(category_id):
    """Delete a category (questionnaire) and all its criteria"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Check if force delete is requested
        force_delete = request.args.get('force', 'false').lower() == 'true'
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if category exists
        cursor.execute("SELECT * FROM evaluation_categories WHERE category_id = %s", (category_id,))
        category = cursor.fetchone()
        
        if not category:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Category not found'}), 404
        
        # Check if there are any evaluation responses using criteria from this category
        cursor.execute("""
            SELECT COUNT(*) as response_count
            FROM evaluation_responses er
            INNER JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
            WHERE ec.category_id = %s
        """, (category_id,))
        
        response_check = cursor.fetchone()
        response_count = response_check['response_count'] if response_check else 0
        
        # If responses exist and force delete is not requested, return warning
        if response_count > 0 and not force_delete:
            cursor.close()
            conn.close()
            return jsonify({
                'success': False,
                'requires_confirmation': True,
                'response_count': response_count,
                'message': f'This category has {response_count} student response(s). Deleting it will remove all evaluation data. Are you sure?'
            }), 409  # Conflict status
        
        # If force delete is requested, delete responses first
        if response_count > 0 and force_delete:
            # Delete all responses linked to criteria in this category
            cursor.execute("""
                DELETE er FROM evaluation_responses er
                INNER JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
                WHERE ec.category_id = %s
            """, (category_id,))
            print(f"Force deleted {response_count} student responses for category {category_id}")
        
        # Delete all criteria in this category
        cursor.execute("DELETE FROM evaluation_criteria WHERE category_id = %s", (category_id,))
        
        # Delete the category
        cursor.execute("DELETE FROM evaluation_categories WHERE category_id = %s", (category_id,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Category and its criteria deleted successfully',
            'deleted_responses': response_count if force_delete else 0
        })
        
    except Exception as e:
        print(f"Error deleting category: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/categories/<int:category_id>/archive', methods=['PUT'])
@login_required
def archive_category(category_id):
    """Archive or unarchive a category"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json() or {}
        is_archived = data.get('is_archived', True)  # Default to archive
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if category exists
        cursor.execute("SELECT * FROM evaluation_categories WHERE category_id = %s", (category_id,))
        category = cursor.fetchone()
        
        if not category:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Category not found'}), 404
        
        # Update archive status
        cursor.execute("""
            UPDATE evaluation_categories 
            SET is_archived = %s 
            WHERE category_id = %s
        """, (is_archived, category_id))
        
        conn.commit()
        
        # Get updated category
        cursor.execute("""
            SELECT 
                category_id,
                name,
                description,
                weight,
                is_archived,
                display_order,
                created_at,
                updated_at
            FROM evaluation_categories
            WHERE category_id = %s
        """, (category_id,))
        
        updated_category = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        # Format dates
        if updated_category:
            if updated_category.get('created_at'):
                updated_category['created_at'] = updated_category['created_at'].isoformat() if hasattr(updated_category['created_at'], 'isoformat') else str(updated_category['created_at'])
            if updated_category.get('updated_at'):
                updated_category['updated_at'] = updated_category['updated_at'].isoformat() if hasattr(updated_category['updated_at'], 'isoformat') else str(updated_category['updated_at'])
        
        action = "archived" if is_archived else "unarchived"
        return jsonify({
            'success': True,
            'message': f'Category {action} successfully',
            'category': updated_category
        })
        
    except Exception as e:
        print(f"Error archiving category: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/categories/reorder', methods=['PUT'])
@login_required
def reorder_categories():
    """Reorder categories by updating their display_order"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        category_order = data.get('category_order', [])
        
        if not category_order:
            return jsonify({'success': False, 'message': 'No category order provided'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Update display order for each category
        for item in category_order:
            category_id = item.get('category_id')
            display_order = item.get('display_order')
            
            if category_id is None or display_order is None:
                continue
            
            cursor.execute("""
                UPDATE evaluation_categories 
                SET display_order = %s 
                WHERE category_id = %s
            """, (display_order, category_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Categories reordered successfully'
        })
        
    except Exception as e:
        print(f"Error reordering categories: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/questionnaires/import', methods=['POST'])
@login_required
def import_questionnaires():
    """Import questionnaire categories and questions from uploaded data"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        import_data = data.get('data', {})
        options = data.get('options', {})
        
        categories = import_data.get('categories', [])
        if not categories:
            return jsonify({'success': False, 'message': 'No categories found in import data'}), 400
        
        overwrite = options.get('overwrite', False)
        validate = options.get('validate', True)
        
        # Validation
        if validate:
            for idx, cat in enumerate(categories):
                if not cat.get('name'):
                    return jsonify({
                        'success': False,
                        'message': f'Category #{idx + 1} is missing a name'
                    }), 400
                if not cat.get('questions') or not isinstance(cat['questions'], list):
                    return jsonify({
                        'success': False,
                        'message': f'Category "{cat.get("name")}" has no questions or invalid format'
                    }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # If overwrite mode, archive all existing categories
        if overwrite:
            cursor.execute("""
                UPDATE evaluation_categories 
                SET is_archived = TRUE 
                WHERE is_archived = FALSE
            """)
        
        # Get the highest display order
        cursor.execute("SELECT MAX(display_order) as max_order FROM evaluation_categories")
        result = cursor.fetchone()
        max_order = result['max_order'] if result and result['max_order'] is not None else 0
        
        imported_categories = 0
        imported_questions = 0
        
        # Import each category
        for cat in categories:
            cat_name = cat.get('name', '').strip()
            cat_description = cat.get('description', '').strip()
            
            # Check if category with same name exists
            cursor.execute("""
                SELECT category_id, is_archived 
                FROM evaluation_categories 
                WHERE name = %s
            """, (cat_name,))
            existing_cat = cursor.fetchone()
            
            if existing_cat:
                # Unarchive if archived
                if existing_cat['is_archived']:
                    cursor.execute("""
                        UPDATE evaluation_categories 
                        SET is_archived = FALSE, description = %s
                        WHERE category_id = %s
                    """, (cat_description, existing_cat['category_id']))
                category_id = existing_cat['category_id']
            else:
                # Create new category
                max_order += 1
                cursor.execute("""
                    INSERT INTO evaluation_categories 
                    (name, description, display_order, is_archived, created_at)
                    VALUES (%s, %s, %s, FALSE, NOW())
                """, (cat_name, cat_description, max_order))
                category_id = cursor.lastrowid
                imported_categories += 1
            
            # Import questions for this category
            questions = cat.get('questions', [])
            for question_text in questions:
                if not question_text or not isinstance(question_text, str):
                    continue
                
                question_text = question_text.strip()
                
                # Check if question already exists in this category
                cursor.execute("""
                    SELECT criteria_id FROM evaluation_criteria 
                    WHERE category_id = %s AND description = %s
                """, (category_id, question_text))
                
                if not cursor.fetchone():
                    # Insert new question
                    cursor.execute("""
                        INSERT INTO evaluation_criteria 
                        (category_id, description, created_at)
                        VALUES (%s, %s, NOW())
                    """, (category_id, question_text))
                    imported_questions += 1
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Import completed successfully',
            'imported': {
                'categories': imported_categories,
                'questions': imported_questions
            }
        })
        
    except Exception as e:
        print(f"Error importing questionnaires: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/questionnaires/export')
@login_required
def export_questionnaires():
    """Export questionnaire categories and questions to file"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from flask import make_response
        import json
        from io import BytesIO
        
        export_format = request.args.get('format', 'json')
        include_active = request.args.get('include_active', 'true').lower() == 'true'
        include_archived = request.args.get('include_archived', 'false').lower() == 'true'
        include_questions = request.args.get('include_questions', 'true').lower() == 'true'
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Build query based on filters
        query = "SELECT * FROM evaluation_categories WHERE 1=1"
        conditions = []
        
        if include_active and not include_archived:
            conditions.append("is_archived = FALSE")
        elif include_archived and not include_active:
            conditions.append("is_archived = TRUE")
        elif not include_active and not include_archived:
            return jsonify({'success': False, 'message': 'Must include at least active or archived categories'}), 400
        
        if conditions:
            query += " AND " + " AND ".join(conditions)
        
        query += " ORDER BY display_order"
        
        cursor.execute(query)
        categories = cursor.fetchall() or []
        
        # Build export data structure
        export_data = {'categories': []}
        
        for cat in categories:
            category_data = {
                'name': cat['name'],
                'description': cat['description'] or '',
                'display_order': cat['display_order'],
                'is_archived': bool(cat['is_archived'])
            }
            
            # Include questions if requested
            if include_questions:
                cursor.execute("""
                    SELECT description 
                    FROM evaluation_criteria 
                    WHERE category_id = %s 
                    ORDER BY `order`, criteria_id
                """, (cat['category_id'],))
                questions = cursor.fetchall() or []
                category_data['questions'] = [q['description'] for q in questions]
            else:
                category_data['questions'] = []
            
            export_data['categories'].append(category_data)
        
        cursor.close()
        conn.close()
        
        # Generate file based on format
        if export_format == 'json':
            # JSON export
            json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
            response = make_response(json_str)
            response.headers['Content-Type'] = 'application/json'
            response.headers['Content-Disposition'] = 'attachment; filename=questionnaires_export.json'
            return response
            
        elif export_format == 'excel':
            # Excel export - same format as CSV (each question on separate row)
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Questionnaires"
                
                # Header styling
                header_fill = PatternFill(start_color="0059cc", end_color="0059cc", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                # Write headers - same as CSV
                headers = ['Category Name', 'Description', 'Display Order', 'Status', 'Question Number', 'Question Text']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Write data - each question on separate row (like CSV)
                row = 2
                for cat in export_data['categories']:
                    if cat['questions']:
                        for idx, question in enumerate(cat['questions'], start=1):
                            ws.cell(row=row, column=1, value=cat['name'])
                            ws.cell(row=row, column=2, value=cat['description'])
                            ws.cell(row=row, column=3, value=cat['display_order'])
                            ws.cell(row=row, column=4, value='Archived' if cat['is_archived'] else 'Active')
                            ws.cell(row=row, column=5, value=idx)
                            ws.cell(row=row, column=6, value=question)
                            row += 1
                    else:
                        # Category with no questions
                        ws.cell(row=row, column=1, value=cat['name'])
                        ws.cell(row=row, column=2, value=cat['description'])
                        ws.cell(row=row, column=3, value=cat['display_order'])
                        ws.cell(row=row, column=4, value='Archived' if cat['is_archived'] else 'Active')
                        ws.cell(row=row, column=5, value='')
                        ws.cell(row=row, column=6, value='No questions')
                        row += 1
                
                # Auto-adjust column widths based on content
                for column_cells in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column_cells[0].column)
                    
                    for cell in column_cells:
                        try:
                            if cell.value:
                                # Calculate length considering line breaks
                                cell_value = str(cell.value)
                                lines = cell_value.split('\n')
                                max_line_length = max(len(line) for line in lines) if lines else len(cell_value)
                                max_length = max(max_length, max_line_length)
                        except:
                            pass
                    
                    # Set column width with padding and limits
                    adjusted_width = min(max_length + 2, 80)  # Max width of 80
                    adjusted_width = max(adjusted_width, 10)   # Min width of 10
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = 'attachment; filename=questionnaires_export.xlsx'
                return response
                
            except ImportError:
                return jsonify({
                    'success': False,
                    'message': 'Excel export requires openpyxl library'
                }), 500
                
        elif export_format == 'csv':
            # CSV export
            import csv
            from io import StringIO
            
            output = StringIO()
            writer = csv.writer(output)
            
            # Write headers
            writer.writerow(['Category Name', 'Description', 'Display Order', 'Status', 'Question Number', 'Question Text'])
            
            # Write data
            for cat in export_data['categories']:
                if cat['questions']:
                    for idx, question in enumerate(cat['questions'], start=1):
                        writer.writerow([
                            cat['name'],
                            cat['description'],
                            cat['display_order'],
                            'Archived' if cat['is_archived'] else 'Active',
                            idx,
                            question
                        ])
                else:
                    writer.writerow([
                        cat['name'],
                        cat['description'],
                        cat['display_order'],
                        'Archived' if cat['is_archived'] else 'Active',
                        '',
                        'No questions'
                    ])
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv'
            response.headers['Content-Disposition'] = 'attachment; filename=questionnaires_export.csv'
            return response
        
        else:
            return jsonify({'success': False, 'message': 'Invalid export format'}), 400
        
    except Exception as e:
        print(f"Error exporting questionnaires: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/rating-distribution')
@login_required
def get_guidance_rating_distribution():
    """Get faculty rating distribution data with optional department filter"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        department = request.args.get('department', None)
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get current active period
        cursor.execute("""
            SELECT period_id, title
            FROM evaluation_periods 
            WHERE status = 'Active' 
            ORDER BY start_date DESC 
            LIMIT 1
        """)
        current_period = cursor.fetchone()
        
        if not current_period:
            cursor.execute("""
                SELECT period_id, title
                FROM evaluation_periods 
                ORDER BY start_date DESC 
                LIMIT 1
            """)
            current_period = cursor.fetchone()
        
        if not current_period:
            return jsonify({
                'success': False,
                'message': 'No evaluation period found'
            }), 404
        
        # Build query with optional department filter
        query = """
            SELECT 
                cs.faculty_id,
                CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                p.name as department,
                AVG(er.rating) as avg_rating
            FROM class_sections cs
            JOIN faculty f ON cs.faculty_id = f.faculty_id
            JOIN programs p ON f.program_id = p.program_id
            JOIN evaluations e ON cs.section_id = e.section_id
            JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            WHERE e.period_id = %s AND e.status = 'Completed'
        """
        
        params = [current_period['period_id']]
        
        if department and department != 'all':
            query += " AND p.name = %s"
            params.append(department)
        
        query += """
            GROUP BY cs.faculty_id, f.first_name, f.last_name, p.name
            HAVING AVG(er.rating) IS NOT NULL
        """
        
        cursor.execute(query, params)
        faculty_ratings = cursor.fetchall() or []
        
        # Define rating categories
        rating_categories = {
            '5.0': 0,
            '4.5-4.9': 0,
            '4.0-4.4': 0,
            '3.5-3.9': 0,
            '3.0-3.4': 0,
            'Below 3.0': 0
        }
        
        # Categorize each faculty member's rating
        for faculty in faculty_ratings:
            avg = float(faculty['avg_rating'] or 0)
            if avg == 5.0:
                rating_categories['5.0'] += 1
            elif avg >= 4.5:
                rating_categories['4.5-4.9'] += 1
            elif avg >= 4.0:
                rating_categories['4.0-4.4'] += 1
            elif avg >= 3.5:
                rating_categories['3.5-3.9'] += 1
            elif avg >= 3.0:
                rating_categories['3.0-3.4'] += 1
            else:
                rating_categories['Below 3.0'] += 1
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'labels': list(rating_categories.keys()),
            'data': list(rating_categories.values()),
            'department': department or 'All Departments',
            'total_faculty': len(faculty_ratings)
        })
        
    except Exception as e:
        print(f"Error in get_guidance_rating_distribution: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@api_bp.route('/guidance/departments-list')
@login_required
def get_guidance_departments_list():
    """Get list of all departments for filtering"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT DISTINCT p.name as department_name
            FROM programs p
            JOIN faculty f ON p.program_id = f.program_id
            WHERE f.status = 'Active'
            ORDER BY p.name
        """)
        
        departments = cursor.fetchall() or []
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'departments': [dept['department_name'] for dept in departments]
        })
        
    except Exception as e:
        print(f"Error in get_guidance_departments_list: {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


# ========================================
# GUIDANCE DATA API ENDPOINTS
# ========================================

@api_bp.route('/guidance/programs', methods=['GET'])
@login_required
def get_guidance_programs():
    """Get all programs/departments for guidance"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                program_id,
                program_code,
                name,
                description
            FROM programs
            ORDER BY name
        """)
        
        programs = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'programs': programs
        })
        
    except Exception as e:
        print(f"Error getting programs: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ========================================
# EVALUATION REPORTS API ENDPOINTS
# ========================================

@api_bp.route('/guidance/reports', methods=['GET'])
@login_required
def get_reports():
    """Get all generated reports with optional filters"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get filter parameters
        report_type = request.args.get('report_type')
        period_id = request.args.get('period_id', type=int)
        program_id = request.args.get('program_id', type=int)
        search = request.args.get('search', '').strip()
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Build query with filters
        query = """
            SELECT 
                r.report_id,
                r.report_name,
                r.report_type,
                r.period_id,
                r.program_id,
                r.faculty_id,
                r.file_format,
                r.file_path,
                r.file_size,
                r.download_count,
                r.created_at,
                r.updated_at,
                ep.title as period_name,
                p.name as program_name,
                CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                u.first_name as generated_by_first_name,
                u.last_name as generated_by_last_name
            FROM generated_reports r
            LEFT JOIN evaluation_periods ep ON r.period_id = ep.period_id
            LEFT JOIN programs p ON r.program_id = p.program_id
            LEFT JOIN faculty f ON r.faculty_id = f.faculty_id
            LEFT JOIN users u ON r.generated_by = u.user_id
            WHERE 1=1
        """
        
        params = []
        
        if report_type:
            query += " AND r.report_type = %s"
            params.append(report_type)
        
        if period_id:
            query += " AND r.period_id = %s"
            params.append(period_id)
        
        if program_id:
            query += " AND r.program_id = %s"
            params.append(program_id)
        
        if search:
            query += " AND r.report_name LIKE %s"
            params.append(f'%{search}%')
        
        query += " ORDER BY r.created_at DESC"
        
        cursor.execute(query, tuple(params))
        reports = cursor.fetchall()
        
        # Format data
        for report in reports:
            if report.get('created_at'):
                report['created_at'] = report['created_at'].isoformat() if hasattr(report['created_at'], 'isoformat') else str(report['created_at'])
            if report.get('updated_at'):
                report['updated_at'] = report['updated_at'].isoformat() if hasattr(report['updated_at'], 'isoformat') else str(report['updated_at'])
            
            # Format file size
            if report.get('file_size'):
                report['file_size_mb'] = round(report['file_size'] / (1024 * 1024), 2)
                report['file_size_formatted'] = format_file_size(report['file_size'])
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'reports': reports
        })
        
    except Exception as e:
        print(f"Error getting reports: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/reports/statistics', methods=['GET'])
@login_required
def get_report_statistics():
    """Get report statistics"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get total reports
        cursor.execute("SELECT COUNT(*) as total_reports FROM generated_reports")
        total = cursor.fetchone()
        
        # Get total downloads
        cursor.execute("SELECT SUM(download_count) as total_downloads FROM generated_reports")
        downloads = cursor.fetchone()
        
        # Get reports this month
        cursor.execute("""
            SELECT COUNT(*) as this_month 
            FROM generated_reports 
            WHERE MONTH(created_at) = MONTH(CURRENT_DATE()) 
            AND YEAR(created_at) = YEAR(CURRENT_DATE())
        """)
        this_month = cursor.fetchone()
        
        # Get total storage used
        cursor.execute("SELECT SUM(file_size) as total_storage FROM generated_reports")
        storage = cursor.fetchone()
        
        # Get reports by type
        cursor.execute("""
            SELECT report_type, COUNT(*) as count
            FROM generated_reports
            GROUP BY report_type
        """)
        by_type = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Format storage
        total_storage_bytes = storage['total_storage'] if storage and storage['total_storage'] else 0
        total_storage_gb = round(total_storage_bytes / (1024 * 1024 * 1024), 2)
        
        return jsonify({
            'success': True,
            'statistics': {
                'total_reports': total['total_reports'] if total else 0,
                'total_downloads': downloads['total_downloads'] if downloads and downloads['total_downloads'] else 0,
                'this_month': this_month['this_month'] if this_month else 0,
                'total_storage_bytes': total_storage_bytes,
                'total_storage_gb': total_storage_gb,
                'total_storage_formatted': format_file_size(total_storage_bytes),
                'by_type': {item['report_type']: item['count'] for item in by_type}
            }
        })
        
    except Exception as e:
        print(f"Error getting report statistics: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/evaluation-periods-list', methods=['GET'])
@login_required
def get_evaluation_periods_for_reports():
    """Get all evaluation periods for reports page with optional academic year filtering"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        academic_year_id = request.args.get('academic_year_id', type=int)
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Build query with optional academic year filter
        if academic_year_id:
            cursor.execute("""
                SELECT 
                    ep.period_id,
                    ep.acad_term_id,
                    ep.title as period_name,
                    ep.start_date,
                    ep.end_date,
                    ep.status,
                    ep.created_at,
                    ep.updated_at,
                    at.acad_year_id,
                    CONCAT(ay.year_code, ' - ', at.term_name) as academic_year,
                    DATEDIFF(ep.end_date, ep.start_date) as duration_days,
                    DATEDIFF(ep.end_date, CURDATE()) as days_remaining,
                    (SELECT COUNT(*) FROM evaluations WHERE period_id = ep.period_id) as total_evaluations,
                    (SELECT COUNT(*) FROM evaluations WHERE period_id = ep.period_id AND status = 'Completed') as completed_evaluations
                FROM evaluation_periods ep
                LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
                LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
                WHERE at.acad_year_id = %s
                ORDER BY ep.start_date DESC
            """, (academic_year_id,))
        else:
            cursor.execute("""
                SELECT 
                    ep.period_id,
                    ep.acad_term_id,
                    ep.title as period_name,
                    ep.start_date,
                    ep.end_date,
                    ep.status,
                    ep.created_at,
                    ep.updated_at,
                    CONCAT(ay.year_code, ' - ', at.term_name) as academic_year,
                    DATEDIFF(ep.end_date, ep.start_date) as duration_days,
                    DATEDIFF(ep.end_date, CURDATE()) as days_remaining,
                    (SELECT COUNT(*) FROM evaluations WHERE period_id = ep.period_id) as total_evaluations,
                    (SELECT COUNT(*) FROM evaluations WHERE period_id = ep.period_id AND status = 'Completed') as completed_evaluations
                FROM evaluation_periods ep
                LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
                LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
                ORDER BY ep.start_date DESC
            """)
        
        periods = cursor.fetchall()
        
        # Calculate additional metrics and format dates
        current_period = None
        for period in periods:
            if period.get('start_date'):
                period['start_date'] = period['start_date'].isoformat() if hasattr(period['start_date'], 'isoformat') else str(period['start_date'])
            if period.get('end_date'):
                period['end_date'] = period['end_date'].isoformat() if hasattr(period['end_date'], 'isoformat') else str(period['end_date'])
            
            # Calculate progress percentage
            if period.get('duration_days') and period['duration_days'] > 0:
                elapsed = period['duration_days'] - (period.get('days_remaining') or 0)
                period['progress_percent'] = round((elapsed / period['duration_days']) * 100)
            else:
                period['progress_percent'] = 0
            
            # Calculate response rate
            if period.get('total_evaluations') and period['total_evaluations'] > 0:
                period['response_rate'] = round((period.get('completed_evaluations', 0) / period['total_evaluations']) * 100)
            else:
                period['response_rate'] = 0
            
            # Identify current period
            if period.get('status') == 'Active' and not current_period:
                current_period = period
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'periods': periods,
            'current_period': current_period
        })
        
    except Exception as e:
        print(f"Error getting evaluation periods: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/evaluation-periods-create', methods=['POST'])
@login_required
def create_evaluation_period():
    """Create a new evaluation period - DEPRECATED - Use /guidance/evaluation-periods POST instead"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        acad_term_id = data.get('acad_term_id')
        title = data.get('title')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # Validate required fields
        if not all([acad_term_id, title, start_date, end_date]):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Check if academic term exists
        cursor.execute("SELECT * FROM academic_terms WHERE acad_term_id = %s", (acad_term_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Invalid academic term'}), 400
        
        # Determine status based on dates
        from datetime import datetime, date
        today = date.today()
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        if start > today:
            status = 'Pending'
        elif start <= today <= end:
            status = 'Active'
        else:
            status = 'Closed'
        
        # Insert evaluation period (without created_by column)
        cursor.execute("""
            INSERT INTO evaluation_periods (acad_term_id, title, start_date, end_date, status)
            VALUES (%s, %s, %s, %s, %s)
        """, (acad_term_id, title, start_date, end_date, status))
        
        conn.commit()
        period_id = cursor.lastrowid
        
        # Email notifications removed - use manual "Send Email Notifications" button instead
        print(f"✅ Evaluation period created with ID: {period_id}")
        print(f"📧 Automatic email notifications disabled. Use 'Send Email Notifications' button.")
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation period created successfully',
            'period_id': period_id
        })
        
    except Exception as e:
        if conn:
            conn.rollback()
        print(f"Error creating evaluation period: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Error creating period: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()


@api_bp.route('/guidance/evaluation-periods-update/<int:period_id>', methods=['PUT'])
@login_required
def update_evaluation_period(period_id):
    """Update an existing evaluation period - DEPRECATED"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        acad_term_id = data.get('acad_term_id')
        title = data.get('title')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # Validate required fields
        if not all([acad_term_id, title, start_date, end_date]):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Check if period exists
        cursor.execute("SELECT * FROM evaluation_periods WHERE period_id = %s", (period_id,))
        period = cursor.fetchone()
        
        if not period:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Evaluation period not found'}), 404
        
        # Check if academic term exists
        cursor.execute("SELECT * FROM academic_terms WHERE acad_term_id = %s", (acad_term_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Invalid academic term'}), 400
        
        # Determine status based on dates
        from datetime import datetime, date
        today = date.today()
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        if start > today:
            status = 'Pending'
        elif start <= today <= end:
            status = 'Active'
        else:
            status = 'Closed'
        
        # Update evaluation period (without updated_by column)
        cursor.execute("""
            UPDATE evaluation_periods 
            SET acad_term_id = %s, title = %s, start_date = %s, end_date = %s, status = %s
            WHERE period_id = %s
        """, (acad_term_id, title, start_date, end_date, status, period_id))
        
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Evaluation period updated successfully'
        })
        
    except Exception as e:
        if conn:
            conn.rollback()
        print(f"Error updating evaluation period: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Error updating period: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()


@api_bp.route('/guidance/programs', methods=['GET'])
@login_required
def get_programs_for_reports():
    """Get all programs/departments for reports page"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                program_id,
                name,
                program_code,
                description
            FROM programs
            ORDER BY name
        """)
        
        programs = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'programs': programs
        })
        
    except Exception as e:
        print(f"Error getting programs: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/reports/<int:report_id>', methods=['DELETE'])
@login_required
def delete_report(report_id):
    """Delete a generated report"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if report exists
        cursor.execute("SELECT * FROM generated_reports WHERE report_id = %s", (report_id,))
        report = cursor.fetchone()
        
        if not report:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Report not found'}), 404
        
        # Delete the report
        cursor.execute("DELETE FROM generated_reports WHERE report_id = %s", (report_id,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # TODO: Delete physical file if it exists
        # import os
        # if report['file_path'] and os.path.exists(report['file_path']):
        #     os.remove(report['file_path'])
        
        return jsonify({
            'success': True,
            'message': 'Report deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting report: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/reports/generate', methods=['POST'])
@login_required
def generate_report():
    """Generate a new evaluation report"""
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        report_type = data.get('report_type')
        period_id = data.get('period_id')
        file_format = data.get('file_format', 'pdf')
        program_id = data.get('program_id')  # Optional
        faculty_id = data.get('faculty_id')  # Optional - for faculty reports
        
        # Validate required fields
        if not report_type or not period_id:
            return jsonify({'success': False, 'message': 'Report type and period are required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get period information
        cursor.execute("""
            SELECT ep.*
            FROM evaluation_periods ep
            WHERE ep.period_id = %s
        """, (period_id,))
        period = cursor.fetchone()
        
        if not period:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Evaluation period not found'}), 404
        
        # Get program name if provided
        program_name = None
        if program_id:
            cursor.execute("SELECT name FROM programs WHERE program_id = %s", (program_id,))
            program_result = cursor.fetchone()
            if program_result:
                program_name = program_result['name']
        
        # Get faculty name if provided
        faculty_name = None
        if faculty_id:
            cursor.execute("""
                SELECT first_name, last_name, f.program_id
                FROM faculty f
                WHERE faculty_id = %s
            """, (faculty_id,))
            faculty_result = cursor.fetchone()
            if faculty_result:
                faculty_name = f"{faculty_result['first_name']} {faculty_result['last_name']}"
                # If faculty_id is provided but program_id is not, use faculty's program
                if not program_id and faculty_result['program_id']:
                    program_id = faculty_result['program_id']
                    cursor.execute("SELECT name FROM programs WHERE program_id = %s", (program_id,))
                    program_result = cursor.fetchone()
                    if program_result:
                        program_name = program_result['name']
        
        # Generate report name
        report_type_names = {
            'summary': 'Summary Report',
            'faculty': 'Faculty Performance Report',
            'department': 'Department Analysis Report',
            'comparative': 'Comparative Analysis Report'
        }
        
        period_name = period.get('title') or f"Period {period_id}"
        report_name = f"{period_name} - {report_type_names.get(report_type, 'Report')}"
        
        # Add specific identifiers to report name
        if faculty_name:
            report_name += f" - {faculty_name}"
        elif program_name:
            report_name += f" - {program_name}"
        
        # Calculate estimated file size (simulated - would be actual size in real implementation)
        import random
        file_size = random.randint(500000, 5000000)  # 500KB to 5MB
        
        # Get current user ID
        user_id = session.get('user_id', 1)
        
        # Insert report record
        cursor.execute("""
            INSERT INTO generated_reports 
            (report_name, report_type, period_id, program_id, faculty_id, file_format, 
             file_size, generated_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (report_name, report_type, period_id, program_id, faculty_id, file_format, 
              file_size, user_id))
        
        report_id = cursor.lastrowid
        
        conn.commit()
        
        # Get the newly created report
        cursor.execute("""
            SELECT 
                r.*,
                ep.title as period_name,
                p.name as program_name,
                f.first_name as faculty_first_name,
                f.last_name as faculty_last_name,
                u.first_name as generated_by_first_name,
                u.last_name as generated_by_last_name
            FROM generated_reports r
            LEFT JOIN evaluation_periods ep ON r.period_id = ep.period_id
            LEFT JOIN programs p ON r.program_id = p.program_id
            LEFT JOIN faculty f ON r.faculty_id = f.faculty_id
            LEFT JOIN users u ON r.generated_by = u.user_id
            WHERE r.report_id = %s
        """, (report_id,))
        
        new_report = cursor.fetchone()
        
        # Format the response
        if new_report.get('created_at'):
            new_report['created_at'] = new_report['created_at'].isoformat() if hasattr(new_report['created_at'], 'isoformat') else str(new_report['created_at'])
        if new_report.get('file_size'):
            new_report['file_size_formatted'] = format_file_size(new_report['file_size'])
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Report generated successfully',
            'report': new_report
        })
        
    except Exception as e:
        print(f"Error generating report: {str(e)}")
        import traceback
        traceback.print_exc()
        # Close connection if it was opened
        try:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()
        except:
            pass
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/reports/<int:report_id>/download', methods=['GET'])
@login_required
def download_report_file(report_id):
    """Download the actual report file"""
    from flask import send_file
    import os
    
    if session.get('role') != 'guidance':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get report information
        cursor.execute("""
            SELECT 
                r.*,
                ep.title as period_name
            FROM generated_reports r
            LEFT JOIN evaluation_periods ep ON r.period_id = ep.period_id
            WHERE r.report_id = %s
        """, (report_id,))
        
        report = cursor.fetchone()
        
        if not report:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Report not found'}), 404
        
        # Increment download count
        cursor.execute("""
            UPDATE generated_reports 
            SET download_count = download_count + 1 
            WHERE report_id = %s
        """, (report_id,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # Generate file path
        file_path = report.get('file_path')
        
        # If file_path is null or file doesn't exist, generate the report on-the-fly
        if not file_path or not os.path.exists(file_path):
            # Generate report dynamically
            file_path = generate_report_file(report)
        
        # Map file format to actual file extension for download
        extension_map = {
            'pdf': 'pdf',
            'excel': 'xlsx',
            'csv': 'csv',
            'powerpoint': 'pptx'
        }
        file_extension = extension_map.get(report['file_format'], report['file_format'])
        
        # Send file
        return send_file(
            file_path,
            as_attachment=True,
            download_name=f"{report['report_name']}.{file_extension}",
            mimetype=get_mimetype(report['file_format'])
        )
        
    except Exception as e:
        print(f"Error downloading report: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


def format_file_size(size_bytes):
    """Format file size in human-readable format"""
    if not size_bytes or size_bytes == 0:
        return "0 B"
    
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    unit_index = 0
    size = float(size_bytes)
    
    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1
    
    if unit_index == 0:
        return f"{int(size)} {units[unit_index]}"
    else:
        return f"{size:.1f} {units[unit_index]}"


def get_mimetype(file_format):
    """Get MIME type for file format"""
    mimetypes = {
        'pdf': 'application/pdf',
        'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'csv': 'text/csv',
        'powerpoint': 'application/vnd.openxmlformats-officedocument.presentationml.presentation'
    }
    return mimetypes.get(file_format, 'application/octet-stream')


def generate_report_file(report):
    """Generate report file dynamically (PDF or Excel)"""
    import os
    from datetime import datetime
    
    # Create reports directory if it doesn't exist
    reports_dir = os.path.join('static', 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    # Generate filename with proper extension
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = "".join(c for c in report['report_name'] if c.isalnum() or c in (' ', '-', '_')).rstrip()
    safe_name = safe_name.replace(' ', '_')
    
    # Map file format to actual file extension
    extension_map = {
        'pdf': 'pdf',
        'excel': 'xlsx',
        'csv': 'csv',
        'powerpoint': 'pptx'
    }
    file_extension = extension_map.get(report['file_format'], report['file_format'])
    filename = f"{safe_name}_{timestamp}.{file_extension}"
    file_path = os.path.join(reports_dir, filename)
    
    # Generate based on format
    if report['file_format'] == 'pdf':
        generate_analytics_pdf_report(report, file_path)
    elif report['file_format'] == 'excel':
        generate_analytics_excel_report(report, file_path)
    elif report['file_format'] == 'csv':
        generate_csv_report(report, file_path)
    else:
        # Fallback: create a simple text file
        with open(file_path, 'w') as f:
            f.write(f"Report: {report['report_name']}\n")
            f.write(f"Type: {report['report_type']}\n")
            f.write(f"Generated: {datetime.now()}\n")
    
    return file_path


def generate_analytics_pdf_report(report, file_path):
    """Generate analytics PDF report using ReportLab"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from datetime import datetime
        
        # Create PDF
        doc = SimpleDocTemplate(file_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0059cc'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph(report['report_name'], title_style))
        story.append(Spacer(1, 0.2 * inch))
        
        # Report Info
        info_data = [
            ['Report Type:', report['report_type'].capitalize()],
            ['Generated:', datetime.now().strftime('%B %d, %Y at %I:%M %p')],
            ['Format:', report['file_format'].upper()],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#495057')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.5 * inch))
        
        # Content
        story.append(Paragraph("Report Summary", styles['Heading2']))
        story.append(Spacer(1, 0.2 * inch))
        
        summary_text = f"""
        This is a {report['report_type']} report generated for the evaluation period. 
        The report contains comprehensive analysis and statistics for faculty evaluation data.
        """
        story.append(Paragraph(summary_text, styles['BodyText']))
        story.append(Spacer(1, 0.3 * inch))
        
        # Sample data table
        story.append(Paragraph("Evaluation Statistics", styles['Heading2']))
        story.append(Spacer(1, 0.2 * inch))
        
        data = [
            ['Metric', 'Value'],
            ['Total Evaluations', 'N/A'],
            ['Average Rating', 'N/A'],
            ['Response Rate', 'N/A'],
            ['Completion Rate', 'N/A'],
        ]
        
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0059cc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        story.append(table)
        
        # Footer
        story.append(Spacer(1, 0.5 * inch))
        footer_text = "Generated by IntellEvalPro - Faculty Evaluation System"
        story.append(Paragraph(footer_text, styles['Italic']))
        
        # Build PDF
        doc.build(story)
        
    except ImportError:
        # Fallback if reportlab is not installed
        generate_simple_text_report(report, file_path)


def generate_analytics_excel_report(report, file_path):
    """Generate analytics Excel report using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Title
        ws['A1'] = report['report_name']
        ws['A1'].font = Font(size=16, bold=True, color="0059cc")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Report Info
        ws['A3'] = 'Report Type:'
        ws['B3'] = report['report_type'].capitalize()
        ws['A4'] = 'Generated:'
        ws['B4'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        ws['A5'] = 'Format:'
        ws['B5'] = report['file_format'].upper()
        
        # Make headers bold
        for row in [3, 4, 5]:
            ws[f'A{row}'].font = Font(bold=True)
        
        # Statistics Header
        ws['A7'] = 'Evaluation Statistics'
        ws['A7'].font = Font(size=14, bold=True)
        
        # Table headers
        ws['A9'] = 'Metric'
        ws['B9'] = 'Value'
        ws['A9'].font = Font(bold=True)
        ws['B9'].font = Font(bold=True)
        ws['A9'].fill = PatternFill(start_color='0059cc', end_color='0059cc', fill_type='solid')
        ws['B9'].fill = PatternFill(start_color='0059cc', end_color='0059cc', fill_type='solid')
        ws['A9'].font = Font(bold=True, color='FFFFFF')
        ws['B9'].font = Font(bold=True, color='FFFFFF')
        
        # Sample data
        data = [
            ['Total Evaluations', 'N/A'],
            ['Average Rating', 'N/A'],
            ['Response Rate', 'N/A'],
            ['Completion Rate', 'N/A'],
        ]
        
        for idx, row in enumerate(data, start=10):
            ws[f'A{idx}'] = row[0]
            ws[f'B{idx}'] = row[1]
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        # Footer
        footer_row = len(data) + 12
        ws[f'A{footer_row}'] = 'Generated by IntellEvalPro - Faculty Evaluation System'
        ws[f'A{footer_row}'].font = Font(italic=True, size=9)
        
        wb.save(file_path)
        
    except ImportError:
        # Fallback if openpyxl is not installed
        generate_simple_text_report(report, file_path)


def generate_csv_report(report, file_path):
    """Generate CSV report"""
    import csv
    from datetime import datetime
    
    with open(file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        
        # Header
        writer.writerow(['Report Name', report['report_name']])
        writer.writerow(['Report Type', report['report_type']])
        writer.writerow(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([])
        
        # Statistics
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Evaluations', 'N/A'])
        writer.writerow(['Average Rating', 'N/A'])
        writer.writerow(['Response Rate', 'N/A'])
        writer.writerow(['Completion Rate', 'N/A'])


def generate_simple_text_report(report, file_path):
    """Fallback: Generate simple text report"""
    from datetime import datetime
    
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(f"{'='*60}\n")
        f.write(f"{report['report_name']}\n")
        f.write(f"{'='*60}\n\n")
        f.write(f"Report Type: {report['report_type']}\n")
        f.write(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}\n")
        f.write(f"Format: {report['file_format']}\n\n")
        f.write(f"{'='*60}\n")
        f.write(f"Evaluation Statistics\n")
        f.write(f"{'='*60}\n\n")
        f.write(f"Total Evaluations: N/A\n")
        f.write(f"Average Rating: N/A\n")
        f.write(f"Response Rate: N/A\n")
        f.write(f"Completion Rate: N/A\n\n")
        f.write(f"{'='*60}\n")
        f.write(f"Generated by IntellEvalPro - Faculty Evaluation System\n")
        f.write(f"{'='*60}\n")


@api_bp.route('/evaluation-periods-detailed')
@login_required
def get_evaluation_periods_detailed():
    """Get all evaluation periods with detailed statistics"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        query = """
            SELECT 
                ep.period_id,
                ep.title,
                ep.start_date,
                ep.end_date,
                ep.status,
                CONCAT(ay.year_code, ' - ', at.term_name) as term_name,
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                SUM(CASE WHEN e.status = 'Completed' THEN 1 ELSE 0 END) as completed_evaluations
            FROM evaluation_periods ep
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            LEFT JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            LEFT JOIN evaluations e ON ep.period_id = e.period_id
            GROUP BY ep.period_id
            ORDER BY ep.start_date DESC
        """
        cursor.execute(query)
        periods = cursor.fetchall()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': periods
        })
        
    except Exception as e:
        print(f"Error getting evaluation periods: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@api_bp.route('/faculty-performance')
@login_required
def get_faculty_performance():
    """Get faculty performance analytics data"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get period filter from query params
        period_id = request.args.get('period_id')
        
        # Build WHERE clause for period filter (safe parameterization)
        if period_id:
            period_filter = "AND ep.period_id = %s"
            period_params = (period_id,)
        else:
            period_filter = ""
            period_params = ()
        
        # Get faculty performance rankings
        ranking_query = f"""
            SELECT 
                f.faculty_id,
                f.first_name,
                f.last_name,
                f.rank,
                p.name as department_name,
                COUNT(DISTINCT e.evaluation_id) as response_count,
                ROUND(AVG(er.rating), 2) as avg_score,
                ROUND(AVG(CASE WHEN er.created_at >= DATE_SUB(NOW(), INTERVAL 6 MONTH) 
                    THEN er.rating END), 2) as recent_avg,
                ROUND(AVG(CASE WHEN er.created_at < DATE_SUB(NOW(), INTERVAL 6 MONTH) 
                    THEN er.rating END), 2) as previous_avg
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id 
                AND e.status = 'Completed'
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            WHERE f.is_archived = FALSE {period_filter}
            GROUP BY f.faculty_id, f.first_name, f.last_name, f.rank, p.name
            HAVING COUNT(DISTINCT e.evaluation_id) > 0
            ORDER BY avg_score DESC, response_count DESC
        """
        if period_params:
            cursor.execute(ranking_query, period_params)
        else:
            cursor.execute(ranking_query)
        rankings = cursor.fetchall()
        
        # Debug: Log the query results
        print(f"DEBUG: Found {len(rankings)} faculty members")
        if len(rankings) == 0:
            # Check if there are any faculty at all
            cursor.execute("SELECT COUNT(*) as count FROM faculty WHERE is_archived = FALSE")
            faculty_count = cursor.fetchone()
            print(f"DEBUG: Total non-archived faculty: {faculty_count['count']}")
            
            # Check if there are evaluations
            cursor.execute("SELECT COUNT(*) as count FROM evaluations WHERE status = 'Completed'")
            eval_count = cursor.fetchone()
            print(f"DEBUG: Total completed evaluations: {eval_count['count']}")
            
            # Check class sections
            cursor.execute("SELECT COUNT(*) as count FROM class_sections")
            section_count = cursor.fetchone()
            print(f"DEBUG: Total class sections: {section_count['count']}")
            
            # Check evaluation responses
            cursor.execute("SELECT COUNT(*) as count FROM evaluation_responses")
            response_count = cursor.fetchone()
            print(f"DEBUG: Total evaluation responses: {response_count['count']}")
            
            # Test simplified query
            cursor.execute("""
                SELECT 
                    f.faculty_id,
                    f.first_name,
                    f.last_name,
                    COUNT(DISTINCT e.evaluation_id) as eval_count,
                    COUNT(DISTINCT er.response_id) as response_count,
                    AVG(er.rating) as avg_rating
                FROM faculty f
                LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                LEFT JOIN evaluations e ON cs.section_id = e.section_id
                LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                WHERE f.is_archived = FALSE
                GROUP BY f.faculty_id, f.first_name, f.last_name
            """)
            simple_results = cursor.fetchall()
            print(f"DEBUG: Simplified query results: {simple_results}")
        
        # Calculate trend for each faculty
        for faculty in rankings:
            if faculty['recent_avg'] and faculty['previous_avg']:
                faculty['trend'] = round(float(faculty['recent_avg']) - float(faculty['previous_avg']), 2)
            else:
                faculty['trend'] = 0
        
        # Get performance distribution
        distribution_query = f"""
            SELECT 
                CASE 
                    WHEN avg_score >= 4.5 THEN 'Excellent'
                    WHEN avg_score >= 3.5 THEN 'Good'
                    WHEN avg_score >= 2.5 THEN 'Average'
                    ELSE 'Poor'
                END as performance_level,
                COUNT(*) as count
            FROM (
                SELECT 
                    AVG(er.rating) as avg_score
                FROM faculty f
                LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                LEFT JOIN evaluations e ON cs.section_id = e.section_id 
                    AND e.status = 'Completed'
                LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
                LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                WHERE f.is_archived = FALSE {period_filter}
                GROUP BY f.faculty_id
                HAVING avg_score IS NOT NULL
            ) as faculty_scores
            GROUP BY performance_level
        """
        if period_params:
            cursor.execute(distribution_query, period_params)
        else:
            cursor.execute(distribution_query)
        distribution = cursor.fetchall()
        
        # Get performance trend over time (last 6 months)
        trend_query = f"""
            SELECT 
                DATE_FORMAT(e.completion_time, '%Y-%m') as month,
                ROUND(AVG(er.rating), 2) as avg_score,
                COUNT(DISTINCT e.evaluation_id) as evaluation_count,
                ROUND(COUNT(DISTINCT e.evaluation_id) * 100.0 / 
                    NULLIF((SELECT COUNT(*) FROM evaluations WHERE period_id = e.period_id), 0), 1) as response_rate
            FROM evaluations e
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            WHERE e.status = 'Completed'
                AND e.completion_time >= DATE_SUB(NOW(), INTERVAL 6 MONTH)
            GROUP BY month
            ORDER BY month
        """
        cursor.execute(trend_query)
        trends = cursor.fetchall()
        
        # Get improvement areas (criteria with lowest average scores)
        improvement_query = f"""
            SELECT 
                ec.description as criteria_name,
                ROUND(AVG(er.rating), 2) as avg_score,
                COUNT(*) as response_count
            FROM evaluation_responses er
            JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
            JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE e.status = 'Completed' {period_filter}
            GROUP BY ec.criteria_id
            ORDER BY avg_score ASC
            LIMIT 5
        """
        if period_params:
            cursor.execute(improvement_query, period_params)
        else:
            cursor.execute(improvement_query)
        improvement_areas = cursor.fetchall()
        
        # Calculate summary statistics
        stats_query = f"""
            SELECT 
                COUNT(DISTINCT faculty_id) as total_faculty,
                SUM(CASE WHEN avg_score >= 4.5 THEN 1 ELSE 0 END) as high_performers,
                SUM(CASE WHEN avg_score >= 3.5 AND avg_score < 4.5 THEN 1 ELSE 0 END) as average_performers,
                SUM(CASE WHEN avg_score < 3.5 THEN 1 ELSE 0 END) as needs_improvement,
                ROUND(AVG(avg_score), 2) as overall_avg
            FROM (
                SELECT 
                    f.faculty_id,
                    AVG(er.rating) as avg_score
                FROM faculty f
                LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
                LEFT JOIN evaluations e ON cs.section_id = e.section_id 
                    AND e.status = 'Completed'
                LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
                LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                WHERE f.is_archived = FALSE {period_filter}
                GROUP BY f.faculty_id
                HAVING avg_score IS NOT NULL
            ) as faculty_scores
        """
        if period_params:
            cursor.execute(stats_query, period_params)
        else:
            cursor.execute(stats_query)
        stats = cursor.fetchone()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': {
                'rankings': rankings,
                'distribution': distribution,
                'trends': trends,
                'improvement_areas': improvement_areas,
                'stats': stats
            }
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error getting faculty performance: {e}")
        print(f"Full traceback:\n{error_details}")
        return jsonify({
            'success': False, 
            'error': str(e),
            'details': error_details if conn else 'Database connection failed'
        }), 500
    finally:
        conn.close()


@api_bp.route('/response-analytics')
@login_required
def get_response_analytics():
    """Get student response analytics data"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get optional period filter
        period_id = request.args.get('period_id', None)
        period_filter = "AND ep.period_id = %s" if period_id else ""
        period_params = (period_id,) if period_id else ()
        
        # 1. Get total responses count
        query_total_responses = f"""
            SELECT COUNT(DISTINCT e.evaluation_id) as total_responses
            FROM evaluations e
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE e.status = 'Completed'
            {period_filter}
        """
        cursor.execute(query_total_responses, period_params)
        total_responses = cursor.fetchone()['total_responses'] or 0
        
        # 2. Calculate response rate
        query_response_rate = f"""
            SELECT 
                COUNT(CASE WHEN e.status = 'Completed' THEN 1 END) as completed,
                COUNT(e.evaluation_id) as total
            FROM evaluations e
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE 1=1
            {period_filter}
        """
        cursor.execute(query_response_rate, period_params)
        rate_data = cursor.fetchone()
        response_rate = round((rate_data['completed'] / rate_data['total'] * 100), 1) if rate_data['total'] > 0 else 0
        
        # 3. Calculate quality score (average rating)
        query_quality = f"""
            SELECT ROUND(AVG(er.rating), 1) as quality_score
            FROM evaluation_responses er
            JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE e.status = 'Completed'
            {period_filter}
        """
        cursor.execute(query_quality, period_params)
        quality_result = cursor.fetchone()
        quality_score = quality_result['quality_score'] if quality_result and quality_result['quality_score'] else 0
        
        # 4. Response rate over time (last 6 weeks)
        query_weekly_trend = f"""
            SELECT 
                WEEK(e.submitted_at) as week_num,
                DATE_FORMAT(e.submitted_at, '%Y-W%u') as week_label,
                COUNT(CASE WHEN e.status = 'Completed' THEN 1 END) as completed,
                COUNT(e.evaluation_id) as total
            FROM evaluations e
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE e.submitted_at >= DATE_SUB(NOW(), INTERVAL 6 WEEK)
            {period_filter}
            GROUP BY week_num, week_label
            ORDER BY week_num ASC
            LIMIT 6
        """
        cursor.execute(query_weekly_trend, period_params)
        weekly_data = cursor.fetchall()
        
        weekly_labels = []
        weekly_rates = []
        for i, week in enumerate(weekly_data, 1):
            weekly_labels.append(f'Week {i}')
            rate = round((week['completed'] / week['total'] * 100), 0) if week['total'] > 0 else 0
            weekly_rates.append(rate)
        
        # 5. Hourly distribution
        query_hourly = f"""
            SELECT 
                HOUR(e.submitted_at) as hour,
                COUNT(e.evaluation_id) as count
            FROM evaluations e
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE e.status = 'Completed'
            AND e.submitted_at IS NOT NULL
            {period_filter}
            GROUP BY hour
            ORDER BY hour
        """
        cursor.execute(query_hourly, period_params)
        hourly_raw = cursor.fetchall()
        
        # Create hourly distribution with proper labels
        hour_labels = ['6AM', '8AM', '10AM', '12PM', '2PM', '4PM', '6PM', '8PM', '10PM']
        hour_mapping = {6: 0, 8: 1, 10: 2, 12: 3, 14: 4, 16: 5, 18: 6, 20: 7, 22: 8}
        hourly_counts = [0] * 9
        
        for row in hourly_raw:
            if row['hour'] in hour_mapping:
                hourly_counts[hour_mapping[row['hour']]] = row['count']
        
        # 6. Department response rates
        query_departments = f"""
            SELECT 
                p.name as department_name,
                COUNT(CASE WHEN e.status = 'Completed' THEN 1 END) as completed,
                COUNT(e.evaluation_id) as total,
                ROUND(COUNT(CASE WHEN e.status = 'Completed' THEN 1 END) / COUNT(e.evaluation_id) * 100, 0) as rate
            FROM evaluations e
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            JOIN courses c ON cs.course_id = c.course_id
            JOIN programs p ON c.program_id = p.program_id
            WHERE 1=1
            {period_filter}
            GROUP BY p.program_id, p.name
            HAVING total > 0
            ORDER BY rate DESC
            LIMIT 5
        """
        cursor.execute(query_departments, period_params)
        departments = cursor.fetchall()
        
        # 7. Response quality metrics
        query_quality_metrics = f"""
            SELECT 
                COUNT(DISTINCT e.evaluation_id) as total_evaluations,
                COUNT(DISTINCT CASE WHEN er.rating IS NOT NULL THEN e.evaluation_id END) as complete_responses,
                COUNT(DISTINCT CASE WHEN e.comments IS NOT NULL AND LENGTH(e.comments) > 50 THEN e.evaluation_id END) as detailed_comments
            FROM evaluations e
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
            WHERE e.status = 'Completed'
            {period_filter}
        """
        cursor.execute(query_quality_metrics, period_params)
        quality_metrics = cursor.fetchone()
        
        complete_pct = round((quality_metrics['complete_responses'] / quality_metrics['total_evaluations'] * 100), 1) if quality_metrics['total_evaluations'] > 0 else 0
        detailed_pct = round((quality_metrics['detailed_comments'] / quality_metrics['total_evaluations'] * 100), 1) if quality_metrics['total_evaluations'] > 0 else 0
        
        # 8. Active students count
        query_active_students = f"""
            SELECT COUNT(DISTINCT e.student_id) as active_students
            FROM evaluations e
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE e.status = 'Completed'
            {period_filter}
        """
        cursor.execute(query_active_students, period_params)
        active_students = cursor.fetchone()['active_students'] or 0
        
        # 9. Text comments count
        query_comments = f"""
            SELECT COUNT(DISTINCT e.evaluation_id) as text_comments
            FROM evaluations e
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE e.status = 'Completed'
            AND e.comments IS NOT NULL
            AND LENGTH(e.comments) > 0
            {period_filter}
        """
        cursor.execute(query_comments, period_params)
        text_comments = cursor.fetchone()['text_comments'] or 0
        
        # 10. Daily activity (last 7 days)
        query_daily = f"""
            SELECT 
                DAYOFWEEK(e.submitted_at) as day_of_week,
                COUNT(e.evaluation_id) as count
            FROM evaluations e
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE e.status = 'Completed'
            AND e.submitted_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)
            {period_filter}
            GROUP BY day_of_week
            ORDER BY day_of_week
        """
        cursor.execute(query_daily, period_params)
        daily_raw = cursor.fetchall()
        
        # Map to Monday-Sunday (day_of_week: 1=Sunday, 2=Monday, etc.)
        daily_activity = [0] * 7  # [Mon, Tue, Wed, Thu, Fri, Sat, Sun]
        for row in daily_raw:
            dow = row['day_of_week']
            # Convert Sunday=1 to index 6, Monday=2 to index 0, etc.
            index = (dow + 5) % 7
            daily_activity[index] = row['count']
        
        # Determine activity levels (0-5 scale)
        max_activity = max(daily_activity) if daily_activity else 1
        daily_levels = []
        for count in daily_activity:
            if max_activity == 0:
                level = 0
            else:
                ratio = count / max_activity
                if ratio == 0:
                    level = 0
                elif ratio < 0.2:
                    level = 1
                elif ratio < 0.4:
                    level = 2
                elif ratio < 0.6:
                    level = 3
                elif ratio < 0.8:
                    level = 4
                else:
                    level = 5
            daily_levels.append({'count': count, 'level': level})
        
        # Calculate sentiment score (positive vs negative ratings)
        query_sentiment = f"""
            SELECT 
                COUNT(CASE WHEN er.rating >= 4 THEN 1 END) as positive,
                COUNT(CASE WHEN er.rating < 3 THEN 1 END) as negative,
                COUNT(er.rating) as total
            FROM evaluation_responses er
            JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            WHERE e.status = 'Completed'
            {period_filter}
        """
        cursor.execute(query_sentiment, period_params)
        sentiment_data = cursor.fetchone()
        
        if sentiment_data['total'] > 0:
            sentiment_score = round(((sentiment_data['positive'] - sentiment_data['negative']) / sentiment_data['total']), 2)
        else:
            sentiment_score = 0
        
        return jsonify({
            'success': True,
            'overview': {
                'total_responses': total_responses,
                'response_rate': response_rate,
                'quality_score': quality_score
            },
            'trends': {
                'weekly_labels': weekly_labels,
                'weekly_rates': weekly_rates
            },
            'hourly': {
                'labels': hour_labels,
                'counts': hourly_counts
            },
            'departments': departments,
            'quality_metrics': {
                'complete_responses': complete_pct,
                'detailed_comments': detailed_pct,
                'constructive_feedback': round((complete_pct + detailed_pct) / 2, 1),  # Estimated
                'rating_consistency': round(quality_score * 20, 1) if quality_score > 0 else 0  # Normalized to 100
            },
            'daily_activity': daily_levels,
            'summary': {
                'active_students': active_students,
                'text_comments': text_comments,
                'sentiment_score': sentiment_score
            }
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error getting response analytics: {e}")
        print(f"Full traceback:\n{error_details}")
        return jsonify({
            'success': False, 
            'error': str(e),
            'details': error_details
        }), 500
    finally:
        conn.close()


# ============================================================================
# EVALUATION MONITORING API ENDPOINTS
# ============================================================================

@api_bp.route('/evaluation-monitoring')
@login_required
def get_evaluation_monitoring():
    """Get evaluation monitoring data by department or section"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    period_id = request.args.get('period_id', type=int)
    view_type = request.args.get('view_type', 'department')
    
    if not period_id:
        return jsonify({'success': False, 'error': 'Period ID is required'}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get summary statistics
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT e.student_id) as total_students,
                SUM(CASE WHEN e.status = 'Completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN e.status = 'Pending' THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN e.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress
            FROM evaluations e
            WHERE e.period_id = %s
        """, (period_id,))
        summary = cursor.fetchone()
        
        # Calculate completion rate
        total_evals = summary['completed'] + summary['pending'] + summary['in_progress']
        completion_rate = round((summary['completed'] / total_evals * 100), 1) if total_evals > 0 else 0
        summary['completion_rate'] = completion_rate
        
        if view_type == 'department':
            # Get department statistics
            cursor.execute("""
                SELECT 
                    COALESCE(p.name, 'Unknown Department') as department_name,
                    COUNT(DISTINCT ss.student_id) as total_students,
                    COUNT(DISTINCT cs.section_id) as total_sections,
                    COUNT(DISTINCT CASE WHEN e.status = 'completed' THEN ss.student_id END) as completed_students,
                    SUM(CASE WHEN e.status = 'completed' THEN 1 ELSE 0 END) as completed,
                    SUM(CASE WHEN e.status = 'pending' THEN 1 ELSE 0 END) as pending,
                    SUM(CASE WHEN e.status = 'in_progress' THEN 1 ELSE 0 END) as in_progress
                FROM section_students ss
                JOIN sections s ON ss.section_id = s.section_id
                JOIN class_sections cs ON s.section_id = cs.section_ref_id
                LEFT JOIN evaluations e ON e.section_id = cs.section_id 
                    AND e.student_id = ss.student_id 
                    AND e.period_id = %s
                LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
                LEFT JOIN programs p ON f.program_id = p.program_id
                GROUP BY p.name
                ORDER BY department_name
            """, (period_id,))
            departments = cursor.fetchall()
            
            result_data = {
                'summary': summary,
                'departments': departments,
                'view_type': 'department'
            }
        else:
            # Get section statistics (aggregated by section, not by individual faculty)
            cursor.execute("""
                SELECT 
                    s.section_id,
                    s.section_name,
                    COUNT(DISTINCT cs.faculty_id) as faculty_count,
                    COUNT(DISTINCT ss.student_id) as total_students,
                    COUNT(DISTINCT CASE WHEN e.status = 'Completed' THEN e.student_id END) as completed_students,
                    COUNT(DISTINCT CASE WHEN e.status = 'Pending' THEN e.student_id END) as pending_students,
                    COUNT(DISTINCT CASE WHEN e.status = 'In Progress' THEN e.student_id END) as in_progress_students,
                    SUM(CASE WHEN e.status = 'Completed' THEN 1 ELSE 0 END) as completed,
                    SUM(CASE WHEN e.status = 'Pending' THEN 1 ELSE 0 END) as pending,
                    SUM(CASE WHEN e.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
                    COALESCE(p.name, 'Unknown') as department_name
                FROM sections s
                LEFT JOIN section_students ss ON s.section_id = ss.section_id
                LEFT JOIN class_sections cs ON s.section_id = cs.section_ref_id
                LEFT JOIN evaluations e ON cs.section_id = e.section_id AND e.period_id = %s
                LEFT JOIN faculty f ON cs.faculty_id = f.faculty_id
                LEFT JOIN programs p ON f.program_id = p.program_id
                WHERE cs.section_id IS NOT NULL
                GROUP BY s.section_id, s.section_name, p.name
                ORDER BY s.section_name
            """, (period_id,))
            sections = cursor.fetchall()
            
            result_data = {
                'summary': summary,
                'sections': sections,
                'view_type': 'section'
            }
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'data': result_data
        })
        
    except Exception as e:
        print(f"Error in evaluation monitoring: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


@api_bp.route('/department-students')
@login_required
def get_department_students():
    """Get list of students in a department with their evaluation status"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    period_id = request.args.get('period_id', type=int)
    department = request.args.get('department')
    
    if not period_id or not department:
        return jsonify({'success': False, 'error': 'Period ID and department are required'}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                st.std_Number as student_number,
                CONCAT(st.std_Firstname, ' ', st.std_Surname) as name,
                s.section_name,
                e.status,
                COUNT(e.evaluation_id) as total,
                SUM(CASE WHEN e.status = 'completed' THEN 1 ELSE 0 END) as completed
            FROM std_info st
            JOIN section_students ss ON st.id = ss.student_id
            JOIN sections s ON ss.section_id = s.section_id
            JOIN class_sections cs ON s.section_id = cs.section_ref_id
            JOIN evaluations e ON cs.section_id = e.section_id AND e.student_id = st.id
            JOIN faculty f ON cs.faculty_id = f.faculty_id
            JOIN programs p ON f.program_id = p.program_id
            WHERE e.period_id = %s AND p.name = %s
            GROUP BY st.id, st.std_Number, name, s.section_name, e.status
            ORDER BY st.std_Number
        """, (period_id, department))
        
        students = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'data': students
        })
        
    except Exception as e:
        print(f"Error getting department students: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


@api_bp.route('/section-students-progress')
@login_required
def get_section_students_progress():
    """Get list of students in a section with their evaluation progress"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    period_id = request.args.get('period_id', type=int)
    section_id = request.args.get('section_id', type=int)
    
    if not period_id or not section_id:
        return jsonify({'success': False, 'error': 'Period ID and section ID are required'}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all students with their evaluation status for this section
        # Including which faculty they need to evaluate
        cursor.execute("""
            SELECT DISTINCT
                st.std_Number as student_number,
                CONCAT(st.std_Firstname, ' ', st.std_Surname) as name,
                CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                sub.subject_code,
                sub.title as subject_name,
                e.status,
                e.completion_time,
                e.start_time
            FROM section_students ss
            JOIN std_info st ON ss.student_id = st.id
            JOIN sections s ON ss.section_id = s.section_id
            JOIN class_sections cs ON s.section_id = cs.section_ref_id
            JOIN faculty f ON cs.faculty_id = f.faculty_id
            JOIN subjects sub ON cs.subject_id = sub.subject_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id 
                AND e.student_id = st.id 
                AND e.period_id = %s
            WHERE s.section_id = %s
            ORDER BY st.std_Number, f.last_name
        """, (period_id, section_id))
        
        students = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'data': students
        })
        
    except Exception as e:
        print(f"Error getting section students: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


# ============================================================================
# STUDENT EVALUATION ENDPOINTS
# ============================================================================

@api_bp.route('/student/my-evaluations')
@login_required
def get_my_evaluations():
    """Get current student's evaluations with filtering"""
    if session.get('role') != 'student':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        user_id = session.get('user_id')
        
        # Get student_id from std_info
        cursor.execute("SELECT id FROM std_info WHERE user_id = %s", (user_id,))
        student_record = cursor.fetchone()
        
        if not student_record:
            return jsonify({'success': False, 'error': 'Student record not found'}), 404
        
        student_id = student_record['id']
        
        # Get filters
        status_filter = request.args.get('status', '')
        academic_year_filter = request.args.get('academic_year', '')
        search_filter = request.args.get('search', '')
        
        # Build query
        query = """
            SELECT 
                e.evaluation_id,
                e.status,
                e.completion_time,
                e.start_time,
                CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                sub.subject_code as course_code,
                sub.title as course_title,
                cs.section_name,
                cs.room,
                ep.title as period_title,
                ep.start_date as period_start,
                ep.end_date as period_end,
                ep.period_id,
                at.acad_term_id,
                ay.acad_year_id,
                ay.year_name
            FROM evaluations e
            JOIN class_sections cs ON e.section_id = cs.section_id
            JOIN faculty f ON cs.faculty_id = f.faculty_id
            JOIN subjects sub ON cs.subject_id = sub.subject_id
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE e.student_id = %s
        """
        params = [student_id]
        
        # Apply filters
        if status_filter:
            query += " AND e.status = %s"
            params.append(status_filter)
        
        if academic_year_filter:
            query += " AND ay.acad_year_id = %s"
            params.append(int(academic_year_filter))
        
        if search_filter:
            query += """ AND (
                CONCAT(f.first_name, ' ', f.last_name) LIKE %s OR
                sub.subject_code LIKE %s OR
                sub.title LIKE %s
            )"""
            search_param = f"%{search_filter}%"
            params.extend([search_param, search_param, search_param])
        
        query += " ORDER BY ep.end_date DESC, e.created_at DESC"
        
        cursor.execute(query, params)
        evaluations = cursor.fetchall()
        
        # Get statistics
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN status = 'Completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN status IN ('Pending', 'In Progress') THEN 1 ELSE 0 END) as pending
            FROM evaluations
            WHERE student_id = %s
        """, (student_id,))
        stats = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'evaluations': evaluations,
            'stats': {
                'total': stats['total'] or 0,
                'completed': stats['completed'] or 0,
                'pending': stats['pending'] or 0
            }
        })
        
    except Exception as e:
        print(f"Error getting student evaluations: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


@api_bp.route('/student/my-courses')
@login_required
def get_my_courses():
    """Get courses for current student's filter dropdown"""
    if session.get('role') != 'student':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        user_id = session.get('user_id')
        
        # Get student_id
        cursor.execute("SELECT id FROM std_info WHERE user_id = %s", (user_id,))
        student_record = cursor.fetchone()
        
        if not student_record:
            return jsonify({'success': False, 'error': 'Student record not found'}), 404
        
        student_id = student_record['id']
        
        # Get unique courses from student's evaluations
        cursor.execute("""
            SELECT DISTINCT 
                sub.subject_id as course_id,
                sub.subject_code as course_code,
                sub.title
            FROM evaluations e
            JOIN class_sections cs ON e.section_id = cs.section_id
            JOIN subjects sub ON cs.subject_id = sub.subject_id
            WHERE e.student_id = %s
            ORDER BY sub.subject_code
        """, (student_id,))
        courses = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'courses': courses
        })
        
    except Exception as e:
        print(f"Error getting student courses: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


# ============================================================================
# ACTIVITY LOGS API ENDPOINTS
# ============================================================================

@api_bp.route('/admin/activity-logs')
@login_required
def get_activity_logs():
    """Get all activity logs for admin"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all activity logs
        cursor.execute("""
            SELECT 
                log_id,
                user_id,
                user_name,
                user_role,
                activity_type,
                description,
                reason,
                target_user,
                ip_address,
                user_agent,
                additional_data,
                timestamp
            FROM activity_logs
            ORDER BY timestamp DESC
        """)
        logs = cursor.fetchall()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'logs': logs
        })
        
    except Exception as e:
        print(f"Error getting activity logs: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


@api_bp.route('/admin/export-logs')
@login_required
def export_activity_logs():
    """Export activity logs to Excel or CSV"""
    from flask import Response
    import csv
    from io import StringIO
    
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    export_format = request.args.get('format', 'csv')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get all activity logs
        cursor.execute("""
            SELECT 
                log_id,
                user_name,
                user_role,
                activity_type,
                description,
                reason,
                target_user,
                ip_address,
                timestamp
            FROM activity_logs
            ORDER BY timestamp DESC
        """)
        logs = cursor.fetchall()
        
        cursor.close()
        
        if export_format == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                from datetime import datetime
                
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Activity Logs"
                
                # Add headers
                headers = ['Log ID', 'User Name', 'Role', 'Activity Type', 'Description', 
                          'Reason', 'Target User', 'IP Address', 'Timestamp']
                ws.append(headers)
                
                # Style headers
                header_fill = PatternFill(start_color='0059cc', end_color='0059cc', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Add data
                for log in logs:
                    ws.append([
                        log.get('log_id'),
                        log.get('user_name'),
                        log.get('user_role'),
                        log.get('activity_type'),
                        log.get('description'),
                        log.get('reason'),
                        log.get('target_user'),
                        log.get('ip_address'),
                        log.get('timestamp')
                    ])
                
                # Save to BytesIO
                from io import BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                return Response(
                    output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=activity_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
                )
            except ImportError:
                # Fallback to CSV if openpyxl not available
                export_format = 'csv'
        
        if export_format == 'csv':
            # Create CSV
            si = StringIO()
            writer = csv.writer(si)
            
            # Write headers
            writer.writerow(['Log ID', 'User Name', 'Role', 'Activity Type', 'Description', 
                           'Reason', 'Target User', 'IP Address', 'Timestamp'])
            
            # Write data
            for log in logs:
                writer.writerow([
                    log.get('log_id'),
                    log.get('user_name'),
                    log.get('user_role'),
                    log.get('activity_type'),
                    log.get('description'),
                    log.get('reason'),
                    log.get('target_user'),
                    log.get('ip_address'),
                    log.get('timestamp')
                ])
            
            output = si.getvalue()
            from datetime import datetime
            
            return Response(
                output,
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=activity_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'}
            )
        
    except Exception as e:
        print(f"Error exporting activity logs: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


def log_activity(user_id=None, user_name=None, user_role=None, activity_type='', 
                description='', reason=None, target_user=None, ip_address=None, 
                additional_data=None):
    """
    Helper function to log activities
    
    Args:
        user_id: User ID (can be None for system events)
        user_name: Name of user performing action
        user_role: Role of user (admin, student, guidance, faculty)
        activity_type: Type of activity (login, logout, retake, evaluation, create, update, delete)
        description: Description of the activity
        reason: Reason for the activity (optional, used for retakes)
        target_user: Target user for the activity (optional)
        ip_address: IP address of the user
        additional_data: Additional JSON data (optional)
    """
    import json
    
    conn = get_db_connection()
    if not conn:
        print("Could not log activity - database connection failed")
        return False
    
    try:
        cursor = conn.cursor()
        
        # Convert additional_data to JSON string if it's a dict
        additional_data_json = None
        if additional_data:
            additional_data_json = json.dumps(additional_data) if isinstance(additional_data, dict) else additional_data
        
        cursor.execute("""
            INSERT INTO activity_logs 
            (user_id, user_name, user_role, activity_type, description, reason, 
             target_user, ip_address, additional_data)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (user_id, user_name, user_role, activity_type, description, reason, 
              target_user, ip_address, additional_data_json))
        
        conn.commit()
        cursor.close()
        
        return True
        
    except Exception as e:
        print(f"Error logging activity: {str(e)}")
        return False
    finally:
        if conn and conn.is_connected():
            conn.close()


# ============================================================================
# USER MANAGEMENT API ENDPOINTS
# ============================================================================

@api_bp.route('/users', methods=['GET'])
@login_required
def get_users():
    """
    Get all users
    Returns list of all users with their details
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({
            'success': False,
            'message': 'Database connection failed'
        }), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT user_id, username, email, first_name, last_name, role, 
                   is_active, is_verified, last_login, created_at, updated_at
            FROM users
            ORDER BY created_at DESC
        """)
        users = cursor.fetchall()
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': users,
            'count': len(users)
        })
        
    except Exception as e:
        print(f"Error fetching users: {e}")
        return jsonify({
            'success': False,
            'message': 'Error fetching users',
            'error': str(e)
        }), 500
    finally:
        conn.close()


@api_bp.route('/users', methods=['POST'])
@login_required
def create_user():
    """
    Create a new user
    Requires: username, email, first_name, last_name, role, password
    """
    from utils.security import generate_password_hash
    from datetime import datetime
    
    data = request.get_json()
    
    # Validate required fields
    required_fields = ['username', 'email', 'first_name', 'last_name', 'role', 'password']
    for field in required_fields:
        if not data.get(field):
            return jsonify({
                'success': False,
                'message': f'Missing required field: {field}'
            }), 400
    
    # Validate role
    valid_roles = ['admin', 'student', 'guidance']
    if data['role'] not in valid_roles:
        return jsonify({
            'success': False,
            'message': 'Invalid role. Must be admin, student, or guidance'
        }), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({
            'success': False,
            'message': 'Database connection failed'
        }), 500
    
    try:
        cursor = conn.cursor()
        
        # Check if username already exists
        cursor.execute("SELECT user_id FROM users WHERE username = %s", (data['username'],))
        if cursor.fetchone():
            return jsonify({
                'success': False,
                'message': 'Username already exists'
            }), 400
        
        # Hash password
        hashed_password = generate_password_hash(data['password'])
        now = datetime.now()
        
        # Insert new user
        cursor.execute("""
            INSERT INTO users 
            (username, password, email, first_name, last_name, role, 
             is_active, is_verified, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, 1, 1, %s, %s)
        """, (data['username'], hashed_password, data['email'], 
              data['first_name'], data['last_name'], data['role'], now, now))
        
        user_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        
        # Log activity
        log_activity(
            user_id=session.get('user_id'),
            user_name=session.get('username'),
            user_role=session.get('role'),
            activity_type='create',
            description=f"Created new user: {data['username']} ({data['role']})",
            ip_address=request.remote_addr,
            target_user=data['username']
        )
        
        return jsonify({
            'success': True,
            'message': 'User created successfully',
            'data': {'user_id': user_id}
        }), 201
        
    except Exception as e:
        print(f"Error creating user: {e}")
        return jsonify({
            'success': False,
            'message': 'Error creating user',
            'error': str(e)
        }), 500
    finally:
        conn.close()


@api_bp.route('/users/<int:user_id>', methods=['PUT'])
@login_required
def update_user(user_id):
    """
    Update a user
    Can update: username, email, first_name, last_name, role, is_active
    """
    from datetime import datetime
    
    data = request.get_json()
    
    # Validate role if provided
    if 'role' in data:
        valid_roles = ['admin', 'student', 'guidance']
        if data['role'] not in valid_roles:
            return jsonify({
                'success': False,
                'message': 'Invalid role. Must be admin, student, or guidance'
            }), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({
            'success': False,
            'message': 'Database connection failed'
        }), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if user exists
        cursor.execute("SELECT * FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({
                'success': False,
                'message': 'User not found'
            }), 404
        
        # Build update query dynamically
        update_fields = []
        update_values = []
        
        allowed_fields = ['username', 'email', 'first_name', 'last_name', 'role', 'is_active']
        for field in allowed_fields:
            if field in data:
                update_fields.append(f"{field} = %s")
                update_values.append(data[field])
        
        if not update_fields:
            return jsonify({
                'success': False,
                'message': 'No valid fields to update'
            }), 400
        
        # Add updated_at timestamp
        update_fields.append("updated_at = %s")
        update_values.append(datetime.now())
        update_values.append(user_id)
        
        # Execute update
        update_query = f"UPDATE users SET {', '.join(update_fields)} WHERE user_id = %s"
        cursor.execute(update_query, update_values)
        conn.commit()
        cursor.close()
        
        # Log activity
        log_activity(
            user_id=session.get('user_id'),
            user_name=session.get('username'),
            user_role=session.get('role'),
            activity_type='update',
            description=f"Updated user: {user['username']}",
            ip_address=request.remote_addr,
            target_user=user['username']
        )
        
        return jsonify({
            'success': True,
            'message': 'User updated successfully'
        })
        
    except Exception as e:
        print(f"Error updating user: {e}")
        return jsonify({
            'success': False,
            'message': 'Error updating user',
            'error': str(e)
        }), 500
    finally:
        conn.close()


@api_bp.route('/users/<int:user_id>', methods=['DELETE'])
@login_required
def delete_user(user_id):
    """
    Delete a user
    Note: This permanently deletes the user from the database
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({
            'success': False,
            'message': 'Database connection failed'
        }), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if user exists
        cursor.execute("SELECT * FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({
                'success': False,
                'message': 'User not found'
            }), 404
        
        # Prevent deleting yourself
        if user_id == session.get('user_id'):
            return jsonify({
                'success': False,
                'message': 'Cannot delete your own account'
            }), 400
        
        # Delete user
        cursor.execute("DELETE FROM users WHERE user_id = %s", (user_id,))
        conn.commit()
        cursor.close()
        
        # Log activity
        log_activity(
            user_id=session.get('user_id'),
            user_name=session.get('username'),
            user_role=session.get('role'),
            activity_type='delete',
            description=f"Deleted user: {user['username']} ({user['role']})",
            ip_address=request.remote_addr,
            target_user=user['username']
        )
        
        return jsonify({
            'success': True,
            'message': 'User deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting user: {e}")
        return jsonify({
            'success': False,
            'message': 'Error deleting user',
            'error': str(e)
        }), 500
    finally:
        conn.close()


@api_bp.route('/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
def reset_user_password(user_id):
    """
    Reset a user's password
    Requires: password
    """
    from utils.security import generate_password_hash
    from datetime import datetime
    
    data = request.get_json()
    
    if not data.get('password'):
        return jsonify({
            'success': False,
            'message': 'Password is required'
        }), 400
    
    if len(data['password']) < 5:
        return jsonify({
            'success': False,
            'message': 'Password must be at least 5 characters'
        }), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({
            'success': False,
            'message': 'Database connection failed'
        }), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Check if user exists
        cursor.execute("SELECT * FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({
                'success': False,
                'message': 'User not found'
            }), 404
        
        # Hash new password
        hashed_password = generate_password_hash(data['password'])
        
        # Update password
        cursor.execute("""
            UPDATE users 
            SET password = %s, updated_at = %s
            WHERE user_id = %s
        """, (hashed_password, datetime.now(), user_id))
        
        conn.commit()
        cursor.close()
        
        # Log activity
        log_activity(
            user_id=session.get('user_id'),
            user_name=session.get('username'),
            user_role=session.get('role'),
            activity_type='update',
            description=f"Reset password for user: {user['username']}",
            ip_address=request.remote_addr,
            target_user=user['username']
        )
        
        return jsonify({
            'success': True,
            'message': 'Password reset successfully'
        })
        
    except Exception as e:
        print(f"Error resetting password: {e}")
        return jsonify({
            'success': False,
            'message': 'Error resetting password',
            'error': str(e)
        }), 500
    finally:
        conn.close()


@api_bp.route('/auth/request-password-reset', methods=['POST'])
def request_password_reset():
    """
    Request a password reset link for a user identified by username (school ID)
    Expects JSON with `username` (school ID)
    This will generate a secure token, store it in the users table and send an email
    """
    from datetime import datetime, timedelta
    import secrets
    from utils.email_utils import send_email

    data = request.get_json() or {}
    username = data.get('username') or data.get('school_id')
    print(f"[Password Reset] Received request for username/school_id: {username}")
    if not username:
        return jsonify({'success': False, 'message': 'School ID is required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()
        if not user:
            # Do not reveal whether user exists to prevent account enumeration
            return jsonify({'success': True, 'message': 'If this School ID is registered, a reset link has been sent to the registered email.'})

        # Generate a secure token and expiry (1 hour)
        token = secrets.token_urlsafe(48)
        expiry = datetime.now() + timedelta(hours=1)

        # Store token in users table
        cursor.execute(
            "UPDATE users SET reset_token = %s, reset_token_expiry = %s WHERE user_id = %s",
            (token, expiry, user['user_id'])
        )
        conn.commit()

        # Build reset URL
        # Use a relative link so it's valid in dev and production; front-end will navigate accordingly
        reset_url = f"{request.url_root.rstrip('/')}/reset-password?token={token}&type=student"

        subject = 'IntellEvalPro Password Reset'
        # Use a polished template with trust signals
        from utils.email_utils import _render_reset_template
        body_html, body_text = _render_reset_template(user.get('first_name') or user.get('username'), reset_url, expiry_hours=1)

        # Attempt to send email (best-effort)
        try:
            if not user.get('email'):
                print(f"[Password Reset] User {username} has no email set; skipping send_email")
                logger.warning(f"User {username} tried to reset password but no email is set for the account")
                send_result = False
            else:
                # Send with friendly sender name
                send_result = send_email(user['email'], subject, body_html, body_text, from_name='IntellEvalPro Support')
            # In debug environments, print the reset URL for inspection
            print(f"[Password Reset] Reset URL for user {username}: {reset_url}")
            print(f"[Password Reset] send_email returned: {send_result}")
        except Exception as e:
            # Log but continue; we don't leak failure info
            logger.exception('Error sending reset email')

        return jsonify({'success': True, 'message': 'If this School ID is registered, a reset link has been sent to the registered email.'})
    except Exception as e:
        print(f"Error requesting password reset: {e}")
        return jsonify({'success': False, 'message': 'Error requesting password reset', 'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@api_bp.route('/auth/confirm-reset', methods=['POST'])
def confirm_password_reset():
    """
    Confirm the password reset using a token. Expects JSON with 'token' and 'password'.
    """
    from datetime import datetime
    from utils.security import generate_password_hash

    data = request.get_json() or {}
    token = data.get('token')
    new_password = data.get('password')

    if not token or not new_password:
        return jsonify({'success': False, 'message': 'Token and password are required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users WHERE reset_token = %s", (token,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'success': False, 'message': 'Invalid or expired token'}), 400

        expiry = user.get('reset_token_expiry')
        if not expiry or expiry < datetime.now():
            return jsonify({'success': False, 'message': 'Token has expired. Please request a new reset link.'}), 400

        # Update password
        hashed = generate_password_hash(new_password)
        cursor.execute(
            "UPDATE users SET password = %s, reset_token = NULL, reset_token_expiry = NULL, updated_at = %s WHERE user_id = %s",
            (hashed, datetime.now(), user['user_id'])
        )
        conn.commit()

        # Log activity
        try:
            log_activity(
                user_id=user['user_id'],
                user_name=user.get('username'),
                user_role=user.get('role'),
                activity_type='update',
                description='User reset password via token',
                ip_address=request.remote_addr
            )
        except Exception:
            pass

        return jsonify({'success': True, 'message': 'Password has been reset successfully'})
    except Exception as e:
        print(f"Error confirming password reset: {e}")
        return jsonify({'success': False, 'message': 'Error resetting password', 'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@api_bp.route('/guidance/export-rankings-pdf')
@login_required
def export_rankings_pdf():
    """Export faculty rankings to PDF with ISO 25010 format - same as department analysis"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from flask import send_file
        import os
        from datetime import datetime
        
        academic_year_id = request.args.get('academic_year_id', type=int)
        period_id = request.args.get('period_id', type=int)
        department_id = request.args.get('department_id', type=int)  # Optional
        
        # Get signature position parameters
        sig_guidance_y = request.args.get('sig_guidance_y', type=float, default=8.33)
        sig_guidance_x = request.args.get('sig_guidance_x', type=float, default=10.0)
        sig_dean_y = request.args.get('sig_dean_y', type=float, default=33.33)
        sig_dean_x = request.args.get('sig_dean_x', type=float, default=10.0)
        sig_president_y = request.args.get('sig_president_y', type=float, default=58.33)
        sig_president_x = request.args.get('sig_president_x', type=float, default=10.0)
        
        if not academic_year_id or not period_id:
            return jsonify({'success': False, 'message': 'Academic year and period are required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get department info if specified
        department = None
        if department_id:
            cursor.execute("""
                SELECT program_id, name, program_code
                FROM programs
                WHERE program_id = %s
            """, (department_id,))
            department = cursor.fetchone()
        
        # Get period and year info
        cursor.execute("""
            SELECT ep.title as period_name, ay.year_name
            FROM evaluation_periods ep
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE ep.period_id = %s AND at.acad_year_id = %s
        """, (period_id, academic_year_id))
        period_info = cursor.fetchone()
        
        if not period_info:
            return jsonify({'success': False, 'message': 'Period information not found'}), 404
        
        # Get faculty rankings data using the same query as the rankings API
        query = """
            SELECT 
                f.faculty_id,
                f.first_name,
                f.last_name,
                f.program_id,
                p.name AS department_name,
                COUNT(DISTINCT e.evaluation_id) AS total_evaluations,
                -- Overall average rating
                ROUND(AVG(overall_avg.avg_rating), 2) AS average_rating,
                -- Learning Delivery (Category ID 2) 
                ROUND(AVG(cat2.avg_rating), 2) AS learning_delivery,
                -- Assessment of Student Learning (Category ID 4)
                ROUND(AVG(cat4.avg_rating), 2) AS assessment_learning,
                -- Student-Teacher Engagement (Category ID 5)
                ROUND(AVG(cat5.avg_rating), 2) AS student_engagement
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
                AND e.period_id = %s
                AND e.status = 'completed'
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            -- Overall average subquery
            LEFT JOIN (
                SELECT 
                    e.evaluation_id,
                    AVG(er.rating) as avg_rating
                FROM evaluations e
                JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                WHERE e.period_id = %s AND e.status = 'completed'
                GROUP BY e.evaluation_id
            ) overall_avg ON e.evaluation_id = overall_avg.evaluation_id
            -- Learning Delivery (Category 2) average
            LEFT JOIN (
                SELECT 
                    e.evaluation_id,
                    AVG(er.rating) as avg_rating
                FROM evaluations e
                JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
                WHERE e.period_id = %s AND e.status = 'completed' AND ec.category_id = 2
                GROUP BY e.evaluation_id
            ) cat2 ON e.evaluation_id = cat2.evaluation_id
            -- Assessment of Student Learning (Category 4) average  
            LEFT JOIN (
                SELECT 
                    e.evaluation_id,
                    AVG(er.rating) as avg_rating
                FROM evaluations e
                JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
                WHERE e.period_id = %s AND e.status = 'completed' AND ec.category_id = 4
                GROUP BY e.evaluation_id
            ) cat4 ON e.evaluation_id = cat4.evaluation_id
            -- Student-Teacher Engagement (Category 5) average
            LEFT JOIN (
                SELECT 
                    e.evaluation_id,
                    AVG(er.rating) as avg_rating
                FROM evaluations e
                JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
                WHERE e.period_id = %s AND e.status = 'completed' AND ec.category_id = 5
                GROUP BY e.evaluation_id
            ) cat5 ON e.evaluation_id = cat5.evaluation_id
            WHERE at.acad_year_id = %s
        """
        
        params = [period_id, period_id, period_id, period_id, period_id, academic_year_id]
        
        if department_id:
            query += " AND f.program_id = %s"
            params.append(department_id)
        
        query += """
            GROUP BY f.faculty_id, f.first_name, f.last_name, f.program_id, p.name
            HAVING COUNT(DISTINCT e.evaluation_id) > 0
            ORDER BY average_rating DESC, total_evaluations DESC
        """
        
        cursor.execute(query, tuple(params))
        rankings = cursor.fetchall()
        
        # Helper function for remarks
        def get_remarks(mean):
            if mean >= 4.5:
                return "OUTSTANDING"
            elif mean >= 3.5:
                return "HIGHLY SATISFACTORY"
            elif mean >= 2.5:
                return "SATISFACTORY"
            elif mean >= 1.5:
                return "NEEDS IMPROVEMENT"
            else:
                return "POOR"
        
        cursor.close()
        conn.close()
        
        # Create PDF
        if department:
            filename = f"Faculty_Rankings_{department['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        else:
            filename = f"Faculty_Rankings_All_Departments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
        filepath = os.path.join('static', 'reports', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=letter,
                              rightMargin=0.5*inch, leftMargin=0.5*inch,
                              topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', fontSize=16, fontName='Times-Bold', 
                                   alignment=TA_CENTER, spaceAfter=6)
        subtitle_style = ParagraphStyle('CustomSubtitle', fontSize=12, fontName='Times-Bold',
                                      alignment=TA_CENTER, spaceAfter=6)
        header_style = ParagraphStyle('CustomHeader', fontSize=10, fontName='Times-Bold',
                                    alignment=TA_LEFT, spaceAfter=6)
        
        story = []
        
        # Add logo and header (same as faculty results)
        logo_path = os.path.join('static', 'images', 'nclogo.png')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=0.8*inch, height=0.8*inch)
            story.append(logo)
        
        # Title (same format as faculty results)
        story.append(Paragraph('NORZAGARAY COLLEGE', title_style))
        story.append(Paragraph('Municipal Compound, Norzagaray, Bulacan', ParagraphStyle('Subtitle', fontSize=10, alignment=TA_CENTER)))
        story.append(Paragraph('GUIDANCE AND COUNSELING CENTER', ParagraphStyle('Subtitle2', fontSize=10, fontName='Times-Bold', alignment=TA_CENTER, spaceAfter=6)))
        story.append(Paragraph('FACULTY TEACHING PERFORMANCE EVALUATION RANKINGS REPORT', ParagraphStyle('Subtitle3', fontSize=9, alignment=TA_CENTER, spaceAfter=3)))
        
        # Remove any status text from period name using regex
        import re
        period_name = re.sub(r'\s*\([^)]*\)\s*$', '', period_info['period_name']).strip()
        story.append(Paragraph(f"{period_name}, A.Y. {period_info['year_name']}", ParagraphStyle('Period', fontSize=9, alignment=TA_CENTER, spaceAfter=12)))
        
        story.append(Spacer(1, 0.2*inch))
        
        # Department info (same format as faculty results)
        if department:
            story.append(Paragraph(f"<b>Department:</b> {department['name']}", header_style))
        else:
            story.append(Paragraph(f"<b>Scope:</b> All Departments", header_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Faculty Rankings Table (same format as faculty results)
        story.append(Paragraph("<b>FACULTY EVALUATION RANKINGS</b>", ParagraphStyle('CategoryHeader', fontSize=11, fontName='Times-Bold', spaceAfter=6)))
        
        # Create paragraph styles for table cells (same as faculty results)
        cell_style = ParagraphStyle('CellStyle', fontSize=8, leading=10)
        remarks_style = ParagraphStyle('RemarksStyle', fontSize=7, leading=9, alignment=TA_CENTER)
        
        # Build table data (same format as faculty results)
        table_data = [['Rank', 'Faculty Name', 'Department', 'Overall Rating', 'Learning Delivery', 'Assessment Learning', 'Student Engagement', 'Evaluations', 'Remarks']]
        
        for idx, faculty in enumerate(rankings, 1):
            overall_rating = faculty['average_rating'] or 0.0
            learning_delivery = faculty['learning_delivery'] or 0.0
            assessment_learning = faculty['assessment_learning'] or 0.0
            student_engagement = faculty['student_engagement'] or 0.0
            
            # Wrap text in Paragraph for proper formatting (same as faculty results)
            remarks_text = get_remarks(overall_rating)
            
            row = [
                str(idx),
                Paragraph(f"{faculty['first_name']} {faculty['last_name']}", cell_style),
                Paragraph(faculty['department_name'] or 'N/A', cell_style),
                f"{overall_rating:.2f}",
                f"{learning_delivery:.2f}",
                f"{assessment_learning:.2f}",
                f"{student_engagement:.2f}",
                str(faculty['total_evaluations']),
                Paragraph(remarks_text, remarks_style)
            ]
            table_data.append(row)
        
        # Create table (same column structure and styling as faculty results)
        table = Table(table_data, colWidths=[0.4*inch, 1.8*inch, 1.3*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 1.0*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 0.15*inch))
        
        # Rating scale (same format as faculty results)
        story.append(Paragraph('<b>RATING SCALE:</b>', ParagraphStyle('RatingHeader', fontSize=10, fontName='Times-Bold', spaceAfter=6)))
        
        # Calculate overall statistics for the summary
        total_faculty = len(rankings)
        top_performer = f"{rankings[0]['first_name']} {rankings[0]['last_name']}" if rankings else 'N/A'
        top_rating = f"{rankings[0]['average_rating']:.2f}" if rankings else 'N/A'
        
        rating_data = [
            ['Rating', 'Equivalent', 'Total Faculty:', str(total_faculty)],
            ['4.50 - 5.00', 'OUTSTANDING', 'Top Performer:', top_performer],
            ['3.50 - 4.49', 'HIGHLY SATISFACTORY', 'Top Rating:', top_rating],
            ['2.50 - 3.49', 'SATISFACTORY', '', ''],
            ['1.50 - 2.49', 'NEEDS IMPROVEMENT', '', ''],
            ['1.00 - 1.49', 'POOR', '', '']
        ]
        
        rating_table = Table(rating_data, colWidths=[1.5*inch, 2*inch, 1.2*inch, 1.5*inch])
        rating_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('SPAN', (2, 3), (3, 5)),
        ]))
        
        story.append(rating_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Signatures with custom positioning (same format as faculty results)
        # Convert percentage to actual spacing (max available space ~4 inches after rating table)
        max_signature_space = 4.0  # inches available for signature positioning
        
        # Calculate vertical spacing based on Y percentages
        guidance_spacing = (sig_guidance_y / 100.0) * max_signature_space
        dean_spacing = ((sig_dean_y - sig_guidance_y) / 100.0) * max_signature_space
        president_spacing = ((sig_president_y - sig_dean_y) / 100.0) * max_signature_space
        
        # Determine horizontal alignment based on X percentage (same as faculty results)
        # 0-33% = LEFT, 33-66% = CENTER, 66-100% = RIGHT
        def get_alignment(x_percent):
            if x_percent < 33:
                return TA_LEFT
            elif x_percent < 66:
                return TA_CENTER
            else:
                return 2  # TA_RIGHT value
        
        guidance_align = get_alignment(sig_guidance_x)
        dean_align = get_alignment(sig_dean_x)
        president_align = get_alignment(sig_president_x)
        
        # Guidance Counselor Signature (changed from Faculty to Guidance Counselor)
        story.append(Spacer(1, guidance_spacing * inch))
        story.append(Paragraph('_' * 50, ParagraphStyle('SigLine1', alignment=guidance_align)))
        story.append(Paragraph('<b>SIGNATURE OF GUIDANCE COUNSELOR</b>', ParagraphStyle('Sig', fontSize=9, spaceAfter=12, alignment=guidance_align)))
        
        # Dean Signature (same as faculty results)
        story.append(Spacer(1, dean_spacing * inch))
        story.append(Paragraph('_' * 50, ParagraphStyle('SigLine2', alignment=dean_align)))
        story.append(Paragraph('<b>SIGNATURE OF COLLEGE DEAN</b>', ParagraphStyle('Sig2', fontSize=9, spaceAfter=12, alignment=dean_align)))
        
        # President Signature (same as faculty results)
        story.append(Spacer(1, president_spacing * inch))
        story.append(Paragraph('<b>NOTED BY:</b>', ParagraphStyle('Noted', fontSize=9, fontName='Times-Bold', spaceAfter=6, alignment=president_align)))
        story.append(Paragraph('<b>MA. LIBERTY DG. PASCUAL, Ph.D.</b>', ParagraphStyle('Name', fontSize=10, fontName='Times-Bold', alignment=president_align)))
        story.append(Paragraph('College President', ParagraphStyle('Title', alignment=president_align)))
        
        # Build PDF
        doc.build(story)
        
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/pdf')
        
    except Exception as e:
        print(f"Error exporting rankings PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/guidance/export-rankings-excel')
@login_required
def export_rankings_excel():
    """Export faculty rankings to Excel with ISO 25010 format - same as department analysis"""
    if session.get('role') not in ['guidance', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from flask import send_file
        import os
        from datetime import datetime
        
        academic_year_id = request.args.get('academic_year_id', type=int)
        period_id = request.args.get('period_id', type=int)
        department_id = request.args.get('department_id', type=int)  # Optional
        
        if not academic_year_id or not period_id:
            return jsonify({'success': False, 'message': 'Academic year and period are required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get department info if specified
        department = None
        if department_id:
            cursor.execute("""
                SELECT program_id, name, program_code
                FROM programs
                WHERE program_id = %s
            """, (department_id,))
            department = cursor.fetchone()
        
        # Get period and year info
        cursor.execute("""
            SELECT ep.title as period_name, ay.year_name
            FROM evaluation_periods ep
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE ep.period_id = %s AND at.acad_year_id = %s
        """, (period_id, academic_year_id))
        period_info = cursor.fetchone()
        
        if not period_info:
            return jsonify({'success': False, 'message': 'Period information not found'}), 404
        
        # Get faculty rankings data using the same query as the PDF export
        query = """
            SELECT 
                f.faculty_id,
                f.first_name,
                f.last_name,
                f.program_id,
                p.name AS department_name,
                COUNT(DISTINCT e.evaluation_id) AS total_evaluations,
                -- Overall average rating
                ROUND(AVG(overall_avg.avg_rating), 2) AS average_rating,
                -- Learning Delivery (Category ID 2) 
                ROUND(AVG(cat2.avg_rating), 2) AS learning_delivery,
                -- Assessment of Student Learning (Category ID 4)
                ROUND(AVG(cat4.avg_rating), 2) AS assessment_learning,
                -- Student-Teacher Engagement (Category ID 5)
                ROUND(AVG(cat5.avg_rating), 2) AS student_engagement
            FROM faculty f
            LEFT JOIN programs p ON f.program_id = p.program_id
            LEFT JOIN class_sections cs ON f.faculty_id = cs.faculty_id
            LEFT JOIN evaluations e ON cs.section_id = e.section_id
                AND e.period_id = %s
                AND e.status = 'completed'
            LEFT JOIN evaluation_periods ep ON e.period_id = ep.period_id
            LEFT JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            -- Overall average subquery
            LEFT JOIN (
                SELECT 
                    e.evaluation_id,
                    AVG(er.rating) as avg_rating
                FROM evaluations e
                JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                WHERE e.period_id = %s AND e.status = 'completed'
                GROUP BY e.evaluation_id
            ) overall_avg ON e.evaluation_id = overall_avg.evaluation_id
            -- Learning Delivery (Category 2) average
            LEFT JOIN (
                SELECT 
                    e.evaluation_id,
                    AVG(er.rating) as avg_rating
                FROM evaluations e
                JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
                WHERE e.period_id = %s AND e.status = 'completed' AND ec.category_id = 2
                GROUP BY e.evaluation_id
            ) cat2 ON e.evaluation_id = cat2.evaluation_id
            -- Assessment of Student Learning (Category 4) average  
            LEFT JOIN (
                SELECT 
                    e.evaluation_id,
                    AVG(er.rating) as avg_rating
                FROM evaluations e
                JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
                WHERE e.period_id = %s AND e.status = 'completed' AND ec.category_id = 4
                GROUP BY e.evaluation_id
            ) cat4 ON e.evaluation_id = cat4.evaluation_id
            -- Student-Teacher Engagement (Category 5) average
            LEFT JOIN (
                SELECT 
                    e.evaluation_id,
                    AVG(er.rating) as avg_rating
                FROM evaluations e
                JOIN evaluation_responses er ON e.evaluation_id = er.evaluation_id
                JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
                WHERE e.period_id = %s AND e.status = 'completed' AND ec.category_id = 5
                GROUP BY e.evaluation_id
            ) cat5 ON e.evaluation_id = cat5.evaluation_id
            WHERE at.acad_year_id = %s
        """
        
        params = [period_id, period_id, period_id, period_id, period_id, academic_year_id]
        
        if department_id:
            query += " AND f.program_id = %s"
            params.append(department_id)
        
        query += """
            GROUP BY f.faculty_id, f.first_name, f.last_name, f.program_id, p.name
            HAVING COUNT(DISTINCT e.evaluation_id) > 0
            ORDER BY average_rating DESC, total_evaluations DESC
        """
        
        cursor.execute(query, tuple(params))
        rankings = cursor.fetchall()
        
        # Helper function for remarks
        def get_remarks(mean):
            if mean >= 4.5:
                return "OUTSTANDING"
            elif mean >= 3.5:
                return "HIGHLY SATISFACTORY"
            elif mean >= 2.5:
                return "SATISFACTORY"
            elif mean >= 1.5:
                return "NEEDS IMPROVEMENT"
            else:
                return "POOR"
        
        cursor.close()
        conn.close()
        
        # Create Excel workbook (same format as faculty results)
        wb = Workbook()
        ws = wb.active
        ws.title = "Faculty Rankings"
        
        # Styles (same as faculty results)
        header_font = Font(name='Times New Roman', size=14, bold=True)
        subheader_font = Font(name='Times New Roman', size=11, bold=True)
        normal_font = Font(name='Times New Roman', size=10)
        bold_font = Font(name='Times New Roman', size=10, bold=True)
        
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        gray_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add logo at the top centered with college name
        row = 1
        try:
            from openpyxl.drawing.image import Image as XLImage
            logo_path = os.path.join('static', 'images', 'nclogo.png')
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                # Make logo appropriately sized (80x80 pixels)
                img.width = 80
                img.height = 80
                # Center the logo horizontally above the college name (around column D-E)
                ws.add_image(img, f'D{row}')
                # Set row heights to accommodate logo
                ws.row_dimensions[row].height = 65
                ws.row_dimensions[row + 1].height = 5
                # Space for the logo
                row += 2
        except Exception as e:
            print(f"Could not add logo: {e}")
            pass
        
        # Header section (same format as faculty results)
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'NORZAGARAY COLLEGE'
        cell.font = Font(name='Times New Roman', size=12, bold=True)
        cell.alignment = center_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'Municipal Compound, Norzagaray, Bulacan'
        cell.font = normal_font
        cell.alignment = center_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'GUIDANCE AND COUNSELING CENTER'
        cell.font = Font(name='Times New Roman', size=10, bold=True)
        cell.alignment = center_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'FACULTY TEACHING PERFORMANCE EVALUATION RANKINGS REPORT'
        cell.font = Font(name='Times New Roman', size=9)
        cell.alignment = center_alignment
        row += 1
        
        # Remove any status text from period name using regex (same as faculty results)
        import re
        period_name = re.sub(r'\s*\([^)]*\)\s*$', '', period_info['period_name']).strip()
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = f"{period_name}, A.Y. {period_info['year_name']}"
        cell.font = Font(name='Times New Roman', size=9)
        cell.alignment = center_alignment
        row += 1
        
        row += 1  # Empty row
        
        # Department info (same format as faculty results)
        if department:
            ws[f'A{row}'] = f"Department: {department['name']}"
        else:
            ws[f'A{row}'] = "Scope: All Departments"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        row += 1  # Empty row
        
        # Rankings table (same format as faculty results categories)
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'FACULTY EVALUATION RANKINGS'
        cell.font = subheader_font
        cell.fill = gray_fill
        cell.alignment = left_alignment
        cell.border = thin_border
        row += 1
        
        # Table headers (same format as faculty results)
        headers = ['Rank', 'Faculty Name', 'Department', 'Overall Rating', 'Learning Delivery', 'Assessment Learning', 'Student Engagement', 'Evaluations', 'Remarks']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = bold_font
            cell.fill = gray_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        row += 1
        
        # Rankings data (same format as faculty results)
        for idx, faculty in enumerate(rankings, 1):
            overall_rating = faculty['average_rating'] or 0.0
            learning_delivery = faculty['learning_delivery'] or 0.0
            assessment_learning = faculty['assessment_learning'] or 0.0
            student_engagement = faculty['student_engagement'] or 0.0
            
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = f"{faculty['first_name']} {faculty['last_name']}"
            ws.cell(row=row, column=3).value = faculty['department_name'] or 'N/A'
            ws.cell(row=row, column=4).value = f"{overall_rating:.2f}"
            ws.cell(row=row, column=5).value = f"{learning_delivery:.2f}"
            ws.cell(row=row, column=6).value = f"{assessment_learning:.2f}"
            ws.cell(row=row, column=7).value = f"{student_engagement:.2f}"
            ws.cell(row=row, column=8).value = faculty['total_evaluations']
            ws.cell(row=row, column=9).value = get_remarks(overall_rating)
            
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.border = thin_border
                if col in [2, 3]:  # Left align text columns
                    cell.alignment = left_alignment
                else:  # Center align numbers and remarks
                    cell.alignment = center_alignment
            
            row += 1
        
        row += 1  # Empty row
        
        # Rating scale (same format as faculty results)
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = 'RATING SCALE:'
        cell.font = subheader_font
        cell.alignment = left_alignment
        row += 1
        
        # Rating table (same format as faculty results)
        total_faculty = len(rankings)
        top_performer = f"{rankings[0]['first_name']} {rankings[0]['last_name']}" if rankings else 'N/A'
        top_rating = f"{rankings[0]['average_rating']:.2f}" if rankings else 'N/A'
        
        rating_headers = [['Rating', 'Equivalent', 'Total Faculty:', str(total_faculty)]]
        rating_data = [
            ['4.50 - 5.00', 'OUTSTANDING', 'Top Performer:', top_performer],
            ['3.50 - 4.49', 'HIGHLY SATISFACTORY', 'Top Rating:', top_rating],
            ['2.50 - 3.49', 'SATISFACTORY', '', ''],
            ['1.50 - 2.49', 'NEEDS IMPROVEMENT', '', ''],
            ['1.00 - 1.49', 'POOR', '', '']
        ]
        
        rating_start_row = row
        for data_row in rating_headers + rating_data:
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = bold_font if row == rating_start_row else normal_font
                cell.fill = gray_fill if row == rating_start_row else PatternFill()
                cell.alignment = center_alignment
                cell.border = thin_border
            row += 1
        
        row += 2  # Empty rows before signatures
        
        # Signature section (same format as faculty results)
        # Signature of Guidance Counselor (changed from Faculty)
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = '________________________________'
        cell.alignment = left_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'SIGNATURE OF GUIDANCE COUNSELOR'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 2
        
        # Signature of College Dean (same as faculty results)
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = '________________________________'
        cell.alignment = left_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'SIGNATURE OF COLLEGE DEAN'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 2
        
        # Noted by section (same as faculty results)
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'NOTED BY:'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'MA. LIBERTY DG. PASCUAL, Ph.D.'
        cell.font = bold_font
        cell.alignment = left_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = 'College President'
        cell.font = normal_font
        cell.alignment = left_alignment
        row += 1
        
        # Auto-adjust column widths (same as faculty results)
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 20
        
        # Save file
        if department:
            filename = f"Faculty_Rankings_{department['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"Faculty_Rankings_All_Departments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
        filepath = os.path.join('static', 'reports', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename, 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        print(f"Error exporting rankings Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ================================
# AI ANALYTICS ENDPOINTS
# ================================

@api_bp.route('/ai-analytics/performance-trends')
@login_required
def ai_analytics_performance_trends():
    """Get faculty performance trends with AI analysis - Shows trends since selected academic year"""
    try:
        faculty_id = request.args.get('faculty_id')
        year_id = request.args.get('year_id')
        period_id = request.args.get('period_id', 'all')
        advanced_mode = request.args.get('advanced_mode', 'false').lower() == 'true'
        
        # Validate required parameters
        if not faculty_id or not year_id:
            return jsonify({
                'success': True,
                'chart_data': {
                    'labels': [],
                    'datasets': []
                },
                'metrics': {
                    'overall_change': 'N/A',
                    'growth_rate': 'N/A',
                    'best_performer': 'N/A'
                },
                'ai_insight': 'Please select both a faculty member and an academic year to view performance trends.'
            })
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Get the selected academic year's start date
        cursor.execute("""
            SELECT start_date FROM academic_years WHERE acad_year_id = %s
        """, (year_id,))
        year_info = cursor.fetchone()
        
        if not year_info:
            return jsonify({'success': False, 'message': 'Academic year not found'}), 404
        
        # Get faculty performance trends since the selected academic year
        query = """
            SELECT 
                CONCAT(ay.year_code, ' - ', at.term_name) as period_label,
                ay.year_name,
                at.term_name,
                AVG(er.rating) as avg_rating,
                COUNT(er.response_id) as response_count,
                ep.start_date
            FROM evaluation_responses er
            JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            JOIN faculty f ON cs.faculty_id = f.faculty_id
            WHERE cs.faculty_id = %s 
              AND er.rating IS NOT NULL 
              AND ay.start_date >= %s
            GROUP BY ay.acad_year_id, ay.year_code, ay.year_name, at.acad_term_id, at.term_name, ep.period_id, ep.start_date
            ORDER BY ay.start_date, at.acad_term_id, ep.start_date
        """
        
        cursor.execute(query, (faculty_id, year_info['start_date']))
        trends_data = cursor.fetchall()
        
        # Get faculty name
        cursor.execute("""
            SELECT CONCAT(first_name, ' ', last_name) as full_name 
            FROM faculty WHERE faculty_id = %s
        """, (faculty_id,))
        faculty_info = cursor.fetchone()
        
        # Check if there's no data
        if not trends_data or len(trends_data) == 0:
            return jsonify({
                'success': True,
                'chart_data': {
                    'labels': [],
                    'datasets': [{
                        'label': faculty_info['full_name'] if faculty_info else 'Faculty Performance',
                        'data': [],
                        'borderColor': '#0059cc',
                        'backgroundColor': 'rgba(0, 89, 204, 0.1)',
                        'borderWidth': 3,
                        'fill': True,
                        'tension': 0.4
                    }]
                },
                'metrics': {
                    'overall_change': 'N/A',
                    'growth_rate': 'N/A',
                    'best_performer': 'N/A'
                },
                'ai_insight': f'No evaluation data available for {faculty_info["full_name"] if faculty_info else "this faculty member"} since the selected academic year. Please ensure evaluations have been completed.'
            })
        
        # Prepare chart data
        labels = [trend['period_label'] for trend in trends_data]
        ratings = [float(trend['avg_rating']) if trend['avg_rating'] else 0 for trend in trends_data]
        
        chart_data = {
            'labels': labels,
            'datasets': [{
                'label': faculty_info['full_name'] if faculty_info else 'Faculty Performance',
                'data': ratings,
                'borderColor': '#0059cc',
                'backgroundColor': 'rgba(0, 89, 204, 0.1)',
                'borderWidth': 3,
                'fill': True,
                'tension': 0.4,
                'pointRadius': 6,
                'pointHoverRadius': 8,
                'pointBackgroundColor': '#0059cc'
            }]
        }
        
        # Calculate metrics
        metrics = {}
        if len(ratings) >= 2:
            latest_rating = ratings[-1]
            first_rating = ratings[0]
            change = latest_rating - first_rating
            growth_rate = f"{(change / first_rating * 100):.1f}%" if first_rating > 0 else "N/A"
            
            metrics = {
                'overall_change': f"{'+' if change > 0 else ''}{change:.2f} (First → Latest)",
                'growth_rate': growth_rate,
                'best_performer': f"{max(ratings):.2f} ({trends_data[ratings.index(max(ratings))]['period_label']})"
            }
        elif len(ratings) == 1:
            metrics = {
                'overall_change': 'N/A (single period)',
                'growth_rate': 'N/A (need multiple periods)',
                'best_performer': f"{ratings[0]:.2f} ({trends_data[0]['period_label']})"
            }
        else:
            metrics = {
                'overall_change': 'N/A',
                'growth_rate': 'N/A',
                'best_performer': 'N/A'
            }
        
        # Generate AI insight
        from utils.ai_support import generate_performance_trend_insight
        ai_insight = generate_performance_trend_insight(trends_data, metrics, advanced_mode)
        
        return jsonify({
            'success': True,
            'chart_data': chart_data,
            'metrics': metrics,
            'ai_insight': ai_insight
        })
        
    except Exception as e:
        print(f"Error in AI analytics performance trends: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@api_bp.route('/ai-analytics/department-trends')
@login_required
def ai_analytics_department_trends():
    """Get department performance trends since selected academic year"""
    try:
        department_id = request.args.get('department_id')
        year_id = request.args.get('year_id')
        period_id = request.args.get('period_id', 'all')
        
        # Validate required parameters
        if not department_id or not year_id:
            return jsonify({
                'success': True,
                'chart_data': {
                    'labels': [],
                    'datasets': []
                },
                'stats': {
                    'top_department': {'name': 'N/A', 'rating': '--'},
                    'average_rating': '--',
                    'low_department': {'name': 'N/A', 'rating': '--'},
                    'most_improved': {'name': 'N/A', 'change': '--'}
                }
            })
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Get the selected academic year's start date
        cursor.execute("""
            SELECT start_date FROM academic_years WHERE acad_year_id = %s
        """, (year_id,))
        year_info = cursor.fetchone()
        
        if not year_info:
            return jsonify({'success': False, 'message': 'Academic year not found'}), 404
        
        # Get department performance trends since the selected academic year
        query = """
            SELECT 
                p.name as department_name,
                p.program_id,
                CONCAT(ay.year_code, ' - ', at.term_name) as period_label,
                AVG(er.rating) as avg_rating,
                COUNT(DISTINCT f.faculty_id) as faculty_count,
                ep.start_date
            FROM evaluation_responses er
            JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            JOIN faculty f ON cs.faculty_id = f.faculty_id
            JOIN programs p ON f.program_id = p.program_id
            JOIN evaluation_periods ep ON e.period_id = ep.period_id
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE er.rating IS NOT NULL 
              AND p.program_id = %s 
              AND ay.start_date >= %s
            GROUP BY p.program_id, p.name, ay.acad_year_id, ay.year_code, at.acad_term_id, at.term_name, ep.period_id, ep.start_date
            ORDER BY ay.start_date, at.acad_term_id, ep.start_date
        """
        
        cursor.execute(query, (department_id, year_info['start_date']))
        trends_data = cursor.fetchall()
        
        # Get department name
        cursor.execute("""
            SELECT name FROM programs WHERE program_id = %s
        """, (department_id,))
        dept_info = cursor.fetchone()
        
        # Check if there's no data
        if not trends_data or len(trends_data) == 0:
            return jsonify({
                'success': True,
                'chart_data': {
                    'labels': [],
                    'datasets': []
                },
                'stats': {
                    'top_department': {'name': 'N/A', 'rating': '--'},
                    'average_rating': '--',
                    'low_department': {'name': 'N/A', 'rating': '--'},
                    'most_improved': {'name': 'N/A', 'change': '--'}
                }
            })
        
        # Prepare chart data
        labels = [trend['period_label'] for trend in trends_data]
        ratings = [float(trend['avg_rating']) if trend['avg_rating'] else 0 for trend in trends_data]
        
        chart_data = {
            'labels': labels,
            'datasets': [{
                'label': dept_info['name'] if dept_info else 'Department Performance',
                'data': ratings,
                'borderColor': '#0059cc',
                'backgroundColor': 'rgba(0, 89, 204, 0.1)',
                'borderWidth': 2,
                'fill': False,
                'tension': 0.4,
                'pointRadius': 4,
                'pointHoverRadius': 6
            }]
        }
        
        # Calculate statistics
        if len(ratings) > 0:
            overall_avg = sum(ratings) / len(ratings)
            max_rating = max(ratings)
            min_rating = min(ratings)
            max_idx = ratings.index(max_rating)
            min_idx = ratings.index(min_rating)
            
            # Calculate improvement
            most_improved = {'name': 'N/A', 'change': '--'}
            if len(ratings) >= 2:
                change = ratings[-1] - ratings[0]
                most_improved = {
                    'name': dept_info['name'] if dept_info else 'Department',
                    'change': f"+{change:.2f}" if change > 0 else f"{change:.2f}"
                }
            
            stats = {
                'top_department': {
                    'name': trends_data[max_idx]['period_label'],
                    'rating': f"{max_rating:.2f}"
                },
                'average_rating': f"{overall_avg:.2f}",
                'low_department': {
                    'name': trends_data[min_idx]['period_label'],
                    'rating': f"{min_rating:.2f}"
                },
                'most_improved': most_improved
            }
        else:
            stats = {
                'top_department': {'name': 'N/A', 'rating': '--'},
                'average_rating': '--',
                'low_department': {'name': 'N/A', 'rating': '--'},
                'most_improved': {'name': 'N/A', 'change': '--'}
            }
        
        return jsonify({
            'success': True,
            'chart_data': chart_data,
            'stats': stats
        })
        
    except Exception as e:
        print(f"Error in AI analytics department trends: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@api_bp.route('/ai-analytics/comparison')
@login_required
def ai_analytics_comparison():
    """Get faculty comparison data with AI analysis"""
    try:
        department_id = request.args.get('department_id', 'all')
        period_id = request.args.get('period_id', 'all')
        advanced_mode = request.args.get('advanced_mode', 'false').lower() == 'true'
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Build period filter
        period_filter = ""
        period_params = []
        
        if period_id != 'all':
            period_filter = "AND e.period_id = %s"
            period_params.append(period_id)
        
        # Get faculty comparison data
        if department_id == 'all':
            query = f"""
                SELECT 
                    CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                    p.name as department,
                    AVG(er.rating) as avg_rating,
                    COUNT(er.response_id) as evaluation_count
                FROM evaluation_responses er
                JOIN evaluations e ON er.evaluation_id = e.evaluation_id
                JOIN class_sections cs ON e.section_id = cs.section_id
                JOIN faculty f ON cs.faculty_id = f.faculty_id
                JOIN programs p ON f.program_id = p.program_id
                WHERE er.rating IS NOT NULL {period_filter}
                GROUP BY f.faculty_id, f.first_name, f.last_name, p.name
                ORDER BY avg_rating DESC
            """
            cursor.execute(query, period_params)
        else:
            query = f"""
                SELECT 
                    CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                    p.name as department,
                    AVG(er.rating) as avg_rating,
                    COUNT(er.response_id) as evaluation_count
                FROM evaluation_responses er
                JOIN evaluations e ON er.evaluation_id = e.evaluation_id
                JOIN class_sections cs ON e.section_id = cs.section_id
                JOIN faculty f ON cs.faculty_id = f.faculty_id
                JOIN programs p ON f.program_id = p.program_id
                WHERE p.program_id = %s AND er.rating IS NOT NULL {period_filter}
                GROUP BY f.faculty_id, f.first_name, f.last_name, p.name
                ORDER BY avg_rating DESC
            """
            params = [department_id] + period_params
            cursor.execute(query, params)
        
        comparison_data = cursor.fetchall()
        
        # Check if there's no data
        if not comparison_data or len(comparison_data) == 0:
            return jsonify({
                'success': True,
                'chart_data': {
                    'labels': [],
                    'datasets': [{
                        'label': 'Average Rating',
                        'data': [],
                        'backgroundColor': [],
                        'borderColor': [],
                        'borderWidth': 2
                    }]
                },
                'rankings': {
                    'top_faculty': [],
                    'bottom_faculty': []
                },
                'ai_insight': 'No faculty evaluation data available for the selected period and department. Please select a different period or ensure evaluations have been completed.'
            })
        
        # Prepare chart data
        faculty_names = [faculty['faculty_name'] for faculty in comparison_data[:10]]  # Top 10
        ratings = [float(faculty['avg_rating']) for faculty in comparison_data[:10]]
        
        # Color code based on performance
        colors = []
        for rating in ratings:
            if rating >= 4.5:
                colors.append('rgba(34, 197, 94, 0.8)')  # Green for excellent
            elif rating >= 4.0:
                colors.append('rgba(59, 130, 246, 0.8)')  # Blue for good
            else:
                colors.append('rgba(239, 68, 68, 0.8)')   # Red for needs improvement
        
        chart_data = {
            'labels': faculty_names,
            'datasets': [{
                'label': 'Average Rating',
                'data': ratings,
                'backgroundColor': colors,
                'borderColor': [color.replace('0.8', '1.0') for color in colors],
                'borderWidth': 2
            }]
        }
        
        # Get top and bottom performers
        top_faculty = comparison_data[:3] if len(comparison_data) >= 3 else comparison_data
        bottom_faculty = comparison_data[-3:] if len(comparison_data) >= 3 else []
        
        rankings = {
            'top_faculty': [{'name': f['faculty_name'], 'rating': f"{f['avg_rating']:.2f}"} for f in top_faculty],
            'bottom_faculty': [{'name': f['faculty_name'], 'rating': f"{f['avg_rating']:.2f}"} for f in bottom_faculty]
        }
        
        # Generate AI insight
        from utils.ai_support import generate_comparison_insight
        ai_insight = generate_comparison_insight(comparison_data, top_faculty, bottom_faculty, advanced_mode)
        
        return jsonify({
            'success': True,
            'chart_data': chart_data,
            'rankings': rankings,
            'ai_insight': ai_insight
        })
        
    except Exception as e:
        print(f"Error in AI analytics comparison: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@api_bp.route('/ai-analytics/question-analysis')
@login_required
def ai_analytics_question_analysis():
    """Get question analysis with AI insights"""
    try:
        period_id = request.args.get('period_id', 'all')
        advanced_mode = request.args.get('advanced_mode', 'false').lower() == 'true'
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Build period filter
        period_filter = ""
        period_params = []
        
        if period_id != 'all':
            period_filter = "AND e.period_id = %s"
            period_params.append(period_id)
        
        # Get question performance data from evaluation criteria
        query = f"""
            SELECT 
                ec.description as question_text,
                AVG(er.rating) as avg_score,
                COUNT(er.response_id) as response_count,
                ec.criteria_id as question_id
            FROM evaluation_responses er
            JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
            WHERE er.rating IS NOT NULL {period_filter}
            GROUP BY ec.criteria_id, ec.description
            ORDER BY avg_score DESC
        """
        cursor.execute(query, period_params)
        
        question_data = cursor.fetchall()
        
        # Prepare chart data
        questions = [q['question_text'][:50] + '...' if len(q['question_text']) > 50 else q['question_text'] for q in question_data]
        scores = [float(q['avg_score']) for q in question_data]
        
        # Color code based on score
        colors = []
        for score in scores:
            if score >= 4.51:
                colors.append('rgba(34, 197, 94, 0.8)')  # Green - Excellent
            elif score >= 3.51:
                colors.append('rgba(59, 130, 246, 0.8)')  # Blue - Good  
            else:
                colors.append('rgba(239, 68, 68, 0.8)')   # Red - Needs Improvement
        
        chart_data = {
            'labels': questions,
            'datasets': [{
                'label': 'Average Score',
                'data': scores,
                'backgroundColor': colors,
                'borderColor': [color.replace('0.8', '1.0') for color in colors],
                'borderWidth': 2
            }]
        }
        
        # Prepare detailed question data
        detailed_questions = []
        for q in question_data:
            detailed_questions.append({
                'question_text': q['question_text'],
                'mean': float(q['avg_score']),
                'response_count': q['response_count']
            })
        
        # Generate AI insight
        from utils.ai_support import generate_question_analysis_insight
        ai_insight = generate_question_analysis_insight(question_data, advanced_mode)
        
        return jsonify({
            'success': True,
            'chart_data': chart_data,
            'question_data': detailed_questions,
            'ai_insight': ai_insight
        })
        
    except Exception as e:
        print(f"Error in AI analytics question analysis: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@api_bp.route('/ai-analytics/engagement')
@login_required
def ai_analytics_engagement():
    """Get student engagement analytics with AI insights"""
    try:
        period_id = request.args.get('period_id', 'all')
        advanced_mode = request.args.get('advanced_mode', 'false').lower() == 'true'
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Build period filter
        period_filter = ""
        period_params = []
        
        if period_id != 'all':
            period_filter = "AND e.period_id = %s"
            period_params.append(period_id)
        
        # Get engagement data by subject/class using correct table relationships
        query = f"""
            SELECT 
                s.title as subject_name,
                f.first_name, f.last_name,
                COUNT(DISTINCT e.student_id) as respondents,
                25 as total_enrolled,  -- Using estimated enrollment for demo
                ROUND((COUNT(DISTINCT e.student_id) / 25) * 100, 1) as engagement_rate
            FROM evaluations e
            JOIN class_sections cs ON e.section_id = cs.section_id
            JOIN subjects s ON cs.subject_id = s.subject_id
            JOIN faculty f ON cs.faculty_id = f.faculty_id
            WHERE e.status = 'Completed' {period_filter}
            GROUP BY s.subject_id, s.title, f.faculty_id, f.first_name, f.last_name
            ORDER BY engagement_rate DESC
        """
        cursor.execute(query, period_params)
        
        engagement_data = cursor.fetchall()
        
        # Calculate overall stats
        total_classes = len(engagement_data)
        high_engagement = len([e for e in engagement_data if e['engagement_rate'] >= 85])
        moderate_engagement = len([e for e in engagement_data if 70 <= e['engagement_rate'] < 85])
        low_engagement = len([e for e in engagement_data if e['engagement_rate'] < 70])
        
        overall_engagement = sum(e['engagement_rate'] for e in engagement_data) / total_classes if total_classes > 0 else 0
        
        # Prepare chart data
        chart_data = {
            'labels': ['High (≥85%)', 'Moderate (70-84%)', 'Low (<70%)'],
            'datasets': [{
                'label': 'Number of Classes',
                'data': [high_engagement, moderate_engagement, low_engagement],
                'backgroundColor': [
                    'rgba(34, 197, 94, 0.8)',   # Green
                    'rgba(251, 191, 36, 0.8)',  # Yellow
                    'rgba(239, 68, 68, 0.8)'    # Red
                ],
                'borderColor': [
                    'rgb(34, 197, 94)',
                    'rgb(251, 191, 36)',
                    'rgb(239, 68, 68)'
                ],
                'borderWidth': 2
            }]
        }
        
        # Prepare stats
        stats = {
            'overall_engagement': f"{overall_engagement:.1f}%",
            'high_engagement': f"{high_engagement} classes",
            'low_engagement': f"{low_engagement} classes"
        }
        
        # Get department engagement data
        dept_query = f"""
            SELECT 
                p.name as department_name,
                COUNT(DISTINCT e.student_id) as total_respondents,
                COUNT(DISTINCT cs.section_id) * 25 as total_enrolled,  -- Estimated
                ROUND((COUNT(DISTINCT e.student_id) / (COUNT(DISTINCT cs.section_id) * 25)) * 100, 1) as engagement_rate
            FROM evaluations e
            JOIN class_sections cs ON e.section_id = cs.section_id
            JOIN faculty f ON cs.faculty_id = f.faculty_id
            JOIN programs p ON f.program_id = p.program_id
            WHERE e.status = 'Completed' {period_filter}
            GROUP BY p.program_id, p.name
            HAVING COUNT(DISTINCT cs.section_id) > 0
            ORDER BY engagement_rate DESC
        """
        cursor.execute(dept_query, period_params)
        
        department_data = cursor.fetchall()
        
        # Prepare department chart data
        dept_chart_data = {
            'labels': [dept['department_name'] for dept in department_data],
            'datasets': [{
                'label': 'Engagement Rate (%)',
                'data': [dept['engagement_rate'] for dept in department_data],
                'backgroundColor': [
                    'rgba(59, 130, 246, 0.8)' if dept['engagement_rate'] >= 80 else
                    'rgba(251, 191, 36, 0.8)' if dept['engagement_rate'] >= 70 else
                    'rgba(239, 68, 68, 0.8)'
                    for dept in department_data
                ],
                'borderColor': [
                    'rgb(59, 130, 246)' if dept['engagement_rate'] >= 80 else
                    'rgb(251, 191, 36)' if dept['engagement_rate'] >= 70 else
                    'rgb(239, 68, 68)'
                    for dept in department_data
                ],
                'borderWidth': 2
            }]
        }
        
        # Calculate department stats
        if department_data:
            highest_dept = max(department_data, key=lambda x: x['engagement_rate'])
            lowest_dept = min(department_data, key=lambda x: x['engagement_rate'])
            avg_engagement = sum(dept['engagement_rate'] for dept in department_data) / len(department_data)
            variance = max(dept['engagement_rate'] for dept in department_data) - min(dept['engagement_rate'] for dept in department_data)
            
            dept_stats = {
                'highest_department': f"{highest_dept['engagement_rate']}%",
                'lowest_department': f"{lowest_dept['engagement_rate']}%",
                'average_engagement': f"{avg_engagement:.1f}%",
                'variance': f"±{variance:.1f}%"
            }
        else:
            dept_stats = {
                'highest_department': '--',
                'lowest_department': '--',
                'average_engagement': '--',
                'variance': '--'
            }
        
        # Prepare department rankings for display
        dept_rankings = [
            {
                'department_name': dept['department_name'],
                'engagement_rate': dept['engagement_rate']
            }
            for dept in department_data
        ]
        
        # Generate AI insight
        from utils.ai_support import generate_engagement_insight
        ai_insight = generate_engagement_insight(engagement_data, stats, advanced_mode)
        
        return jsonify({
            'success': True,
            'chart_data': chart_data,
            'stats': stats,
            'department_data': {
                'chart_data': dept_chart_data,
                'stats': dept_stats,
                'rankings': dept_rankings
            },
            'ai_insight': ai_insight
        })
        
    except Exception as e:
        print(f"Error in AI analytics engagement: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@api_bp.route('/ai-analytics/improvement-opportunities')
@login_required
def ai_analytics_improvement_opportunities():
    """Get improvement opportunities with AI recommendations"""
    try:
        faculty_id = request.args.get('faculty_id')
        period_id = request.args.get('period_id', 'all')
        advanced_mode = request.args.get('advanced_mode', 'false').lower() == 'true'
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Build filters
        filters = ["er.rating IS NOT NULL"]
        query_params = []
        
        # Faculty filter (REQUIRED)
        if faculty_id:
            filters.append("f.faculty_id = %s")
            query_params.append(faculty_id)
        else:
            # If no faculty selected, return empty results
            return jsonify({
                'success': True,
                'improvement_data': [],
                'ai_insight': 'Please select a faculty member to view improvement opportunities.'
            })
        
        # Period filter
        if period_id != 'all':
            filters.append("e.period_id = %s")
            query_params.append(period_id)
        
        where_clause = " AND ".join(filters)
        
        # Get lowest scoring areas for selected faculty
        query = f"""
            SELECT 
                CONCAT(f.first_name, ' ', f.last_name) as faculty_name,
                p.name as department,
                ec.description as criteria,
                AVG(er.rating) as avg_score,
                COUNT(er.response_id) as response_count
            FROM evaluation_responses er
            JOIN evaluation_criteria ec ON er.criteria_id = ec.criteria_id
            JOIN evaluations e ON er.evaluation_id = e.evaluation_id
            JOIN class_sections cs ON e.section_id = cs.section_id
            JOIN faculty f ON cs.faculty_id = f.faculty_id
            JOIN programs p ON f.program_id = p.program_id
            WHERE {where_clause}
            GROUP BY f.faculty_id, f.first_name, f.last_name, p.name, ec.criteria_id, ec.description
            HAVING avg_score < 4.0
            ORDER BY avg_score ASC
            LIMIT 10
        """
        cursor.execute(query, query_params)
        
        improvement_data = cursor.fetchall()
        
        # Check if faculty has any evaluation data
        if not improvement_data:
            # Check if faculty exists and has any evaluations
            check_query = """
                SELECT COUNT(*) as eval_count
                FROM evaluations e
                JOIN class_sections cs ON e.section_id = cs.section_id
                JOIN faculty f ON cs.faculty_id = f.faculty_id
                WHERE f.faculty_id = %s AND e.status = 'Completed'
            """
            cursor.execute(check_query, [faculty_id])
            result = cursor.fetchone()
            
            if result['eval_count'] == 0:
                # Faculty has no completed evaluations
                return jsonify({
                    'success': True,
                    'improvement_data': [],
                    'ai_insight': 'This faculty member has not been evaluated yet. No improvement data available.',
                    'no_data': True
                })
            else:
                # Faculty has evaluations but all scores are above 4.0 (excellent performance)
                return jsonify({
                    'success': True,
                    'improvement_data': [],
                    'ai_insight': '🎉 Excellent performance! This faculty member has no areas scoring below 4.0. All evaluation criteria show strong performance.',
                    'no_data': False
                })
        
        # Format improvement data
        formatted_data = []
        for item in improvement_data:
            formatted_data.append({
                'faculty_name': item['faculty_name'],
                'department': item['department'],
                'criteria': item['criteria'][:60] + '...' if len(item['criteria']) > 60 else item['criteria'],
                'score': float(item['avg_score']),
                'response_count': item['response_count']
            })
        
        # Generate AI insight and recommendations
        from utils.ai_support import generate_improvement_opportunities_insight
        ai_insight = generate_improvement_opportunities_insight(improvement_data, advanced_mode)
        
        return jsonify({
            'success': True,
            'improvement_data': formatted_data,
            'ai_insight': ai_insight
        })
        
    except Exception as e:
        print(f"Error in AI analytics improvement opportunities: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@api_bp.route('/ai-analytics/generate-training-plan', methods=['POST'])
@login_required
def generate_training_plan():
    """Generate AI-powered training plan"""
    try:
        from utils.ai_support import generate_comprehensive_training_plan
        training_plan = generate_comprehensive_training_plan()
        
        return jsonify({
            'success': True,
            'training_plan': training_plan
        })
        
    except Exception as e:
        print(f"Error generating training plan: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@api_bp.route('/ai-analytics/send-to-dean', methods=['POST'])
@login_required
def send_report_to_dean():
    """Send analytics report to dean"""
    try:
        # This would integrate with email system
        # For now, return success message
        return jsonify({
            'success': True,
            'message': 'Analytics report sent to Dean successfully'
        })
        
    except Exception as e:
        print(f"Error sending report to dean: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


# API endpoints for filter dropdowns
@api_bp.route('/faculty/list')
@login_required
def get_faculty_list():
    """Get list of all active faculty for filter dropdowns"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                faculty_id,
                CONCAT(first_name, ' ', last_name) as full_name,
                first_name,
                last_name
            FROM faculty
            WHERE is_archived = 0
            ORDER BY last_name, first_name
        """)
        
        faculty = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'faculty': faculty
        })
        
    except Exception as e:
        print(f"Error loading faculty list: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@api_bp.route('/programs/list')
@login_required
def get_programs_list():
    """Get list of all programs/departments for filter dropdowns"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                program_id,
                program_code,
                name
            FROM programs
            ORDER BY name
        """)
        
        programs = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'programs': programs
        })
        
    except Exception as e:
        print(f"Error loading programs list: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@api_bp.route('/evaluation-periods/list')
@login_required
def get_evaluation_periods_list():
    """Get list of all evaluation periods for filter dropdowns"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                ep.period_id,
                CONCAT(ay.year_name, ' - ', at.term_name, ' - ', ep.title) as period_name,
                ep.status
            FROM evaluation_periods ep
            JOIN academic_terms at ON ep.acad_term_id = at.acad_term_id
            JOIN academic_years ay ON at.acad_year_id = ay.acad_year_id
            WHERE ep.is_archived = 0
            ORDER BY ep.start_date DESC
        """)
        
        periods = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'periods': periods
        })
        
    except Exception as e:
        print(f"Error loading evaluation periods: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@api_bp.route('/academic-years/list')
@login_required
def get_academic_years_list():
    """Get list of all academic years for filter dropdowns"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                acad_year_id,
                year_code,
                year_name,
                start_date,
                end_date
            FROM academic_years
            ORDER BY start_date DESC
        """)
        
        academic_years = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'academic_years': academic_years
        })
        
    except Exception as e:
        print(f"Error loading academic years: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@api_bp.route('/ai-analytics/export-report')
@login_required
def export_analytics_report():
    """Export comprehensive analytics report"""
    try:
        from flask import send_file
        import os
        from datetime import datetime
        
        # Generate report file (placeholder)
        filename = f"AI_Analytics_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join('static', 'reports', filename)
        
        # Create placeholder file
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'w') as f:
            f.write("AI Analytics Report - Placeholder")
        
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"Error exporting analytics report: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# EXPIRED EVALUATIONS MANAGEMENT
# ============================================

@api_bp.route('/expired-evaluations/mark', methods=['POST'])
@login_required
def mark_expired_evaluations_api():
    """Auto-mark evaluations as expired based on timer and period end"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    from utils.expired_evaluations import mark_expired_evaluations
    
    result = mark_expired_evaluations()
    return jsonify(result)


@api_bp.route('/expired-evaluations')
@login_required
def get_expired_evaluations_api():
    """Get list of expired evaluations"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    from utils.expired_evaluations import get_expired_evaluations
    
    period_id = request.args.get('period_id', type=int)
    student_id = request.args.get('student_id', type=int)
    
    result = get_expired_evaluations(period_id, student_id)
    return jsonify(result)


@api_bp.route('/expired-evaluations/summary')
@login_required
def get_expired_evaluations_summary_api():
    """Get summary statistics of expired evaluations"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    from utils.expired_evaluations import get_expired_evaluations_summary
    
    period_id = request.args.get('period_id', type=int)
    
    result = get_expired_evaluations_summary(period_id)
    return jsonify(result)


@api_bp.route('/expired-evaluations/<int:evaluation_id>/reset', methods=['POST'])
@login_required
def reset_expired_evaluation_api(evaluation_id):
    """Reset a single expired evaluation to allow retake"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    from utils.expired_evaluations import reset_expired_evaluation
    
    result = reset_expired_evaluation(evaluation_id)
    
    if result['success']:
        return jsonify(result), 200
    else:
        return jsonify(result), 400


@api_bp.route('/expired-evaluations/reset-multiple', methods=['POST'])
@login_required
def reset_multiple_expired_evaluations_api():
    """Reset multiple expired evaluations at once"""
    if session.get('role') not in ['admin', 'guidance']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    from utils.expired_evaluations import reset_multiple_expired_evaluations
    
    data = request.get_json()
    evaluation_ids = data.get('evaluation_ids', [])
    
    if not evaluation_ids:
        return jsonify({'success': False, 'error': 'No evaluation IDs provided'}), 400
    
    result = reset_multiple_expired_evaluations(evaluation_ids)
    return jsonify(result)
